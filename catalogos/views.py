import json
import logging
import zipfile
from datetime import datetime, time, timedelta
from functools import wraps
from io import BytesIO
from urllib.parse import quote

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Count, ExpressionWrapper, F, FloatField, Max, OuterRef, Q, Subquery, Sum
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce, Upper
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from produccion.models import ESTADOS, Viga

from core.excepciones import ErrorDeDominio
from core.estados import clase as clase_de_estado
from core.servicios import almacen as servicio_almacen
from core.servicios import cierres as servicio_cierres

logger = logging.getLogger("mes.catalogos")

STATUS_COLORS = {
    "Corte": "#f39c12",
    "Espera Armado": "#d35400",
    "Armado": "#3498db",
    "Espera Soldadura": "#5d6d7e",
    "Soldadura": "#9b59b6",
    "Espera Pintura": "#7f8c8d",
    "Pintura": "#16a085",
    "Terminado": "#2dce89",
    "Enviado": "#11cdef",
}

DECOTE_DAYS = 5
CIERRE_PENDIENTE_LABEL = "Terminado (bloqueo pend.)"
CIERRE_FINAL_LABEL = "Terminado"
CIERRE_VENTANA_MINUTOS = 10

from .models import (
    Colaborador,
    CortaClienteProyecto,
    CortaPiezaCatalogo,
    EquipoTrabajo,
    HerrAsignacion,
    HerrAvanceCambio,
    HerrCliente,
    HerrEstadoCambio,
    HerrOrdenItem,
    HerrOrdenAsignacion,
    HerrOrdenProduccion,
    HerrPiezaCatalogo,
    HerrProduccion,
    HerrSolicitudProduccion,
    LogisticaAcuseEntrega,
    LogisticaEnvio,
    LogisticaEnvioCorta,
    LogisticaEnvioItem,
    LogisticaExpediente,
    LogisticaExpedienteDescarga,
    LogisticaMovimiento,
    LogisticaMovimientoCorta,
    LogisticaStock,
    LogisticaStockCorta,
    LaserAsignacion,
    LaserEstadoCambio,
    LaserMaterialPlaca,
    LaserOrdenAsignacion,
    LaserOrdenItem,
    LaserOrdenProduccion,
    LaserProduccion,
    MaquinaFalla,
    MaquinaFallaTipo,
    Maquina,
    MaquinaParo,
    MaquinaParoMotivo,
    PlantaEvento,
    Proyecto,
    PedidoProduccion,
    PedidoProduccionItem,
    RobotOrdenItem,
    RobotOrdenProduccion,
    RobotPiezaCatalogo,
    RobotProduccion,
    RobotOrdenAsignacion,
    VigaAsignacion,
)

from produccion.views import _build_participantes_payload


def _finalize_herreria_cierres_expirados():
    """Compatibilidad: la consolidación vive ahora en core.servicios.cierres.

    Se conserva la llamada desde la pantalla de control para que las órdenes
    vencidas aparezcan cerradas aunque el trabajo programado no esté todavía
    configurado en el servidor. Una vez en marcha el temporizador, estas dos
    llamadas se pueden quitar y la pantalla dejará de hacer trabajo de fondo
    antes de pintar nada.
    """
    return servicio_cierres.consolidar_linea("herreria")


def _finalize_corta_cierres_expirados():
    """Ver _finalize_herreria_cierres_expirados."""
    return servicio_cierres.consolidar_linea("corta")


ESTADOS_EQUIPO = [s for s in ESTADOS if (not s.startswith("Espera") and s != "Enviado")]


def _norm_text(value: str) -> str:
    return (value or "").strip()


def _norm_upper(value: str) -> str:
    return _norm_text(value).upper()


def _get_or_create_corta_cliente_proyecto(raw: str, *, tipo_default: str = "cliente") -> CortaClienteProyecto:
    nombre = _norm_text(raw)
    nombre_norm = _norm_upper(raw)
    if not nombre_norm:
        raise ValueError("Nombre requerido.")
    obj = CortaClienteProyecto.objects.filter(nombre_normalizado=nombre_norm).first()
    if obj:
        if obj.nombre != nombre:
            obj.nombre = nombre
            obj.save(update_fields=["nombre", "nombre_normalizado", "actualizado_en"])
        return obj
    return CortaClienteProyecto.objects.create(nombre=nombre, nombre_normalizado=nombre_norm, tipo=tipo_default, activo=True)


def _get_or_create_herr_cliente(raw: str) -> HerrCliente:
    nombre = _norm_text(raw)
    nombre_norm = _norm_upper(raw)
    if not nombre_norm:
        raise ValueError("Nombre requerido.")
    obj = HerrCliente.objects.filter(nombre_normalizado=nombre_norm).first()
    if obj:
        if obj.nombre != nombre:
            obj.nombre = nombre
            obj.save(update_fields=["nombre", "nombre_normalizado", "actualizado_en"])
        return obj
    return HerrCliente.objects.create(nombre=nombre, nombre_normalizado=nombre_norm, activo=True)


def _is_admin_user(user) -> bool:
    try:
        return bool(
            user
            and user.is_authenticated
            and (
                getattr(user, "is_superuser", False)
                or getattr(user, "is_staff", False)
                or user.groups.filter(name__in=["admin_general", "ingenieria_civil"]).exists()
            )
        )
    except Exception:
        return False


def admin_required(view_func):
    """Exige sesión iniciada y rol de administración.

    Las pantallas de configuración de planta (proyectos, equipos,
    colaboradores y maquinaria) ya se muestran solo a `is_admin` en la
    navbar, pero sus vistas nunca aplicaron esa restricción: cualquiera
    en la red del taller podía llamarlas, incluidos los borrados por POST.
    Este decorador aplica en el servidor la regla que la interfaz ya daba
    por supuesta.
    """

    @login_required
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not _is_admin_user(request.user):
            messages.error(request, "No tienes permiso para acceder a esa sección.")
            return redirect("/")
        return view_func(request, *args, **kwargs)

    return _wrapped


def _can_pedidos(user) -> bool:
    try:
        return bool(
            user
            and user.is_authenticated
            and (
                getattr(user, "is_superuser", False)
                or user.groups.filter(name__in=["admin_general", "ingenieria_civil", "pedidos_ventas"]).exists()
            )
        )
    except Exception:
        return False


def _can_pedidos_logistica(user) -> bool:
    try:
        return bool(
            user
            and user.is_authenticated
            and (
                _can_pedidos(user)
                or user.groups.filter(name__in=["corte_laser", "corte_laser_supervision"]).exists()
                or _is_admin_user(user)
            )
        )
    except Exception:
        return False


def _safe_next_url(raw: str) -> str:
    val = (raw or "").strip()
    return val if (val.startswith("/") and not val.startswith("//")) else ""


def _pedido_is_completo(items: list[PedidoProduccionItem]) -> bool:
    has_any = False
    for it in items:
        if (getattr(it, "estado_herreria", "") or "").strip() == "Cancelado":
            continue
        has_any = True
        total = int(getattr(it, "cantidad_total", 0) or 0)
        apartado = int(getattr(it, "apartado", 0) or 0)
        enviado = int(getattr(it, "enviado", 0) or 0)
        saldo = max(0, total - apartado - enviado)
        if saldo != 0:
            return False
    return has_any


def _ensure_logistica_expediente(pedido: PedidoProduccion, actor: str, now: datetime) -> LogisticaExpediente:
    exp, created = LogisticaExpediente.objects.get_or_create(pedido=pedido)
    if created or not getattr(exp, "generado_en", None):
        exp.generado_en = now
        exp.generado_por = (actor or "").strip()
        exp.save(update_fields=["generado_en", "generado_por", "actualizado_en"])
    return exp


def _corta_is_completo(orden: LaserOrdenProduccion) -> bool:
    total = int(getattr(orden, "total_piezas", 0) or 0)
    apartado = int(getattr(orden, "apartado", 0) or 0)
    enviado = int(getattr(orden, "enviado", 0) or 0)
    return bool(max(0, total - apartado - enviado) == 0 and total > 0)


def _ensure_logistica_expediente_corta(orden: LaserOrdenProduccion, actor: str, now: datetime) -> LogisticaExpediente:
    exp, created = LogisticaExpediente.objects.get_or_create(orden_corta=orden)
    if created or not getattr(exp, "generado_en", None):
        exp.generado_en = now
        exp.generado_por = (actor or "").strip()
        exp.save(update_fields=["generado_en", "generado_por", "actualizado_en"])
    return exp


def _xlsx_bytes_for_expediente(
    pedido: PedidoProduccion,
    items: list[PedidoProduccionItem],
    envios: list[LogisticaEnvio],
    envios_items: list[LogisticaEnvioItem],
    movimientos: list[LogisticaMovimiento],
    pdf_rows: list[dict],
) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Pedido"
    ws0.append(["Campo", "Valor"])
    ws0.append(["Folio", getattr(pedido, "folio", "")])
    ws0.append(["Cliente", getattr(pedido, "cliente", "")])
    ws0.append(["Teléfono", getattr(pedido, "telefono", "")])
    ws0.append(["Fecha compromiso", getattr(pedido, "fecha_compromiso", None)])
    ws0.append(["Estado", getattr(pedido, "estado", "")])
    ws0.append(["Comentarios", getattr(pedido, "comentarios", "")])
    ws0.append(["Creado", timezone.localtime(getattr(pedido, "creado_en")).strftime("%Y-%m-%d %H:%M") if getattr(pedido, "creado_en", None) else ""])

    ws1 = wb.create_sheet("Items")
    ws1.append(["ID", "Producto", "Total", "Apartado", "Enviado", "Saldo", "Estado herrería", "Aceptado por", "Aceptado en"])
    for it in items:
        total = int(getattr(it, "cantidad_total", 0) or 0)
        apartado = int(getattr(it, "apartado", 0) or 0)
        enviado = int(getattr(it, "enviado", 0) or 0)
        saldo = max(0, total - apartado - enviado)
        aceptado_en = getattr(it, "aceptado_en", None)
        ws1.append(
            [
                int(it.id),
                getattr(getattr(it, "producto", None), "nombre", "") or "",
                total,
                apartado,
                enviado,
                saldo,
                getattr(it, "estado_herreria", "") or "",
                getattr(it, "aceptado_por", "") or "",
                timezone.localtime(aceptado_en).strftime("%Y-%m-%d %H:%M") if aceptado_en else "",
            ]
        )

    ws2 = wb.create_sheet("Envios")
    ws2.append(["Envio ID", "Fecha", "Actor", "PDF", "Item ID", "Producto", "Cantidad", "Revertido"])
    for ei in envios_items:
        envio = getattr(ei, "envio", None)
        ws2.append(
            [
                int(getattr(envio, "id", 0) or 0),
                getattr(envio, "fecha", None),
                getattr(envio, "actor", "") or "",
                "",
                int(getattr(ei, "pedido_item_id", 0) or 0),
                getattr(getattr(ei, "producto", None), "nombre", "") or "",
                int(getattr(ei, "cantidad", 0) or 0),
                int(getattr(ei, "revertido", 0) or 0),
            ]
        )

    ws3 = wb.create_sheet("Movimientos")
    ws3.append(["ID", "Creado", "Tipo", "Producto", "Cantidad", "Item ID", "Actor"])
    for m in movimientos:
        created = getattr(m, "creado_en", None)
        ws3.append(
            [
                int(m.id),
                timezone.localtime(created).strftime("%Y-%m-%d %H:%M") if created else "",
                getattr(m, "tipo", "") or "",
                getattr(getattr(m, "producto", None), "nombre", "") or "",
                int(getattr(m, "cantidad", 0) or 0),
                int(getattr(m, "pedido_item_id", 0) or 0) if getattr(m, "pedido_item_id", None) else "",
                getattr(m, "actor", "") or "",
            ]
        )

    ws4 = wb.create_sheet("PDFs")
    ws4.append(["Envio ID", "Fecha", "Nombre en ZIP", "Estado"])
    for row in pdf_rows:
        ws4.append([row.get("envio_id"), row.get("fecha"), row.get("zip_name"), row.get("status")])

    for ws in [ws0, ws1, ws2, ws3, ws4]:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            letter = get_column_letter(col)
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[letter].width = min(70, max(12, max_len + 2))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_pedido_expediente_zip_bytes(pedido_id: int) -> bytes:
    pedido = get_object_or_404(PedidoProduccion.objects.using("mes"), pk=pedido_id)
    items = list(PedidoProduccionItem.objects.using("mes").select_related("producto").filter(pedido_id=int(pedido.id)).order_by("id"))
    envios = list(LogisticaEnvio.objects.using("mes").filter(pedido_id=int(pedido.id)).order_by("fecha", "id"))
    envios_items = list(
        LogisticaEnvioItem.objects.using("mes")
        .select_related("envio", "producto", "pedido_item")
        .filter(envio__pedido_id=int(pedido.id))
        .order_by("envio__fecha", "id")
    )
    movimientos = list(
        LogisticaMovimiento.objects.using("mes")
        .select_related("producto")
        .filter(pedido_item__pedido_id=int(pedido.id))
        .order_by("creado_en", "id")
    )

    root = f"Expediente_{(getattr(pedido, 'folio', '') or str(pedido.id)).strip()}"
    pdf_rows: list[dict] = []
    pdf_files: list[tuple[str, bytes]] = []
    used_pdf_names: dict[str, int] = {}
    for e in envios:
        base = ""
        try:
            base = f"Envio_{e.fecha.strftime('%Y-%m-%d')}.pdf" if getattr(e, "fecha", None) else "Envio.pdf"
        except Exception:
            base = "Envio.pdf"
        n = used_pdf_names.get(base, 0) + 1
        used_pdf_names[base] = n
        zip_name = base if n == 1 else base.replace(".pdf", f"_{n}.pdf")
        if getattr(e, "comprobante_pdf", None):
            try:
                with e.comprobante_pdf.open("rb") as fh:
                    pdf_files.append((f"{root}/PDFs/{zip_name}", fh.read()))
                pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "OK"})
            except Exception:
                pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "NO_SE_PUDO_LEER"})
        else:
            pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "SIN_PDF"})

    xlsx_bytes = _xlsx_bytes_for_expediente(pedido, items, envios, envios_items, movimientos, pdf_rows)
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{root}/", "")
        zf.writestr(f"{root}/{root}.xlsx", xlsx_bytes)
        for path, b in pdf_files:
            zf.writestr(path, b)
    return zip_buf.getvalue()


def _xlsx_bytes_for_expediente_corta(
    orden: LaserOrdenProduccion,
    envios: list[LogisticaEnvioCorta],
    movimientos: list[LogisticaMovimientoCorta],
    pdf_rows: list[dict],
) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Orden"
    ws0.append(["Campo", "Valor"])
    ws0.append(["Folio", (getattr(orden, "folio_externo", "") or "").strip() or (getattr(orden, "codigo", "") or "").strip() or getattr(orden, "folio", "")])
    ws0.append(["Cliente/Proyecto", getattr(getattr(orden, "corta_cliente_proyecto", None), "nombre", "") or ""])
    ws0.append(["Producto", _corta_producto_from_orden(orden)])
    ws0.append(["Total", int(getattr(orden, "total_piezas", 0) or 0)])
    ws0.append(["Apartado", int(getattr(orden, "apartado", 0) or 0)])
    ws0.append(["Enviado", int(getattr(orden, "enviado", 0) or 0)])
    ws0.append(["Saldo", max(0, int(getattr(orden, "total_piezas", 0) or 0) - int(getattr(orden, "apartado", 0) or 0) - int(getattr(orden, "enviado", 0) or 0))])
    ws0.append(["Estado", getattr(orden, "estado", "") or ""])
    ws0.append(["Etapa", getattr(orden, "estado_etapa", "") or ""])
    ws0.append(["Observaciones", getattr(orden, "observaciones", "") or ""])
    ws0.append(["Creado", timezone.localtime(getattr(orden, "creado_en")).strftime("%Y-%m-%d %H:%M") if getattr(orden, "creado_en", None) else ""])

    ws1 = wb.create_sheet("Envios")
    ws1.append(["Envio ID", "Fecha", "Actor", "Producto", "Cantidad", "Revertido", "PDF"])
    for e in envios:
        ws1.append(
            [
                int(e.id),
                getattr(e, "fecha", None),
                getattr(e, "actor", "") or "",
                getattr(e, "producto", "") or "",
                int(getattr(e, "cantidad", 0) or 0),
                int(getattr(e, "revertido", 0) or 0),
                "",
            ]
        )

    ws2 = wb.create_sheet("Movimientos")
    ws2.append(["ID", "Creado", "Tipo", "Producto", "Cantidad", "Actor"])
    for m in movimientos:
        created = getattr(m, "creado_en", None)
        ws2.append(
            [
                int(m.id),
                timezone.localtime(created).strftime("%Y-%m-%d %H:%M") if created else "",
                getattr(m, "tipo", "") or "",
                getattr(m, "producto", "") or "",
                int(getattr(m, "cantidad", 0) or 0),
                getattr(m, "actor", "") or "",
            ]
        )

    ws3 = wb.create_sheet("PDFs")
    ws3.append(["Envio ID", "Fecha", "Nombre en ZIP", "Estado"])
    for row in pdf_rows:
        ws3.append([row.get("envio_id"), row.get("fecha"), row.get("zip_name"), row.get("status")])

    for ws in [ws0, ws1, ws2, ws3]:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            letter = get_column_letter(col)
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[letter].width = min(70, max(12, max_len + 2))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_corta_expediente_zip_bytes(orden_id: int) -> bytes:
    orden = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_related("corta_cliente_proyecto"), pk=int(orden_id))
    envios = list(LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(orden.id)).order_by("fecha", "id"))
    movimientos = list(LogisticaMovimientoCorta.objects.using("mes").filter(orden_id=int(orden.id)).order_by("creado_en", "id"))

    root = f"Expediente_{((getattr(orden, 'folio_externo', '') or getattr(orden, 'codigo', '') or getattr(orden, 'folio', '') or str(orden.id)).strip())}"
    pdf_rows: list[dict] = []
    pdf_files: list[tuple[str, bytes]] = []
    used_pdf_names: dict[str, int] = {}
    for e in envios:
        base = ""
        try:
            base = f"Envio_{e.fecha.strftime('%Y-%m-%d')}.pdf" if getattr(e, "fecha", None) else "Envio.pdf"
        except Exception:
            base = "Envio.pdf"
        n = used_pdf_names.get(base, 0) + 1
        used_pdf_names[base] = n
        zip_name = base if n == 1 else base.replace(".pdf", f"_{n}.pdf")
        if getattr(e, "comprobante_pdf", None):
            try:
                with e.comprobante_pdf.open("rb") as fh:
                    pdf_files.append((f"{root}/PDFs/{zip_name}", fh.read()))
                pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "OK"})
            except Exception:
                pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "NO_SE_PUDO_LEER"})
        else:
            pdf_rows.append({"envio_id": int(e.id), "fecha": e.fecha, "zip_name": zip_name, "status": "SIN_PDF"})

    xlsx_bytes = _xlsx_bytes_for_expediente_corta(orden, envios, movimientos, pdf_rows)
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{root}/", "")
        zf.writestr(f"{root}/{root}.xlsx", xlsx_bytes)
        for path, b in pdf_files:
            zf.writestr(path, b)
    return zip_buf.getvalue()


def _is_robotica_user(user) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name="robotica").exists())
    except Exception:
        return False


def _is_herreria_user(user) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name="herreria").exists())
    except Exception:
        return False


def _is_herreria_supervision(user) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name="herreria_supervision").exists())
    except Exception:
        return False


def _can_herreria(user) -> bool:
    return bool(_is_admin_user(user) or _is_herreria_user(user) or _is_herreria_supervision(user))


def _can_herreria_catalogo(user) -> bool:
    return bool(_can_herreria(user) or _can_pedidos(user))


def _is_corte_laser_user(user) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name="corte_laser").exists())
    except Exception:
        return False


def _is_corte_laser_supervision(user) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name="corte_laser_supervision").exists())
    except Exception:
        return False


def _can_corte_laser(user) -> bool:
    return bool(_is_admin_user(user) or _is_corte_laser_user(user) or _is_corte_laser_supervision(user))


def _sync_equipo_integrantes(equipo_id: int):
    if not equipo_id:
        return
    cnt = int(Colaborador.objects.filter(equipo_id=equipo_id, activo=True).count())
    EquipoTrabajo.objects.filter(id=equipo_id).update(integrantes=cnt)


class ProyectoForm(forms.ModelForm):
    class Meta:
        model = Proyecto
        fields = ["nombre", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")


class EquipoTrabajoForm(forms.ModelForm):
    estados = forms.MultipleChoiceField(
        choices=[(s, s) for s in ESTADOS_EQUIPO],
        required=True,
        widget=forms.CheckboxSelectMultiple,
    )

    class Meta:
        model = EquipoTrabajo
        fields = ["nombre", "integrantes", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["integrantes"].widget.attrs["readonly"] = "readonly"
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            elif name == "estados":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")

        if self.instance and getattr(self.instance, "pk", None):
            self.initial["estados"] = self.instance.estados
            self.initial["integrantes"] = int(Colaborador.objects.filter(equipo=self.instance, activo=True).count())

    def clean_integrantes(self):
        instance = getattr(self, "instance", None)
        if instance and getattr(instance, "pk", None):
            return int(Colaborador.objects.filter(equipo=instance, activo=True).count())
        return 0

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.estados_json = json.dumps(self.cleaned_data.get("estados") or [])
        instance.integrantes = int(self.cleaned_data.get("integrantes") or 0)
        if commit:
            instance.save()
        return instance


class ColaboradorForm(forms.ModelForm):
    class Meta:
        model = Colaborador
        fields = ["equipo", "rol", "nombre", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        active_qs = EquipoTrabajo.objects.filter(activo=True)
        if self.instance and getattr(self.instance, "pk", None) and getattr(self.instance, "equipo_id", None):
            active_qs = EquipoTrabajo.objects.filter(pk=self.instance.equipo_id) | active_qs
        self.fields["equipo"].queryset = active_qs.order_by("area", "sub_area", "nombre")
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def clean(self):
        cleaned = super().clean()
        if self.instance and getattr(self.instance, "pk", None):
            if VigaAsignacion.objects.filter(colaborador=self.instance).exists():
                if cleaned.get("equipo") and cleaned.get("equipo").id != self.instance.equipo_id:
                    self.add_error("equipo", "No se puede cambiar el equipo: el colaborador ya tiene asignaciones. Usa dar de baja y crea otro si es necesario.")
                if cleaned.get("rol") and cleaned.get("rol") != self.instance.rol:
                    self.add_error("rol", "No se puede cambiar el rol: el colaborador ya tiene asignaciones.")
        return cleaned


@admin_required
@require_http_methods(["GET", "POST"])
def proyectos(request):
    if request.method == "POST":
        if request.POST.get("action") == "delete":
            pk = int(request.POST.get("pk") or 0)
            if pk:
                try:
                    Proyecto.objects.filter(pk=pk).delete()
                    messages.success(request, "Proyecto eliminado.")
                except ProtectedError as e:
                    protected = list(getattr(e, "protected_objects", []) or [])
                    protected_str = ", ".join(str(o) for o in sorted(protected, key=lambda x: str(x))[:5])
                    extra = f" (+{len(protected) - 5} más)" if len(protected) > 5 else ""
                    detail = f" ({protected_str}{extra})" if protected_str else ""
                    messages.error(
                        request,
                        "No se puede eliminar: el proyecto está en uso"
                        f"{detail}. Mejor desactiva o cambia el proyecto en esas órdenes.",
                    )
            return redirect("catalogos:proyectos")

        if request.POST.get("action") == "toggle":
            pk = request.POST.get("pk")
            activo = request.POST.get("activo") == "1"
            Proyecto.objects.filter(pk=pk).update(activo=activo)
            if activo:
                messages.success(request, "Proyecto activado. Se mostrará de nuevo en Producción.")
            else:
                messages.success(request, "Proyecto desactivado. Se ocultará en Producción, pero no se elimina.")
            return redirect("catalogos:proyectos")

        form = ProyectoForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Proyecto guardado.")
            except Exception:
                messages.error(request, "No se pudo guardar. Revisa si ya existe con otro formato.")
            return redirect("catalogos:proyectos")
    else:
        form = ProyectoForm(initial={"activo": True})

    proyectos_qs = Proyecto.objects.all().order_by("nombre")
    return render(
        request,
        "catalogos/proyectos.html",
        {"form": form, "proyectos": proyectos_qs},
    )


@admin_required
@require_http_methods(["GET", "POST"])
def equipos(request):
    if request.method == "POST":
        if request.POST.get("form_type") == "colaborador_create":
            colab_form = ColaboradorForm(request.POST)
            form = EquipoTrabajoForm(initial={"activo": True, "integrantes": 0})
            if colab_form.is_valid():
                colab = colab_form.save()
                _sync_equipo_integrantes(colab.equipo_id)
                messages.success(request, "Colaborador guardado.")
                return redirect("catalogos:equipos")
        else:
            form = EquipoTrabajoForm(request.POST)
            colab_form = ColaboradorForm(initial={"activo": True})
            if form.is_valid():
                form.save()
                messages.success(request, "Equipo guardado.")
                return redirect("catalogos:equipos")
    else:
        form = EquipoTrabajoForm(initial={"activo": True, "integrantes": 0})
        colab_form = ColaboradorForm(initial={"activo": True})

    equipos_qs = EquipoTrabajo.objects.all().order_by("area", "sub_area", "nombre")
    colaboradores = (
        Colaborador.objects.select_related("equipo")
        .order_by("equipo__nombre", "rol", "nombre")
    )
    colabs_by_equipo = {}
    for c in colaboradores:
        colabs_by_equipo.setdefault(c.equipo_id, []).append(c)
    active_counts = dict(
        Colaborador.objects.filter(activo=True)
        .values_list("equipo_id")
        .annotate(cnt=Count("id"))
        .values_list("equipo_id", "cnt")
    )
    equipos_list = list(equipos_qs)
    for e in equipos_list:
        e.colaboradores_list = colabs_by_equipo.get(e.id, [])
        e.colaboradores_activos = int(active_counts.get(e.id, 0))
    return render(
        request,
        "catalogos/equipos.html",
        {
            "form": form,
            "equipos": equipos_list,
            "colab_form": colab_form,
        },
    )


@admin_required
@require_http_methods(["POST"])
def equipo_toggle(request, pk: int):
    equipo = get_object_or_404(EquipoTrabajo, pk=pk)
    equipo.activo = not bool(equipo.activo)
    equipo.save(update_fields=["activo", "actualizado_en"])
    messages.success(request, "Equipo actualizado.")
    return redirect("catalogos:equipos")


@admin_required
@require_http_methods(["POST"])
def equipo_delete(request, pk: int):
    equipo = get_object_or_404(EquipoTrabajo, pk=pk)
    if Colaborador.objects.filter(equipo=equipo).exists():
        messages.error(request, "No se puede eliminar: el equipo tiene colaboradores. Primero desactívalos o muévelos.")
        return redirect("catalogos:equipos")
    equipo.delete()
    messages.success(request, "Equipo eliminado.")
    return redirect("catalogos:equipos")


@admin_required
@require_http_methods(["POST"])
def colaborador_toggle(request, pk: int):
    colab = get_object_or_404(Colaborador, pk=pk)
    colab.activo = not bool(colab.activo)
    colab.save(update_fields=["activo", "actualizado_en"])
    _sync_equipo_integrantes(colab.equipo_id)
    messages.success(request, "Colaborador actualizado.")
    return redirect("catalogos:equipos")


@admin_required
@require_http_methods(["POST"])
def colaborador_delete(request, pk: int):
    colab = get_object_or_404(Colaborador, pk=pk)
    equipo_id = colab.equipo_id
    try:
        colab.delete()
    except ProtectedError:
        messages.error(request, "No se puede eliminar: el colaborador ya tiene asignaciones. Usa desactivar.")
        return redirect("catalogos:equipos")
    _sync_equipo_integrantes(equipo_id)
    messages.success(request, "Colaborador eliminado.")
    return redirect("catalogos:equipos")


@admin_required
@require_http_methods(["GET", "POST"])
def colaborador_editar(request, pk: int):
    colab = get_object_or_404(Colaborador, pk=pk)
    original_equipo_id = colab.equipo_id
    if request.method == "POST":
        form = ColaboradorForm(request.POST, instance=colab)
        if form.is_valid():
            updated = form.save()
            _sync_equipo_integrantes(original_equipo_id)
            _sync_equipo_integrantes(updated.equipo_id)
            messages.success(request, "Colaborador actualizado.")
            return redirect("catalogos:equipos")
    else:
        form = ColaboradorForm(instance=colab)
    return render(request, "catalogos/colaborador_editar.html", {"form": form, "colaborador": colab})


class MaquinaForm(forms.ModelForm):
    class Meta:
        model = Maquina
        fields = ["nombre", "tipo", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.es_robot = False
        if commit:
            instance.save()
        return instance


@admin_required
@require_http_methods(["GET", "POST"])
def maquinas(request):
    if request.method == "POST":
        form = MaquinaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina guardada.")
            return redirect("catalogos:maquinas")
    else:
        form = MaquinaForm(initial={"activo": True})

    maquinas_qs = Maquina.objects.filter(es_robot=False).order_by("tipo", "nombre")
    return render(request, "catalogos/maquinas.html", {"form": form, "maquinas": maquinas_qs})


@admin_required
@require_http_methods(["POST"])
def maquina_toggle(request, pk: int):
    m = get_object_or_404(Maquina, pk=pk, es_robot=False)
    m.activo = not bool(m.activo)
    m.save(update_fields=["activo", "actualizado_en"])
    messages.success(request, "Máquina actualizada.")
    return redirect("catalogos:maquinas")


@admin_required
@require_http_methods(["POST"])
def maquina_delete(request, pk: int):
    m = get_object_or_404(Maquina, pk=pk, es_robot=False)
    try:
        m.delete()
    except ProtectedError:
        messages.error(request, "No se puede eliminar: la máquina ya tiene asignaciones.")
        return redirect("catalogos:maquinas")
    messages.success(request, "Máquina eliminada.")
    return redirect("catalogos:maquinas")


@admin_required
@require_http_methods(["GET", "POST"])
def maquina_editar(request, pk: int):
    m = get_object_or_404(Maquina, pk=pk, es_robot=False)
    if request.method == "POST":
        form = MaquinaForm(request.POST, instance=m)
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina actualizada.")
            return redirect("catalogos:maquinas")
    else:
        form = MaquinaForm(instance=m)
    return render(request, "catalogos/maquina_editar.html", {"form": form, "maquina": m})


class RobotForm(forms.ModelForm):
    class Meta:
        model = Maquina
        fields = ["nombre", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.tipo = "Soldadura"
        instance.es_robot = True
        if commit:
            instance.save()
        return instance


class RobotPiezaForm(forms.ModelForm):
    class Meta:
        model = RobotPiezaCatalogo
        fields = ["nombre", "peso_kg", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")


class RobotProduccionForm(forms.Form):
    fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    robot = forms.ModelChoiceField(queryset=Maquina.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    operadores = forms.ModelMultipleChoiceField(
        queryset=Colaborador.objects.none(),
        widget=forms.SelectMultiple(attrs={"class": "form-control", "size": "10"}),
    )
    pieza = forms.ModelChoiceField(queryset=RobotPiezaCatalogo.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    cantidad = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    corte_maquina = forms.ModelChoiceField(
        queryset=Maquina.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    observaciones = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control", "placeholder": "Opcional"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["robot"].queryset = Maquina.objects.filter(activo=True, es_robot=True, tipo="Soldadura").order_by("nombre")
        self.fields["operadores"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")
        self.fields["pieza"].queryset = RobotPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        self.fields["corte_maquina"].queryset = Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre")

    def clean_robot(self):
        robot = self.cleaned_data.get("robot")
        if robot and (not getattr(robot, "es_robot", False) or getattr(robot, "tipo", "") != "Soldadura"):
            raise forms.ValidationError("Selecciona un robot válido.")
        return robot

    def clean_corte_maquina(self):
        m = self.cleaned_data.get("corte_maquina")
        if m and (getattr(m, "tipo", "") != "Corte" or getattr(m, "es_robot", False)):
            raise forms.ValidationError("Selecciona una máquina de corte válida.")
        return m


class RobotOrdenProduccionForm(forms.ModelForm):
    class Meta:
        model = RobotOrdenProduccion
        fields = ["proyecto", "nombre", "producto", "cantidad_objetivo", "observaciones"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["proyecto"].required = True
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")
        self.fields["nombre"].required = True
        if getattr(self.instance, "id", None) and (not (self.instance.nombre or "").strip()):
            self.initial.setdefault("nombre", (self.instance.producto or "").strip())


class RobotOrdenItemForm(forms.ModelForm):
    class Meta:
        model = RobotOrdenItem
        fields = ["etapa", "pieza", "pieza_custom_nombre", "pieza_custom_peso_kg", "cantidad_requerida", "robot_sugerido"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["pieza"].queryset = RobotPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        self.fields["robot_sugerido"].queryset = Maquina.objects.filter(activo=True, es_robot=True).order_by("nombre")
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")
        self.fields["robot_sugerido"].required = False

    def clean(self):
        cleaned = super().clean()
        pieza = cleaned.get("pieza")
        custom_nombre = (cleaned.get("pieza_custom_nombre") or "").strip()
        custom_peso = float(cleaned.get("pieza_custom_peso_kg") or 0.0)
        if pieza and custom_nombre:
            raise forms.ValidationError("Selecciona catálogo o pieza única, no ambos.")
        if not pieza and not custom_nombre:
            raise forms.ValidationError("Selecciona una pieza de catálogo o captura una pieza única.")
        if not pieza and custom_peso <= 0:
            raise forms.ValidationError("Peso (kg) requerido para pieza única.")
        return cleaned


class RobotOrdenRegistroProduccionForm(forms.Form):
    fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    orden_item = forms.ModelChoiceField(queryset=RobotOrdenItem.objects.none(), widget=forms.Select(attrs={"class": "form-control"}))
    cantidad = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    robot = forms.ModelChoiceField(queryset=Maquina.objects.none(), widget=forms.Select(attrs={"class": "form-control"}))
    operador = forms.ModelMultipleChoiceField(
        queryset=Colaborador.objects.none(),
        widget=forms.SelectMultiple(attrs={"class": "form-control", "size": "10"}),
    )
    corte_maquina = forms.ModelChoiceField(
        queryset=Maquina.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "form-control"}),
    )
    observaciones = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, orden: RobotOrdenProduccion, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._orden = orden
        self.fields["orden_item"].queryset = RobotOrdenItem.objects.filter(orden=orden).select_related("pieza").order_by("id")
        self.fields["robot"].queryset = Maquina.objects.filter(activo=True, es_robot=True).order_by("nombre")
        self.fields["operador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")
        self.fields["corte_maquina"].queryset = Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre")

    def clean_orden_item(self):
        item = self.cleaned_data.get("orden_item")
        if item and int(item.orden_id) != int(self._orden.id):
            raise forms.ValidationError("Selección inválida.")
        return item


class HerrPiezaForm(forms.ModelForm):
    class Meta:
        model = HerrPiezaCatalogo
        fields = ["nombre", "peso_kg", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            else:
                field.widget.attrs.setdefault("class", "form-control")


class LaserMaterialPlacaForm(forms.ModelForm):
    largo_cm = forms.FloatField(min_value=0.0, widget=forms.NumberInput(attrs={"step": "0.1"}))
    ancho_cm = forms.FloatField(min_value=0.0, widget=forms.NumberInput(attrs={"step": "0.1"}))

    class Meta:
        model = LaserMaterialPlaca
        fields = ["categoria_material", "tipo_material", "nombre", "calibre", "espesor_mm", "peso_kg", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        inst = getattr(self, "instance", None)
        if inst and getattr(inst, "pk", None):
            try:
                self.fields["largo_cm"].initial = float(getattr(inst, "largo_mm", 0) or 0) / 10.0
            except Exception:
                self.fields["largo_cm"].initial = 0.0
            try:
                self.fields["ancho_cm"].initial = float(getattr(inst, "ancho_mm", 0) or 0) / 10.0
            except Exception:
                self.fields["ancho_cm"].initial = 0.0
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            elif isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def clean(self):
        cleaned = super().clean()
        try:
            largo_cm = float(cleaned.get("largo_cm") or 0.0)
        except Exception:
            largo_cm = 0.0
        try:
            ancho_cm = float(cleaned.get("ancho_cm") or 0.0)
        except Exception:
            ancho_cm = 0.0
        if largo_cm <= 0:
            self.add_error("largo_cm", "Largo (cm) requerido.")
        if ancho_cm <= 0:
            self.add_error("ancho_cm", "Ancho (cm) requerido.")
        return cleaned

    def save(self, commit=True):
        inst = super().save(commit=False)
        largo_cm = float(self.cleaned_data.get("largo_cm") or 0.0)
        ancho_cm = float(self.cleaned_data.get("ancho_cm") or 0.0)
        inst.largo_mm = int(round(largo_cm * 10.0))
        inst.ancho_mm = int(round(ancho_cm * 10.0))
        if commit:
            inst.save()
            self.save_m2m()
        return inst


class CortaClienteProyectoForm(forms.ModelForm):
    class Meta:
        model = CortaClienteProyecto
        fields = ["nombre", "tipo", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if "tipo" in self.fields:
            self.fields["tipo"].choices = [("cliente", "Cliente"), ("proyecto", "Proyecto")]
            if getattr(getattr(self, "instance", None), "tipo", "") == "mixto":
                self.initial["tipo"] = "cliente"
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            elif isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")


class CortaPiezaCatalogoForm(forms.ModelForm):
    class Meta:
        model = CortaPiezaCatalogo
        fields = ["nombre", "pdf", "dxf", "activo"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if name == "activo":
                field.widget.attrs["class"] = "form-check-input"
            elif isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")


class HerrOrdenProduccionForm(forms.ModelForm):
    class Meta:
        model = HerrOrdenProduccion
        fields = ["proyecto", "nombre", "cantidad_objetivo", "observaciones"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["proyecto"].required = True
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")
        self.fields["nombre"].required = True


class HerrOrdenItemForm(forms.ModelForm):
    class Meta:
        model = HerrOrdenItem
        fields = ["etapa", "pieza", "pieza_custom_nombre", "pieza_custom_peso_kg", "cantidad_requerida"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["pieza"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def clean(self):
        cleaned = super().clean()
        pieza = cleaned.get("pieza")
        custom_nombre = (cleaned.get("pieza_custom_nombre") or "").strip()
        custom_peso = float(cleaned.get("pieza_custom_peso_kg") or 0.0)
        if pieza and custom_nombre:
            raise forms.ValidationError("Selecciona catálogo o pieza única, no ambos.")
        if not pieza and not custom_nombre:
            raise forms.ValidationError("Selecciona una pieza de catálogo o captura una pieza única.")
        if not pieza and custom_peso <= 0:
            raise forms.ValidationError("Peso (kg) requerido para pieza única.")
        return cleaned


class HerrOrdenRegistroProduccionForm(forms.Form):
    fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    orden_item = forms.ModelChoiceField(queryset=HerrOrdenItem.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    cantidad = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    operador = forms.ModelMultipleChoiceField(
        queryset=Colaborador.objects.none(),
        widget=forms.SelectMultiple(attrs={"class": "form-select", "size": "10"}),
    )
    observaciones = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, orden: HerrOrdenProduccion, *args, etapa_permitida: str | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self._orden = orden
        qs = HerrOrdenItem.objects.filter(orden=orden).select_related("pieza").order_by("id")
        if etapa_permitida:
            qs = qs.filter(etapa=etapa_permitida)
        self.fields["orden_item"].queryset = qs
        self.fields["operador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")

    def clean_orden_item(self):
        item = self.cleaned_data.get("orden_item")
        if item and int(item.orden_id) != int(self._orden.id):
            raise forms.ValidationError("Selección inválida.")
        return item


class HerrOrdenAsignacionForm(forms.Form):
    etapa = forms.ChoiceField(
        choices=[("Corte", "Corte"), ("Armado", "Armado"), ("Soldadura", "Soldadura"), ("Pintura", "Pintura")],
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    colaborador = forms.ModelChoiceField(queryset=Colaborador.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["colaborador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")


class LaserOrdenProduccionForm(forms.ModelForm):
    cliente_proyecto = forms.CharField(
        required=True,
        widget=forms.TextInput(attrs={"class": "form-control", "placeholder": "Escribe el cliente o proyecto"}),
    )

    class Meta:
        model = LaserOrdenProduccion
        fields = [
            "nombre",
            "material",
            "pieza_ancho_mm",
            "pieza_alto_mm",
            "cantidad_objetivo",
            "observaciones",
        ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        materiales_qs = LaserMaterialPlaca.objects.filter(activo=True).order_by(
            "categoria_material",
            "tipo_material",
            "nombre",
            "calibre",
            "espesor_mm",
            "largo_mm",
            "ancho_mm",
            "id",
        )
        materiales = list(materiales_qs)
        self.fields["material"].queryset = materiales_qs
        name_key_counts: dict[tuple[str, str, str], int] = {}
        for m in materiales:
            k = (
                (getattr(m, "categoria_material", "") or "").strip().upper(),
                (getattr(m, "tipo_material", "") or "").strip().upper(),
                (getattr(m, "nombre", "") or "").strip().upper(),
            )
            name_key_counts[k] = name_key_counts.get(k, 0) + 1
        name_key_seen: dict[tuple[str, str, str], int] = {}
        material_pos: dict[int, int] = {}
        for m in materiales:
            k = (
                (getattr(m, "categoria_material", "") or "").strip().upper(),
                (getattr(m, "tipo_material", "") or "").strip().upper(),
                (getattr(m, "nombre", "") or "").strip().upper(),
            )
            name_key_seen[k] = name_key_seen.get(k, 0) + 1
            material_pos[int(m.id)] = int(name_key_seen[k])

        def _material_label(obj: LaserMaterialPlaca) -> str:
            categoria = (getattr(obj, "categoria_material", "") or "").strip()
            tipo = (getattr(obj, "tipo_material", "") or "").strip()
            nombre = (getattr(obj, "nombre", "") or "").strip()
            calibre = (getattr(obj, "calibre", "") or "").strip()
            espesor = float(getattr(obj, "espesor_mm", 0.0) or 0.0)
            largo = int(getattr(obj, "largo_mm", 0) or 0)
            ancho = int(getattr(obj, "ancho_mm", 0) or 0)
            parts: list[str] = []
            if categoria:
                parts.append(categoria)
            if tipo:
                parts.append(tipo)
            parts.append(nombre if nombre else f"Material #{int(obj.id)}")
            if calibre:
                parts.append(f"Cédula {calibre}")
            if espesor > 0:
                parts.append(f"{espesor:.2f}mm")
            if largo > 0 and ancho > 0:
                parts.append(f"{(largo / 10.0):.1f}x{(ancho / 10.0):.1f}cm")
            try:
                peso = float(getattr(obj, "peso_kg", 0.0) or 0.0)
            except Exception:
                peso = 0.0
            if peso > 0:
                parts.append(f"{peso:.3f}kg")
            key = ((categoria or "").strip().upper(), (tipo or "").strip().upper(), (nombre or "").strip().upper())
            if name_key_counts.get(key, 0) > 1:
                parts.append(f"(#{material_pos.get(int(obj.id), 1)})")
            return " · ".join(parts)

        self.fields["material"].label_from_instance = _material_label
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")
        self.fields["nombre"].required = True
        self.fields["pieza_ancho_mm"].required = True
        self.fields["pieza_alto_mm"].required = True

        orden = getattr(self, "instance", None)
        if orden and getattr(orden, "corta_cliente_proyecto_id", None):
            self.fields["cliente_proyecto"].initial = getattr(getattr(orden, "corta_cliente_proyecto", None), "nombre", "") or ""

    def clean(self):
        cleaned = super().clean()
        cp = _norm_text(cleaned.get("cliente_proyecto") or "")
        if not cp:
            self.add_error("cliente_proyecto", "Cliente o proyecto requerido.")
        try:
            ancho = int(cleaned.get("pieza_ancho_mm") or 0)
        except Exception:
            ancho = 0
        try:
            alto = int(cleaned.get("pieza_alto_mm") or 0)
        except Exception:
            alto = 0
        if ancho <= 0:
            self.add_error("pieza_ancho_mm", "Ancho requerido.")
        if alto <= 0:
            self.add_error("pieza_alto_mm", "Largo requerido.")
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        raw = _norm_text(self.cleaned_data.get("cliente_proyecto") or "")
        try:
            instance.corta_cliente_proyecto = _get_or_create_corta_cliente_proyecto(raw)
        except Exception:
            instance.corta_cliente_proyecto = None
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class LaserOrdenItemForm(forms.ModelForm):
    class Meta:
        model = LaserOrdenItem
        fields = ["etapa", "item_nombre", "item_peso_kg", "cantidad_requerida"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            if isinstance(field.widget, forms.Select):
                field.widget.attrs.setdefault("class", "form-select")
            else:
                field.widget.attrs.setdefault("class", "form-control")

    def clean(self):
        cleaned = super().clean()
        nombre = (cleaned.get("item_nombre") or "").strip()
        try:
            peso = float(cleaned.get("item_peso_kg") or 0.0)
        except Exception:
            peso = 0.0
        if not nombre:
            raise forms.ValidationError("Nombre requerido.")
        if peso <= 0:
            raise forms.ValidationError("Peso (kg) requerido.")
        return cleaned


class LaserOrdenRegistroProduccionForm(forms.Form):
    fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    orden_item = forms.ModelChoiceField(queryset=LaserOrdenItem.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    cantidad = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    operador = forms.ModelMultipleChoiceField(
        queryset=Colaborador.objects.none(),
        widget=forms.SelectMultiple(attrs={"class": "form-select", "size": "10"}),
    )
    observaciones = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, orden: LaserOrdenProduccion, *args, etapa_permitida: str | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self._orden = orden
        qs = LaserOrdenItem.objects.filter(orden=orden).order_by("id")
        if etapa_permitida:
            qs = qs.filter(etapa=etapa_permitida)
        self.fields["orden_item"].queryset = qs
        self.fields["operador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")

    def clean_orden_item(self):
        item = self.cleaned_data.get("orden_item")
        if item and int(item.orden_id) != int(self._orden.id):
            raise forms.ValidationError("Selección inválida.")
        return item


class LaserOrdenAsignacionForm(forms.Form):
    etapa = forms.ChoiceField(
        choices=[("Corte", "Corte"), ("Armado", "Armado"), ("Soldadura", "Soldadura"), ("Pintura", "Pintura")],
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    colaborador = forms.ModelChoiceField(queryset=Colaborador.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["colaborador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")


class RobotOrdenAsignacionForm(forms.Form):
    etapa = forms.ChoiceField(
        choices=[("Corte", "Corte"), ("Armado", "Armado"), ("Soldadura", "Soldadura"), ("Pintura", "Pintura")],
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    colaborador = forms.ModelChoiceField(queryset=Colaborador.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["colaborador"].queryset = Colaborador.objects.filter(activo=True).order_by("nombre")


class HerrIndividualForm(forms.Form):
    proyecto = forms.ModelChoiceField(queryset=Proyecto.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    nombre = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    pieza_custom_nombre = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza_custom_peso_kg = forms.FloatField(required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))
    observaciones = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["pieza"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")

    def clean(self):
        cleaned = super().clean()
        pieza = cleaned.get("pieza")
        custom_nombre = (cleaned.get("pieza_custom_nombre") or "").strip()
        custom_peso = float(cleaned.get("pieza_custom_peso_kg") or 0.0) if cleaned.get("pieza_custom_peso_kg") is not None else 0.0
        if pieza and custom_nombre:
            raise forms.ValidationError("Selecciona catálogo o pieza única, no ambos.")
        if not pieza and not custom_nombre:
            raise forms.ValidationError("Selecciona una pieza de catálogo o captura una pieza única.")
        if not pieza and custom_peso <= 0:
            raise forms.ValidationError("Peso (kg/pza) requerido para pieza única.")
        return cleaned


class HerrTrabajoCreateForm(forms.Form):
    TIPO_CHOICES = [("PIEZA", "Pieza individual"), ("OP", "Orden grande (OP)")]

    tipo = forms.ChoiceField(choices=TIPO_CHOICES, widget=forms.Select(attrs={"class": "form-select"}))
    proyecto = forms.ModelChoiceField(queryset=Proyecto.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    nombre = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    descripcion = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    cantidad_objetivo = forms.IntegerField(min_value=1, required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))

    pieza = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    pieza_custom_nombre = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza_custom_peso_kg = forms.FloatField(required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))
    plano_pdf = forms.FileField(required=False)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["pieza"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")

    def clean(self):
        cleaned = super().clean()
        tipo = cleaned.get("tipo")
        pieza = cleaned.get("pieza")
        custom_nombre = (cleaned.get("pieza_custom_nombre") or "").strip()
        custom_peso = float(cleaned.get("pieza_custom_peso_kg") or 0.0) if cleaned.get("pieza_custom_peso_kg") is not None else 0.0
        if pieza and custom_nombre:
            raise forms.ValidationError("Selecciona catálogo o pieza única, no ambos.")
        if not pieza and not custom_nombre:
            raise forms.ValidationError("Selecciona una pieza de catálogo o captura una pieza única.")
        if not pieza and custom_peso <= 0:
            raise forms.ValidationError("Peso (kg/pza) requerido para pieza única.")
        if tipo == "OP":
            cant = cleaned.get("cantidad_objetivo") or 0
            if int(cant) <= 0:
                raise forms.ValidationError("Cantidad objetivo requerida para OP.")
        return cleaned


class HerrTrabajoEditForm(forms.Form):
    proyecto = forms.ModelChoiceField(queryset=Proyecto.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    nombre = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    descripcion = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    plano_pdf = forms.FileField(required=False)

    pieza = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    pieza_custom_nombre = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza_custom_peso_kg = forms.FloatField(required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))

    def __init__(self, orden: HerrOrdenProduccion, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._orden = orden
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["pieza"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")

        self.initial.setdefault("proyecto", orden.proyecto_id)
        self.initial.setdefault("nombre", (orden.nombre or "").strip())
        self.initial.setdefault("descripcion", (orden.descripcion or "").strip())

        it = HerrOrdenItem.objects.filter(orden=orden).order_by("id").first()
        if it:
            self.initial.setdefault("pieza", it.pieza_id)
            self.initial.setdefault("pieza_custom_nombre", (it.pieza_custom_nombre or "").strip())
            self.initial.setdefault("pieza_custom_peso_kg", float(it.pieza_custom_peso_kg or 0.0))

    def clean(self):
        cleaned = super().clean()
        pieza = cleaned.get("pieza")
        custom_nombre = (cleaned.get("pieza_custom_nombre") or "").strip()
        custom_peso = float(cleaned.get("pieza_custom_peso_kg") or 0.0) if cleaned.get("pieza_custom_peso_kg") is not None else 0.0
        if pieza and custom_nombre:
            raise forms.ValidationError("Selecciona catálogo o pieza única, no ambos.")
        if not pieza and not custom_nombre:
            raise forms.ValidationError("Selecciona una pieza de catálogo o captura una pieza única.")
        if not pieza and custom_peso <= 0:
            raise forms.ValidationError("Peso (kg/pza) requerido para pieza única.")
        return cleaned


class HerrSolicitudForm(forms.Form):
    producto = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    cantidad = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    fecha_compromiso = forms.DateField(required=False, widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    cliente = forms.ModelChoiceField(queryset=HerrCliente.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    cliente_nuevo = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["producto"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        self.fields["cliente"].queryset = HerrCliente.objects.filter(activo=True).order_by("nombre")
        self.initial.setdefault("fecha_compromiso", timezone.localdate())

    def clean(self):
        cleaned = super().clean()
        cliente = cleaned.get("cliente")
        cliente_nuevo = (cleaned.get("cliente_nuevo") or "").strip()
        if cliente and cliente_nuevo:
            raise forms.ValidationError("Selecciona cliente o captura cliente nuevo, no ambos.")
        if not cliente and not cliente_nuevo:
            raise forms.ValidationError("Selecciona un cliente o captura un cliente nuevo.")
        if not cleaned.get("fecha_compromiso"):
            cleaned["fecha_compromiso"] = timezone.localdate()
        return cleaned


class PedidoProduccionForm(forms.Form):
    cliente = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    telefono = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    fecha_compromiso = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    comentarios = forms.CharField(required=False, widget=forms.Textarea(attrs={"class": "form-control", "rows": 3}))


class PedidoProduccionEditForm(forms.Form):
    cliente = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    telefono = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    fecha_compromiso = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    comentarios = forms.CharField(required=False, widget=forms.Textarea(attrs={"class": "form-control", "rows": 3}))


class PedidoProduccionItemEditForm(forms.Form):
    item_id = forms.IntegerField(required=False, widget=forms.HiddenInput())
    producto = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    cantidad_total = forms.IntegerField(min_value=1, required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        item = kwargs.pop("item", None)
        super().__init__(*args, **kwargs)
        self.fields["producto"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        if item is not None:
            self.fields["item_id"].initial = int(getattr(item, "id", 0) or 0)
            self.fields["producto"].initial = getattr(item, "producto", None)
            self.fields["producto"].disabled = True
            self.fields["cantidad_total"].initial = int(getattr(item, "cantidad_total", 0) or 0)


class PedidoProduccionItemForm(forms.Form):
    producto = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    cantidad_total = forms.IntegerField(min_value=1, required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["producto"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")


def _next_pedido_folio() -> str:
    last = PedidoProduccion.objects.order_by("-id").only("id").first()
    nxt = int(getattr(last, "id", 0) or 0) + 1
    return f"ORD-{nxt:05d}"


def _get_or_create_stock_locked(producto: HerrPiezaCatalogo) -> LogisticaStock:
    st = LogisticaStock.objects.select_for_update().filter(producto=producto).first()
    if st:
        return st
    return LogisticaStock.objects.create(producto=producto, stock=0)


def _corta_producto_from_orden(orden: LaserOrdenProduccion) -> str:
    pieza = getattr(orden, "corta_pieza", None)
    if pieza:
        nombre = (getattr(pieza, "nombre", "") or "").strip()
        if nombre:
            return nombre
    producto = (getattr(orden, "descripcion", "") or "").strip()
    if producto:
        return producto
    producto = (getattr(orden, "nombre", "") or "").strip()
    if producto:
        return producto
    return (getattr(orden, "codigo", "") or "").strip()


def _get_or_create_stock_corta_locked(producto: str) -> LogisticaStockCorta:
    key = (producto or "").strip().upper()
    st = LogisticaStockCorta.objects.select_for_update().filter(producto_normalizado=key).first()
    if st:
        return st
    return LogisticaStockCorta.objects.create(producto=(producto or "").strip(), stock=0)


def _seed_paros_catalogos():
    motivos = [
        ("Energía", True),
        ("Falla", True),
        ("Falta de material", False),
        ("Falta de planos", False),
        ("Falta de suministros", False),
        ("Set-up / cambio de modelo", False),
        ("Espera de calidad", False),
        ("Logística interna", False),
        ("Mantenimiento planeado", False),
        ("Seguridad / bloqueo", False),
        ("Limpieza / orden", False),
        ("Sin operador", False),
    ]
    fallas = [
        ("Eléctrica", True),
        ("Mecánica", True),
        ("Neumática / hidráulica", False),
        ("Software / PLC / HMI", True),
        ("Herramental / desgaste", False),
        ("Calibración / ajuste", False),
        ("Lubricación / refrigeración", False),
        ("Atoro / colisión", False),
        ("Calidad por máquina", False),
    ]
    for nombre, sistema in motivos:
        MaquinaParoMotivo.objects.get_or_create(
            nombre_normalizado=nombre.upper(),
            defaults={"nombre": nombre, "activo": True, "es_sistema": bool(sistema)},
        )
    for nombre, sistema in fallas:
        MaquinaFallaTipo.objects.get_or_create(
            nombre_normalizado=nombre.upper(),
            defaults={"nombre": nombre, "activo": True, "es_sistema": bool(sistema)},
        )


class ParoStartForm(forms.Form):
    motivo = forms.ModelChoiceField(queryset=MaquinaParoMotivo.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    comentario = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["motivo"].queryset = MaquinaParoMotivo.objects.filter(activo=True).order_by("nombre")


class FallaStartForm(forms.Form):
    tipo = forms.ModelChoiceField(queryset=MaquinaFallaTipo.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    comentario = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["tipo"].queryset = MaquinaFallaTipo.objects.filter(activo=True).order_by("nombre")


class ComentarioForm(forms.Form):
    comentario = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))


class EnergiaForm(forms.Form):
    inicio_fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    inicio_hora = forms.TimeField(widget=forms.TimeInput(attrs={"type": "time", "class": "form-control"}))
    fin_fecha = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    fin_hora = forms.TimeField(widget=forms.TimeInput(attrs={"type": "time", "class": "form-control"}))
    comentario = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    aplicar_maquinas = forms.BooleanField(required=False)

    def clean(self):
        cleaned = super().clean()
        try:
            inicio = timezone.make_aware(datetime.combine(cleaned["inicio_fecha"], cleaned["inicio_hora"]), timezone.get_default_timezone())
            fin = timezone.make_aware(datetime.combine(cleaned["fin_fecha"], cleaned["fin_hora"]), timezone.get_default_timezone())
        except Exception:
            return cleaned
        if fin < inicio:
            raise forms.ValidationError("La hora fin no puede ser menor a la hora inicio.")
        cleaned["inicio_dt"] = inicio
        cleaned["fin_dt"] = fin
        return cleaned

@login_required
@require_http_methods(["GET", "POST"])
def robotica(request):
    if not (_is_admin_user(request.user) or _is_robotica_user(request.user)):
        return redirect("/")
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "robot_create":
            robot_form = RobotForm(request.POST)
            if robot_form.is_valid():
                robot_form.save()
                messages.success(request, "Robot guardado.")
                return redirect("catalogos:robotica")
        elif action == "robot_toggle":
            pk = int(request.POST.get("pk") or 0)
            r = get_object_or_404(Maquina, pk=pk, es_robot=True)
            r.activo = not bool(r.activo)
            r.save(update_fields=["activo", "actualizado_en"])
            messages.success(request, "Robot actualizado.")
            return redirect("catalogos:robotica")
        elif action == "robot_delete":
            pk = int(request.POST.get("pk") or 0)
            r = get_object_or_404(Maquina, pk=pk, es_robot=True)
            try:
                r.delete()
            except ProtectedError:
                messages.error(request, "No se puede eliminar: el robot ya tiene producción o asignaciones.")
                return redirect("catalogos:robotica")
            messages.success(request, "Robot eliminado.")
            return redirect("catalogos:robotica")
        elif action == "pieza_create":
            pieza_form = RobotPiezaForm(request.POST)
            if pieza_form.is_valid():
                pieza_form.save()
                messages.success(request, "Pieza de robótica guardada.")
                return redirect("catalogos:robotica")
        elif action == "pieza_toggle":
            pk = int(request.POST.get("pk") or 0)
            p = get_object_or_404(RobotPiezaCatalogo, pk=pk)
            p.activo = not bool(p.activo)
            p.save(update_fields=["activo", "actualizado_en"])
            messages.success(request, "Pieza actualizada.")
            return redirect("catalogos:robotica")
        elif action == "pieza_delete":
            pk = int(request.POST.get("pk") or 0)
            p = get_object_or_404(RobotPiezaCatalogo, pk=pk)
            try:
                p.delete()
            except ProtectedError:
                messages.error(request, "No se puede eliminar: la pieza ya tiene producción registrada.")
                return redirect("catalogos:robotica")
            messages.success(request, "Pieza eliminada.")
            return redirect("catalogos:robotica")
        elif action == "prod_create":
            prod_form = RobotProduccionForm(request.POST)
            if prod_form.is_valid():
                operadores = list(prod_form.cleaned_data.get("operadores") or [])
                cantidad_total = int(prod_form.cleaned_data.get("cantidad") or 0)
                if cantidad_total <= 0 or not operadores:
                    messages.error(request, "Debes seleccionar operador(es) y una cantidad válida.")
                    return redirect("catalogos:robotica")

                n_ops = len(operadores)
                base = cantidad_total // n_ops
                rem = cantidad_total % n_ops
                created = 0
                for i, op in enumerate(operadores):
                    qty = base + (1 if i < rem else 0)
                    if qty <= 0:
                        continue
                    RobotProduccion.objects.create(
                        fecha=prod_form.cleaned_data["fecha"],
                        robot=prod_form.cleaned_data["robot"],
                        operador=op,
                        pieza=prod_form.cleaned_data["pieza"],
                        cantidad=int(qty),
                        corte_maquina=prod_form.cleaned_data.get("corte_maquina"),
                        observaciones=prod_form.cleaned_data.get("observaciones") or "",
                    )
                    created += 1
                messages.success(request, f"Producción robótica guardada ({created} operador(es)).")
                return redirect("catalogos:robotica")
        elif action == "prod_delete":
            pk = int(request.POST.get("pk") or 0)
            RobotProduccion.objects.filter(pk=pk).delete()
            messages.success(request, "Registro eliminado.")
            return redirect("catalogos:robotica")

    robot_form = RobotForm(initial={"activo": True})
    pieza_form = RobotPiezaForm(initial={"activo": True, "peso_kg": 0.0})
    prod_form = RobotProduccionForm(initial={"fecha": timezone.localdate(), "cantidad": 1})

    robots_qs = Maquina.objects.filter(es_robot=True).order_by("-activo", "nombre")
    piezas_qs = RobotPiezaCatalogo.objects.all().order_by("-activo", "nombre")
    prod_recent = (
        RobotProduccion.objects.select_related("pieza", "robot", "operador", "corte_maquina")
        .order_by("-fecha", "-id")[:200]
    )

    today = timezone.localdate()
    raw_week_start = (request.GET.get("week_start") or "").strip()
    if raw_week_start:
        try:
            week_start = datetime.strptime(raw_week_start, "%Y-%m-%d").date()
        except Exception:
            week_start = today
    else:
        delta = (today.weekday() - 3) % 7
        week_start = today - timedelta(days=delta)
    week_end = week_start + timedelta(days=7)

    qs_week = RobotProduccion.objects.filter(fecha__gte=week_start, fecha__lt=week_end).select_related("pieza", "robot", "operador")
    expr_kg = ExpressionWrapper(F("cantidad") * F("pieza__peso_kg"), output_field=FloatField())
    total_kg = float(qs_week.aggregate(total=Sum(expr_kg))["total"] or 0.0)
    total_ton = total_kg / 1000.0
    total_piezas = int(qs_week.aggregate(total=Sum("cantidad"))["total"] or 0)
    ops_count = int(qs_week.values("operador_id").distinct().count())
    ton_por_operador = (total_ton / ops_count) if ops_count else 0.0

    daily_rows = qs_week.values("fecha").annotate(kg=Sum(expr_kg), piezas=Sum("cantidad")).order_by("fecha")
    day_map = {r["fecha"]: r for r in daily_rows}
    daily = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        row = day_map.get(d) or {}
        daily.append(
            {
                "date": d.isoformat(),
                "piezas": int(row.get("piezas") or 0),
                "ton": round(float(row.get("kg") or 0.0) / 1000.0, 3),
            }
        )

    top_robots = list(
        qs_week.values("robot__nombre")
        .annotate(kg=Sum(expr_kg))
        .order_by("-kg")[:10]
    )
    top_piezas = list(
        qs_week.values("pieza__nombre")
        .annotate(kg=Sum(expr_kg))
        .order_by("-kg")[:10]
    )

    weekly = {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "ton": round(total_ton, 3),
        "piezas": total_piezas,
        "operadores": ops_count,
        "ton_por_operador": round(ton_por_operador, 3),
        "daily": daily,
        "top_robots": [{"robot": r["robot__nombre"], "ton": round(float(r["kg"] or 0.0) / 1000.0, 3)} for r in top_robots],
        "top_piezas": [{"pieza": r["pieza__nombre"], "ton": round(float(r["kg"] or 0.0) / 1000.0, 3)} for r in top_piezas],
    }

    return render(
        request,
        "catalogos/robotica.html",
        {
            "robot_form": robot_form,
            "pieza_form": pieza_form,
            "prod_form": prod_form,
            "robots": robots_qs,
            "piezas": piezas_qs,
            "produccion": prod_recent,
            "weekly": weekly,
            "weekly_payload": {
                "labels": [d["date"] for d in daily],
                "ton": [d["ton"] for d in daily],
                "piezas": [d["piezas"] for d in daily],
            },
            "weekly_selected_start": week_start.isoformat(),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def robot_editar(request, pk: int):
    if not (_is_admin_user(request.user) or _is_robotica_user(request.user)):
        return redirect("/")
    robot = get_object_or_404(Maquina, pk=pk, es_robot=True)
    if request.method == "POST":
        form = RobotForm(request.POST, instance=robot)
        if form.is_valid():
            form.save()
            messages.success(request, "Robot actualizado.")
            return redirect("catalogos:robotica")
    else:
        form = RobotForm(instance=robot)
    return render(request, "catalogos/robot_editar.html", {"form": form, "robot": robot})


@login_required
@require_http_methods(["GET", "POST"])
def robot_pieza_editar(request, pk: int):
    if not (_is_admin_user(request.user) or _is_robotica_user(request.user)):
        return redirect("/")
    pieza = get_object_or_404(RobotPiezaCatalogo, pk=pk)
    if request.method == "POST":
        form = RobotPiezaForm(request.POST, instance=pieza)
        if form.is_valid():
            form.save()
            messages.success(request, "Pieza actualizada.")
            return redirect("catalogos:robotica")
    else:
        form = RobotPiezaForm(instance=pieza)
    return render(request, "catalogos/robot_pieza_editar.html", {"form": form, "pieza": pieza})


@login_required
@require_http_methods(["GET", "POST"])
def robotica_ordenes(request):
    if not (_is_admin_user(request.user) or _is_robotica_user(request.user)):
        return redirect("/")

    form = RobotOrdenProduccionForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_create":
            form = RobotOrdenProduccionForm(request.POST)
            if form.is_valid():
                orden = form.save()
                messages.success(request, "Orden creada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))

    ordenes_abiertas = list(RobotOrdenProduccion.objects.filter(estado="Abierta").select_related("proyecto").order_by("-id")[:200])
    ordenes_recientes = list(
        RobotOrdenProduccion.objects.exclude(estado="Abierta").select_related("proyecto").order_by("-id")[:50]
    )
    orden_ids = [int(o.id) for o in (ordenes_abiertas + ordenes_recientes)]
    if orden_ids:
        req_rows = (
            RobotOrdenItem.objects.filter(orden_id__in=orden_ids)
            .values("orden_id", "etapa")
            .annotate(req=Coalesce(Sum("cantidad_requerida"), 0))
        )
        prod_rows = (
            RobotProduccion.objects.filter(orden_item__orden_id__in=orden_ids)
            .values("orden_item__orden_id", "orden_item__etapa")
            .annotate(prod=Coalesce(Sum("cantidad"), 0))
        )
        req_map = {}
        for r in req_rows:
            req_map[(int(r["orden_id"]), str(r["etapa"]))] = int(r["req"] or 0)
        prod_map = {}
        for r in prod_rows:
            prod_map[(int(r["orden_item__orden_id"]), str(r["orden_item__etapa"]))] = int(r["prod"] or 0)
        etapas = ["Corte", "Armado", "Soldadura", "Pintura"]
        for o in (ordenes_abiertas + ordenes_recientes):
            sp = {}
            for e in etapas:
                req = req_map.get((int(o.id), e), 0)
                prod = prod_map.get((int(o.id), e), 0)
                sp[e] = 0.0 if req <= 0 else min(100.0, (prod / req) * 100.0)
            o.stage_pcts = sp
    return render(
        request,
        "catalogos/robotica_ordenes.html",
        {"form": form, "ordenes_abiertas": ordenes_abiertas, "ordenes_recientes": ordenes_recientes},
    )


@login_required
@require_http_methods(["GET", "POST"])
def robotica_orden_detalle(request, pk: int):
    if not (_is_admin_user(request.user) or _is_robotica_user(request.user)):
        return redirect("/")

    orden = get_object_or_404(RobotOrdenProduccion, pk=pk)
    edit_form = RobotOrdenProduccionForm(instance=orden)
    item_form = RobotOrdenItemForm()
    prod_form = RobotOrdenRegistroProduccionForm(orden, initial={"fecha": timezone.localdate(), "cantidad": 1})
    assign_form = RobotOrdenAsignacionForm()
    has_prod = RobotProduccion.objects.filter(orden_item__orden=orden).exists()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_update":
            if orden.estado != "Abierta":
                messages.error(request, "La orden no se puede editar (cerrada/cancelada).")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
            edit_form = RobotOrdenProduccionForm(request.POST, instance=orden)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "Orden actualizada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "orden_delete":
            if has_prod:
                messages.error(request, "No se puede eliminar: la orden ya tiene producción registrada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
            try:
                orden.delete()
            except Exception:
                messages.error(request, "No se pudo eliminar la orden.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
            messages.success(request, "Orden eliminada.")
            return redirect("catalogos:robotica_ordenes")
        elif action == "item_add":
            item_form = RobotOrdenItemForm(request.POST)
            if item_form.is_valid():
                try:
                    RobotOrdenItem.objects.create(
                        orden=orden,
                        etapa=item_form.cleaned_data["etapa"],
                        pieza=item_form.cleaned_data.get("pieza"),
                        pieza_custom_nombre=(item_form.cleaned_data.get("pieza_custom_nombre") or "").strip(),
                        pieza_custom_peso_kg=float(item_form.cleaned_data.get("pieza_custom_peso_kg") or 0.0),
                        cantidad_requerida=item_form.cleaned_data["cantidad_requerida"],
                        robot_sugerido=item_form.cleaned_data.get("robot_sugerido"),
                    )
                    messages.success(request, "Pieza requerida agregada.")
                    return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
                except Exception:
                    messages.error(request, "No se pudo agregar la pieza (puede que ya exista en la orden).")
        elif action == "item_delete":
            item_id = int(request.POST.get("item_id") or 0)
            if item_id:
                RobotOrdenItem.objects.filter(pk=item_id, orden=orden).delete()
                messages.success(request, "Pieza removida de la orden.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "prod_add":
            prod_form = RobotOrdenRegistroProduccionForm(orden, request.POST)
            if prod_form.is_valid():
                item = prod_form.cleaned_data["orden_item"]
                pieza_fk = item.pieza if getattr(item, "pieza_id", None) else None
                custom_nombre = "" if pieza_fk else (item.pieza_custom_nombre or "").strip()
                custom_peso = 0.0 if pieza_fk else float(item.pieza_custom_peso_kg or 0.0)
                operadores = list(prod_form.cleaned_data.get("operador") or [])
                cantidad_total = int(prod_form.cleaned_data["cantidad"] or 0)
                if cantidad_total <= 0 or not operadores:
                    messages.error(request, "Debes seleccionar operador(es) y una cantidad válida.")
                    return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))

                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""

                n_ops = len(operadores)
                base = cantidad_total // n_ops
                rem = cantidad_total % n_ops
                for i, op in enumerate(operadores):
                    qty = base + (1 if i < rem else 0)
                    if qty <= 0:
                        continue
                    RobotProduccion.objects.create(
                        fecha=prod_form.cleaned_data["fecha"],
                        pieza=pieza_fk,
                        pieza_custom_nombre=custom_nombre,
                        pieza_custom_peso_kg=custom_peso,
                        cantidad=int(qty),
                        robot=prod_form.cleaned_data["robot"],
                        operador=op,
                        orden_item=item,
                        corte_maquina=prod_form.cleaned_data.get("corte_maquina"),
                        observaciones=prod_form.cleaned_data.get("observaciones") or "",
                    )
                    RobotOrdenAsignacion.objects.get_or_create(
                        orden=orden,
                        etapa=str(getattr(item, "etapa", "") or "").strip() or "Corte",
                        colaborador=op,
                        defaults={"asignado_por": actor},
                    )
                messages.success(request, "Producción registrada en la orden.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "asig_add":
            assign_form = RobotOrdenAsignacionForm(request.POST)
            if assign_form.is_valid():
                etapa = assign_form.cleaned_data["etapa"]
                col = assign_form.cleaned_data["colaborador"]
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                RobotOrdenAsignacion.objects.get_or_create(
                    orden=orden,
                    etapa=etapa,
                    colaborador=col,
                    defaults={"asignado_por": actor},
                )
                messages.success(request, "Asignación guardada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "asig_delete":
            asig_id = int(request.POST.get("asig_id") or 0)
            if asig_id:
                RobotOrdenAsignacion.objects.filter(id=asig_id, orden=orden).delete()
                messages.success(request, "Asignación eliminada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "orden_close":
            if orden.estado == "Abierta":
                orden.estado = "Cerrada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cerrada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))
        elif action == "orden_cancel":
            if orden.estado == "Abierta":
                orden.estado = "Cancelada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cancelada.")
                return redirect("catalogos:robotica_orden_detalle", pk=int(orden.id))

    items = list(RobotOrdenItem.objects.filter(orden=orden).select_related("pieza", "robot_sugerido").order_by("id"))
    prod_rows = (
        RobotProduccion.objects.filter(orden_item__orden=orden)
        .values("orden_item_id")
        .annotate(producidas=Coalesce(Sum("cantidad"), 0))
    )
    prod_map = {int(r["orden_item_id"]): r for r in prod_rows if r.get("orden_item_id")}
    resumen = []
    pcts = []
    for it in items:
        row = prod_map.get(int(it.id), {})
        producidas = int(row.get("producidas") or 0)
        requeridas = int(it.cantidad_requerida or 0)
        faltan = max(0, requeridas - producidas)
        pct = 0.0 if requeridas <= 0 else min(100.0, (producidas / requeridas) * 100.0)
        pcts.append(pct)
        kg = float(it.pieza_peso_kg or 0.0) * producidas
        resumen.append(
            {
                "item": it,
                "requeridas": requeridas,
                "producidas": producidas,
                "faltan": faltan,
                "pct": round(pct, 2),
                "kg": round(float(kg or 0.0), 3),
            }
        )
    avance_pct = round(min(pcts) if pcts else 0.0, 2)

    historial = list(
        RobotProduccion.objects.filter(orden_item__orden=orden)
        .select_related("pieza", "robot", "operador", "corte_maquina", "orden_item")
        .order_by("-fecha", "-id")[:250]
    )
    asignaciones = list(RobotOrdenAsignacion.objects.filter(orden=orden).select_related("colaborador").order_by("-asignado_en"))
    return render(
        request,
        "catalogos/robotica_orden_detalle.html",
        {
            "orden": orden,
            "avance_pct": avance_pct,
            "edit_form": edit_form,
            "can_delete": (not has_prod),
            "item_form": item_form,
            "prod_form": prod_form,
            "assign_form": assign_form,
            "resumen": resumen,
            "historial": historial,
            "asignaciones": asignaciones,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria(request):
    if not _can_herreria(request.user):
        return redirect("/")
    return redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["GET", "POST"])
def herreria_catalogo(request):
    if not _can_herreria_catalogo(request.user):
        return redirect("/")

    next_url = _safe_next_url(request.GET.get("next") or "")
    can_herreria_nav = _can_herreria(request.user)
    if next_url:
        back_url = next_url
        back_label = "Volver"
    else:
        back_url = reverse("catalogos:herreria") if can_herreria_nav else reverse("catalogos:pedidos_ordenes")
        back_label = "Volver" if can_herreria_nav else "Órdenes"

    pieza_form = HerrPiezaForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "pieza_create":
            pieza_form = HerrPiezaForm(request.POST)
            if pieza_form.is_valid():
                pieza_form.save()
                messages.success(request, "Pieza de herrería guardada.")
                if next_url:
                    return redirect(f"{reverse('catalogos:herreria_catalogo')}?next={quote(next_url)}")
                return redirect("catalogos:herreria_catalogo")
        if action == "pieza_delete":
            pid = int(request.POST.get("pieza_id") or 0)
            if not pid:
                messages.error(request, "No se indicó qué pieza eliminar.")
            elif HerrOrdenItem.objects.filter(pieza_id=pid).exists():
                messages.error(request, "No se puede eliminar: la pieza ya está usada en una orden.")
            else:
                try:
                    borradas, _ = HerrPiezaCatalogo.objects.filter(id=pid).delete()
                    if borradas:
                        messages.success(request, "Pieza eliminada.")
                    else:
                        messages.error(request, "Esa pieza ya no existe.")
                except ProtectedError:
                    messages.error(
                        request,
                        "No se puede eliminar: la pieza tiene stock o movimientos de logística. "
                        "Desactívala en su lugar.",
                    )
            if next_url:
                return redirect(f"{reverse('catalogos:herreria_catalogo')}?next={quote(next_url)}")
            return redirect("catalogos:herreria_catalogo")

    piezas_qs = list(HerrPiezaCatalogo.objects.order_by("-activo", "nombre")[:500])
    return render(
        request,
        "catalogos/herreria.html",
        {
            "pieza_form": pieza_form,
            "piezas": piezas_qs,
            "can_herreria_nav": can_herreria_nav,
            "back_url": back_url,
            "back_label": back_label,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria_control(request):
    if not _can_herreria(request.user):
        return redirect("/")
    _finalize_herreria_cierres_expirados()
    is_admin = _is_admin_user(request.user)
    can_accept_solicitudes = bool(is_admin or _is_herreria_supervision(request.user))
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "pedido_item_accept":
            if not can_accept_solicitudes:
                return redirect("catalogos:herreria_control")
            item_id = int(request.POST.get("pedido_item_id") or 0)
            material = (request.POST.get("hay_material") or "").strip().lower()
            if item_id:
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                now = timezone.now()
                with transaction.atomic(using="mes"):
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update().select_related("pedido", "producto"), pk=item_id)
                    if it.estado_herreria != "Pendiente":
                        messages.error(request, "El registro ya no está pendiente.")
                    elif material == "no":
                        it.estado_herreria = "Espera de material"
                        it.aceptado_por = actor
                        it.aceptado_en = now
                        it.espera_material_entrada_en = now
                        it.save(update_fields=["estado_herreria", "aceptado_por", "aceptado_en", "espera_material_entrada_en", "actualizado_en"])
                        messages.success(request, "Enviado a Espera de Material.")
                    else:
                        qty = max(1, int(it.cantidad_total or 1))
                        prod = it.producto
                        kg_total = float(getattr(prod, "peso_kg", 0.0) or 0.0) * float(qty)
                        codigo = f"{it.pedido.folio}-{int(it.id):03d}"
                        orden = HerrOrdenProduccion.objects.create(
                            proyecto=None,
                            cliente_herreria=None,
                            codigo=codigo,
                            pieza_no=1,
                            total_piezas=qty,
                            cantidad_objetivo=qty,
                            es_op=bool(int(qty or 0) >= 2),
                            es_individual=bool(int(qty or 0) < 2),
                            cantidad_producida=0,
                            cantidad_pintada=0,
                            cantidad_terminada=0,
                            nombre=codigo,
                            descripcion=(getattr(prod, "nombre", "") or "").strip(),
                            fecha_compromiso=it.pedido.fecha_compromiso,
                            prioridad=3,
                            peso_kg=float(round(float(kg_total or 0.0), 3)),
                            ultimo_cambio=now,
                            estado_etapa="Espera de corte",
                            observaciones=f"{it.pedido.cliente} {it.pedido.telefono}".strip(),
                        )
                        HerrOrdenItem.objects.create(
                            orden=orden,
                            etapa="Corte",
                            pieza=prod,
                            pieza_custom_nombre="",
                            pieza_custom_peso_kg=0.0,
                            cantidad_requerida=qty,
                        )
                        it.estado_herreria = "En producción"
                        it.orden_herreria = orden
                        it.aceptado_por = actor
                        it.aceptado_en = now
                        it.save(update_fields=["estado_herreria", "orden_herreria", "aceptado_por", "aceptado_en", "actualizado_en"])
                        messages.success(request, "Aceptado y enviado a producción.")
                return redirect("catalogos:herreria_control")
        elif action == "pedido_item_start":
            if not can_accept_solicitudes:
                return redirect("catalogos:herreria_control")
            item_id = int(request.POST.get("pedido_item_id") or 0)
            if item_id:
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                now = timezone.now()
                with transaction.atomic(using="mes"):
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update().select_related("pedido", "producto"), pk=item_id)
                    if it.estado_herreria != "Espera de material":
                        messages.error(request, "El registro ya no está en Espera de Material.")
                    else:
                        qty = max(1, int(it.cantidad_total or 1))
                        prod = it.producto
                        kg_total = float(getattr(prod, "peso_kg", 0.0) or 0.0) * float(qty)
                        codigo = f"{it.pedido.folio}-{int(it.id):03d}"
                        orden = HerrOrdenProduccion.objects.create(
                            proyecto=None,
                            cliente_herreria=None,
                            codigo=codigo,
                            pieza_no=1,
                            total_piezas=qty,
                            cantidad_objetivo=qty,
                            es_op=bool(int(qty or 0) >= 2),
                            es_individual=bool(int(qty or 0) < 2),
                            cantidad_producida=0,
                            cantidad_pintada=0,
                            cantidad_terminada=0,
                            nombre=codigo,
                            descripcion=(getattr(prod, "nombre", "") or "").strip(),
                            fecha_compromiso=it.pedido.fecha_compromiso,
                            prioridad=3,
                            peso_kg=float(round(float(kg_total or 0.0), 3)),
                            ultimo_cambio=now,
                            estado_etapa="Espera de corte",
                            observaciones=f"{it.pedido.cliente} {it.pedido.telefono}".strip(),
                        )
                        HerrOrdenItem.objects.create(
                            orden=orden,
                            etapa="Corte",
                            pieza=prod,
                            pieza_custom_nombre="",
                            pieza_custom_peso_kg=0.0,
                            cantidad_requerida=qty,
                        )
                        it.estado_herreria = "En producción"
                        it.orden_herreria = orden
                        it.espera_material_salida_en = now
                        it.save(update_fields=["estado_herreria", "orden_herreria", "espera_material_salida_en", "actualizado_en"])
                        messages.success(request, "Enviado a producción.")
                return redirect("catalogos:herreria_control")
        elif action == "pedido_item_delete":
            if not is_admin:
                return redirect("catalogos:herreria_control")
            item_id = int(request.POST.get("pedido_item_id") or 0)
            if item_id:
                with transaction.atomic(using="mes"):
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update(), pk=item_id)
                    if it.estado_herreria not in {"Pendiente", "Espera de material"} or it.orden_herreria_id:
                        messages.error(request, "No se puede eliminar: ya está en producción.")
                    else:
                        it.estado_herreria = "Cancelado"
                        it.save(update_fields=["estado_herreria", "actualizado_en"])
                        messages.success(request, "Eliminado.")
                return redirect("catalogos:herreria_control")
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    estados_op = ["Espera de corte", "Corte", "Soldadura"]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
        CIERRE_PENDIENTE_LABEL: "#f1c40f",
        "Enviado": "#11cdef",
    }

    estado = (request.GET.get("estado") or "Todos").strip() or "Todos"
    proyecto = (request.GET.get("proyecto") or "Todos").strip() or "Todos"
    q = (request.GET.get("q") or "").strip()
    order = (request.GET.get("order") or "ultimo_mov_desc").strip() or "ultimo_mov_desc"
    tipo = (request.GET.get("tipo") or "todas").strip().lower() or "todas"
    if tipo not in {"todas", "ventas", "obra"}:
        tipo = "todas"
    filters = {"estado": estado, "proyecto": proyecto, "q": q, "order": order, "tipo": tipo}

    qs = (
        HerrOrdenProduccion.objects.select_related("proyecto", "cliente_herreria")
        .filter(estado="Abierta")
    )
    if estado and estado not in {"Todos", ""}:
        qs = qs.filter(estado_etapa=estado)
    if proyecto and proyecto not in {"Todos", ""}:
        qs = qs.filter(proyecto__nombre=proyecto)
    if q:
        qs = qs.filter(
            Q(codigo__icontains=q)
            | Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(proyecto__nombre__icontains=q)
        )

    order_map = {
        "ultimo_mov_desc": ["-ultimo_cambio", "-id"],
        "ultimo_mov_asc": ["ultimo_cambio", "id"],
        "prioridad_asc": ["prioridad", "-ultimo_cambio", "-id"],
        "prioridad_desc": ["-prioridad", "-ultimo_cambio", "-id"],
        "fecha_asc": ["fecha_compromiso", "-prioridad", "-ultimo_cambio", "-id"],
        "fecha_desc": ["-fecha_compromiso", "-prioridad", "-ultimo_cambio", "-id"],
        "estado": ["estado_etapa", "-ultimo_cambio", "-id"],
        "proyecto": ["proyecto__nombre", "-ultimo_cambio", "-id"],
    }
    qs = qs.order_by(*order_map.get(order, order_map["ultimo_mov_desc"]))

    vigas_all = list(qs[:2000])
    orden_ids_all = [int(v.id) for v in vigas_all if v and getattr(v, "id", None)]
    orden_to_pedido_id: dict[int, int] = {}
    pedido_ids = set()
    if orden_ids_all:
        for oid, pid in (
            PedidoProduccionItem.objects.filter(orden_herreria_id__in=orden_ids_all)
            .exclude(estado_herreria="Cancelado")
            .values_list("orden_herreria_id", "pedido_id")
        ):
            if oid and pid:
                orden_to_pedido_id[int(oid)] = int(pid)
                pedido_ids.add(int(pid))
    items_by_pedido_id: dict[int, list[PedidoProduccionItem]] = {}
    if pedido_ids:
        for it in PedidoProduccionItem.objects.filter(pedido_id__in=list(pedido_ids)).exclude(estado_herreria="Cancelado"):
            items_by_pedido_id.setdefault(int(getattr(it, "pedido_id", 0) or 0), []).append(it)
    pedidos_enviados_ids: set[int] = set()
    if items_by_pedido_id:
        for pid, items in items_by_pedido_id.items():
            if not items:
                continue
            if not _pedido_is_completo(items):
                continue
            apartado_only = all(
                int(getattr(x, "enviado", 0) or 0) == 0
                for x in items
                if int(getattr(x, "apartado", 0) or 0) > 0
            )
            if not apartado_only:
                pedidos_enviados_ids.add(int(pid))
    ordenes_enviadas_ids = {int(oid) for oid, pid in orden_to_pedido_id.items() if int(pid) in pedidos_enviados_ids}
    linked_ids = set(int(k) for k in orden_to_pedido_id.keys())
    obra_enviado_ids = set(
        int(x)
        for x in HerrOrdenProduccion.objects.filter(estado="Abierta", estado_etapa="Enviado")
        .exclude(id__in=list(linked_ids) if linked_ids else [])
        .values_list("id", flat=True)[:5000]
        if x
    )

    excluded_ids = set(int(x) for x in ordenes_enviadas_ids) | set(int(x) for x in obra_enviado_ids)
    vigas = [v for v in vigas_all if int(getattr(v, "id", 0) or 0) not in excluded_ids]
    ids = [int(v.id) for v in vigas]
    for v in vigas:
        v.internal_id = int(v.id)
        v.codigo_viga = v.codigo or v.folio
        v.pieza_no = int(v.pieza_no or 1)
        v.total_piezas = int(v.total_piezas or 1)
        v.es_op = bool(int(v.total_piezas or 0) >= 2)
        objetivo = int(v.total_piezas or 1)
        soldadas = max(0, min(int(getattr(v, "cantidad_producida", 0) or 0), objetivo))
        pintadas = max(0, min(int(getattr(v, "cantidad_pintada", 0) or 0), objetivo))
        terminadas = max(0, min(int(getattr(v, "cantidad_terminada", 0) or 0), objetivo))
        v.op_soldadas = soldadas
        v.op_pintadas = pintadas
        v.op_terminadas = terminadas
        v.op_soldadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (soldadas / objetivo) * 100.0), 2)
        v.op_pintadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (pintadas / objetivo) * 100.0), 2)
        v.op_terminadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (terminadas / objetivo) * 100.0), 2)
        v.avance_pct = v.op_terminadas_pct
        v.cliente_nombre = getattr(getattr(v, "cliente_herreria", None), "nombre", "") or ""
        v.proyecto_nombre = getattr(getattr(v, "proyecto", None), "nombre", "") or v.cliente_nombre
        v.estado = (v.estado_etapa or "").strip()
        v.estado_color = status_colors.get(v.estado, "#8898aa")
        v.op_can_avance = bool(v.es_op) and v.estado in {"Soldadura", CIERRE_PENDIENTE_LABEL}
        v.cierre_pendiente_activo = False
        v.cierre_pendiente_hasta_str = ""
        if v.estado == CIERRE_PENDIENTE_LABEL and getattr(v, "cierre_pendiente_hasta", None):
            hasta = timezone.localtime(getattr(v, "cierre_pendiente_hasta"))
            v.cierre_pendiente_hasta_str = hasta.strftime("%Y-%m-%d %H:%M")
            v.cierre_pendiente_activo = bool(hasta > timezone.localtime(timezone.now()))
        v.plano_url = v.plano_pdf.url if getattr(v, "plano_pdf", None) else ""
        v.ultimo_mov_str = timezone.localtime(v.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if getattr(v, "ultimo_cambio", None) else ""
        estados_for_next = estados_op if v.es_op else estados
        try:
            idx = estados_for_next.index(v.estado)
        except Exception:
            idx = -1
        v.next_estado = "" if idx < 0 or idx >= (len(estados_for_next) - 1) else estados_for_next[idx + 1]
        v.show_badge = ""
        if v.es_op and v.estado == "Soldadura":
            v.next_estado = ""
            v.show_badge = "Avance habilitado"

    asigns = (
        HerrAsignacion.objects.filter(orden_id__in=ids, vigente=True, etapa__in=["Corte", "Soldadura", "Pintura"])
        .values_list("orden_id", "etapa", "rol", "colaborador_id", "maquina_id")
    )
    asign_map = {}
    for oid, etapa, rol, cid, mid in asigns:
        entry = asign_map.setdefault(int(oid), {}).setdefault(str(etapa), {"ids": [], "operadores": [], "maquinas": []})
        if mid:
            entry["maquinas"].append(int(mid))
        if cid:
            if (rol or "") == "Operador":
                entry["operadores"].append(int(cid))
            else:
                entry["ids"].append(int(cid))
    for v in vigas:
        m = asign_map.get(int(v.id), {})
        v.asig_corte = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("ids", []))
        v.asig_corte_maquinas = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("maquinas", []))
        v.asig_corte_operadores = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("operadores", []))
        v.asig_soldadura = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("ids", []))
        v.asig_pintura = ",".join(str(x) for x in (m.get("Pintura", {}) or {}).get("ids", []))
        v.asig_soldadura_maquinas = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("maquinas", []))
        v.asig_soldadura_operadores = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("operadores", []))

    maquina_ids = set()
    for v in vigas:
        for part in (getattr(v, "asig_corte_maquinas", "") or "").split(","):
            part = (part or "").strip()
            if part.isdigit():
                maquina_ids.add(int(part))
    maquinas_info = {}
    maquinas_estado = {}
    if maquina_ids:
        for m in Maquina.objects.filter(id__in=list(maquina_ids)).values("id", "nombre", "tipo", "es_robot"):
            maquinas_info[int(m["id"])] = {"nombre": m["nombre"], "tipo": m["tipo"], "es_robot": bool(m["es_robot"])}
        for p in MaquinaParo.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("motivo"):
            maquinas_estado[int(p.maquina_id)] = {
                "paro": True,
                "falla": False,
                "motivo": getattr(getattr(p, "motivo", None), "nombre", "") or "",
                "paro_inicio": timezone.localtime(p.inicio).strftime("%Y-%m-%d %H:%M") if getattr(p, "inicio", None) else "",
            }
        for f in MaquinaFalla.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("tipo"):
            entry = maquinas_estado.setdefault(int(f.maquina_id), {"paro": False, "falla": False, "motivo": "", "paro_inicio": ""})
            entry["falla"] = True
            entry["tipo_falla"] = getattr(getattr(f, "tipo", None), "nombre", "") or ""
            entry["falla_inicio"] = timezone.localtime(f.inicio).strftime("%Y-%m-%d %H:%M") if getattr(f, "inicio", None) else ""

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    participantes_payload = _build_participantes_payload()

    hoy = timezone.localdate()
    cutoff = hoy - timedelta(days=DECOTE_DAYS)

    enviados_recent_sales: list[HerrOrdenProduccion] = []
    decote_vigas_sales: list[HerrOrdenProduccion] = []
    if pedidos_enviados_ids and ordenes_enviadas_ids:
        last_map = {
            int(r["pedido_id"]): r["last_fecha"]
            for r in LogisticaEnvio.objects.filter(pedido_id__in=list(pedidos_enviados_ids))
            .values("pedido_id")
            .annotate(last_fecha=Max("fecha"))
        }
        exp_map = {int(e.pedido_id): e for e in LogisticaExpediente.objects.filter(pedido_id__in=list(pedidos_enviados_ids))}
        env_qs = HerrOrdenProduccion.objects.select_related("proyecto").filter(
            estado="Abierta",
            id__in=list(ordenes_enviadas_ids),
        )
        if proyecto and proyecto not in {"Todos", ""}:
            env_qs = env_qs.filter(proyecto__nombre=proyecto)
        if q:
            env_qs = env_qs.filter(
                Q(codigo__icontains=q)
                | Q(nombre__icontains=q)
                | Q(descripcion__icontains=q)
                | Q(proyecto__nombre__icontains=q)
            )
        for v in list(env_qs[:2000]):
            pid = int(orden_to_pedido_id.get(int(getattr(v, "id", 0) or 0), 0) or 0)
            last_fecha = last_map.get(int(pid))
            v.enviado_fecha = last_fecha
            v.enviado_dias = int((hoy - last_fecha).days) if last_fecha else 0
            exp = exp_map.get(int(pid))
            v.expediente_generado = bool(exp and getattr(exp, "generado_en", None))
            v.expediente_descargado = bool(exp and int(getattr(exp, "descargas_count", 0) or 0) > 0)
            v.__dict__["expediente_descargas_count"] = int(getattr(exp, "descargas_count", 0) or 0) if exp else 0
            can_decote = bool(v.expediente_generado and v.expediente_descargado)
            is_decote_age = bool(last_fecha and (hoy - last_fecha).days >= DECOTE_DAYS)
            v.can_delete_decote = bool(last_fecha and (hoy - last_fecha).days >= 5)
            if is_decote_age and can_decote:
                decote_vigas_sales.append(v)
            else:
                enviados_recent_sales.append(v)
        enviados_recent_sales.sort(key=lambda x: (getattr(x, "enviado_fecha", None) or datetime.min.date()), reverse=True)
        decote_vigas_sales.sort(key=lambda x: (getattr(x, "enviado_fecha", None) or datetime.max.date()))

    enviados_recent_obra: list[HerrOrdenProduccion] = []
    decote_vigas_obra: list[HerrOrdenProduccion] = []
    linked_ids = set(int(k) for k in orden_to_pedido_id.keys())
    obra_qs = HerrOrdenProduccion.objects.select_related("proyecto").filter(estado="Abierta", estado_etapa="Enviado")
    if linked_ids:
        obra_qs = obra_qs.exclude(id__in=list(linked_ids))
    if proyecto and proyecto not in {"Todos", ""}:
        obra_qs = obra_qs.filter(proyecto__nombre=proyecto)
    if q:
        obra_qs = obra_qs.filter(
            Q(codigo__icontains=q)
            | Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(proyecto__nombre__icontains=q)
        )
    for v in list(obra_qs.order_by("-ultimo_cambio", "-id")[:2000]):
        enviado_dt = getattr(v, "ultimo_cambio", None)
        enviado_fecha = timezone.localdate(enviado_dt) if enviado_dt else hoy
        v.enviado_fecha = enviado_fecha
        v.enviado_dias = int((hoy - enviado_fecha).days) if enviado_fecha else 0
        v.can_delete_decote = bool(enviado_fecha and (hoy - enviado_fecha).days >= 5)
        if v.enviado_dias >= DECOTE_DAYS:
            decote_vigas_obra.append(v)
        else:
            enviados_recent_obra.append(v)

    def _decorate_enviado(items):
        for v in items:
            v.internal_id = int(v.id)
            v.codigo_viga = v.codigo or v.folio
            v.pieza_no = int(getattr(v, "pieza_no", 1) or 1)
            v.total_piezas = int(getattr(v, "total_piezas", 1) or 1)
            v.proyecto_nombre = getattr(getattr(v, "proyecto", None), "nombre", "") or ""
            v.estado = "Enviado"
            v.estado_color = status_colors.get("Enviado", "#11cdef")
            v.enviado_fecha = getattr(v, "enviado_fecha", None)
            v.enviado_dias = int((hoy - v.enviado_fecha).days) if v.enviado_fecha else 0
            v.enviado_str = v.enviado_fecha.strftime("%Y-%m-%d") if v.enviado_fecha else ""

    _decorate_enviado(enviados_recent_sales)
    _decorate_enviado(decote_vigas_sales)
    _decorate_enviado(enviados_recent_obra)
    _decorate_enviado(decote_vigas_obra)

    if tipo == "ventas":
        enviados_recent = []
        decote_vigas = []
    elif tipo == "obra":
        enviados_recent = list(enviados_recent_obra)
        decote_vigas = list(decote_vigas_obra)
    else:
        enviados_recent = list(enviados_recent_sales) + list(enviados_recent_obra)
        decote_vigas = list(decote_vigas_sales) + list(decote_vigas_obra)

    enviados_total = len(enviados_recent)
    decote_total = len(decote_vigas)
    enviados_by_project: dict[str, list[HerrOrdenProduccion]] = {}
    for v in enviados_recent:
        key = getattr(v, "proyecto_nombre", "") or "-"
        enviados_by_project.setdefault(str(key), []).append(v)
    enviados_por_proyecto = sorted(enviados_by_project.items(), key=lambda t: str(t[0] or "").lower())
    pedidos_pendientes = list(
        PedidoProduccionItem.objects.select_related("pedido", "producto")
        .filter(estado_herreria="Pendiente", pedido__estado="Activa")
        .order_by("-creado_en")[:500]
    )
    espera_material = list(
        PedidoProduccionItem.objects.select_related("pedido", "producto")
        .filter(estado_herreria="Espera de material", pedido__estado="Activa")
        .order_by("espera_material_entrada_en", "creado_en")[:500]
    )
    vigas_op = [v for v in vigas if bool(getattr(v, "es_op", False))]
    op_ids = [int(v.id) for v in vigas_op if v and getattr(v, "id", None)]
    ventas_op_ids = set()
    if op_ids:
        ventas_op_ids = set(
            int(x)
            for x in PedidoProduccionItem.objects.filter(orden_herreria_id__in=op_ids).values_list("orden_herreria_id", flat=True)
            if x
        )
    vigas_op_ventas = [v for v in vigas_op if int(getattr(v, "id", 0) or 0) in ventas_op_ids]
    vigas_op = [v for v in vigas_op if int(getattr(v, "id", 0) or 0) not in ventas_op_ids]
    ventas_items_map = {}
    if ventas_op_ids:
        ventas_items = list(
            PedidoProduccionItem.objects.select_related("pedido", "producto")
            .filter(orden_herreria_id__in=list(ventas_op_ids))
            .exclude(estado_herreria="Cancelado")
        )
        for it in ventas_items:
            ventas_items_map[int(getattr(it, "orden_herreria_id", 0) or 0)] = it
    acuse_sum_map = {}
    ventas_item_ids = [int(getattr(ventas_items_map.get(int(getattr(v, "id", 0) or 0)), "id", 0) or 0) for v in vigas_op_ventas]
    ventas_item_ids = [x for x in ventas_item_ids if x > 0]
    if ventas_item_ids:
        acuse_sum_map = {
            int(r["pedido_item_id"]): int(r["total"] or 0)
            for r in LogisticaAcuseEntrega.objects.filter(pedido_item_id__in=ventas_item_ids)
            .values("pedido_item_id")
            .annotate(total=Sum("cantidad"))
        }
    for v in vigas_op_ventas:
        oid = int(getattr(v, "id", 0) or 0)
        it = ventas_items_map.get(oid)
        v.pedido_item_id = int(getattr(it, "id", 0) or 0) if it else 0
        v.pedido_folio = getattr(getattr(it, "pedido", None), "folio", "") if it else ""
        v.pedido_cliente = getattr(getattr(it, "pedido", None), "cliente", "") if it else ""
        v.producto_nombre = getattr(getattr(it, "producto", None), "nombre", "") if it else ""
        v.total_linea = int(getattr(it, "cantidad_total", 0) or 0) if it else 0
        v.entregado_logistica = int(acuse_sum_map.get(int(v.pedido_item_id), 0) or 0)
        terminadas = int(getattr(v, "op_terminadas", 0) or 0)
        v.disponible_entrega = max(0, terminadas - int(v.entregado_logistica or 0))
        v.acuse_enabled = bool(v.pedido_item_id and v.disponible_entrega > 0)
    vigas_individual = [v for v in vigas if not bool(getattr(v, "es_op", False))]
    if tipo == "ventas":
        vigas_op = []
        vigas_individual = []
    elif tipo == "obra":
        vigas_op_ventas = []

    base_url = reverse("catalogos:herreria_control")
    qd = request.GET.copy()
    qd["tipo"] = "todas"
    tab_todas_url = base_url + (f"?{qd.urlencode()}" if qd else "")
    qd["tipo"] = "ventas"
    tab_ventas_url = base_url + (f"?{qd.urlencode()}" if qd else "")
    qd["tipo"] = "obra"
    tab_obra_url = base_url + (f"?{qd.urlencode()}" if qd else "")
    tipo_tabs = [
        {"key": "ventas", "label": "Órdenes de ventas", "url": tab_ventas_url},
        {"key": "obra", "label": "Obra", "url": tab_obra_url},
        {"key": "todas", "label": "Todas", "url": tab_todas_url},
    ]
    return render(
        request,
        "catalogos/herreria_list.html",
        {
            "vigas": vigas,
            "vigas_op": vigas_op,
            "vigas_op_ventas": vigas_op_ventas,
            "vigas_individual": vigas_individual,
            "mode": "admin",
            "reset_url": reverse("catalogos:herreria_control"),
            "estados_filtro": ["Todos", *estados],
            "estados_operables": estados,
            "estados_order": estados,
            "proyectos": ["Todos", *projects],
            "filters": filters,
            "tipo_tabs": tipo_tabs,
            "status_colors": status_colors,
            "participantes_payload": participantes_payload,
            "maquinas_info_payload": maquinas_info,
            "maquinas_estado_payload": maquinas_estado,
            "decote_days": DECOTE_DAYS,
            "enviados_total": enviados_total,
            "enviados_por_proyecto": enviados_por_proyecto,
            "decote_total": decote_total,
            "decote_vigas": decote_vigas,
            "is_admin": is_admin,
            "can_accept_solicitudes": can_accept_solicitudes,
            "pedidos_pendientes": pedidos_pendientes,
            "espera_material": espera_material,
        },
    )


class HerrVigaLikeForm(forms.Form):
    codigo_viga = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    proyecto = forms.ModelChoiceField(queryset=Proyecto.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    descripcion = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza = forms.ModelChoiceField(queryset=HerrPiezaCatalogo.objects.none(), required=False, widget=forms.Select(attrs={"class": "form-select"}))
    fecha_compromiso = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    estado = forms.ChoiceField(choices=[], widget=forms.Select(attrs={"class": "form-select"}))
    prioridad = forms.ChoiceField(
        choices=[("1", "1"), ("2", "2"), ("3", "3"), ("4", "4"), ("5", "5")],
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    peso_kg = forms.FloatField(required=False, widget=forms.NumberInput(attrs={"class": "form-control"}))
    total_piezas = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    observaciones = forms.CharField(required=False, widget=forms.Textarea(attrs={"class": "form-control", "rows": 4}))

    def __init__(self, *args, estados: list[str], **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["estado"].choices = [(s, s) for s in estados]
        self.fields["proyecto"].queryset = Proyecto.objects.filter(activo=True).order_by("nombre")
        self.fields["pieza"].queryset = HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre")

    def clean_codigo_viga(self):
        s = (self.cleaned_data.get("codigo_viga") or "").strip()
        if not s:
            raise forms.ValidationError("Código requerido.")
        return s

    def clean(self):
        cleaned = super().clean()
        pieza = cleaned.get("pieza")
        kg = cleaned.get("peso_kg")
        if not pieza:
            try:
                v = float(kg or 0.0)
            except Exception:
                v = 0.0
            if v <= 0:
                self.add_error("peso_kg", "KG requerido (o selecciona pieza de catálogo).")
        return cleaned


class LaserVigaLikeForm(forms.Form):
    folio_externo = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    pieza = forms.ModelChoiceField(queryset=CortaPiezaCatalogo.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    cliente_proyecto = forms.CharField(widget=forms.TextInput(attrs={"class": "form-control"}))
    correo = forms.CharField(required=False, widget=forms.EmailInput(attrs={"class": "form-control"}))
    telefono = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    descripcion = forms.CharField(required=False, widget=forms.TextInput(attrs={"class": "form-control"}))
    material = forms.ModelChoiceField(queryset=LaserMaterialPlaca.objects.none(), widget=forms.Select(attrs={"class": "form-select"}))
    pieza_ancho_mm = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    pieza_alto_mm = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    fecha_compromiso = forms.DateField(widget=forms.DateInput(attrs={"type": "date", "class": "form-control"}))
    estado = forms.ChoiceField(choices=[], widget=forms.Select(attrs={"class": "form-select"}))
    prioridad = forms.ChoiceField(
        choices=[("1", "1"), ("2", "2"), ("3", "3"), ("4", "4"), ("5", "5")],
        widget=forms.Select(attrs={"class": "form-select"}),
    )
    total_piezas = forms.IntegerField(min_value=1, widget=forms.NumberInput(attrs={"class": "form-control"}))
    observaciones = forms.CharField(required=False, widget=forms.Textarea(attrs={"class": "form-control", "rows": 4}))

    def __init__(self, *args, estados: list[str], **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["estado"].choices = [(s, s) for s in estados]
        self.fields["pieza"].queryset = CortaPiezaCatalogo.objects.filter(activo=True).order_by("nombre")
        materiales_qs = LaserMaterialPlaca.objects.filter(activo=True).order_by(
            "categoria_material",
            "tipo_material",
            "nombre",
            "calibre",
            "espesor_mm",
            "largo_mm",
            "ancho_mm",
            "id",
        )
        materiales = list(materiales_qs)
        self.fields["material"].queryset = materiales_qs
        name_key_counts: dict[tuple[str, str, str], int] = {}
        for m in materiales:
            k = (
                (getattr(m, "categoria_material", "") or "").strip().upper(),
                (getattr(m, "tipo_material", "") or "").strip().upper(),
                (getattr(m, "nombre", "") or "").strip().upper(),
            )
            name_key_counts[k] = name_key_counts.get(k, 0) + 1
        name_key_seen: dict[tuple[str, str, str], int] = {}
        material_pos: dict[int, int] = {}
        for m in materiales:
            k = (
                (getattr(m, "categoria_material", "") or "").strip().upper(),
                (getattr(m, "tipo_material", "") or "").strip().upper(),
                (getattr(m, "nombre", "") or "").strip().upper(),
            )
            name_key_seen[k] = name_key_seen.get(k, 0) + 1
            material_pos[int(m.id)] = int(name_key_seen[k])

        def _material_label(obj: LaserMaterialPlaca) -> str:
            categoria = (getattr(obj, "categoria_material", "") or "").strip()
            tipo = (getattr(obj, "tipo_material", "") or "").strip()
            nombre = (getattr(obj, "nombre", "") or "").strip()
            calibre = (getattr(obj, "calibre", "") or "").strip()
            espesor = float(getattr(obj, "espesor_mm", 0.0) or 0.0)
            largo = int(getattr(obj, "largo_mm", 0) or 0)
            ancho = int(getattr(obj, "ancho_mm", 0) or 0)
            parts: list[str] = []
            if categoria:
                parts.append(categoria)
            if tipo:
                parts.append(tipo)
            parts.append(nombre if nombre else f"Material #{int(obj.id)}")
            if calibre:
                parts.append(f"Cédula {calibre}")
            if espesor > 0:
                parts.append(f"{espesor:.2f}mm")
            if largo > 0 and ancho > 0:
                parts.append(f"{(largo / 10.0):.1f}x{(ancho / 10.0):.1f}cm")
            try:
                peso = float(getattr(obj, "peso_kg", 0.0) or 0.0)
            except Exception:
                peso = 0.0
            if peso > 0:
                parts.append(f"{peso:.3f}kg")
            key = ((categoria or "").strip().upper(), (tipo or "").strip().upper(), (nombre or "").strip().upper())
            if name_key_counts.get(key, 0) > 1:
                parts.append(f"(#{material_pos.get(int(obj.id), 1)})")
            return " · ".join(parts)

        self.fields["material"].label_from_instance = _material_label

    def clean_folio_externo(self):
        s = (self.cleaned_data.get("folio_externo") or "").strip()
        if not s:
            raise forms.ValidationError("Folio requerido.")
        return s

    def clean(self):
        cleaned = super().clean()
        cp = _norm_text(cleaned.get("cliente_proyecto") or "")
        pieza = cleaned.get("pieza")
        material = cleaned.get("material")
        ancho = int(cleaned.get("pieza_ancho_mm") or 0)
        alto = int(cleaned.get("pieza_alto_mm") or 0)
        if not cp:
            self.add_error("cliente_proyecto", "Cliente o proyecto requerido.")
        if not pieza:
            self.add_error("pieza", "Pieza requerida.")
        if not material:
            self.add_error("material", "Material requerido.")
        if ancho <= 0:
            self.add_error("pieza_ancho_mm", "Ancho requerido.")
        if alto <= 0:
            self.add_error("pieza_alto_mm", "Largo requerido.")
        return cleaned


@login_required
@require_http_methods(["GET", "POST"])
def herreria_create(request):
    if not _can_herreria(request.user):
        return redirect("/")

    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }

    corte_operadores = list(Colaborador.objects.filter(activo=True, rol="Operador").order_by("nombre"))
    corte_maquinas = list(Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre"))
    selected_corte_operadores = [int(x) for x in request.POST.getlist("corte_operador_ids") if (x or "").isdigit()]
    selected_corte_maquinas = [int(x) for x in request.POST.getlist("corte_maquina_ids") if (x or "").isdigit()]

    piezas_catalogo = list(HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre"))
    piezas_catalogo_weight = {int(p.id): float(p.peso_kg or 0.0) for p in piezas_catalogo}
    piezas_catalogo_name = {int(p.id): (p.nombre or "").strip() for p in piezas_catalogo}

    if request.method == "POST":
        form = HerrVigaLikeForm(request.POST, estados=estados)
        if form.is_valid():
            codigo = form.cleaned_data["codigo_viga"]
            if HerrOrdenProduccion.objects.filter(codigo_normalizado=codigo.upper()).exists():
                form.add_error("codigo_viga", "Código ya existe.")
            else:
                total = int(form.cleaned_data["total_piezas"] or 1)
                pieza_obj = form.cleaned_data.get("pieza")
                if pieza_obj:
                    kg_pieza = float(getattr(pieza_obj, "peso_kg", 0.0) or 0.0)
                    pieza_nombre = (getattr(pieza_obj, "nombre", "") or "").strip()
                else:
                    kg_pieza = float(form.cleaned_data.get("peso_kg") or 0.0)
                    pieza_nombre = ""
                descripcion = (form.cleaned_data.get("descripcion") or "").strip()
                if not descripcion:
                    descripcion = pieza_nombre or descripcion
                orden = HerrOrdenProduccion.objects.create(
                    proyecto=form.cleaned_data["proyecto"],
                    codigo=codigo,
                    pieza_no=1,
                    total_piezas=total,
                    cantidad_objetivo=total,
                    es_op=bool(int(total or 0) >= 2),
                    cantidad_producida=0,
                    nombre=codigo,
                    descripcion=descripcion,
                    fecha_compromiso=form.cleaned_data["fecha_compromiso"],
                    prioridad=int(form.cleaned_data["prioridad"] or 3),
                    estado_etapa=form.cleaned_data["estado"],
                    peso_kg=float(kg_pieza) * float(total),
                    ultimo_cambio=timezone.now(),
                    observaciones=(form.cleaned_data.get("observaciones") or "").strip(),
                    plano_pdf=request.FILES.get("plano_pdf"),
                )
                HerrOrdenItem.objects.create(
                    orden=orden,
                    etapa="Corte",
                    pieza=pieza_obj if pieza_obj else None,
                    pieza_custom_nombre=("" if pieza_obj else (orden.descripcion or orden.nombre or "").strip()),
                    pieza_custom_peso_kg=(0.0 if pieza_obj else float(kg_pieza or 0.0)),
                    cantidad_requerida=int(total),
                )
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                HerrAsignacion.objects.filter(orden=orden, etapa="Corte", vigente=True).update(vigente=False)
                for cid in selected_corte_operadores:
                    HerrAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Operador",
                        colaborador_id=int(cid),
                        vigente=True,
                        asignado_por=actor,
                    )
                for mid in selected_corte_maquinas:
                    HerrAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Máquina",
                        maquina_id=int(mid),
                        vigente=True,
                        asignado_por=actor,
                    )
                messages.success(request, "Pieza creada.")
                return redirect("catalogos:herreria_control")
    else:
        form = HerrVigaLikeForm(
            initial={"fecha_compromiso": timezone.localdate(), "estado": "Espera de corte", "prioridad": "3", "total_piezas": 1},
            estados=estados,
        )

    return render(
        request,
        "catalogos/herreria_form.html",
        {
            "form": form,
            "mode": "admin",
            "reset_url": reverse("catalogos:herreria_control"),
            "status_colors": status_colors,
            "plano_url": "",
            "corte_operadores": corte_operadores,
            "corte_maquinas": corte_maquinas,
            "selected_corte_operadores": selected_corte_operadores,
            "selected_corte_maquinas": selected_corte_maquinas,
            "piezas_catalogo": piezas_catalogo,
            "piezas_catalogo_weight": piezas_catalogo_weight,
            "piezas_catalogo_name": piezas_catalogo_name,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria_update(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }

    corte_operadores = list(Colaborador.objects.filter(activo=True, rol="Operador").order_by("nombre"))
    corte_maquinas = list(Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre"))
    current_asigs = list(
        HerrAsignacion.objects.filter(orden=orden, vigente=True, etapa="Corte").values_list("rol", "colaborador_id", "maquina_id")
    )
    selected_corte_operadores = [int(cid) for rol, cid, _mid in current_asigs if (rol or "") == "Operador" and cid]
    selected_corte_maquinas = [int(mid) for rol, _cid, mid in current_asigs if (rol or "") == "Máquina" and mid]

    kg_pieza = 0.0
    if int(orden.total_piezas or 1) > 0:
        kg_pieza = float(orden.peso_kg or 0.0) / float(int(orden.total_piezas or 1))

    piezas_catalogo = list(HerrPiezaCatalogo.objects.filter(activo=True).order_by("nombre"))
    piezas_catalogo_weight = {int(p.id): float(p.peso_kg or 0.0) for p in piezas_catalogo}
    piezas_catalogo_name = {int(p.id): (p.nombre or "").strip() for p in piezas_catalogo}

    if request.method == "POST":
        form = HerrVigaLikeForm(request.POST, estados=estados)
        if form.is_valid():
            codigo = form.cleaned_data["codigo_viga"]
            if HerrOrdenProduccion.objects.exclude(id=orden.id).filter(codigo_normalizado=codigo.upper()).exists():
                form.add_error("codigo_viga", "Código ya existe.")
            else:
                total = int(form.cleaned_data["total_piezas"] or 1)
                pieza_obj = form.cleaned_data.get("pieza")
                if pieza_obj:
                    kg_pieza = float(getattr(pieza_obj, "peso_kg", 0.0) or 0.0)
                    pieza_nombre = (getattr(pieza_obj, "nombre", "") or "").strip()
                else:
                    kg_pieza = float(form.cleaned_data.get("peso_kg") or 0.0)
                    pieza_nombre = ""
                orden.proyecto = form.cleaned_data["proyecto"]
                orden.codigo = codigo
                orden.pieza_no = 1
                orden.total_piezas = total
                orden.cantidad_objetivo = total
                orden.es_op = bool(int(total or 0) >= 2)
                orden.nombre = codigo
                descripcion = (form.cleaned_data.get("descripcion") or "").strip()
                if not descripcion:
                    descripcion = pieza_nombre or descripcion
                orden.descripcion = descripcion
                orden.fecha_compromiso = form.cleaned_data["fecha_compromiso"]
                orden.prioridad = int(form.cleaned_data["prioridad"] or 3)
                orden.estado_etapa = form.cleaned_data["estado"]
                orden.peso_kg = float(kg_pieza) * float(total)
                orden.observaciones = (form.cleaned_data.get("observaciones") or "").strip()
                plano = request.FILES.get("plano_pdf")
                if plano:
                    orden.plano_pdf = plano
                orden.save()
                it = HerrOrdenItem.objects.filter(orden=orden).order_by("id").first()
                if it:
                    it.pieza = pieza_obj if pieza_obj else None
                    it.pieza_custom_nombre = ("" if pieza_obj else (orden.descripcion or orden.nombre or "").strip())
                    it.pieza_custom_peso_kg = (0.0 if pieza_obj else float(kg_pieza or 0.0))
                    it.cantidad_requerida = int(total)
                    it.save()
                messages.success(request, "Pieza actualizada.")
                return redirect("catalogos:herreria_control")
    else:
        it = HerrOrdenItem.objects.filter(orden=orden).order_by("id").first()
        pieza_id = it.pieza_id if it else None
        form = HerrVigaLikeForm(
            initial={
                "codigo_viga": orden.codigo,
                "proyecto": orden.proyecto_id,
                "descripcion": orden.descripcion,
                "pieza": pieza_id,
                "fecha_compromiso": orden.fecha_compromiso,
                "estado": orden.estado_etapa,
                "prioridad": str(int(orden.prioridad or 3)),
                "peso_kg": round(float(kg_pieza or 0.0), 3),
                "total_piezas": int(orden.total_piezas or 1),
                "observaciones": orden.observaciones,
            },
            estados=estados,
        )

    plano_url = orden.plano_pdf.url if getattr(orden, "plano_pdf", None) else ""
    return render(
        request,
        "catalogos/herreria_form.html",
        {
            "form": form,
            "mode": "admin",
            "reset_url": reverse("catalogos:herreria_control"),
            "status_colors": status_colors,
            "plano_url": plano_url,
            "corte_operadores": corte_operadores,
            "corte_maquinas": corte_maquinas,
            "selected_corte_operadores": selected_corte_operadores,
            "selected_corte_maquinas": selected_corte_maquinas,
            "piezas_catalogo": piezas_catalogo,
            "piezas_catalogo_weight": piezas_catalogo_weight,
            "piezas_catalogo_name": piezas_catalogo_name,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria_delete(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    if request.method == "POST":
        HerrProduccion.objects.filter(orden_item__orden=orden).delete()
        HerrAsignacion.objects.filter(orden=orden).delete()
        orden.delete()
        messages.success(request, "Eliminado.")
        return redirect("catalogos:herreria_control")
    return render(request, "catalogos/herreria_confirm_delete.html", {"orden": orden})


@login_required
@require_http_methods(["POST"])
def herreria_update_meta_json(request, pk: int):
    if not _can_herreria(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    fecha = (request.POST.get("fecha_compromiso") or "").strip()
    prioridad = (request.POST.get("prioridad") or "").strip()
    try:
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Fecha inválida."}, status=400)
    try:
        pr = int(prioridad or 0)
    except Exception:
        pr = 0
    if pr not in {1, 2, 3, 4, 5}:
        return JsonResponse({"ok": False, "error": "Prioridad inválida."}, status=400)
    orden.fecha_compromiso = fecha_dt
    orden.prioridad = pr
    orden.save(update_fields=["fecha_compromiso", "prioridad", "actualizado_en"])
    return JsonResponse({"ok": True, "id": int(orden.id), "fecha_compromiso": fecha_dt.isoformat(), "prioridad": pr})


@login_required
@require_http_methods(["POST"])
def herreria_update_avance_json(request, pk: int):
    if not _can_herreria(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    raw_sold = (request.POST.get("cantidad_soldada") or request.POST.get("cantidad_producida") or "").strip()
    raw_pint = (request.POST.get("cantidad_pintada") or "").strip()
    raw_term = (request.POST.get("cantidad_terminada") or "").strip()
    recibido_por = (request.POST.get("recibido_por") or "").strip()
    pedido_item_id = 0
    try:
        pedido_item_id = int((request.POST.get("pedido_item_id") or "").strip() or 0)
    except Exception:
        pedido_item_id = 0
    try:
        sold = int(raw_sold or 0)
    except Exception:
        sold = -1
    try:
        pint = int(raw_pint or 0)
    except Exception:
        pint = -1
    try:
        term = int(raw_term or 0)
    except Exception:
        term = -1

    with transaction.atomic(using="mes"):
        orden = get_object_or_404(HerrOrdenProduccion.objects.select_for_update(), pk=pk)
        if orden.estado != "Abierta":
            return JsonResponse({"ok": False, "error": "La orden está cerrada/cancelada."}, status=409)
        if int(getattr(orden, "total_piezas", 0) or 0) < 2:
            return JsonResponse({"ok": False, "error": "Este registro no es una orden grande."}, status=409)
        now = timezone.now()
        etapa = (getattr(orden, "estado_etapa", "") or "").strip()
        if etapa == CIERRE_PENDIENTE_LABEL:
            hasta = getattr(orden, "cierre_pendiente_hasta", None)
            if not hasta or hasta <= now:
                orden.estado_etapa = CIERRE_FINAL_LABEL
                orden.ultimo_cambio = now
                orden.cierre_bloqueado_en = now
                orden.cierre_pendiente_hasta = None
                orden.save(
                    update_fields=["estado_etapa", "ultimo_cambio", "cierre_bloqueado_en", "cierre_pendiente_hasta", "actualizado_en"]
                )
                HerrEstadoCambio.objects.create(
                    orden=orden,
                    estado_anterior=CIERRE_PENDIENTE_LABEL,
                    estado_nuevo=CIERRE_FINAL_LABEL,
                    fecha_operacion=timezone.localdate(),
                    actor_username="system",
                    comentario="auto_bloqueo",
                )
                return JsonResponse({"ok": False, "error": "Orden bloqueada: cierre confirmado."}, status=409)
        if etapa not in {"Soldadura", CIERRE_PENDIENTE_LABEL}:
            return JsonResponse({"ok": False, "error": "El avance se habilita a partir de Soldadura."}, status=409)

        objetivo = int(getattr(orden, "total_piezas", 0) or 0)
        if objetivo <= 0:
            return JsonResponse({"ok": False, "error": "Cantidad objetivo inválida."}, status=400)
        if sold < 0 or pint < 0 or term < 0:
            return JsonResponse({"ok": False, "error": "Cantidades inválidas."}, status=400)
        if sold > objetivo or pint > objetivo or term > objetivo:
            return JsonResponse({"ok": False, "error": "Cantidades inválidas."}, status=400)

        prev_sold = int(getattr(orden, "cantidad_producida", 0) or 0)
        prev_pint = int(getattr(orden, "cantidad_pintada", 0) or 0)
        prev_term = int(getattr(orden, "cantidad_terminada", 0) or 0)
        delta_term = int(term) - int(prev_term)
        pedido_item = None
        is_venta = False
        if delta_term != 0:
            if pedido_item_id > 0:
                pedido_item = (
                    PedidoProduccionItem.objects.using("mes")
                    .select_for_update()
                    .filter(id=int(pedido_item_id), orden_herreria_id=int(orden.id))
                    .order_by("-id")
                    .first()
                )
            if not pedido_item:
                pedido_item = (
                    PedidoProduccionItem.objects.using("mes")
                    .select_for_update()
                    .filter(orden_herreria_id=int(orden.id))
                    .order_by("-id")
                    .first()
                )
            is_venta = bool(pedido_item)
            if delta_term > 0 and is_venta and not recibido_por:
                transaction.set_rollback(True, using="mes")
                return JsonResponse({"ok": False, "error": "Recibe (Logística) requerido para generar acuse."}, status=400)
        orden.cantidad_producida = sold
        orden.cantidad_pintada = pint
        orden.cantidad_terminada = term
        orden.ultimo_cambio = now
        orden.save(update_fields=["cantidad_producida", "cantidad_pintada", "cantidad_terminada", "ultimo_cambio", "actualizado_en"])

        if sold != prev_sold or pint != prev_pint or term != prev_term:
            HerrAvanceCambio.objects.create(
                orden=orden,
                fecha_operacion=timezone.localdate(),
                soldadas_prev=max(0, int(prev_sold)),
                soldadas_new=max(0, int(sold)),
                pintadas_prev=max(0, int(prev_pint)),
                pintadas_new=max(0, int(pint)),
                terminadas_prev=max(0, int(prev_term)),
                terminadas_new=max(0, int(term)),
                actor_username=actor,
            )

        acuse_url = ""
        if delta_term != 0 and is_venta:
            it = HerrOrdenItem.objects.filter(orden=orden, pieza__isnull=False).select_related("pieza").order_by("id").first()
            if it and getattr(it, "pieza", None):
                stock = _get_or_create_stock_locked(it.pieza)
                new_stock = int(stock.stock or 0) + int(delta_term)
                stock.stock = max(0, new_stock)
                stock.save(update_fields=["stock", "actualizado_en"])
                LogisticaMovimiento.objects.create(
                    tipo="stock_in" if delta_term > 0 else "ajuste",
                    producto=it.pieza,
                    pedido_item=pedido_item if pedido_item else None,
                    cantidad=int(delta_term),
                    actor=actor,
                )
                if delta_term > 0 and pedido_item:
                    acuse = LogisticaAcuseEntrega.objects.using("mes").create(
                        pedido_item=pedido_item,
                        cantidad=int(delta_term),
                        fecha=timezone.localdate(),
                        entregado_por=(actor or "").strip(),
                        recibido_por=recibido_por,
                    )
                    acuse_url = reverse("catalogos:herreria_acuse_print", kwargs={"pk": int(acuse.id)})

        reload_ui = False
        if int(term) == int(objetivo) and (getattr(orden, "estado_etapa", "") or "").strip() not in {CIERRE_PENDIENTE_LABEL, CIERRE_FINAL_LABEL}:
            prev_etapa = (getattr(orden, "estado_etapa", "") or "").strip()
            orden.estado_etapa = CIERRE_PENDIENTE_LABEL
            orden.ultimo_cambio = now
            orden.cierre_pendiente_en = now
            orden.cierre_pendiente_hasta = now + timedelta(minutes=CIERRE_VENTANA_MINUTOS)
            orden.cierre_pendiente_por = actor
            orden.cierre_estado_previo = prev_etapa
            orden.save(
                update_fields=[
                    "estado_etapa",
                    "ultimo_cambio",
                    "cierre_pendiente_en",
                    "cierre_pendiente_hasta",
                    "cierre_pendiente_por",
                    "cierre_estado_previo",
                    "actualizado_en",
                ]
            )
            HerrEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=prev_etapa,
                estado_nuevo=CIERRE_PENDIENTE_LABEL,
                fecha_operacion=timezone.localdate(),
                actor_username=actor,
                comentario="cierre_pendiente",
            )
            reload_ui = True
        elif int(term) < int(objetivo) and (getattr(orden, "estado_etapa", "") or "").strip() == CIERRE_PENDIENTE_LABEL:
            prev_etapa = CIERRE_PENDIENTE_LABEL
            target = (getattr(orden, "cierre_estado_previo", "") or "").strip() or "Soldadura"
            orden.estado_etapa = target
            orden.ultimo_cambio = now
            orden.cierre_revertido_en = now
            orden.cierre_revertido_por = actor
            orden.cierre_pendiente_hasta = None
            orden.save(
                update_fields=[
                    "estado_etapa",
                    "ultimo_cambio",
                    "cierre_revertido_en",
                    "cierre_revertido_por",
                    "cierre_pendiente_hasta",
                    "actualizado_en",
                ]
            )
            HerrEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=prev_etapa,
                estado_nuevo=target,
                fecha_operacion=timezone.localdate(),
                actor_username=actor,
                comentario="auto_reabrir_por_correccion",
            )
            reload_ui = True

    sold_pct = 0.0 if objetivo <= 0 else round(min(100.0, (sold / objetivo) * 100.0), 2)
    pint_pct = 0.0 if objetivo <= 0 else round(min(100.0, (pint / objetivo) * 100.0), 2)
    term_pct = 0.0 if objetivo <= 0 else round(min(100.0, (term / objetivo) * 100.0), 2)
    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "soldadas": sold,
            "pintadas": pint,
            "terminadas": term,
            "objetivo": objetivo,
            "soldadas_pct": sold_pct,
            "pintadas_pct": pint_pct,
            "terminadas_pct": term_pct,
            "reload": reload_ui,
            "acuse_url": acuse_url,
        }
    )


@login_required
@require_http_methods(["POST"])
def herreria_revertir_cierre(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    next_url = _safe_next_url(request.POST.get("next") or "")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    now = timezone.now()
    with transaction.atomic(using="mes"):
        orden = get_object_or_404(HerrOrdenProduccion.objects.using("mes").select_for_update(), pk=pk)
        if orden.estado != "Abierta":
            messages.error(request, "La orden está cerrada/cancelada.")
            return redirect(next_url) if next_url else redirect("catalogos:herreria_control")
        if int(getattr(orden, "total_piezas", 0) or 0) < 2:
            messages.error(request, "No aplica: no es una orden grande.")
            return redirect(next_url) if next_url else redirect("catalogos:herreria_control")
        etapa = (getattr(orden, "estado_etapa", "") or "").strip()
        if etapa != CIERRE_PENDIENTE_LABEL:
            messages.error(request, "No hay cierre pendiente para revertir.")
            return redirect(next_url) if next_url else redirect("catalogos:herreria_control")
        hasta = getattr(orden, "cierre_pendiente_hasta", None)
        if not hasta or hasta <= now:
            orden.estado_etapa = CIERRE_FINAL_LABEL
            orden.ultimo_cambio = now
            orden.cierre_bloqueado_en = now
            orden.cierre_pendiente_hasta = None
            orden.save(
                update_fields=["estado_etapa", "ultimo_cambio", "cierre_bloqueado_en", "cierre_pendiente_hasta", "actualizado_en"]
            )
            HerrEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=CIERRE_PENDIENTE_LABEL,
                estado_nuevo=CIERRE_FINAL_LABEL,
                fecha_operacion=timezone.localdate(),
                actor_username="system",
                comentario="auto_bloqueo",
            )
            messages.error(request, "La ventana expiró; la orden ya quedó bloqueada.")
            return redirect(next_url) if next_url else redirect("catalogos:herreria_control")
        target = (getattr(orden, "cierre_estado_previo", "") or "").strip() or "Soldadura"
        orden.estado_etapa = target
        orden.ultimo_cambio = now
        orden.cierre_revertido_en = now
        orden.cierre_revertido_por = actor
        orden.cierre_pendiente_hasta = None
        orden.save(
            update_fields=[
                "estado_etapa",
                "ultimo_cambio",
                "cierre_revertido_en",
                "cierre_revertido_por",
                "cierre_pendiente_hasta",
                "actualizado_en",
            ]
        )
        HerrEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=CIERRE_PENDIENTE_LABEL,
            estado_nuevo=target,
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="revertir_cierre",
        )
    messages.success(request, "Cierre revertido.")
    return redirect(next_url) if next_url else redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["POST"])
def herreria_asignaciones(request, pk: int):
    if not _can_herreria(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    corte_operador_ids = [int(x) for x in request.POST.getlist("corte_operador_ids") if (x or "").isdigit()]
    corte_maquina_ids = [int(x) for x in request.POST.getlist("corte_maquina_ids") if (x or "").isdigit()]
    soldadura_ids = [int(x) for x in request.POST.getlist("soldadura_ids") if (x or "").isdigit()]
    pintura_ids = [int(x) for x in request.POST.getlist("pintura_ids") if (x or "").isdigit()]
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    HerrAsignacion.objects.filter(orden=orden, etapa__in=["Corte", "Soldadura", "Pintura"], vigente=True).update(vigente=False)
    for cid in corte_operador_ids:
        HerrAsignacion.objects.create(orden=orden, etapa="Corte", rol="Operador", colaborador_id=cid, vigente=True, asignado_por=actor)
    for mid in corte_maquina_ids:
        HerrAsignacion.objects.create(orden=orden, etapa="Corte", rol="Máquina", maquina_id=mid, vigente=True, asignado_por=actor)
    for cid in soldadura_ids:
        HerrAsignacion.objects.create(orden=orden, etapa="Soldadura", rol="Asignado", colaborador_id=cid, vigente=True, asignado_por=actor)
    for cid in pintura_ids:
        HerrAsignacion.objects.create(orden=orden, etapa="Pintura", rol="Asignado", colaborador_id=cid, vigente=True, asignado_por=actor)

    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "corte_operador_ids": corte_operador_ids,
            "corte_maquina_ids": corte_maquina_ids,
            "soldadura_ids": soldadura_ids,
            "pintura_ids": pintura_ids,
        }
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria_pieza_editar(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    pieza = get_object_or_404(HerrPiezaCatalogo, pk=pk)
    if request.method == "POST":
        form = HerrPiezaForm(request.POST, instance=pieza)
        if form.is_valid():
            form.save()
            messages.success(request, "Pieza actualizada.")
            return redirect("catalogos:herreria_catalogo")
    else:
        form = HerrPiezaForm(instance=pieza)
    return render(request, "catalogos/herreria_pieza_editar.html", {"form": form, "pieza": pieza})


@login_required
@require_http_methods(["GET", "POST"])
def herreria_ordenes(request):
    if not _can_herreria(request.user):
        return redirect("/")
    return redirect("catalogos:herreria_control")

    form = HerrOrdenProduccionForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_create":
            form = HerrOrdenProduccionForm(request.POST)
            if form.is_valid():
                orden = form.save(commit=False)
                orden.estado_etapa = "Espera de corte"
                orden.es_individual = False
                orden.save()
                messages.success(request, "Orden creada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        if action == "etapa_next":
            oid = int(request.POST.get("orden_id") or 0)
            o = HerrOrdenProduccion.objects.filter(id=oid, es_individual=False).first()
            if o and o.estado == "Abierta":
                estados = [
                    "Espera de corte",
                    "Corte",
                    "Espera de armado",
                    "Armado",
                    "Espera de soldadura",
                    "Soldadura",
                    "Espera de pintura",
                    "Pintura",
                    "Terminado",
                ]
                try:
                    idx = estados.index((o.estado_etapa or "").strip())
                except Exception:
                    idx = -1
                nxt = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
                if nxt:
                    o.estado_etapa = nxt
                    o.save(update_fields=["estado_etapa", "actualizado_en"])
                    messages.success(request, f"Actualizado a {nxt}.")
            return redirect("catalogos:herreria_ordenes")

    ordenes_abiertas = list(
        HerrOrdenProduccion.objects.filter(estado="Abierta", es_individual=False).select_related("proyecto").order_by("-id")[:200]
    )
    ordenes_recientes = list(
        HerrOrdenProduccion.objects.exclude(estado="Abierta").filter(es_individual=False).select_related("proyecto").order_by("-id")[:50]
    )
    orden_ids = [int(o.id) for o in (ordenes_abiertas + ordenes_recientes)]
    if orden_ids:
        req_rows = (
            HerrOrdenItem.objects.filter(orden_id__in=orden_ids)
            .values("orden_id", "etapa")
            .annotate(req=Coalesce(Sum("cantidad_requerida"), 0))
        )
        prod_rows = (
            HerrProduccion.objects.filter(orden_item__orden_id__in=orden_ids)
            .values("orden_item__orden_id", "orden_item__etapa")
            .annotate(prod=Coalesce(Sum("cantidad"), 0))
        )
        req_map = {(int(r["orden_id"]), str(r["etapa"])): int(r["req"] or 0) for r in req_rows}
        prod_map = {(int(r["orden_item__orden_id"]), str(r["orden_item__etapa"])): int(r["prod"] or 0) for r in prod_rows}
        etapas = ["Corte", "Armado", "Soldadura", "Pintura"]
        for o in (ordenes_abiertas + ordenes_recientes):
            sp = {}
            for e in etapas:
                req = req_map.get((int(o.id), e), 0)
                prod = prod_map.get((int(o.id), e), 0)
                sp[e] = 0.0 if req <= 0 else min(100.0, (prod / req) * 100.0)
            o.stage_pcts = sp
            estados = [
                "Espera de corte",
                "Corte",
                "Espera de armado",
                "Armado",
                "Espera de soldadura",
                "Soldadura",
                "Espera de pintura",
                "Pintura",
                "Terminado",
            ]
            try:
                idx = estados.index((o.estado_etapa or "").strip())
            except Exception:
                idx = -1
            o.next_estado = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
    return render(
        request,
        "catalogos/herreria_ordenes.html",
        {"form": form, "ordenes_abiertas": ordenes_abiertas, "ordenes_recientes": ordenes_recientes},
    )


@login_required
@require_http_methods(["GET", "POST"])
def herreria_orden_detalle(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    return redirect("catalogos:herreria_control")

    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    edit_form = HerrOrdenProduccionForm(instance=orden)
    item_form = HerrOrdenItemForm()
    assign_form = HerrOrdenAsignacionForm()
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    try:
        idx = estados.index((orden.estado_etapa or "").strip())
    except Exception:
        idx = -1
    next_estado = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
    etapa_activa = (orden.estado_etapa or "").strip()
    etapa_permitida = etapa_activa if etapa_activa in {"Corte", "Armado", "Soldadura", "Pintura"} else None
    prod_form = HerrOrdenRegistroProduccionForm(
        orden, initial={"fecha": timezone.localdate(), "cantidad": 1}, etapa_permitida=etapa_permitida
    )
    has_prod = HerrProduccion.objects.filter(orden_item__orden=orden).exists()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_update":
            if orden.estado != "Abierta":
                messages.error(request, "La orden no se puede editar (cerrada/cancelada).")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
            edit_form = HerrOrdenProduccionForm(request.POST, instance=orden)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "Orden actualizada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "orden_delete":
            if has_prod:
                messages.error(request, "No se puede eliminar: la orden ya tiene producción registrada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
            try:
                orden.delete()
            except Exception:
                messages.error(request, "No se pudo eliminar la orden.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
            messages.success(request, "Orden eliminada.")
            return redirect("catalogos:herreria_ordenes")
        elif action == "item_add":
            item_form = HerrOrdenItemForm(request.POST)
            if item_form.is_valid():
                try:
                    HerrOrdenItem.objects.create(
                        orden=orden,
                        etapa=item_form.cleaned_data["etapa"],
                        pieza=item_form.cleaned_data.get("pieza"),
                        pieza_custom_nombre=(item_form.cleaned_data.get("pieza_custom_nombre") or "").strip(),
                        pieza_custom_peso_kg=float(item_form.cleaned_data.get("pieza_custom_peso_kg") or 0.0),
                        cantidad_requerida=item_form.cleaned_data["cantidad_requerida"],
                    )
                    messages.success(request, "Pieza requerida agregada.")
                    return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
                except Exception:
                    messages.error(request, "No se pudo agregar la pieza (puede que ya exista en la orden).")
        elif action == "item_delete":
            item_id = int(request.POST.get("item_id") or 0)
            if item_id:
                HerrOrdenItem.objects.filter(pk=item_id, orden=orden).delete()
                messages.success(request, "Pieza removida de la orden.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "prod_add":
            if not etapa_permitida:
                messages.error(request, "La orden no está en una etapa de producción activa.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
            prod_form = HerrOrdenRegistroProduccionForm(orden, request.POST, etapa_permitida=etapa_permitida)
            if prod_form.is_valid():
                item = prod_form.cleaned_data["orden_item"]
                operadores = list(prod_form.cleaned_data.get("operador") or [])
                cantidad_total = int(prod_form.cleaned_data["cantidad"] or 0)
                if cantidad_total <= 0 or not operadores:
                    messages.error(request, "Debes seleccionar operador(es) y una cantidad válida.")
                    return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))

                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""

                n_ops = len(operadores)
                base = cantidad_total // n_ops
                rem = cantidad_total % n_ops
                for i, op in enumerate(operadores):
                    qty = base + (1 if i < rem else 0)
                    if qty <= 0:
                        continue
                    HerrProduccion.objects.create(
                        fecha=prod_form.cleaned_data["fecha"],
                        cantidad=int(qty),
                        operador=op,
                        orden_item=item,
                        observaciones=prod_form.cleaned_data.get("observaciones") or "",
                    )
                    HerrOrdenAsignacion.objects.get_or_create(
                        orden=orden,
                        etapa=str(getattr(item, "etapa", "") or "").strip() or str(etapa_permitida),
                        colaborador=op,
                        defaults={"asignado_por": actor},
                    )
                messages.success(request, "Producción registrada en la orden.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "etapa_next":
            if orden.estado != "Abierta":
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
            if next_estado:
                orden.estado_etapa = next_estado
                orden.save(update_fields=["estado_etapa", "actualizado_en"])
                messages.success(request, f"Actualizado a {next_estado}.")
            return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "asig_add":
            assign_form = HerrOrdenAsignacionForm(request.POST)
            if assign_form.is_valid():
                etapa = assign_form.cleaned_data["etapa"]
                col = assign_form.cleaned_data["colaborador"]
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                HerrOrdenAsignacion.objects.get_or_create(
                    orden=orden,
                    etapa=etapa,
                    colaborador=col,
                    defaults={"asignado_por": actor},
                )
                messages.success(request, "Asignación guardada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "asig_delete":
            asig_id = int(request.POST.get("asig_id") or 0)
            if asig_id:
                HerrOrdenAsignacion.objects.filter(id=asig_id, orden=orden).delete()
                messages.success(request, "Asignación eliminada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "orden_close":
            if orden.estado == "Abierta":
                orden.estado = "Cerrada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cerrada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))
        elif action == "orden_cancel":
            if orden.estado == "Abierta":
                orden.estado = "Cancelada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cancelada.")
                return redirect("catalogos:herreria_orden_detalle", pk=int(orden.id))

    items = list(HerrOrdenItem.objects.filter(orden=orden).select_related("pieza").order_by("id"))
    prod_rows = (
        HerrProduccion.objects.filter(orden_item__orden=orden)
        .values("orden_item_id")
        .annotate(producidas=Coalesce(Sum("cantidad"), 0))
    )
    prod_map = {int(r["orden_item_id"]): r for r in prod_rows if r.get("orden_item_id")}
    resumen = []
    pcts = []
    for it in items:
        row = prod_map.get(int(it.id), {})
        producidas = int(row.get("producidas") or 0)
        requeridas = int(it.cantidad_requerida or 0)
        faltan = max(0, requeridas - producidas)
        pct = 0.0 if requeridas <= 0 else min(100.0, (producidas / requeridas) * 100.0)
        pcts.append(pct)
        kg = float(it.pieza_peso_kg or 0.0) * producidas
        resumen.append(
            {
                "item": it,
                "requeridas": requeridas,
                "producidas": producidas,
                "faltan": faltan,
                "pct": round(pct, 2),
                "kg": round(float(kg or 0.0), 3),
            }
        )
    avance_pct = round(min(pcts) if pcts else 0.0, 2)

    historial = list(
        HerrProduccion.objects.filter(orden_item__orden=orden)
        .select_related("operador", "orden_item")
        .order_by("-fecha", "-id")[:250]
    )
    asignaciones = list(HerrOrdenAsignacion.objects.filter(orden=orden).select_related("colaborador").order_by("-asignado_en"))
    return render(
        request,
        "catalogos/herreria_orden_detalle.html",
        {
            "orden": orden,
            "avance_pct": avance_pct,
            "edit_form": edit_form,
            "can_delete": (not has_prod),
            "item_form": item_form,
            "prod_form": prod_form,
            "assign_form": assign_form,
            "resumen": resumen,
            "historial": historial,
            "asignaciones": asignaciones,
            "next_estado": next_estado,
        },
    )


@login_required
@require_http_methods(["POST"])
def herreria_change_status_json(request, pk: int):
    if not _can_herreria(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    estado_nuevo = (request.POST.get("estado_nuevo") or "").strip()
    if not estado_nuevo:
        return JsonResponse({"ok": False, "error": "Estado inválido."}, status=400)
    is_op = bool(int(getattr(orden, "total_piezas", 0) or 0) >= 2)
    allowed = {
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    }
    allowed_op = {"Espera de corte", "Corte", "Soldadura"}
    if is_op:
        if estado_nuevo not in allowed_op:
            return JsonResponse({"ok": False, "error": "Estado inválido para una orden OP."}, status=409)
    else:
        if estado_nuevo not in allowed:
            return JsonResponse({"ok": False, "error": "Estado inválido."}, status=400)
    if orden.estado != "Abierta":
        return JsonResponse({"ok": False, "error": "La orden está cerrada/cancelada."}, status=409)
    if is_op and estado_nuevo in {"Espera de pintura", "Pintura", "Terminado"}:
        return JsonResponse({"ok": False, "error": "Esta orden usa avance y no pasa a pintura/terminado."}, status=409)
    fecha_operacion = (request.POST.get("fecha_operacion") or "").strip()
    if not fecha_operacion:
        return JsonResponse({"ok": False, "error": "Fecha requerida."}, status=400)
    try:
        fecha_dt = datetime.strptime(fecha_operacion, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Fecha inválida."}, status=400)

    estados_order = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    cur = (orden.estado_etapa or "").strip()
    try:
        idx_cur = estados_order.index(cur)
    except Exception:
        idx_cur = -1
    try:
        idx_new = estados_order.index(estado_nuevo)
    except Exception:
        idx_new = -1
    retro = bool(idx_cur >= 0 and idx_new >= 0 and idx_new < idx_cur)
    if retro:
        motivo = (request.POST.get("motivo_retroceso") or "").strip()
        if motivo not in {"error_dedo", "retrabajo"}:
            return JsonResponse({"ok": False, "error": "Motivo del retroceso requerido."}, status=400)

    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    now = timezone.now()
    with transaction.atomic(using="mes"):
        orden = get_object_or_404(HerrOrdenProduccion.objects.select_for_update(), pk=pk)
        prev = (orden.estado_etapa or "").strip()
        orden.estado_etapa = estado_nuevo
        orden.ultimo_cambio = now
        orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
        if int(getattr(orden, "total_piezas", 0) or 0) < 2 and prev != "Terminado" and estado_nuevo == "Terminado":
            it = HerrOrdenItem.objects.filter(orden=orden, pieza__isnull=False).select_related("pieza").order_by("id").first()
            if it and getattr(it, "pieza", None):
                qty = max(0, int(getattr(orden, "total_piezas", 0) or 0))
                if qty > 0:
                    servicio_almacen.registrar_entrada(it.pieza, qty, actor)
    HerrEstadoCambio.objects.create(
        orden=orden,
        estado_anterior=prev,
        estado_nuevo=estado_nuevo,
        fecha_operacion=fecha_dt,
        actor_username=actor,
        motivo_retroceso=(request.POST.get("motivo_retroceso") or "").strip() if retro else "",
        comentario=(request.POST.get("comentario") or "").strip(),
    )
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }
    try:
        idx = estados_order.index(estado_nuevo)
    except Exception:
        idx = -1
    if is_op:
        estados_next = ["Espera de corte", "Corte", "Soldadura"]
        try:
            idx2 = estados_next.index(estado_nuevo)
        except Exception:
            idx2 = -1
        next_estado = "" if idx2 < 0 or idx2 >= (len(estados_next) - 1) else estados_next[idx2 + 1]
        if estado_nuevo == "Soldadura":
            next_estado = ""
    else:
        next_estado = "" if idx < 0 or idx >= (len(estados_order) - 1) else estados_order[idx + 1]
    ultimo_mov = timezone.localtime(orden.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if orden.ultimo_cambio else ""
    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "estado": estado_nuevo,
            "estado_color": status_colors.get(estado_nuevo, "#8898aa"),
            "estado_clase": clase_de_estado(estado_nuevo),
            "next_estado": next_estado,
            "ultimo_mov": ultimo_mov,
        }
    )


@login_required
@require_http_methods(["POST"])
def herreria_enviar(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:herreria_control")
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    linked = (
        PedidoProduccionItem.objects.filter(orden_herreria_id=int(getattr(orden, "id", 0) or 0))
        .exclude(estado_herreria="Cancelado")
        .exists()
    )
    if linked:
        return redirect("catalogos:herreria_control")
    if orden.estado != "Abierta" or (orden.estado_etapa or "").strip() != "Terminado":
        return redirect("catalogos:herreria_control")
    now = timezone.now()
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        prev = (orden.estado_etapa or "").strip()
        orden.estado_etapa = "Enviado"
        orden.ultimo_cambio = now
        orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
        HerrEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=prev,
            estado_nuevo="Enviado",
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="",
        )
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["POST"])
def herreria_regresar_produccion(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:herreria_control")
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    linked = (
        PedidoProduccionItem.objects.filter(orden_herreria_id=int(getattr(orden, "id", 0) or 0))
        .exclude(estado_herreria="Cancelado")
        .exists()
    )
    if linked:
        return redirect("catalogos:herreria_control")
    if orden.estado != "Abierta" or (orden.estado_etapa or "").strip() != "Enviado":
        return redirect("catalogos:herreria_control")
    now = timezone.now()
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        prev = (orden.estado_etapa or "").strip()
        orden.estado_etapa = "Terminado"
        orden.ultimo_cambio = now
        orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
        HerrEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=prev,
            estado_nuevo="Terminado",
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="",
        )
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["POST"])
def herreria_delete_decote(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:herreria_control")
    orden = get_object_or_404(HerrOrdenProduccion, pk=pk)
    if orden.estado != "Abierta":
        return redirect("catalogos:herreria_control")
    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    pedido_id = (
        PedidoProduccionItem.objects.filter(orden_herreria_id=int(orden.id))
        .exclude(estado_herreria="Cancelado")
        .values_list("pedido_id", flat=True)
        .first()
    )
    if pedido_id:
        items = list(PedidoProduccionItem.objects.filter(pedido_id=int(pedido_id)).exclude(estado_herreria="Cancelado"))
        if not items or not _pedido_is_completo(items):
            return redirect("catalogos:herreria_control")
        apartado_only = all(
            int(getattr(x, "enviado", 0) or 0) == 0
            for x in items
            if int(getattr(x, "apartado", 0) or 0) > 0
        )
        if apartado_only:
            return redirect("catalogos:herreria_control")
        last_enviado = (
            LogisticaEnvio.objects.filter(pedido_id=int(pedido_id))
            .aggregate(mx=Max("fecha"))
            .get("mx")
        )
        if not last_enviado or last_enviado > cutoff:
            return redirect("catalogos:herreria_control")
        exp = LogisticaExpediente.objects.filter(pedido_id=int(pedido_id)).first()
        if not exp or not getattr(exp, "generado_en", None) or int(getattr(exp, "descargas_count", 0) or 0) <= 0:
            return redirect("catalogos:herreria_control")
        with transaction.atomic(using="mes"):
            HerrProduccion.objects.filter(orden_item__orden=orden).delete()
            HerrAsignacion.objects.filter(orden=orden).delete()
            orden.delete()
    else:
        if (orden.estado_etapa or "").strip() != "Enviado":
            return redirect("catalogos:herreria_control")
        ult = getattr(orden, "ultimo_cambio", None)
        ult_fecha = timezone.localdate(ult) if ult else today
        if not ult_fecha or ult_fecha > cutoff:
            return redirect("catalogos:herreria_control")
        with transaction.atomic(using="mes"):
            HerrProduccion.objects.filter(orden_item__orden=orden).delete()
            HerrAsignacion.objects.filter(orden=orden).delete()
            orden.delete()
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["POST"])
def herreria_delete_decote_all(request):
    if not _can_herreria(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:herreria_control")
    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    filters = {
        "proyecto": (request.POST.get("proyecto") or "").strip(),
        "q": (request.POST.get("q") or "").strip(),
    }
    base_qs = HerrOrdenProduccion.objects.filter(estado="Abierta")
    if filters["proyecto"] and filters["proyecto"] != "Todos":
        base_qs = base_qs.filter(proyecto__nombre=filters["proyecto"])
    if filters["q"]:
        q = filters["q"]
        base_qs = base_qs.filter(
            Q(codigo__icontains=q)
            | Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(proyecto__nombre__icontains=q)
        )
    orden_ids = list(base_qs.values_list("id", flat=True)[:2000])
    if orden_ids:
        orden_to_pedido = {}
        pedido_ids = set()
        for oid, pid in (
            PedidoProduccionItem.objects.filter(orden_herreria_id__in=orden_ids)
            .exclude(estado_herreria="Cancelado")
            .values_list("orden_herreria_id", "pedido_id")
        ):
            if oid and pid:
                orden_to_pedido[int(oid)] = int(pid)
                pedido_ids.add(int(pid))
        items_by_pedido_id: dict[int, list[PedidoProduccionItem]] = {}
        if pedido_ids:
            for it in PedidoProduccionItem.objects.filter(pedido_id__in=list(pedido_ids)).exclude(estado_herreria="Cancelado"):
                items_by_pedido_id.setdefault(int(getattr(it, "pedido_id", 0) or 0), []).append(it)
        pedidos_enviados = set()
        for pid, items in items_by_pedido_id.items():
            if not items or not _pedido_is_completo(items):
                continue
            apartado_only = all(
                int(getattr(x, "enviado", 0) or 0) == 0
                for x in items
                if int(getattr(x, "apartado", 0) or 0) > 0
            )
            if not apartado_only:
                pedidos_enviados.add(int(pid))
        decote_ids: list[int] = []
        if pedidos_enviados:
            last_map = {
                int(r["pedido_id"]): r["last_fecha"]
                for r in LogisticaEnvio.objects.filter(pedido_id__in=list(pedidos_enviados))
                .values("pedido_id")
                .annotate(last_fecha=Max("fecha"))
            }
            exp_map = {int(e.pedido_id): e for e in LogisticaExpediente.objects.filter(pedido_id__in=list(pedidos_enviados))}
            decote_pedidos = set()
            for pid in pedidos_enviados:
                last_fecha = last_map.get(int(pid))
                if not last_fecha or last_fecha > cutoff:
                    continue
                exp = exp_map.get(int(pid))
                if not exp or not getattr(exp, "generado_en", None) or int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                    continue
                decote_pedidos.add(int(pid))
            decote_ids = [int(oid) for oid, pid in orden_to_pedido.items() if int(pid) in decote_pedidos]

        obra_ids = list(
            HerrOrdenProduccion.objects.filter(id__in=orden_ids, estado="Abierta", estado_etapa="Enviado", ultimo_cambio__date__lte=cutoff)
            .exclude(id__in=list(orden_to_pedido.keys()) if orden_to_pedido else [])
            .values_list("id", flat=True)[:2000]
        )
        all_ids = list({int(x) for x in (decote_ids + obra_ids) if x})
        if all_ids:
            with transaction.atomic(using="mes"):
                HerrProduccion.objects.filter(orden_item__orden_id__in=all_ids).delete()
                HerrAsignacion.objects.filter(orden_id__in=all_ids).delete()
                HerrOrdenProduccion.objects.filter(id__in=all_ids).delete()
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:herreria_control")


@login_required
@require_http_methods(["POST"])
def herreria_acuse_create(request):
    if not _can_herreria(request.user):
        return redirect("/")
    item_id = int(request.POST.get("pedido_item_id") or 0)
    cantidad = int(request.POST.get("cantidad") or 0)
    recibido_por = (request.POST.get("recibido_por") or "").strip()
    if item_id <= 0 or cantidad <= 0:
        return redirect("catalogos:herreria_control")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        it = (
            PedidoProduccionItem.objects.using("mes")
            .select_for_update()
            .select_related("pedido", "producto")
            .filter(id=item_id)
            .first()
        )
        if not it or not getattr(it, "pedido_id", None):
            return redirect("catalogos:herreria_control")
        if (getattr(it, "pedido", None) is None) or (getattr(it.pedido, "estado", "") or "").strip() != "Activa":
            return redirect("catalogos:herreria_control")
        if not getattr(it, "orden_herreria_id", None):
            return redirect("catalogos:herreria_control")
        orden = (
            HerrOrdenProduccion.objects.using("mes")
            .select_for_update()
            .filter(id=int(getattr(it, "orden_herreria_id") or 0))
            .first()
        )
        terminadas = int(getattr(orden, "cantidad_terminada", 0) or 0) if orden else 0
        entregado = int(
            LogisticaAcuseEntrega.objects.using("mes").filter(pedido_item_id=int(it.id)).aggregate(total=Sum("cantidad")).get("total")
            or 0
        )
        disponible = max(0, terminadas - entregado)
        if cantidad > disponible:
            return redirect("catalogos:herreria_control")
        acuse = LogisticaAcuseEntrega.objects.using("mes").create(
            pedido_item=it,
            cantidad=int(cantidad),
            fecha=timezone.localdate(),
            entregado_por=(actor or "").strip(),
            recibido_por=recibido_por,
        )
    return redirect("catalogos:herreria_acuse_print", pk=int(acuse.id))


@login_required
@require_http_methods(["GET"])
def herreria_acuse_print(request, pk: int):
    if not _can_herreria(request.user):
        return redirect("/")
    acuse = (
        LogisticaAcuseEntrega.objects.using("mes")
        .select_related("pedido_item", "pedido_item__pedido", "pedido_item__producto", "pedido_item__orden_herreria")
        .filter(id=int(pk))
        .first()
    )
    if not acuse:
        return redirect("catalogos:herreria_control")
    it = getattr(acuse, "pedido_item", None)
    pedido = getattr(it, "pedido", None) if it else None
    producto = getattr(it, "producto", None) if it else None
    orden = getattr(it, "orden_herreria", None) if it else None
    total_linea = int(getattr(it, "cantidad_total", 0) or 0) if it else 0
    terminadas = int(getattr(orden, "cantidad_terminada", 0) or 0) if orden else 0
    entregado_total = int(
        LogisticaAcuseEntrega.objects.using("mes")
        .filter(pedido_item_id=int(getattr(it, "id", 0) or 0))
        .aggregate(total=Sum("cantidad"))
        .get("total")
        or 0
    )
    entregado_antes = max(0, entregado_total - int(getattr(acuse, "cantidad", 0) or 0))
    entregado_despues = entregado_total
    pendiente = max(0, total_linea - entregado_despues)
    return render(
        request,
        "catalogos/herreria_acuse.html",
        {
            "acuse": acuse,
            "pedido": pedido,
            "producto": producto,
            "orden": orden,
            "total_linea": total_linea,
            "terminadas": terminadas,
            "entregado_antes": entregado_antes,
            "entregado_despues": entregado_despues,
            "pendiente": pendiente,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    return redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_materiales(request):
    if not _can_corte_laser(request.user):
        return redirect("/")

    solo = (request.GET.get("solo") or "").strip() in {"1", "true", "True", "yes", "YES"}
    saved = (request.GET.get("saved") or "").strip() in {"1", "true", "True", "yes", "YES"}
    embedded = (request.GET.get("embedded") or "").strip() in {"1", "true", "True", "yes", "YES"}
    next_url = _safe_next_url(request.GET.get("next") or request.POST.get("next") or "")

    edit_id = int(request.GET.get("edit") or 0)
    edit_obj = LaserMaterialPlaca.objects.filter(id=edit_id).first() if edit_id else None
    material_form = LaserMaterialPlacaForm(instance=edit_obj) if edit_obj else LaserMaterialPlacaForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "material_create":
            material_form = LaserMaterialPlacaForm(request.POST)
            if material_form.is_valid():
                try:
                    material_form.save()
                    messages.success(request, "Material guardado.")
                    return redirect("catalogos:corte_laser_materiales")
                except IntegrityError:
                    messages.error(request, "Ese material ya existe con la misma cédula/espesor/tamaño. Revisa y vuelve a intentar.")
        if action == "material_save":
            mid = int(request.POST.get("material_id") or 0)
            obj = LaserMaterialPlaca.objects.filter(id=mid).first() if mid else None
            if not obj:
                messages.error(request, "Material inválido.")
                return redirect("catalogos:corte_laser_materiales")
            material_form = LaserMaterialPlacaForm(request.POST, instance=obj)
            if material_form.is_valid():
                try:
                    material_form.save()
                    messages.success(request, "Material actualizado.")
                    if solo:
                        url = reverse("catalogos:corte_laser_materiales")
                        qs = f"?edit={int(obj.id)}&solo=1&saved=1"
                        if embedded:
                            qs += "&embedded=1"
                        if next_url:
                            qs += f"&next={quote(next_url)}"
                        return redirect(url + qs)
                    if next_url:
                        return redirect(next_url)
                    return redirect("catalogos:corte_laser_materiales")
                except IntegrityError:
                    messages.error(request, "No se pudo actualizar: ya existe otro material con la misma cédula/espesor/tamaño.")
        if action == "material_delete":
            mid = int(request.POST.get("material_id") or 0)
            if not mid:
                messages.error(request, "No se indicó qué material eliminar.")
            elif LaserOrdenProduccion.objects.filter(material_id=mid).exists():
                messages.error(request, "No se puede eliminar: el material ya está usado en un pedido.")
            else:
                try:
                    borrados, _ = LaserMaterialPlaca.objects.filter(id=mid).delete()
                    if borrados:
                        messages.success(request, "Material eliminado.")
                    else:
                        messages.error(request, "Ese material ya no existe.")
                except ProtectedError:
                    messages.error(
                        request,
                        "No se puede eliminar: el material tiene registros dependientes. "
                        "Desactívalo en su lugar.",
                    )
            return redirect("catalogos:corte_laser_materiales")

    materiales_qs = list(
        LaserMaterialPlaca.objects.order_by(
            "-activo",
            "categoria_material",
            "tipo_material",
            "nombre",
            "calibre",
            "espesor_mm",
            "largo_mm",
            "ancho_mm",
            "id",
        )[:2000]
    )
    nom_seen: set[tuple[str, str, str]] = set()
    nomenclaturas_payload: list[dict] = []
    for m in materiales_qs:
        try:
            m.largo_cm_show = float(getattr(m, "largo_mm", 0) or 0) / 10.0
        except Exception:
            m.largo_cm_show = 0.0
        try:
            m.ancho_cm_show = float(getattr(m, "ancho_mm", 0) or 0) / 10.0
        except Exception:
            m.ancho_cm_show = 0.0
        key = ((getattr(m, "categoria_material", "") or "").strip(), (getattr(m, "tipo_material", "") or "").strip(), (getattr(m, "nombre", "") or "").strip())
        if key not in nom_seen and key[2]:
            nom_seen.add(key)
            nomenclaturas_payload.append({"categoria": key[0], "tipo": key[1], "nombre": key[2]})
    return render(
        request,
        "catalogos/corte_laser_materiales.html",
        {
            "material_form": material_form,
            "materiales": materiales_qs,
            "nomenclaturas_payload": nomenclaturas_payload,
            "edit_obj": edit_obj,
            "solo": solo,
            "saved": saved,
            "embedded": embedded,
            "next_url": next_url,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_catalogos(request):
    if not _can_corte_laser(request.user):
        return redirect("/")

    edit_id = int(request.GET.get("edit") or 0)
    edit_obj = CortaClienteProyecto.objects.filter(id=edit_id).first() if edit_id else None
    form = CortaClienteProyectoForm(instance=edit_obj) if edit_obj else CortaClienteProyectoForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action in {"create", "save"}:
            cp_id = int(request.POST.get("cp_id") or 0)
            obj = CortaClienteProyecto.objects.filter(id=cp_id).first() if cp_id else None
            form = CortaClienteProyectoForm(request.POST)
            if form.is_valid():
                nombre = _norm_text(form.cleaned_data.get("nombre") or "")
                nombre_norm = _norm_upper(nombre)
                if not nombre_norm:
                    messages.error(request, "Nombre requerido.")
                    return redirect("catalogos:corte_laser_catalogos")
                existing = CortaClienteProyecto.objects.filter(nombre_normalizado=nombre_norm).first()
                if existing and obj and int(existing.id) != int(obj.id):
                    messages.error(request, "Ya existe ese nombre.")
                    return redirect("catalogos:corte_laser_catalogos")
                target = obj or existing
                if target:
                    target.nombre = nombre
                    target.nombre_normalizado = nombre_norm
                    target.tipo = form.cleaned_data.get("tipo") or target.tipo
                    target.activo = bool(form.cleaned_data.get("activo"))
                    target.save(update_fields=["nombre", "nombre_normalizado", "tipo", "activo", "actualizado_en"])
                    messages.success(request, "Actualizado.")
                else:
                    CortaClienteProyecto.objects.create(
                        nombre=nombre,
                        nombre_normalizado=nombre_norm,
                        tipo=form.cleaned_data.get("tipo") or "cliente",
                        activo=bool(form.cleaned_data.get("activo")),
                    )
                    messages.success(request, "Guardado.")
            else:
                messages.error(request, "Revisa los datos del formulario.")
            return redirect("catalogos:corte_laser_catalogos")
        if action == "toggle":
            cid = int(request.POST.get("cp_id") or 0)
            if cid:
                obj = CortaClienteProyecto.objects.filter(id=cid).first()
                if obj:
                    obj.activo = not bool(obj.activo)
                    obj.save(update_fields=["activo", "actualizado_en"])
            return redirect("catalogos:corte_laser_catalogos")
        if action == "delete":
            cid = int(request.POST.get("cp_id") or 0)
            if cid:
                used = LaserOrdenProduccion.objects.filter(corta_cliente_proyecto_id=cid).exists()
                if used:
                    messages.error(request, "No se puede eliminar: ya está usado en órdenes. Mejor desactiva.")
                else:
                    CortaClienteProyecto.objects.filter(id=cid).delete()
                    messages.success(request, "Eliminado.")
            return redirect("catalogos:corte_laser_catalogos")

    rows = list(CortaClienteProyecto.objects.order_by("-activo", "tipo", "nombre")[:2000])
    return render(
        request,
        "catalogos/corte_laser_catalogos.html",
        {"form": form, "rows": rows, "edit_obj": edit_obj},
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_piezas(request):
    if not _can_corte_laser(request.user):
        return redirect("/")

    edit_id = int(request.GET.get("edit") or 0)
    edit_obj = CortaPiezaCatalogo.objects.filter(id=edit_id).first() if edit_id else None
    form = CortaPiezaCatalogoForm(instance=edit_obj) if edit_obj else CortaPiezaCatalogoForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action in {"create", "save"}:
            pieza_id = int(request.POST.get("pieza_id") or 0)
            obj = CortaPiezaCatalogo.objects.filter(id=pieza_id).first() if pieza_id else None
            form = CortaPiezaCatalogoForm(request.POST, request.FILES, instance=obj) if obj else CortaPiezaCatalogoForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, "Guardado.")
                    return redirect("catalogos:corte_laser_piezas")
                except IntegrityError:
                    messages.error(request, "Ya existe una pieza con ese nombre.")
        if action == "toggle":
            pid = int(request.POST.get("pieza_id") or 0)
            if pid:
                obj = CortaPiezaCatalogo.objects.filter(id=pid).first()
                if obj:
                    obj.activo = not bool(obj.activo)
                    obj.save(update_fields=["activo", "actualizado_en"])
            return redirect("catalogos:corte_laser_piezas")
        if action == "delete":
            pid = int(request.POST.get("pieza_id") or 0)
            if not pid:
                messages.error(request, "No se indicó qué pieza eliminar.")
            elif LaserOrdenProduccion.objects.filter(corta_pieza_id=pid).exists():
                messages.error(request, "No se puede eliminar: ya está usada en pedidos. Mejor desactiva.")
            else:
                try:
                    borradas, _ = CortaPiezaCatalogo.objects.filter(id=pid).delete()
                    if borradas:
                        messages.success(request, "Eliminado.")
                    else:
                        messages.error(request, "Esa pieza ya no existe.")
                except ProtectedError:
                    messages.error(
                        request,
                        "No se puede eliminar: la pieza tiene registros dependientes. "
                        "Desactívala en su lugar.",
                    )
            return redirect("catalogos:corte_laser_piezas")

    rows = list(CortaPiezaCatalogo.objects.order_by("-activo", "nombre")[:2000])
    return render(
        request,
        "catalogos/corte_laser_piezas.html",
        {"form": form, "rows": rows, "edit_obj": edit_obj},
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_control(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    _finalize_corta_cierres_expirados()
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    estados_op = [
        "Espera de corte",
        "Corte",
        "Soldadura",
    ]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
        CIERRE_PENDIENTE_LABEL: "#f1c40f",
        "Enviado": "#11cdef",
    }

    estado = (request.GET.get("estado") or "Todos").strip() or "Todos"
    cliente = (request.GET.get("cliente") or request.GET.get("proyecto") or "Todos").strip() or "Todos"
    q = (request.GET.get("q") or "").strip()
    order = (request.GET.get("order") or "ultimo_mov_desc").strip() or "ultimo_mov_desc"
    filters = {"estado": estado, "cliente": cliente, "q": q, "order": order}

    qs = (
        LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto", "material")
        .filter(estado="Abierta")
    )
    if estado and estado not in {"Todos", ""}:
        qs = qs.filter(estado_etapa=estado)
    if cliente and cliente not in {"Todos", ""}:
        qs = qs.filter(corta_cliente_proyecto__nombre=cliente)
    if q:
        qs = qs.filter(
            Q(folio_externo__icontains=q)
            | Q(codigo__icontains=q)
            | Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(corta_cliente_proyecto__nombre__icontains=q)
        )

    order_map = {
        "ultimo_mov_desc": ["-ultimo_cambio", "-id"],
        "ultimo_mov_asc": ["ultimo_cambio", "id"],
        "prioridad_asc": ["prioridad", "-ultimo_cambio", "-id"],
        "prioridad_desc": ["-prioridad", "-ultimo_cambio", "-id"],
        "fecha_asc": ["fecha_compromiso", "-prioridad", "-ultimo_cambio", "-id"],
        "fecha_desc": ["-fecha_compromiso", "-prioridad", "-ultimo_cambio", "-id"],
        "estado": ["estado_etapa", "-ultimo_cambio", "-id"],
        "cliente": ["corta_cliente_proyecto__nombre", "-ultimo_cambio", "-id"],
    }
    qs = qs.order_by(*order_map.get(order, order_map["ultimo_mov_desc"]))

    vigas_all = list(qs[:2000])
    ordenes_enviadas_ids: set[int] = set()
    for v in vigas_all:
        total = int(getattr(v, "total_piezas", 0) or 0)
        apartado = int(getattr(v, "apartado", 0) or 0)
        enviado = int(getattr(v, "enviado", 0) or 0)
        saldo = max(0, total - apartado - enviado)
        if total > 0 and saldo == 0 and enviado > 0:
            ordenes_enviadas_ids.add(int(getattr(v, "id", 0) or 0))
    vigas = [v for v in vigas_all if int(getattr(v, "id", 0) or 0) not in ordenes_enviadas_ids]
    def _calc_peso_total(material: LaserMaterialPlaca, ancho_mm: int, alto_mm: int, qty: int) -> float:
        try:
            plate_area = float(int(getattr(material, "largo_mm", 0) or 0)) * float(int(getattr(material, "ancho_mm", 0) or 0))
            if not (plate_area > 0):
                return 0.0
            margin = 30.0
            rect_area = (float(max(0, int(ancho_mm or 0))) + (2.0 * margin)) * (float(max(0, int(alto_mm or 0))) + (2.0 * margin))
            peso_placa = float(getattr(material, "peso_kg", 0.0) or 0.0)
            if not (peso_placa > 0):
                esp = float(getattr(material, "espesor_mm", 0.0) or 0.0)
                largo = float(int(getattr(material, "largo_mm", 0) or 0))
                ancho = float(int(getattr(material, "ancho_mm", 0) or 0))
                if esp > 0 and largo > 0 and ancho > 0:
                    cat = (getattr(material, "categoria_material", "") or "").upper()
                    tip = (getattr(material, "tipo_material", "") or "").upper()
                    dens = 7.85
                    if ("ALUMIN" in cat) or ("ALUMIN" in tip):
                        dens = 2.7
                    elif ("INOX" in cat) or ("INOXID" in cat) or ("INOX" in tip) or ("INOXID" in tip):
                        dens = 8.0
                    vol_mm3 = largo * ancho * esp
                    vol_cm3 = vol_mm3 / 1000.0
                    peso_placa = (vol_cm3 * dens) / 1000.0
            if not (peso_placa > 0):
                return 0.0
            per_piece = (rect_area / plate_area) * peso_placa
            total = float(per_piece) * float(max(0, int(qty or 0)))
            return float(round(total, 3))
        except Exception:
            return 0.0

    ids = [int(v.id) for v in vigas]
    for v in vigas:
        v.internal_id = int(v.id)
        v.codigo_viga = (getattr(v, "folio_externo", "") or "").strip() or v.codigo or v.folio
        v.pieza_no = int(v.pieza_no or 1)
        v.total_piezas = int(v.total_piezas or 1)
        v.es_op = bool(int(v.total_piezas or 0) >= 2)
        objetivo = int(v.total_piezas or 1)
        soldadas = max(0, min(int(getattr(v, "cantidad_producida", 0) or 0), objetivo))
        pintadas = max(0, min(int(getattr(v, "cantidad_pintada", 0) or 0), objetivo))
        terminadas = max(0, min(int(getattr(v, "cantidad_terminada", 0) or 0), objetivo))
        v.op_soldadas = soldadas
        v.op_pintadas = pintadas
        v.op_terminadas = terminadas
        v.op_soldadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (soldadas / objetivo) * 100.0), 2)
        v.op_pintadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (pintadas / objetivo) * 100.0), 2)
        v.op_terminadas_pct = 0.0 if not v.es_op or objetivo <= 0 else round(min(100.0, (terminadas / objetivo) * 100.0), 2)
        v.avance_pct = v.op_terminadas_pct
        v.proyecto_nombre = getattr(getattr(v, "corta_cliente_proyecto", None), "nombre", "") or ""
        v.estado = (v.estado_etapa or "").strip()
        v.estado_color = status_colors.get(v.estado, "#8898aa")
        v.op_can_avance = bool(v.es_op) and v.estado in {"Soldadura", CIERRE_PENDIENTE_LABEL}
        v.cierre_pendiente_activo = False
        v.cierre_pendiente_hasta_str = ""
        if v.estado == CIERRE_PENDIENTE_LABEL and getattr(v, "cierre_pendiente_hasta", None):
            hasta = timezone.localtime(getattr(v, "cierre_pendiente_hasta"))
            v.cierre_pendiente_hasta_str = hasta.strftime("%Y-%m-%d %H:%M")
            v.cierre_pendiente_activo = bool(hasta > timezone.localtime(timezone.now()))
        v.plano_url = v.archivo.url if getattr(v, "archivo", None) else ""
        v.dxf_url = v.archivo_dxf.url if getattr(v, "archivo_dxf", None) else ""
        v.ultimo_mov_str = timezone.localtime(v.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if getattr(v, "ultimo_cambio", None) else ""
        try:
            if getattr(v, "material_id", None) and getattr(v, "material", None):
                calc = _calc_peso_total(v.material, int(getattr(v, "pieza_ancho_mm", 0) or 0), int(getattr(v, "pieza_alto_mm", 0) or 0), int(v.total_piezas or 1))
                stored = float(getattr(v, "peso_kg", 0.0) or 0.0)
                if calc > 0 and abs(calc - stored) > 0.001:
                    LaserOrdenProduccion.objects.filter(id=int(v.id)).update(peso_kg=calc)
                    v.peso_kg = calc
        except Exception:
            logger.exception("Error ignorado en corte_laser_control()")
        estados_flow = estados_op if bool(getattr(v, "es_op", False)) else estados
        try:
            idx = estados_flow.index(v.estado)
        except Exception:
            idx = -1
        v.next_estado = "" if idx < 0 or idx >= (len(estados_flow) - 1) else estados_flow[idx + 1]

    asigns = (
        LaserAsignacion.objects.filter(orden_id__in=ids, vigente=True, etapa__in=["Corte", "Soldadura", "Pintura"])
        .values_list("orden_id", "etapa", "rol", "colaborador_id", "maquina_id")
    )
    asign_map = {}
    for oid, etapa, rol, cid, mid in asigns:
        entry = asign_map.setdefault(int(oid), {}).setdefault(str(etapa), {"ids": [], "operadores": [], "maquinas": []})
        if mid:
            entry["maquinas"].append(int(mid))
        if cid:
            if (rol or "") == "Operador":
                entry["operadores"].append(int(cid))
            else:
                entry["ids"].append(int(cid))
    for v in vigas:
        m = asign_map.get(int(v.id), {})
        v.asig_corte = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("ids", []))
        v.asig_corte_maquinas = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("maquinas", []))
        v.asig_corte_operadores = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("operadores", []))
        v.asig_soldadura = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("ids", []))
        v.asig_pintura = ",".join(str(x) for x in (m.get("Pintura", {}) or {}).get("ids", []))
        v.asig_soldadura_maquinas = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("maquinas", []))
        v.asig_soldadura_operadores = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("operadores", []))

    maquina_ids = set()
    for v in vigas:
        for part in (getattr(v, "asig_corte_maquinas", "") or "").split(","):
            part = (part or "").strip()
            if part.isdigit():
                maquina_ids.add(int(part))
    maquinas_info = {}
    maquinas_estado = {}
    if maquina_ids:
        for m in Maquina.objects.filter(id__in=list(maquina_ids)).values("id", "nombre", "tipo", "es_robot"):
            maquinas_info[int(m["id"])] = {"nombre": m["nombre"], "tipo": m["tipo"], "es_robot": bool(m["es_robot"])}
        for p in MaquinaParo.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("motivo"):
            maquinas_estado[int(p.maquina_id)] = {
                "paro": True,
                "falla": False,
                "motivo": getattr(getattr(p, "motivo", None), "nombre", "") or "",
                "paro_inicio": timezone.localtime(p.inicio).strftime("%Y-%m-%d %H:%M") if getattr(p, "inicio", None) else "",
            }
        for f in MaquinaFalla.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("tipo"):
            entry = maquinas_estado.setdefault(int(f.maquina_id), {"paro": False, "falla": False, "motivo": "", "paro_inicio": ""})
            entry["falla"] = True
            entry["tipo_falla"] = getattr(getattr(f, "tipo", None), "nombre", "") or ""
            entry["falla_inicio"] = timezone.localtime(f.inicio).strftime("%Y-%m-%d %H:%M") if getattr(f, "inicio", None) else ""

    clientes_proyectos = list(CortaClienteProyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    participantes_payload = _build_participantes_payload()

    hoy = timezone.localdate()
    cutoff = hoy - timedelta(days=DECOTE_DAYS)
    enviados_recent: list[LaserOrdenProduccion] = []
    decote_vigas: list[LaserOrdenProduccion] = []
    if ordenes_enviadas_ids:
        last_map = {
            int(r["orden_id"]): r["last_fecha"]
            for r in LogisticaEnvioCorta.objects.filter(orden_id__in=list(ordenes_enviadas_ids))
            .values("orden_id")
            .annotate(last_fecha=Max("fecha"))
        }
        exp_map = {
            int(e.orden_corta_id): e for e in LogisticaExpediente.objects.filter(orden_corta_id__in=list(ordenes_enviadas_ids))
        }
        env_qs = LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto").filter(
            estado="Abierta",
            id__in=list(ordenes_enviadas_ids),
        )
        if cliente and cliente not in {"Todos", ""}:
            env_qs = env_qs.filter(corta_cliente_proyecto__nombre=cliente)
        if q:
            env_qs = env_qs.filter(
                Q(folio_externo__icontains=q)
                | Q(codigo__icontains=q)
                | Q(nombre__icontains=q)
                | Q(descripcion__icontains=q)
                | Q(corta_cliente_proyecto__nombre__icontains=q)
            )
        for v in list(env_qs[:2000]):
            last_fecha = last_map.get(int(getattr(v, "id", 0) or 0))
            v.enviado_fecha = last_fecha
            v.enviado_dias = int((hoy - last_fecha).days) if last_fecha else 0
            exp = exp_map.get(int(getattr(v, "id", 0) or 0))
            v.expediente_generado = bool(exp and getattr(exp, "generado_en", None))
            v.expediente_descargado = bool(exp and int(getattr(exp, "descargas_count", 0) or 0) > 0)
            v.__dict__["expediente_descargas_count"] = int(getattr(exp, "descargas_count", 0) or 0) if exp else 0
            can_decote = bool(v.expediente_generado and v.expediente_descargado)
            is_decote_age = bool(last_fecha and (hoy - last_fecha).days >= DECOTE_DAYS)
            v.can_delete_decote = bool(last_fecha and (hoy - last_fecha).days >= 5)
            if is_decote_age and can_decote:
                decote_vigas.append(v)
            else:
                enviados_recent.append(v)
        enviados_recent.sort(key=lambda x: (getattr(x, "enviado_fecha", None) or datetime.min.date()), reverse=True)
        decote_vigas.sort(key=lambda x: (getattr(x, "enviado_fecha", None) or datetime.max.date()))

    def _decorate_enviado(items):
        for v in items:
            v.internal_id = int(v.id)
            v.codigo_viga = (getattr(v, "folio_externo", "") or "").strip() or v.codigo or v.folio
            v.pieza_no = int(getattr(v, "pieza_no", 1) or 1)
            v.total_piezas = int(getattr(v, "total_piezas", 1) or 1)
            v.proyecto_nombre = getattr(getattr(v, "corta_cliente_proyecto", None), "nombre", "") or ""
            v.estado = "Enviado"
            v.estado_color = status_colors.get("Enviado", "#11cdef")
            v.plano_url = v.archivo.url if getattr(v, "archivo", None) else ""
            v.dxf_url = v.archivo_dxf.url if getattr(v, "archivo_dxf", None) else ""
            v.enviado_fecha = getattr(v, "enviado_fecha", None)
            v.enviado_dias = int((hoy - v.enviado_fecha).days) if v.enviado_fecha else 0
            v.enviado_str = v.enviado_fecha.strftime("%Y-%m-%d") if v.enviado_fecha else ""

    _decorate_enviado(enviados_recent)
    _decorate_enviado(decote_vigas)
    enviados_total = len(enviados_recent)
    decote_total = len(decote_vigas)
    enviados_by_cp = {}
    for v in enviados_recent:
        key = v.proyecto_nombre or "-"
        enviados_by_cp.setdefault(key, []).append(v)
    enviados_por_proyecto = sorted(enviados_by_cp.items(), key=lambda t: str(t[0] or "").lower())
    return render(
        request,
        "catalogos/corte_laser_list.html",
        {
            "vigas": vigas,
            "mode": "admin",
            "reset_url": reverse("catalogos:corte_laser_control"),
            "estados_filtro": ["Todos", *estados, CIERRE_PENDIENTE_LABEL],
            "estados_operables": estados,
            "estados_order": estados,
            "estados_order_op": estados_op,
            "estados_operables_op": estados_op,
            "clientes_proyectos": ["Todos", *clientes_proyectos],
            "filters": filters,
            "status_colors": status_colors,
            "participantes_payload": participantes_payload,
            "maquinas_info_payload": maquinas_info,
            "maquinas_estado_payload": maquinas_estado,
            "decote_days": DECOTE_DAYS,
            "enviados_total": enviados_total,
            "enviados_por_proyecto": enviados_por_proyecto,
            "decote_total": decote_total,
            "decote_vigas": decote_vigas,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_create(request):
    if not _can_corte_laser(request.user):
        return redirect("/")

    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    estados_op = [
        "Espera de corte",
        "Corte",
        "Soldadura",
    ]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }

    corte_operadores = list(Colaborador.objects.filter(activo=True, rol="Operador").order_by("nombre"))
    corte_maquinas = list(Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre"))
    selected_corte_operadores = [int(x) for x in request.POST.getlist("corte_operador_ids") if (x or "").isdigit()]
    selected_corte_maquinas = [int(x) for x in request.POST.getlist("corte_maquina_ids") if (x or "").isdigit()]

    def _calc_peso_total(material: LaserMaterialPlaca, ancho_mm: int, alto_mm: int, qty: int) -> float:
        try:
            plate_area = float(int(getattr(material, "largo_mm", 0) or 0)) * float(int(getattr(material, "ancho_mm", 0) or 0))
            if not (plate_area > 0):
                return 0.0
            margin = 30.0
            rect_area = float(max(0, ancho_mm)) + (2.0 * margin)
            rect_area *= float(max(0, alto_mm)) + (2.0 * margin)
            peso_placa = float(getattr(material, "peso_kg", 0.0) or 0.0)
            if not (peso_placa > 0):
                esp = float(getattr(material, "espesor_mm", 0.0) or 0.0)
                largo = float(int(getattr(material, "largo_mm", 0) or 0))
                ancho = float(int(getattr(material, "ancho_mm", 0) or 0))
                if esp > 0 and largo > 0 and ancho > 0:
                    cat = (getattr(material, "categoria_material", "") or "").upper()
                    tip = (getattr(material, "tipo_material", "") or "").upper()
                    dens = 7.85
                    if ("ALUMIN" in cat) or ("ALUMIN" in tip):
                        dens = 2.7
                    elif ("INOX" in cat) or ("INOXID" in cat) or ("INOX" in tip) or ("INOXID" in tip):
                        dens = 8.0
                    vol_mm3 = largo * ancho * esp
                    vol_cm3 = vol_mm3 / 1000.0
                    peso_placa = (vol_cm3 * dens) / 1000.0
            per_piece = (rect_area / plate_area) * peso_placa
            total = float(per_piece) * float(max(0, int(qty or 0)))
            return float(round(total, 3))
        except Exception:
            return 0.0

    if request.method == "POST":
        form = LaserVigaLikeForm(request.POST, estados=estados)
        if form.is_valid():
            folio_externo = form.cleaned_data["folio_externo"]
            pieza = form.cleaned_data["pieza"]
            codigo = (getattr(pieza, "nombre", "") or "").strip()
            total = int(form.cleaned_data["total_piezas"] or 1)
            is_op = bool(int(total or 0) >= 2)
            estado_etapa = (form.cleaned_data.get("estado") or "").strip()
            if is_op and estado_etapa not in set(estados_op):
                form.add_error("estado", "Estado inválido para orden grande.")
            elif LaserOrdenProduccion.objects.filter(folio_externo_normalizado=folio_externo.upper()).exists():
                form.add_error("folio_externo", "Folio ya existe.")
            else:
                descripcion = (form.cleaned_data.get("descripcion") or "").strip()
                if not descripcion:
                    descripcion = codigo
                material = form.cleaned_data["material"]
                ancho_mm = int(form.cleaned_data["pieza_ancho_mm"] or 0)
                alto_mm = int(form.cleaned_data["pieza_alto_mm"] or 0)
                peso_total = _calc_peso_total(material, ancho_mm, alto_mm, total)
                orden = LaserOrdenProduccion.objects.create(
                    corta_cliente_proyecto=_get_or_create_corta_cliente_proyecto(form.cleaned_data["cliente_proyecto"]),
                    folio_externo=folio_externo,
                    codigo=codigo,
                    corta_pieza=pieza,
                    correo=(form.cleaned_data.get("correo") or "").strip(),
                    telefono=(form.cleaned_data.get("telefono") or "").strip(),
                    pieza_no=1,
                    total_piezas=total,
                    cantidad_objetivo=total,
                    es_op=bool(int(total or 0) >= 2),
                    cantidad_producida=0,
                    nombre=codigo,
                    descripcion=descripcion,
                    fecha_compromiso=form.cleaned_data["fecha_compromiso"],
                    prioridad=int(form.cleaned_data["prioridad"] or 3),
                    estado_etapa=estado_etapa,
                    material=material,
                    pieza_ancho_mm=ancho_mm,
                    pieza_alto_mm=alto_mm,
                    peso_kg=float(peso_total or 0.0),
                    ultimo_cambio=timezone.now(),
                    observaciones=(form.cleaned_data.get("observaciones") or "").strip(),
                    archivo=request.FILES.get("archivo"),
                    archivo_dxf=request.FILES.get("archivo_dxf"),
                )
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                LaserEstadoCambio.objects.create(
                    orden=orden,
                    estado_anterior="",
                    estado_nuevo=(orden.estado_etapa or "").strip(),
                    fecha_operacion=timezone.localdate(),
                    actor_username=actor,
                )
                per_piece = 0.0
                if int(orden.total_piezas or 0) > 0:
                    per_piece = float(orden.peso_kg or 0.0) / float(int(orden.total_piezas or 1))
                LaserOrdenItem.objects.create(
                    orden=orden,
                    etapa="Corte",
                    item_nombre=(orden.descripcion or orden.nombre or "").strip(),
                    item_peso_kg=float(per_piece or 0.0),
                    cantidad_requerida=int(total),
                )
                LaserAsignacion.objects.filter(orden=orden, etapa="Corte", vigente=True).update(vigente=False)
                for cid in selected_corte_operadores:
                    LaserAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Operador",
                        colaborador_id=int(cid),
                        vigente=True,
                        asignado_por=actor,
                    )
                for mid in selected_corte_maquinas:
                    LaserAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Máquina",
                        maquina_id=int(mid),
                        vigente=True,
                        asignado_por=actor,
                    )
                messages.success(request, "Pedido creado.")
                return redirect("catalogos:corte_laser_control")
    else:
        form = LaserVigaLikeForm(
            initial={
                "fecha_compromiso": timezone.localdate(),
                "estado": "Espera de corte",
                "prioridad": "3",
                "total_piezas": 1,
                "pieza_ancho_mm": 0,
                "pieza_alto_mm": 0,
            },
            estados=estados,
        )

    return render(
        request,
        "catalogos/corte_laser_form.html",
        {
            "form": form,
            "mode": "admin",
            "reset_url": reverse("catalogos:corte_laser_control"),
            "status_colors": status_colors,
            "archivo_url": "",
            "dxf_url": "",
            "clientes_proyectos": list(
                CortaClienteProyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre")[:2000]
            ),
            "materiales_payload": list(
                LaserMaterialPlaca.objects.filter(activo=True).values(
                    "id",
                    "categoria_material",
                    "tipo_material",
                    "nombre",
                    "calibre",
                    "espesor_mm",
                    "largo_mm",
                    "ancho_mm",
                    "peso_kg",
                )
            ),
            "corte_operadores": corte_operadores,
            "corte_maquinas": corte_maquinas,
            "selected_corte_operadores": selected_corte_operadores,
            "selected_corte_maquinas": selected_corte_maquinas,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_update(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    embedded = str(request.GET.get("embedded") or request.POST.get("embedded") or "").strip().lower() in {"1", "true", "yes", "on"}
    saved = str(request.GET.get("saved") or "").strip().lower() in {"1", "true", "yes", "on"}
    next_url = _safe_next_url(request.GET.get("next") or request.POST.get("next") or "")
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    estados_op = [
        "Espera de corte",
        "Corte",
        "Soldadura",
    ]
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }

    corte_operadores = list(Colaborador.objects.filter(activo=True, rol="Operador").order_by("nombre"))
    corte_maquinas = list(Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre"))
    current_asigs = list(
        LaserAsignacion.objects.filter(orden=orden, vigente=True, etapa="Corte").values_list("rol", "colaborador_id", "maquina_id")
    )
    selected_corte_operadores = [int(cid) for rol, cid, _mid in current_asigs if (rol or "") == "Operador" and cid]
    selected_corte_maquinas = [int(mid) for rol, _cid, mid in current_asigs if (rol or "") == "Máquina" and mid]

    def _calc_peso_total(material: LaserMaterialPlaca, ancho_mm: int, alto_mm: int, qty: int) -> float:
        try:
            plate_area = float(int(getattr(material, "largo_mm", 0) or 0)) * float(int(getattr(material, "ancho_mm", 0) or 0))
            if not (plate_area > 0):
                return 0.0
            margin = 30.0
            rect_area = float(max(0, ancho_mm)) + (2.0 * margin)
            rect_area *= float(max(0, alto_mm)) + (2.0 * margin)
            peso_placa = float(getattr(material, "peso_kg", 0.0) or 0.0)
            if not (peso_placa > 0):
                esp = float(getattr(material, "espesor_mm", 0.0) or 0.0)
                largo = float(int(getattr(material, "largo_mm", 0) or 0))
                ancho = float(int(getattr(material, "ancho_mm", 0) or 0))
                if esp > 0 and largo > 0 and ancho > 0:
                    cat = (getattr(material, "categoria_material", "") or "").upper()
                    tip = (getattr(material, "tipo_material", "") or "").upper()
                    dens = 7.85
                    if ("ALUMIN" in cat) or ("ALUMIN" in tip):
                        dens = 2.7
                    elif ("INOX" in cat) or ("INOXID" in cat) or ("INOX" in tip) or ("INOXID" in tip):
                        dens = 8.0
                    vol_mm3 = largo * ancho * esp
                    vol_cm3 = vol_mm3 / 1000.0
                    peso_placa = (vol_cm3 * dens) / 1000.0
            per_piece = (rect_area / plate_area) * peso_placa
            total = float(per_piece) * float(max(0, int(qty or 0)))
            return float(round(total, 3))
        except Exception:
            return 0.0

    if request.method == "POST":
        form = LaserVigaLikeForm(request.POST, estados=estados)
        if form.is_valid():
            folio_externo = form.cleaned_data["folio_externo"]
            pieza = form.cleaned_data["pieza"]
            codigo = (getattr(pieza, "nombre", "") or "").strip()
            total = int(form.cleaned_data["total_piezas"] or 1)
            is_op = bool(int(total or 0) >= 2)
            estado_etapa = (form.cleaned_data.get("estado") or "").strip()
            if is_op and estado_etapa not in set(estados_op):
                form.add_error("estado", "Estado inválido para orden grande.")
            elif LaserOrdenProduccion.objects.exclude(id=orden.id).filter(folio_externo_normalizado=folio_externo.upper()).exists():
                form.add_error("folio_externo", "Folio ya existe.")
            else:
                orden.corta_cliente_proyecto = _get_or_create_corta_cliente_proyecto(form.cleaned_data["cliente_proyecto"])
                orden.folio_externo = folio_externo
                orden.codigo = codigo
                orden.corta_pieza = pieza
                orden.correo = (form.cleaned_data.get("correo") or "").strip()
                orden.telefono = (form.cleaned_data.get("telefono") or "").strip()
                orden.pieza_no = 1
                orden.total_piezas = total
                orden.cantidad_objetivo = total
                orden.es_op = bool(int(total or 0) >= 2)
                orden.nombre = codigo
                descripcion = (form.cleaned_data.get("descripcion") or "").strip()
                if not descripcion:
                    descripcion = codigo
                orden.descripcion = descripcion
                orden.fecha_compromiso = form.cleaned_data["fecha_compromiso"]
                orden.prioridad = int(form.cleaned_data["prioridad"] or 3)
                orden.estado_etapa = estado_etapa
                material = form.cleaned_data["material"]
                orden.material = material
                orden.pieza_ancho_mm = int(form.cleaned_data["pieza_ancho_mm"] or 0)
                orden.pieza_alto_mm = int(form.cleaned_data["pieza_alto_mm"] or 0)
                orden.peso_kg = float(_calc_peso_total(material, int(orden.pieza_ancho_mm or 0), int(orden.pieza_alto_mm or 0), total) or 0.0)
                orden.observaciones = (form.cleaned_data.get("observaciones") or "").strip()
                estado_prev = (LaserOrdenProduccion.objects.filter(id=orden.id).values_list("estado_etapa", flat=True).first() or "").strip()
                archivo = request.FILES.get("archivo")
                if archivo:
                    orden.archivo = archivo
                archivo_dxf = request.FILES.get("archivo_dxf")
                if archivo_dxf:
                    orden.archivo_dxf = archivo_dxf
                orden.save()
                if estado_prev and estado_prev != (orden.estado_etapa or "").strip():
                    actor = ""
                    try:
                        actor = request.user.get_username() or ""
                    except Exception:
                        actor = ""
                    LaserEstadoCambio.objects.create(
                        orden=orden,
                        estado_anterior=estado_prev,
                        estado_nuevo=(orden.estado_etapa or "").strip(),
                        fecha_operacion=timezone.localdate(),
                        actor_username=actor,
                        comentario="manual_edit",
                    )
                it = LaserOrdenItem.objects.filter(orden=orden).order_by("id").first()
                if it:
                    per_piece = 0.0
                    if int(orden.total_piezas or 0) > 0:
                        per_piece = float(orden.peso_kg or 0.0) / float(int(orden.total_piezas or 1))
                    it.item_nombre = (orden.descripcion or orden.nombre or "").strip()
                    it.item_peso_kg = float(per_piece or 0.0)
                    it.cantidad_requerida = int(total)
                    it.save()
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                LaserAsignacion.objects.filter(orden=orden, etapa="Corte", vigente=True).update(vigente=False)
                selected_corte_operadores = [int(x) for x in request.POST.getlist("corte_operador_ids") if (x or "").isdigit()]
                selected_corte_maquinas = [int(x) for x in request.POST.getlist("corte_maquina_ids") if (x or "").isdigit()]
                for cid in selected_corte_operadores:
                    LaserAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Operador",
                        colaborador_id=int(cid),
                        vigente=True,
                        asignado_por=actor,
                    )
                for mid in selected_corte_maquinas:
                    LaserAsignacion.objects.create(
                        orden=orden,
                        etapa="Corte",
                        rol="Máquina",
                        maquina_id=int(mid),
                        vigente=True,
                        asignado_por=actor,
                    )
                messages.success(request, "Pedido actualizado.")
                if embedded:
                    u = reverse("catalogos:corte_laser_update", args=[int(orden.id)])
                    u = u + "?embedded=1&saved=1"
                    if next_url:
                        u = u + "&next=" + quote(next_url)
                    return redirect(u)
                return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
    else:
        form = LaserVigaLikeForm(
            initial={
                "folio_externo": getattr(orden, "folio_externo", "") or "",
                "pieza": getattr(orden, "corta_pieza_id", None),
                "cliente_proyecto": getattr(getattr(orden, "corta_cliente_proyecto", None), "nombre", "") or "",
                "correo": getattr(orden, "correo", "") or "",
                "telefono": getattr(orden, "telefono", "") or "",
                "descripcion": orden.descripcion,
                "material": orden.material_id,
                "pieza_ancho_mm": int(getattr(orden, "pieza_ancho_mm", 0) or 0),
                "pieza_alto_mm": int(getattr(orden, "pieza_alto_mm", 0) or 0),
                "fecha_compromiso": orden.fecha_compromiso,
                "estado": orden.estado_etapa,
                "prioridad": str(int(orden.prioridad or 3)),
                "total_piezas": int(orden.total_piezas or 1),
                "observaciones": orden.observaciones,
            },
            estados=estados,
        )

    archivo_url = orden.archivo.url if getattr(orden, "archivo", None) else ""
    dxf_url = orden.archivo_dxf.url if getattr(orden, "archivo_dxf", None) else ""
    return render(
        request,
        "catalogos/corte_laser_form.html",
        {
            "form": form,
            "mode": "admin",
            "reset_url": reverse("catalogos:corte_laser_control"),
            "embedded": embedded,
            "saved": saved,
            "next_url": next_url,
            "status_colors": status_colors,
            "archivo_url": archivo_url,
            "dxf_url": dxf_url,
            "clientes_proyectos": list(
                CortaClienteProyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre")[:2000]
            ),
            "materiales_payload": list(
                LaserMaterialPlaca.objects.filter(activo=True).values(
                    "id",
                    "categoria_material",
                    "tipo_material",
                    "nombre",
                    "calibre",
                    "espesor_mm",
                    "largo_mm",
                    "ancho_mm",
                    "peso_kg",
                )
            ),
            "corte_operadores": corte_operadores,
            "corte_maquinas": corte_maquinas,
            "selected_corte_operadores": selected_corte_operadores,
            "selected_corte_maquinas": selected_corte_maquinas,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_delete(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    next_url = _safe_next_url(request.GET.get("next") or request.POST.get("next") or "")
    if request.method == "POST":
        with transaction.atomic(using="mes"):
            orden_locked = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_for_update(), pk=pk)
            if LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(orden_locked.id)).exists():
                messages.error(request, "No se puede eliminar: el pedido ya tiene envíos en Logística (historial).")
                return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
            try:
                LaserProduccion.objects.using("mes").filter(orden_item__orden_id=int(orden_locked.id)).delete()
                LaserAsignacion.objects.using("mes").filter(orden_id=int(orden_locked.id)).delete()
                orden_locked.delete(using="mes")
            except ProtectedError:
                messages.error(request, "No se puede eliminar: el pedido tiene registros protegidos (envíos/historial).")
                return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
        messages.success(request, "Eliminado.")
        return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
    return render(request, "catalogos/corte_laser_confirm_delete.html", {"orden": orden})


@login_required
@require_http_methods(["POST"])
def corte_laser_update_meta_json(request, pk: int):
    if not _can_corte_laser(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    fecha = (request.POST.get("fecha_compromiso") or "").strip()
    prioridad = (request.POST.get("prioridad") or "").strip()
    if not fecha:
        return JsonResponse({"ok": False, "error": "Fecha inválida."}, status=400)
    try:
        fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Fecha inválida."}, status=400)
    try:
        pr = int(prioridad or 0)
    except Exception:
        pr = 0
    if pr not in {1, 2, 3, 4, 5}:
        return JsonResponse({"ok": False, "error": "Prioridad inválida."}, status=400)
    orden.fecha_compromiso = fecha_dt
    orden.prioridad = pr
    orden.save(update_fields=["fecha_compromiso", "prioridad", "actualizado_en"])
    return JsonResponse({"ok": True, "id": int(orden.id), "fecha_compromiso": fecha_dt.isoformat(), "prioridad": pr})


@login_required
@require_http_methods(["POST"])
def corte_laser_update_avance_json(request, pk: int):
    if not _can_corte_laser(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    raw_sold = (request.POST.get("cantidad_soldada") or request.POST.get("cantidad_producida") or "").strip()
    raw_pint = (request.POST.get("cantidad_pintada") or "").strip()
    raw_term = (request.POST.get("cantidad_terminada") or "").strip()
    try:
        sold = int(raw_sold or 0)
    except Exception:
        sold = -1
    try:
        pint = int(raw_pint or 0)
    except Exception:
        pint = -1
    try:
        term = int(raw_term or 0)
    except Exception:
        term = -1
    with transaction.atomic(using="mes"):
        orden = get_object_or_404(LaserOrdenProduccion.objects.select_for_update(), pk=pk)
        if orden.estado != "Abierta":
            return JsonResponse({"ok": False, "error": "La orden está cerrada/cancelada."}, status=409)
        if int(getattr(orden, "total_piezas", 0) or 0) < 2:
            return JsonResponse({"ok": False, "error": "Este registro no es una orden grande."}, status=409)
        now = timezone.now()
        etapa = (getattr(orden, "estado_etapa", "") or "").strip()
        if etapa == CIERRE_PENDIENTE_LABEL:
            hasta = getattr(orden, "cierre_pendiente_hasta", None)
            if not hasta or hasta <= now:
                orden.estado_etapa = CIERRE_FINAL_LABEL
                orden.ultimo_cambio = now
                orden.cierre_bloqueado_en = now
                orden.cierre_pendiente_hasta = None
                orden.save(
                    update_fields=["estado_etapa", "ultimo_cambio", "cierre_bloqueado_en", "cierre_pendiente_hasta", "actualizado_en"]
                )
                LaserEstadoCambio.objects.create(
                    orden=orden,
                    estado_anterior=CIERRE_PENDIENTE_LABEL,
                    estado_nuevo=CIERRE_FINAL_LABEL,
                    fecha_operacion=timezone.localdate(),
                    actor_username="system",
                    comentario="auto_bloqueo",
                )
                return JsonResponse({"ok": False, "error": "Orden bloqueada: cierre confirmado."}, status=409)
        if etapa not in {"Soldadura", CIERRE_PENDIENTE_LABEL}:
            return JsonResponse({"ok": False, "error": "El avance se habilita a partir de Soldadura."}, status=409)

        objetivo = int(getattr(orden, "total_piezas", 0) or 0)
        if objetivo <= 0:
            return JsonResponse({"ok": False, "error": "Cantidad objetivo inválida."}, status=400)
        if sold < 0 or pint < 0 or term < 0:
            return JsonResponse({"ok": False, "error": "Cantidades inválidas."}, status=400)
        if sold > objetivo or pint > objetivo or term > objetivo:
            return JsonResponse({"ok": False, "error": "Cantidades inválidas."}, status=400)

        prev_term = int(getattr(orden, "cantidad_terminada", 0) or 0)
        orden.cantidad_producida = sold
        orden.cantidad_pintada = pint
        orden.cantidad_terminada = term
        orden.ultimo_cambio = now
        orden.save(update_fields=["cantidad_producida", "cantidad_pintada", "cantidad_terminada", "ultimo_cambio", "actualizado_en"])

        delta_term = int(term) - int(prev_term)
        if delta_term != 0:
            producto = _corta_producto_from_orden(orden)
            if producto:
                stock = _get_or_create_stock_corta_locked(producto)
                new_stock = int(stock.stock or 0) + int(delta_term)
                stock.stock = max(0, new_stock)
                stock.save(update_fields=["stock", "actualizado_en"])
                LogisticaMovimientoCorta.objects.create(
                    tipo="stock_in" if delta_term > 0 else "ajuste",
                    producto=producto,
                    orden=orden,
                    cantidad=int(delta_term),
                    actor=actor,
                )

        reload_ui = False
        if int(term) == int(objetivo) and (getattr(orden, "estado_etapa", "") or "").strip() not in {CIERRE_PENDIENTE_LABEL, CIERRE_FINAL_LABEL}:
            prev_etapa = (getattr(orden, "estado_etapa", "") or "").strip()
            orden.estado_etapa = CIERRE_PENDIENTE_LABEL
            orden.ultimo_cambio = now
            orden.cierre_pendiente_en = now
            orden.cierre_pendiente_hasta = now + timedelta(minutes=CIERRE_VENTANA_MINUTOS)
            orden.cierre_pendiente_por = actor
            orden.cierre_estado_previo = prev_etapa
            orden.save(
                update_fields=[
                    "estado_etapa",
                    "ultimo_cambio",
                    "cierre_pendiente_en",
                    "cierre_pendiente_hasta",
                    "cierre_pendiente_por",
                    "cierre_estado_previo",
                    "actualizado_en",
                ]
            )
            LaserEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=prev_etapa,
                estado_nuevo=CIERRE_PENDIENTE_LABEL,
                fecha_operacion=timezone.localdate(),
                actor_username=actor,
                comentario="cierre_pendiente",
            )
            reload_ui = True
        elif int(term) < int(objetivo) and (getattr(orden, "estado_etapa", "") or "").strip() == CIERRE_PENDIENTE_LABEL:
            prev_etapa = CIERRE_PENDIENTE_LABEL
            target = (getattr(orden, "cierre_estado_previo", "") or "").strip() or "Soldadura"
            orden.estado_etapa = target
            orden.ultimo_cambio = now
            orden.cierre_revertido_en = now
            orden.cierre_revertido_por = actor
            orden.cierre_pendiente_hasta = None
            orden.save(
                update_fields=[
                    "estado_etapa",
                    "ultimo_cambio",
                    "cierre_revertido_en",
                    "cierre_revertido_por",
                    "cierre_pendiente_hasta",
                    "actualizado_en",
                ]
            )
            LaserEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=prev_etapa,
                estado_nuevo=target,
                fecha_operacion=timezone.localdate(),
                actor_username=actor,
                comentario="auto_reabrir_por_correccion",
            )
            reload_ui = True

    sold_pct = 0.0 if objetivo <= 0 else round(min(100.0, (sold / objetivo) * 100.0), 2)
    pint_pct = 0.0 if objetivo <= 0 else round(min(100.0, (pint / objetivo) * 100.0), 2)
    term_pct = 0.0 if objetivo <= 0 else round(min(100.0, (term / objetivo) * 100.0), 2)
    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "soldadas": sold,
            "pintadas": pint,
            "terminadas": term,
            "objetivo": objetivo,
            "soldadas_pct": sold_pct,
            "pintadas_pct": pint_pct,
            "terminadas_pct": term_pct,
            "reload": reload_ui,
        }
    )


@login_required
@require_http_methods(["POST"])
def corte_laser_revertir_cierre(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    next_url = _safe_next_url(request.POST.get("next") or "")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    now = timezone.now()
    with transaction.atomic(using="mes"):
        orden = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_for_update(), pk=pk)
        if orden.estado != "Abierta":
            messages.error(request, "La orden está cerrada/cancelada.")
            return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
        if int(getattr(orden, "total_piezas", 0) or 0) < 2:
            messages.error(request, "No aplica: no es una orden grande.")
            return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
        etapa = (getattr(orden, "estado_etapa", "") or "").strip()
        if etapa != CIERRE_PENDIENTE_LABEL:
            messages.error(request, "No hay cierre pendiente para revertir.")
            return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
        hasta = getattr(orden, "cierre_pendiente_hasta", None)
        if not hasta or hasta <= now:
            orden.estado_etapa = CIERRE_FINAL_LABEL
            orden.ultimo_cambio = now
            orden.cierre_bloqueado_en = now
            orden.cierre_pendiente_hasta = None
            orden.save(
                update_fields=["estado_etapa", "ultimo_cambio", "cierre_bloqueado_en", "cierre_pendiente_hasta", "actualizado_en"]
            )
            LaserEstadoCambio.objects.create(
                orden=orden,
                estado_anterior=CIERRE_PENDIENTE_LABEL,
                estado_nuevo=CIERRE_FINAL_LABEL,
                fecha_operacion=timezone.localdate(),
                actor_username="system",
                comentario="auto_bloqueo",
            )
            messages.error(request, "La ventana expiró; la orden ya quedó bloqueada.")
            return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")
        target = (getattr(orden, "cierre_estado_previo", "") or "").strip() or "Soldadura"
        orden.estado_etapa = target
        orden.ultimo_cambio = now
        orden.cierre_revertido_en = now
        orden.cierre_revertido_por = actor
        orden.cierre_pendiente_hasta = None
        orden.save(
            update_fields=[
                "estado_etapa",
                "ultimo_cambio",
                "cierre_revertido_en",
                "cierre_revertido_por",
                "cierre_pendiente_hasta",
                "actualizado_en",
            ]
        )
        LaserEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=CIERRE_PENDIENTE_LABEL,
            estado_nuevo=target,
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="revertir_cierre",
        )
    messages.success(request, "Cierre revertido.")
    return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["POST"])
def corte_laser_asignaciones(request, pk: int):
    if not _can_corte_laser(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    corte_operador_ids = [int(x) for x in request.POST.getlist("corte_operador_ids") if (x or "").isdigit()]
    corte_maquina_ids = [int(x) for x in request.POST.getlist("corte_maquina_ids") if (x or "").isdigit()]
    soldadura_ids = [int(x) for x in request.POST.getlist("soldadura_ids") if (x or "").isdigit()]
    pintura_ids = [int(x) for x in request.POST.getlist("pintura_ids") if (x or "").isdigit()]
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    LaserAsignacion.objects.filter(orden=orden, etapa__in=["Corte", "Soldadura", "Pintura"], vigente=True).update(vigente=False)
    for cid in corte_operador_ids:
        LaserAsignacion.objects.create(orden=orden, etapa="Corte", rol="Operador", colaborador_id=cid, vigente=True, asignado_por=actor)
    for mid in corte_maquina_ids:
        LaserAsignacion.objects.create(orden=orden, etapa="Corte", rol="Máquina", maquina_id=mid, vigente=True, asignado_por=actor)
    for cid in soldadura_ids:
        LaserAsignacion.objects.create(orden=orden, etapa="Soldadura", rol="Asignado", colaborador_id=cid, vigente=True, asignado_por=actor)
    for cid in pintura_ids:
        LaserAsignacion.objects.create(orden=orden, etapa="Pintura", rol="Asignado", colaborador_id=cid, vigente=True, asignado_por=actor)

    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "corte_operador_ids": corte_operador_ids,
            "corte_maquina_ids": corte_maquina_ids,
            "soldadura_ids": soldadura_ids,
            "pintura_ids": pintura_ids,
        }
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_ordenes(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    return redirect("catalogos:corte_laser_control")

    form = LaserOrdenProduccionForm()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_create":
            form = LaserOrdenProduccionForm(request.POST)
            if form.is_valid():
                orden = form.save(commit=False)
                orden.estado_etapa = "Espera de corte"
                orden.es_individual = False
                orden.ultimo_cambio = timezone.now()
                orden.save()
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                LaserEstadoCambio.objects.create(
                    orden=orden,
                    estado_anterior="",
                    estado_nuevo=(orden.estado_etapa or "").strip(),
                    fecha_operacion=timezone.localdate(),
                    actor_username=actor,
                )
                messages.success(request, "Orden creada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        if action == "etapa_next":
            oid = int(request.POST.get("orden_id") or 0)
            o = LaserOrdenProduccion.objects.filter(id=oid, es_individual=False).first()
            if o and o.estado == "Abierta":
                estados = [
                    "Espera de corte",
                    "Corte",
                    "Espera de armado",
                    "Armado",
                    "Espera de soldadura",
                    "Soldadura",
                    "Espera de pintura",
                    "Pintura",
                    "Terminado",
                ]
                try:
                    idx = estados.index((o.estado_etapa or "").strip())
                except Exception:
                    idx = -1
                nxt = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
                if nxt:
                    prev = (o.estado_etapa or "").strip()
                    o.estado_etapa = nxt
                    o.ultimo_cambio = timezone.now()
                    o.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
                    actor = ""
                    try:
                        actor = request.user.get_username() or ""
                    except Exception:
                        actor = ""
                    LaserEstadoCambio.objects.create(
                        orden=o,
                        estado_anterior=prev,
                        estado_nuevo=nxt,
                        fecha_operacion=timezone.localdate(),
                        actor_username=actor,
                    )
                    messages.success(request, f"Actualizado a {nxt}.")
            return redirect("catalogos:corte_laser_ordenes")

    ordenes_abiertas = list(
        LaserOrdenProduccion.objects.filter(estado="Abierta", es_individual=False)
        .select_related("corta_cliente_proyecto")
        .order_by("-id")[:200]
    )
    ordenes_recientes = list(
        LaserOrdenProduccion.objects.exclude(estado="Abierta")
        .filter(es_individual=False)
        .select_related("corta_cliente_proyecto")
        .order_by("-id")[:50]
    )
    orden_ids = [int(o.id) for o in (ordenes_abiertas + ordenes_recientes)]
    if orden_ids:
        req_rows = (
            LaserOrdenItem.objects.filter(orden_id__in=orden_ids)
            .values("orden_id", "etapa")
            .annotate(req=Coalesce(Sum("cantidad_requerida"), 0))
        )
        prod_rows = (
            LaserProduccion.objects.filter(orden_item__orden_id__in=orden_ids)
            .values("orden_item__orden_id", "orden_item__etapa")
            .annotate(prod=Coalesce(Sum("cantidad"), 0))
        )
        req_map = {(int(r["orden_id"]), str(r["etapa"])): int(r["req"] or 0) for r in req_rows}
        prod_map = {(int(r["orden_item__orden_id"]), str(r["orden_item__etapa"])): int(r["prod"] or 0) for r in prod_rows}
        etapas = ["Corte", "Armado", "Soldadura", "Pintura"]
        for o in (ordenes_abiertas + ordenes_recientes):
            sp = {}
            for e in etapas:
                req = req_map.get((int(o.id), e), 0)
                prod = prod_map.get((int(o.id), e), 0)
                sp[e] = 0.0 if req <= 0 else min(100.0, (prod / req) * 100.0)
            o.stage_pcts = sp
            estados = [
                "Espera de corte",
                "Corte",
                "Espera de armado",
                "Armado",
                "Espera de soldadura",
                "Soldadura",
                "Espera de pintura",
                "Pintura",
                "Terminado",
            ]
            try:
                idx = estados.index((o.estado_etapa or "").strip())
            except Exception:
                idx = -1
            o.next_estado = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
    return render(
        request,
        "catalogos/corte_laser_ordenes.html",
        {
            "form": form,
            "ordenes_abiertas": ordenes_abiertas,
            "ordenes_recientes": ordenes_recientes,
            "clientes_proyectos": list(
                CortaClienteProyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre")[:2000]
            ),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def corte_laser_orden_detalle(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    return redirect("catalogos:corte_laser_control")

    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    edit_form = LaserOrdenProduccionForm(instance=orden)
    item_form = LaserOrdenItemForm()
    assign_form = LaserOrdenAsignacionForm()
    estados = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    try:
        idx = estados.index((orden.estado_etapa or "").strip())
    except Exception:
        idx = -1
    next_estado = "" if idx < 0 or idx >= (len(estados) - 1) else estados[idx + 1]
    etapa_activa = (orden.estado_etapa or "").strip()
    etapa_permitida = etapa_activa if etapa_activa in {"Corte", "Armado", "Soldadura", "Pintura"} else None
    prod_form = LaserOrdenRegistroProduccionForm(
        orden, initial={"fecha": timezone.localdate(), "cantidad": 1}, etapa_permitida=etapa_permitida
    )
    has_prod = LaserProduccion.objects.filter(orden_item__orden=orden).exists()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "orden_update":
            if orden.estado != "Abierta":
                messages.error(request, "La orden no se puede editar (cerrada/cancelada).")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
            edit_form = LaserOrdenProduccionForm(request.POST, instance=orden)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "Orden actualizada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "orden_delete":
            if has_prod:
                messages.error(request, "No se puede eliminar: la orden ya tiene producción registrada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
            try:
                orden.delete()
            except Exception:
                messages.error(request, "No se pudo eliminar la orden.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
            messages.success(request, "Orden eliminada.")
            return redirect("catalogos:corte_laser_ordenes")
        elif action == "item_add":
            item_form = LaserOrdenItemForm(request.POST)
            if item_form.is_valid():
                try:
                    LaserOrdenItem.objects.create(
                        orden=orden,
                        etapa=item_form.cleaned_data["etapa"],
                        item_nombre=(item_form.cleaned_data.get("item_nombre") or "").strip(),
                        item_peso_kg=float(item_form.cleaned_data.get("item_peso_kg") or 0.0),
                        cantidad_requerida=item_form.cleaned_data["cantidad_requerida"],
                    )
                    messages.success(request, "Item agregado.")
                    return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
                except Exception:
                    messages.error(request, "No se pudo agregar el item.")
        elif action == "item_delete":
            item_id = int(request.POST.get("item_id") or 0)
            if item_id:
                LaserOrdenItem.objects.filter(pk=item_id, orden=orden).delete()
                messages.success(request, "Item removido de la orden.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "prod_add":
            if not etapa_permitida:
                messages.error(request, "La orden no está en una etapa de producción activa.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
            prod_form = LaserOrdenRegistroProduccionForm(orden, request.POST, etapa_permitida=etapa_permitida)
            if prod_form.is_valid():
                item = prod_form.cleaned_data["orden_item"]
                operadores = list(prod_form.cleaned_data.get("operador") or [])
                cantidad_total = int(prod_form.cleaned_data["cantidad"] or 0)
                if cantidad_total <= 0 or not operadores:
                    messages.error(request, "Debes seleccionar operador(es) y una cantidad válida.")
                    return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))

                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""

                n_ops = len(operadores)
                base = cantidad_total // n_ops
                rem = cantidad_total % n_ops
                for i, op in enumerate(operadores):
                    qty = base + (1 if i < rem else 0)
                    if qty <= 0:
                        continue
                    LaserProduccion.objects.create(
                        fecha=prod_form.cleaned_data["fecha"],
                        cantidad=int(qty),
                        operador=op,
                        orden_item=item,
                        observaciones=prod_form.cleaned_data.get("observaciones") or "",
                    )
                    LaserOrdenAsignacion.objects.get_or_create(
                        orden=orden,
                        etapa=str(getattr(item, "etapa", "") or "").strip() or str(etapa_permitida),
                        colaborador=op,
                        defaults={"asignado_por": actor},
                    )
                messages.success(request, "Producción registrada en la orden.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "etapa_next":
            if orden.estado != "Abierta":
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
            if next_estado:
                orden.estado_etapa = next_estado
                orden.save(update_fields=["estado_etapa", "actualizado_en"])
                messages.success(request, f"Actualizado a {next_estado}.")
            return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "asig_add":
            assign_form = LaserOrdenAsignacionForm(request.POST)
            if assign_form.is_valid():
                etapa = assign_form.cleaned_data["etapa"]
                col = assign_form.cleaned_data["colaborador"]
                actor = ""
                try:
                    actor = request.user.get_username() or ""
                except Exception:
                    actor = ""
                LaserOrdenAsignacion.objects.get_or_create(
                    orden=orden,
                    etapa=etapa,
                    colaborador=col,
                    defaults={"asignado_por": actor},
                )
                messages.success(request, "Asignación guardada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "asig_delete":
            asig_id = int(request.POST.get("asig_id") or 0)
            if asig_id:
                LaserOrdenAsignacion.objects.filter(id=asig_id, orden=orden).delete()
                messages.success(request, "Asignación eliminada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "orden_close":
            if orden.estado == "Abierta":
                orden.estado = "Cerrada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cerrada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))
        elif action == "orden_cancel":
            if orden.estado == "Abierta":
                orden.estado = "Cancelada"
                orden.save(update_fields=["estado", "actualizado_en"])
                messages.success(request, "Orden cancelada.")
                return redirect("catalogos:corte_laser_orden_detalle", pk=int(orden.id))

    items = list(LaserOrdenItem.objects.filter(orden=orden).order_by("id"))
    prod_rows = (
        LaserProduccion.objects.filter(orden_item__orden=orden)
        .values("orden_item_id")
        .annotate(producidas=Coalesce(Sum("cantidad"), 0))
    )
    prod_map = {int(r["orden_item_id"]): r for r in prod_rows if r.get("orden_item_id")}
    resumen = []
    pcts = []
    for it in items:
        row = prod_map.get(int(it.id), {})
        producidas = int(row.get("producidas") or 0)
        requeridas = int(it.cantidad_requerida or 0)
        faltan = max(0, requeridas - producidas)
        pct = 0.0 if requeridas <= 0 else min(100.0, (producidas / requeridas) * 100.0)
        pcts.append(pct)
        kg = float(it.pieza_peso_kg or 0.0) * producidas
        resumen.append(
            {
                "item": it,
                "requeridas": requeridas,
                "producidas": producidas,
                "faltan": faltan,
                "pct": round(pct, 2),
                "kg": round(float(kg or 0.0), 3),
            }
        )
    avance_pct = round(min(pcts) if pcts else 0.0, 2)

    historial = list(
        LaserProduccion.objects.filter(orden_item__orden=orden)
        .select_related("operador", "orden_item")
        .order_by("-fecha", "-id")[:250]
    )
    asignaciones = list(LaserOrdenAsignacion.objects.filter(orden=orden).select_related("colaborador").order_by("-asignado_en"))
    return render(
        request,
        "catalogos/corte_laser_orden_detalle.html",
        {
            "orden": orden,
            "avance_pct": avance_pct,
            "edit_form": edit_form,
            "can_delete": (not has_prod),
            "item_form": item_form,
            "prod_form": prod_form,
            "assign_form": assign_form,
            "resumen": resumen,
            "historial": historial,
            "asignaciones": asignaciones,
            "next_estado": next_estado,
            "clientes_proyectos": list(
                CortaClienteProyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre")[:2000]
            ),
        },
    )


@login_required
@require_http_methods(["POST"])
def corte_laser_change_status_json(request, pk: int):
    if not _can_corte_laser(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    estado_nuevo = (request.POST.get("estado_nuevo") or "").strip()
    if not estado_nuevo:
        return JsonResponse({"ok": False, "error": "Estado inválido."}, status=400)
    allowed_full = {
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    }
    allowed_op = {
        "Espera de corte",
        "Corte",
        "Soldadura",
    }
    is_op = bool(int(getattr(orden, "total_piezas", 0) or 0) >= 2)
    allowed = allowed_op if is_op else allowed_full
    if estado_nuevo not in allowed:
        return JsonResponse({"ok": False, "error": "Estado inválido."}, status=400)
    if orden.estado != "Abierta":
        return JsonResponse({"ok": False, "error": "La orden está cerrada/cancelada."}, status=409)
    fecha_operacion = (request.POST.get("fecha_operacion") or "").strip()
    if not fecha_operacion:
        return JsonResponse({"ok": False, "error": "Fecha requerida."}, status=400)
    try:
        fecha_dt = datetime.strptime(fecha_operacion, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Fecha inválida."}, status=400)

    estados_order = [
        "Espera de corte",
        "Corte",
        "Soldadura",
    ] if is_op else [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
        "Pintura",
        "Terminado",
    ]
    cur = (orden.estado_etapa or "").strip()
    try:
        idx_cur = estados_order.index(cur)
    except Exception:
        idx_cur = -1
    try:
        idx_new = estados_order.index(estado_nuevo)
    except Exception:
        idx_new = -1
    retro = bool(idx_cur >= 0 and idx_new >= 0 and idx_new < idx_cur)
    if retro:
        motivo = (request.POST.get("motivo_retroceso") or "").strip()
        if motivo not in {"error_dedo", "retrabajo"}:
            return JsonResponse({"ok": False, "error": "Motivo del retroceso requerido."}, status=400)
    else:
        motivo = ""

    comentario = (request.POST.get("comentario") or "").strip()
    estado_prev = (orden.estado_etapa or "").strip()
    orden.estado_etapa = estado_nuevo
    orden.ultimo_cambio = timezone.now()
    orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    LaserEstadoCambio.objects.create(
        orden=orden,
        estado_anterior=estado_prev,
        estado_nuevo=estado_nuevo,
        fecha_operacion=fecha_dt,
        actor_username=actor,
        motivo_retroceso=motivo,
        comentario=comentario,
    )
    status_colors = {
        "Espera de corte": "#ff8f00",
        "Corte": "#f39c12",
        "Espera de armado": "#d35400",
        "Armado": "#3498db",
        "Espera de soldadura": "#5d6d7e",
        "Soldadura": "#9b59b6",
        "Espera de pintura": "#7f8c8d",
        "Pintura": "#16a085",
        "Terminado": "#2dce89",
    }
    try:
        idx = estados_order.index(estado_nuevo)
    except Exception:
        idx = -1
    next_estado = "" if idx < 0 or idx >= (len(estados_order) - 1) else estados_order[idx + 1]
    ultimo_mov = timezone.localtime(orden.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if orden.ultimo_cambio else ""
    return JsonResponse(
        {
            "ok": True,
            "id": int(orden.id),
            "estado": estado_nuevo,
            "estado_color": status_colors.get(estado_nuevo, "#8898aa"),
            "estado_clase": clase_de_estado(estado_nuevo),
            "next_estado": next_estado,
            "ultimo_mov": ultimo_mov,
        }
    )


@login_required
@require_http_methods(["POST"])
def corte_laser_enviar(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:corte_laser_control")
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    if orden.estado != "Abierta" or (orden.estado_etapa or "").strip() != "Terminado":
        return redirect("catalogos:corte_laser_control")
    now = timezone.now()
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        prev = (orden.estado_etapa or "").strip()
        orden.estado_etapa = "Enviado"
        orden.ultimo_cambio = now
        orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
        LaserEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=prev,
            estado_nuevo="Enviado",
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="",
        )
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["POST"])
def corte_laser_regresar_produccion(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:corte_laser_control")
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    if orden.estado != "Abierta" or (orden.estado_etapa or "").strip() != "Enviado":
        return redirect("catalogos:corte_laser_control")
    now = timezone.now()
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        prev = (orden.estado_etapa or "").strip()
        orden.estado_etapa = "Terminado"
        orden.ultimo_cambio = now
        orden.save(update_fields=["estado_etapa", "ultimo_cambio", "actualizado_en"])
        LaserEstadoCambio.objects.create(
            orden=orden,
            estado_anterior=prev,
            estado_nuevo="Terminado",
            fecha_operacion=timezone.localdate(),
            actor_username=actor,
            comentario="",
        )
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["POST"])
def corte_laser_delete_decote(request, pk: int):
    if not _can_corte_laser(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:corte_laser_control")
    orden = get_object_or_404(LaserOrdenProduccion, pk=pk)
    if orden.estado != "Abierta":
        return redirect("catalogos:corte_laser_control")
    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    total = int(getattr(orden, "total_piezas", 0) or 0)
    apartado = int(getattr(orden, "apartado", 0) or 0)
    enviado = int(getattr(orden, "enviado", 0) or 0)
    saldo = max(0, total - apartado - enviado)
    if not (total > 0 and saldo == 0 and enviado > 0):
        return redirect("catalogos:corte_laser_control")
    last_enviado = (
        LogisticaEnvioCorta.objects.filter(orden_id=int(orden.id))
        .aggregate(mx=Max("fecha"))
        .get("mx")
    )
    if not last_enviado or last_enviado > cutoff:
        return redirect("catalogos:corte_laser_control")
    exp = LogisticaExpediente.objects.filter(orden_corta_id=int(orden.id)).first()
    if not exp or not getattr(exp, "generado_en", None) or int(getattr(exp, "descargas_count", 0) or 0) <= 0:
        return redirect("catalogos:corte_laser_control")
    with transaction.atomic(using="mes"):
        LaserProduccion.objects.filter(orden_item__orden=orden).delete()
        LaserAsignacion.objects.filter(orden=orden).delete()
        orden.delete()
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["POST"])
def corte_laser_delete_decote_all(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    if not _is_admin_user(request.user):
        return redirect("catalogos:corte_laser_control")
    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    filters = {
        "cliente": (request.POST.get("cliente") or "").strip(),
        "q": (request.POST.get("q") or "").strip(),
    }
    base_qs = LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto").filter(estado="Abierta")
    if filters["cliente"] and filters["cliente"] != "Todos":
        base_qs = base_qs.filter(corta_cliente_proyecto__nombre=filters["cliente"])
    if filters["q"]:
        q = filters["q"]
        base_qs = base_qs.filter(
            Q(folio_externo__icontains=q)
            | Q(codigo__icontains=q)
            | Q(nombre__icontains=q)
            | Q(descripcion__icontains=q)
            | Q(corta_cliente_proyecto__nombre__icontains=q)
        )
    ordenes = list(base_qs[:2000])
    orden_ids = []
    for o in ordenes:
        total = int(getattr(o, "total_piezas", 0) or 0)
        apartado = int(getattr(o, "apartado", 0) or 0)
        enviado = int(getattr(o, "enviado", 0) or 0)
        saldo = max(0, total - apartado - enviado)
        if total > 0 and saldo == 0 and enviado > 0:
            orden_ids.append(int(getattr(o, "id", 0) or 0))
    if orden_ids:
        last_map = {
            int(r["orden_id"]): r["last_fecha"]
            for r in LogisticaEnvioCorta.objects.filter(orden_id__in=orden_ids).values("orden_id").annotate(last_fecha=Max("fecha"))
        }
        exp_map = {int(e.orden_corta_id): e for e in LogisticaExpediente.objects.filter(orden_corta_id__in=orden_ids)}
        decote_ids = []
        for oid in orden_ids:
            last_fecha = last_map.get(int(oid))
            if not last_fecha or last_fecha > cutoff:
                continue
            exp = exp_map.get(int(oid))
            if not exp or not getattr(exp, "generado_en", None) or int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                continue
            decote_ids.append(int(oid))
        if decote_ids:
            with transaction.atomic(using="mes"):
                LaserProduccion.objects.filter(orden_item__orden_id__in=decote_ids).delete()
                LaserAsignacion.objects.filter(orden_id__in=decote_ids).delete()
                LaserOrdenProduccion.objects.filter(id__in=decote_ids).delete()
    next_url = _safe_next_url(request.POST.get("next") or "")
    return redirect(next_url) if next_url else redirect("catalogos:corte_laser_control")


@login_required
@require_http_methods(["GET"])
def corte_laser_reportes(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    stages = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
    ]

    def _to_thursday_start(d):
        try:
            delta = (int(d.weekday() or 0) - 3) % 7
        except Exception:
            delta = 0
        return d - timedelta(days=delta)

    def _effective_dt(fecha_operacion, creado_en):
        if creado_en:
            dt_local = timezone.localtime(creado_en)
            t = dt_local.time().replace(microsecond=0)
            d = fecha_operacion or dt_local.date()
        else:
            t = time.min
            d = fecha_operacion or timezone.localdate()
        try:
            return timezone.make_aware(datetime.combine(d, t), timezone.get_default_timezone())
        except Exception:
            return timezone.now()

    def _durations_for_order(order_id: int):
        rows = list(
            LaserEstadoCambio.objects.filter(orden_id=int(order_id))
            .values("estado_nuevo", "fecha_operacion", "creado_en")
            .order_by("fecha_operacion", "creado_en", "id")[:2000]
        )
        points = []
        for r in rows:
            st = (r.get("estado_nuevo") or "").strip()
            if st not in stages and st != "Terminado":
                continue
            points.append({"st": st, "dt": _effective_dt(r.get("fecha_operacion"), r.get("creado_en"))})
        points.sort(key=lambda x: x["dt"])
        out = {s: 0.0 for s in stages}
        prev = None
        for p in points:
            if prev is not None:
                sec = max(0.0, float((p["dt"] - prev["dt"]).total_seconds() or 0.0))
                if prev["st"] in out:
                    out[prev["st"]] = float(out.get(prev["st"], 0.0)) + sec
            prev = p
        return out

    today = timezone.localdate()
    selected_raw = (request.GET.get("week_start") or "").strip()
    base = today
    if selected_raw:
        try:
            base = datetime.strptime(selected_raw, "%Y-%m-%d").date()
        except Exception:
            base = today
    week_start = _to_thursday_start(base)
    week_end = week_start + timedelta(days=7)

    term_rows = list(
        LaserEstadoCambio.objects.filter(
            estado_nuevo="Terminado",
            fecha_operacion__gte=week_start,
            fecha_operacion__lt=week_end,
        )
        .order_by("orden_id", "-fecha_operacion", "-creado_en")
        .values("orden_id", "fecha_operacion", "creado_en")[:10000]
    )
    term_map = {}
    for r in term_rows:
        oid = int(r.get("orden_id") or 0)
        if oid <= 0:
            continue
        if oid in term_map:
            continue
        term_map[oid] = r

    term_ids = list(term_map.keys())
    orders = []
    total_kg = 0.0
    total_piezas = 0
    if term_ids:
        qs = LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto", "material").filter(id__in=term_ids)
        o_map = {int(o.id): o for o in qs}
        for oid, r in term_map.items():
            o = o_map.get(int(oid))
            if not o:
                continue
            kg = float(getattr(o, "peso_kg", 0.0) or 0.0)
            piezas = int(getattr(o, "total_piezas", 0) or 0)
            term_dt = _effective_dt(r.get("fecha_operacion"), r.get("creado_en"))
            orders.append(
                {
                    "id": int(o.id),
                    "folio": getattr(o, "folio", ""),
                    "codigo": (getattr(o, "codigo", "") or "").strip() or getattr(o, "folio", ""),
                    "cliente": getattr(getattr(o, "corta_cliente_proyecto", None), "nombre", "") or "",
                    "piezas": piezas,
                    "kg": round(kg, 3),
                    "terminado_dt": term_dt,
                    "terminado_str": term_dt.strftime("%Y-%m-%d %H:%M"),
                }
            )
            total_kg += kg
            total_piezas += piezas
    orders.sort(key=lambda x: x.get("terminado_dt") or timezone.now(), reverse=True)

    stage_avg = []
    count_for_avg = 0
    stage_sum = {s: 0.0 for s in stages}
    for row in orders[:300]:
        dur = _durations_for_order(int(row["id"]))
        count_for_avg += 1
        for s in stages:
            stage_sum[s] = float(stage_sum.get(s, 0.0)) + float(dur.get(s, 0.0))
    for s in stages:
        avg_h = (float(stage_sum.get(s, 0.0)) / float(count_for_avg) / 3600.0) if count_for_avg else 0.0
        stage_avg.append({"etapa": s, "horas_promedio": round(avg_h, 2)})

    export_static = bool(getattr(request, "_corte_laser_reportes_export_static", False))
    export_detail_map = {}
    if export_static and term_ids:
        try:
            max_ids = term_ids[:2000]
            o_map = {int(o["id"]): o for o in orders if isinstance(o, dict) and o.get("id")}
            sec_map = {int(oid): {s: 0.0 for s in stages} for oid in max_ids}
            prev_oid = None
            prev_stage = ""
            prev_dt = None
            rows = (
                LaserEstadoCambio.objects.filter(orden_id__in=max_ids)
                .values("id", "orden_id", "estado_nuevo", "fecha_operacion", "creado_en")
                .order_by("orden_id", "fecha_operacion", "creado_en", "id")[:200000]
            )
            for r in rows:
                oid = int(r.get("orden_id") or 0)
                st = (r.get("estado_nuevo") or "").strip()
                if oid <= 0:
                    continue
                if st not in stages and st != "Terminado":
                    continue
                dt = _effective_dt(r.get("fecha_operacion"), r.get("creado_en"))
                if prev_oid is None or oid != prev_oid:
                    prev_oid = oid
                    prev_stage = st
                    prev_dt = dt
                    continue
                if prev_dt and dt and prev_stage in stages:
                    sec = max(0.0, float((dt - prev_dt).total_seconds() or 0.0))
                    sec_map[oid][prev_stage] = float(sec_map[oid].get(prev_stage, 0.0)) + sec
                prev_stage = st
                prev_dt = dt
            for oid in max_ids:
                o = o_map.get(int(oid)) or {}
                etapas = []
                for s in stages:
                    sec = float((sec_map.get(int(oid)) or {}).get(s, 0.0) or 0.0)
                    etapas.append({"etapa": s, "horas": round(sec / 3600.0, 2)})
                export_detail_map[str(int(oid))] = {
                    "ok": True,
                    "orden": {
                        "id": int(o.get("id") or oid),
                        "folio": o.get("folio") or "",
                        "codigo": o.get("codigo") or "",
                        "cliente": o.get("cliente") or "",
                        "piezas": int(o.get("piezas") or 0),
                        "kg": float(o.get("kg") or 0.0),
                    },
                    "etapas": etapas,
                }
        except Exception:
            export_detail_map = {}

    ctx = {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "week_selected_start": week_start.isoformat(),
        "totales": {
            "ordenes": int(len(orders)),
            "piezas": int(total_piezas),
            "kg": round(float(total_kg or 0.0), 3),
            "ton": round(float(total_kg or 0.0) / 1000.0, 3),
        },
        "orders": orders,
        "stages": stages,
        "stage_avg": stage_avg,
        "export_static": export_static,
        "export_detail_map": export_detail_map,
    }
    if getattr(request, "_corte_laser_reportes_context_only", False):
        return ctx
    return render(request, "catalogos/corte_laser_reportes.html", ctx)


@login_required
@require_http_methods(["GET"])
def corte_laser_reportes_terminado_json(request, pk: int):
    if not _can_corte_laser(request.user):
        return JsonResponse({"ok": False, "error": "No autorizado."}, status=403)
    orden = get_object_or_404(LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto", "material"), pk=int(pk))
    stages = [
        "Espera de corte",
        "Corte",
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
    ]

    def _effective_dt(fecha_operacion, creado_en):
        if creado_en:
            dt_local = timezone.localtime(creado_en)
            t = dt_local.time().replace(microsecond=0)
            d = fecha_operacion or dt_local.date()
        else:
            t = time.min
            d = fecha_operacion or timezone.localdate()
        try:
            return timezone.make_aware(datetime.combine(d, t), timezone.get_default_timezone())
        except Exception:
            return timezone.now()

    rows = list(
        LaserEstadoCambio.objects.filter(orden_id=int(orden.id))
        .values("estado_nuevo", "fecha_operacion", "creado_en")
        .order_by("fecha_operacion", "creado_en", "id")[:2000]
    )
    points = []
    for r in rows:
        st = (r.get("estado_nuevo") or "").strip()
        if st not in stages and st != "Terminado":
            continue
        points.append({"st": st, "dt": _effective_dt(r.get("fecha_operacion"), r.get("creado_en"))})
    points.sort(key=lambda x: x["dt"])
    sec_map = {s: 0.0 for s in stages}
    prev = None
    for p in points:
        if prev is not None:
            sec = max(0.0, float((p["dt"] - prev["dt"]).total_seconds() or 0.0))
            if prev["st"] in sec_map:
                sec_map[prev["st"]] = float(sec_map.get(prev["st"], 0.0)) + sec
        prev = p

    etapas = []
    for s in stages:
        sec = float(sec_map.get(s, 0.0) or 0.0)
        horas = sec / 3600.0
        etapas.append({"etapa": s, "segundos": round(sec, 0), "horas": round(horas, 2)})

    return JsonResponse(
        {
            "ok": True,
            "orden": {
                "id": int(orden.id),
                "folio": getattr(orden, "folio", ""),
                "codigo": (getattr(orden, "codigo", "") or "").strip() or getattr(orden, "folio", ""),
                "cliente": getattr(getattr(orden, "corta_cliente_proyecto", None), "nombre", "") or "",
                "piezas": int(getattr(orden, "total_piezas", 0) or 0),
                "kg": round(float(getattr(orden, "peso_kg", 0.0) or 0.0), 3),
            },
            "etapas": etapas,
        }
    )


@login_required
def corte_laser_reportes_export_html(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    try:
        from produccion.views import _inline_html_assets
    except Exception:
        _inline_html_assets = (lambda x: x)
    request._corte_laser_reportes_export_static = True
    resp = corte_laser_reportes(request)
    html = resp.content.decode("utf-8", errors="ignore")
    html = _inline_html_assets(html)
    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    out = HttpResponse(html, content_type="text/html; charset=utf-8")
    out["Content-Disposition"] = f'attachment; filename="corta_reportes_{ts}.html"'
    return out


@login_required
def corte_laser_reportes_export_xlsx(request):
    if not _can_corte_laser(request.user):
        return redirect("/")
    import io

    request._corte_laser_reportes_context_only = True
    ctx = corte_laser_reportes(request)
    if not isinstance(ctx, dict):
        return redirect("catalogos:corte_laser_reportes")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    stages = list(ctx.get("stages") or [])
    tot = ctx.get("totales") or {}
    orders = list(ctx.get("orders") or [])
    start = str(ctx.get("week_start") or "")
    end = str(ctx.get("week_end") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    header = Font(bold=True)
    center = Alignment(vertical="center")

    ws.append(["Corta.mx · Reportes", ""])
    ws["A1"].font = bold
    ws.append(["Periodo", f"{start} → {end}".strip()])
    ws.append([])

    ws.append(["KPI", "Valor", "Unidad"])
    for cell in ws[ws.max_row]:
        cell.font = header
        cell.alignment = center
    ws.append(["Órdenes terminadas", int(tot.get("ordenes") or 0), "ordenes"])
    ws.append(["Piezas terminadas", int(tot.get("piezas") or 0), "pzs"])
    ws.append(["Kg terminados", float(tot.get("kg") or 0.0), "kg"])
    ws.append(["Ton terminados", float(tot.get("ton") or 0.0), "ton"])

    ws.append([])
    ws.append(["Promedio horas por etapa", "", ""])
    ws["A" + str(ws.max_row)].font = bold
    ws.append(["Etapa", "Horas promedio"])
    for cell in ws[ws.max_row]:
        cell.font = header
        cell.alignment = center
    for r in (ctx.get("stage_avg") or []):
        ws.append([r.get("etapa") or "", float(r.get("horas_promedio") or 0.0)])

    ws2 = wb.create_sheet("Terminados")
    ws2.append(["Folio", "Código", "Cliente/Proyecto", "Piezas", "Kg", "Terminado"] + [f"Horas {s}" for s in stages])
    for cell in ws2[1]:
        cell.font = header
        cell.alignment = center

    def _effective_dt(fecha_operacion, creado_en):
        if creado_en:
            dt_local = timezone.localtime(creado_en)
            t = dt_local.time().replace(microsecond=0)
            d = fecha_operacion or dt_local.date()
        else:
            t = time.min
            d = fecha_operacion or timezone.localdate()
        try:
            return timezone.make_aware(datetime.combine(d, t), timezone.get_default_timezone())
        except Exception:
            return timezone.now()

    def _hours_by_stage(order_id: int):
        rows = list(
            LaserEstadoCambio.objects.filter(orden_id=int(order_id))
            .values("estado_nuevo", "fecha_operacion", "creado_en")
            .order_by("fecha_operacion", "creado_en", "id")[:2000]
        )
        points = []
        for r in rows:
            st = (r.get("estado_nuevo") or "").strip()
            if st not in stages and st != "Terminado":
                continue
            points.append({"st": st, "dt": _effective_dt(r.get("fecha_operacion"), r.get("creado_en"))})
        points.sort(key=lambda x: x["dt"])
        sec_map = {s: 0.0 for s in stages}
        prev = None
        for p in points:
            if prev is not None:
                sec = max(0.0, float((p["dt"] - prev["dt"]).total_seconds() or 0.0))
                if prev["st"] in sec_map:
                    sec_map[prev["st"]] = float(sec_map.get(prev["st"], 0.0)) + sec
            prev = p
        return {s: round(float(sec_map.get(s, 0.0)) / 3600.0, 2) for s in stages}

    for o in orders[:2000]:
        h_map = _hours_by_stage(int(o.get("id") or 0))
        row = [
            o.get("folio") or "",
            o.get("codigo") or "",
            o.get("cliente") or "",
            int(o.get("piezas") or 0),
            float(o.get("kg") or 0.0),
            o.get("terminado_str") or "",
        ]
        for s in stages:
            row.append(float(h_map.get(s) or 0.0))
        ws2.append(row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    out = HttpResponse(bio.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    out["Content-Disposition"] = f'attachment; filename="corta_reportes_{ts}.xlsx"'
    return out


@login_required
@require_http_methods(["GET", "POST"])
def paros(request):
    _seed_paros_catalogos()
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    next_val = (request.POST.get("next") or "").strip()
    next_url = next_val if (next_val.startswith("/") and not next_val.startswith("//")) else ""
    redirect_back = (lambda: redirect(next_url)) if next_url else (lambda: redirect("catalogos:paros"))

    paro_form = ParoStartForm()
    falla_form = FallaStartForm()
    comentario_form = ComentarioForm()
    now = timezone.localtime(timezone.now())
    energia_form = EnergiaForm(
        initial={
            "inicio_fecha": now.date(),
            "inicio_hora": (now - timedelta(minutes=10)).time().replace(second=0, microsecond=0),
            "fin_fecha": now.date(),
            "fin_hora": now.time().replace(second=0, microsecond=0),
            "aplicar_maquinas": True,
        }
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        maquina_id = int(request.POST.get("maquina_id") or 0)
        if action == "start_paro":
            paro_form = ParoStartForm(request.POST)
            if not (paro_form.is_valid() and maquina_id):
                messages.error(request, "Faltan datos para registrar el paro: elige máquina y motivo.")
                return redirect_back()
            # El bloqueo evita que dos operadores registren a la vez dos paros
            # abiertos sobre la misma máquina, que es lo que la comprobación
            # exists() por sí sola no impide.
            with transaction.atomic(using="mes"):
                maquina = get_object_or_404(
                    Maquina.objects.using("mes").select_for_update(), pk=maquina_id
                )
                if MaquinaParo.objects.using("mes").filter(maquina=maquina, fin__isnull=True).exists():
                    messages.error(request, "Esa máquina ya está en paro.")
                else:
                    MaquinaParo.objects.create(
                        maquina=maquina,
                        motivo=paro_form.cleaned_data["motivo"],
                        inicio=timezone.now(),
                        comentario=(paro_form.cleaned_data.get("comentario") or "").strip(),
                        registrado_por=actor,
                    )
                    messages.success(request, "Paro registrado.")
            return redirect_back()
        elif action == "end_paro":
            comentario_form = ComentarioForm(request.POST)
            if comentario_form.is_valid() and maquina_id:
                maquina = get_object_or_404(Maquina, pk=maquina_id)
                if MaquinaFalla.objects.filter(maquina=maquina, fin__isnull=True).exists():
                    messages.error(request, "Primero resuelve la falla antes de reanudar.")
                    return redirect_back()
                paro = MaquinaParo.objects.filter(maquina=maquina, fin__isnull=True).first()
                if not paro:
                    messages.error(request, "No hay un paro activo para esa máquina.")
                    return redirect_back()
                paro.fin = timezone.now()
                c = (comentario_form.cleaned_data.get("comentario") or "").strip()
                if c:
                    paro.comentario = f"{paro.comentario} | {c}".strip(" |") if paro.comentario else c
                paro.registrado_por = actor or paro.registrado_por
                paro.save(update_fields=["fin", "comentario", "registrado_por", "actualizado_en"])
                messages.success(request, "Máquina reanudada.")
                return redirect_back()
        elif action == "start_falla":
            falla_form = FallaStartForm(request.POST)
            if falla_form.is_valid() and maquina_id:
                maquina = get_object_or_404(Maquina, pk=maquina_id)
                if MaquinaFalla.objects.filter(maquina=maquina, fin__isnull=True).exists():
                    messages.error(request, "Esa máquina ya tiene una falla activa.")
                    return redirect_back()
                paro = MaquinaParo.objects.filter(maquina=maquina, fin__isnull=True).first()
                if not paro:
                    motivo_falla, _ = MaquinaParoMotivo.objects.get_or_create(
                        nombre_normalizado="FALLA",
                        defaults={"nombre": "Falla", "activo": True, "es_sistema": True},
                    )
                    paro = MaquinaParo.objects.create(
                        maquina=maquina,
                        motivo=motivo_falla,
                        inicio=timezone.now(),
                        comentario="",
                        registrado_por=actor,
                    )
                MaquinaFalla.objects.create(
                    maquina=maquina,
                    tipo=falla_form.cleaned_data["tipo"],
                    inicio=timezone.now(),
                    comentario=(falla_form.cleaned_data.get("comentario") or "").strip(),
                    registrado_por=actor,
                    paro=paro,
                )
                messages.success(request, "Falla registrada.")
                return redirect_back()
        elif action == "end_falla":
            comentario_form = ComentarioForm(request.POST)
            if comentario_form.is_valid() and maquina_id:
                maquina = get_object_or_404(Maquina, pk=maquina_id)
                falla = MaquinaFalla.objects.filter(maquina=maquina, fin__isnull=True).first()
                if not falla:
                    messages.error(request, "No hay una falla activa para esa máquina.")
                    return redirect_back()
                falla.fin = timezone.now()
                c = (comentario_form.cleaned_data.get("comentario") or "").strip()
                if c:
                    falla.comentario = f"{falla.comentario} | {c}".strip(" |") if falla.comentario else c
                falla.registrado_por = actor or falla.registrado_por
                falla.save(update_fields=["fin", "comentario", "registrado_por", "actualizado_en"])
                messages.success(request, "Falla resuelta.")
                return redirect_back()
        elif action == "energia":
            energia_form = EnergiaForm(request.POST)
            if energia_form.is_valid():
                inicio = energia_form.cleaned_data["inicio_dt"]
                fin = energia_form.cleaned_data["fin_dt"]
                comentario = (energia_form.cleaned_data.get("comentario") or "").strip()
                ev = PlantaEvento.objects.create(
                    tipo="Energia",
                    inicio=inicio.astimezone(timezone.utc),
                    fin=fin.astimezone(timezone.utc),
                    comentario=comentario,
                    registrado_por=actor,
                )
                if energia_form.cleaned_data.get("aplicar_maquinas"):
                    motivo_energia, _ = MaquinaParoMotivo.objects.get_or_create(
                        nombre_normalizado="ENERGÍA",
                        defaults={"nombre": "Energía", "activo": True, "es_sistema": True},
                    )
                    machines = list(Maquina.objects.filter(activo=True))
                    inicio_utc = ev.inicio
                    fin_utc = ev.fin
                    for m in machines:
                        overlap = MaquinaParo.objects.filter(
                            maquina=m,
                            inicio__lt=fin_utc,
                        ).filter(Q(fin__isnull=True) | Q(fin__gt=inicio_utc))
                        if overlap.exists():
                            continue
                        MaquinaParo.objects.create(
                            maquina=m,
                            motivo=motivo_energia,
                            inicio=inicio_utc,
                            fin=fin_utc,
                            comentario=comentario,
                            registrado_por=actor,
                            evento_planta=ev,
                        )
                messages.success(request, "Corte de energía registrado.")
                return redirect_back()

    maquinas = list(Maquina.objects.filter(activo=True).order_by("tipo", "-es_robot", "nombre"))
    open_paros = {
        int(p.maquina_id): p
        for p in MaquinaParo.objects.filter(fin__isnull=True).select_related("maquina", "motivo")
    }
    open_fallas = {
        int(f.maquina_id): f
        for f in MaquinaFalla.objects.filter(fin__isnull=True).select_related("maquina", "tipo")
    }
    rows = []
    for m in maquinas:
        rows.append(
            {
                "maquina": m,
                "paro": open_paros.get(int(m.id)),
                "falla": open_fallas.get(int(m.id)),
            }
        )

    motivos = list(MaquinaParoMotivo.objects.filter(activo=True).order_by("nombre"))
    tipos = list(MaquinaFallaTipo.objects.filter(activo=True).order_by("nombre"))
    energia_recent = list(PlantaEvento.objects.filter(tipo="Energia").order_by("-inicio")[:20])
    recent_paros = list(MaquinaParo.objects.select_related("maquina", "motivo").order_by("-inicio")[:30])
    recent_fallas = list(MaquinaFalla.objects.select_related("maquina", "tipo").order_by("-inicio")[:30])
    focus_id = int((request.GET.get("focus") or "0").strip() or 0)
    next_url = (request.GET.get("next") or "").strip()
    if not (next_url.startswith("/") and not next_url.startswith("//")):
        next_url = ""
    return render(
        request,
        "catalogos/paros.html",
        {
            "rows": rows,
            "paro_form": paro_form,
            "falla_form": falla_form,
            "comentario_form": comentario_form,
            "energia_form": energia_form,
            "motivos": motivos,
            "tipos": tipos,
            "energia_recent": energia_recent,
            "recent_paros": recent_paros,
            "recent_fallas": recent_fallas,
            "is_admin": _is_admin_user(request.user),
            "focus_id": focus_id,
            "next_url": next_url,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def paros_motivos(request):
    if not _is_admin_user(request.user):
        return redirect("/")
    _seed_paros_catalogos()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        pk = int(request.POST.get("pk") or 0)
        if action == "create":
            nombre = (request.POST.get("nombre") or "").strip()
            if not nombre:
                messages.error(request, "Nombre requerido.")
                return redirect("catalogos:paros_motivos")
            try:
                MaquinaParoMotivo.objects.create(nombre=nombre, nombre_normalizado=nombre.upper(), activo=True, es_sistema=False)
                messages.success(request, "Motivo creado.")
            except Exception:
                messages.error(request, "No se pudo crear (puede que ya exista).")
            return redirect("catalogos:paros_motivos")
        if pk and action == "toggle":
            m = get_object_or_404(MaquinaParoMotivo, pk=pk)
            m.activo = not bool(m.activo)
            m.save(update_fields=["activo", "actualizado_en"])
            messages.success(request, "Motivo actualizado.")
            return redirect("catalogos:paros_motivos")
        if pk and action == "delete":
            m = get_object_or_404(MaquinaParoMotivo, pk=pk)
            if MaquinaParo.objects.filter(motivo=m).exists():
                messages.error(request, "No se puede eliminar: ya está usado. Desactívalo.")
                return redirect("catalogos:paros_motivos")
            try:
                m.delete()
                messages.success(request, "Motivo eliminado.")
            except Exception:
                messages.error(request, "No se pudo eliminar.")
            return redirect("catalogos:paros_motivos")
    motivos = list(MaquinaParoMotivo.objects.all().order_by("-activo", "-es_sistema", "nombre"))
    return render(request, "catalogos/paros_motivos.html", {"motivos": motivos})


@login_required
@require_http_methods(["GET", "POST"])
def paros_fallas(request):
    if not _is_admin_user(request.user):
        return redirect("/")
    _seed_paros_catalogos()
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        pk = int(request.POST.get("pk") or 0)
        if action == "create":
            nombre = (request.POST.get("nombre") or "").strip()
            if not nombre:
                messages.error(request, "Nombre requerido.")
                return redirect("catalogos:paros_fallas")
            try:
                MaquinaFallaTipo.objects.create(nombre=nombre, nombre_normalizado=nombre.upper(), activo=True, es_sistema=False)
                messages.success(request, "Tipo de falla creado.")
            except Exception:
                messages.error(request, "No se pudo crear (puede que ya exista).")
            return redirect("catalogos:paros_fallas")
        if pk and action == "toggle":
            t = get_object_or_404(MaquinaFallaTipo, pk=pk)
            t.activo = not bool(t.activo)
            t.save(update_fields=["activo", "actualizado_en"])
            messages.success(request, "Tipo actualizado.")
            return redirect("catalogos:paros_fallas")
        if pk and action == "delete":
            t = get_object_or_404(MaquinaFallaTipo, pk=pk)
            if MaquinaFalla.objects.filter(tipo=t).exists():
                messages.error(request, "No se puede eliminar: ya está usado. Desactívalo.")
                return redirect("catalogos:paros_fallas")
            try:
                t.delete()
                messages.success(request, "Tipo eliminado.")
            except Exception:
                messages.error(request, "No se pudo eliminar.")
            return redirect("catalogos:paros_fallas")
    tipos = list(MaquinaFallaTipo.objects.all().order_by("-activo", "-es_sistema", "nombre"))
    return render(request, "catalogos/paros_fallas.html", {"tipos": tipos})


@admin_required
@require_http_methods(["GET", "POST"])
def equipo_editar(request, pk: int):
    equipo = get_object_or_404(EquipoTrabajo, pk=pk)
    if request.method == "POST":
        form = EquipoTrabajoForm(request.POST, instance=equipo)
        if form.is_valid():
            form.save()
            messages.success(request, "Equipo actualizado.")
            return redirect("catalogos:equipos")
    else:
        form = EquipoTrabajoForm(instance=equipo)
    return render(request, "catalogos/equipo_editar.html", {"form": form, "equipo": equipo})


@admin_required
@require_http_methods(["GET"])
def proyecto_detalle(request, pk: int):
    proyecto = get_object_or_404(Proyecto, pk=pk)

    estados_index = {s: i for i, s in enumerate(ESTADOS)}
    total_steps = max(len(ESTADOS) - 1, 1)

    qs = (
        Viga.objects.annotate(proyecto_norm=Upper("proyecto"))
        .filter(proyecto_norm=proyecto.nombre_normalizado)
        .order_by("estado", "prioridad", "fecha_compromiso", "codigo_viga", "pieza_no")
    )
    vigas = list(qs[:2000])
    for v in vigas:
        idx = estados_index.get(v.estado, 0)
        v.avance_pct = round(idx / total_steps * 100.0, 1)
        v.avance_css = f"{v.avance_pct:.1f}"
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")

    return render(
        request,
        "catalogos/proyecto_detalle.html",
        {
            "proyecto": proyecto,
            "vigas": vigas,
            "total": qs.count(),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def pedidos_ordenes(request):
    if not _can_pedidos(request.user):
        return redirect("/")

    ItemFormSet = forms.formset_factory(PedidoProduccionItemForm, extra=6)
    pedido_form = PedidoProduccionForm(initial={"fecha_compromiso": timezone.localdate()})
    formset = ItemFormSet()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "pedido_create":
            pedido_form = PedidoProduccionForm(request.POST)
            formset = ItemFormSet(request.POST)
            if pedido_form.is_valid() and formset.is_valid():
                items = []
                for f in formset:
                    prod = f.cleaned_data.get("producto")
                    qty = int(f.cleaned_data.get("cantidad_total") or 0)
                    if not prod or qty <= 0:
                        continue
                    items.append((prod, qty))
                if not items:
                    pedido_form.add_error(None, "Agrega al menos un producto.")
                else:
                    with transaction.atomic(using="mes"):
                        folio = _next_pedido_folio()
                        pedido = PedidoProduccion.objects.using("mes").create(
                            folio=folio,
                            cliente=(pedido_form.cleaned_data.get("cliente") or "").strip(),
                            telefono=(pedido_form.cleaned_data.get("telefono") or "").strip(),
                            fecha_compromiso=pedido_form.cleaned_data["fecha_compromiso"],
                            comentarios=(pedido_form.cleaned_data.get("comentarios") or "").strip(),
                            estado="Activa",
                        )
                        for prod, qty in items:
                            PedidoProduccionItem.objects.using("mes").create(
                                pedido=pedido,
                                producto=prod,
                                cantidad_total=max(1, int(qty)),
                                apartado=0,
                                enviado=0,
                                estado_herreria="Pendiente",
                                aceptado_por="",
                            )
                    messages.success(request, f"Orden creada: {pedido.folio}. Enviada a Herrería.")
                    return redirect("catalogos:pedidos_ordenes")

    pedidos = list(PedidoProduccion.objects.order_by("-creado_en")[:200])
    pedido_ids = [int(p.id) for p in pedidos]
    items = []
    if pedido_ids:
        items = list(PedidoProduccionItem.objects.filter(pedido_id__in=pedido_ids).select_related("producto", "pedido").order_by("-creado_en")[:2000])
    items_by_pedido = {}
    for it in items:
        items_by_pedido.setdefault(int(it.pedido_id), []).append(it)
    item_ids = [int(it.id) for it in items if it and getattr(it, "id", None)]
    envio_item_ids = set()
    if item_ids:
        envio_item_ids = set(
            int(x)
            for x in LogisticaEnvioItem.objects.filter(pedido_item_id__in=item_ids).values_list("pedido_item_id", flat=True)
            if x
        )
    for p in pedidos:
        p.items_list = items_by_pedido.get(int(p.id), [])
        reasons = []
        for it in p.items_list:
            if int(getattr(it, "apartado", 0) or 0) > 0:
                reasons.append("tiene apartado")
            if int(getattr(it, "enviado", 0) or 0) > 0:
                reasons.append("tiene enviado")
            if int(getattr(it, "id", 0) or 0) in envio_item_ids:
                reasons.append("tiene envíos registrados")
            if bool(getattr(it, "orden_herreria_id", None)):
                reasons.append("ya fue aceptada en Herrería")
            if (getattr(it, "estado_herreria", "") or "").strip() in {"Espera de material", "En producción"}:
                reasons.append("está en Herrería (en proceso)")
        clean = (len(set(reasons)) == 0)
        is_cancelada = (p.estado or "") == "Cancelada"
        p.can_delete = bool(is_cancelada)
        p.can_delete_enabled = bool(is_cancelada and clean)
        p.delete_reason = ", ".join(sorted(set(reasons))) if is_cancelada and (not clean) else ""
        p.can_cancel = (p.estado or "") == "Activa"

    return render(
        request,
        "catalogos/pedidos_ordenes.html",
        {
            "pedido_form": pedido_form,
            "formset": formset,
            "pedidos": pedidos,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def pedidos_orden_editar(request, pk: int):
    if not _can_pedidos(request.user):
        return redirect("/")
    pedido = get_object_or_404(PedidoProduccion, pk=int(pk))
    items = list(PedidoProduccionItem.objects.filter(pedido=pedido).select_related("producto", "orden_herreria").order_by("id"))

    class _BaseItemFormSet(forms.BaseFormSet):
        def __init__(self, *args, **kwargs):
            self._items = list(kwargs.pop("items") or [])
            super().__init__(*args, **kwargs)

        def get_form_kwargs(self, index):
            kw = super().get_form_kwargs(index)
            if index is not None and index < len(self._items):
                kw["item"] = self._items[index]
            return kw

    ItemFormSet = forms.formset_factory(PedidoProduccionItemEditForm, formset=_BaseItemFormSet, extra=3, can_delete=True)
    form = PedidoProduccionEditForm(
        initial={
            "cliente": (pedido.cliente or "").strip(),
            "telefono": (pedido.telefono or "").strip(),
            "fecha_compromiso": pedido.fecha_compromiso,
            "comentarios": (getattr(pedido, "comentarios", "") or "").strip(),
        }
    )
    formset = ItemFormSet(items=items)

    if request.method == "POST":
        form = PedidoProduccionEditForm(request.POST)
        formset = ItemFormSet(request.POST, items=items)
        if form.is_valid() and formset.is_valid():
            new_cliente = (form.cleaned_data.get("cliente") or "").strip()
            new_tel = (form.cleaned_data.get("telefono") or "").strip()
            new_fecha = form.cleaned_data.get("fecha_compromiso")
            new_comentarios = (form.cleaned_data.get("comentarios") or "").strip()

            with transaction.atomic(using="mes"):
                pedido = get_object_or_404(PedidoProduccion.objects.select_for_update(), pk=int(pk))
                locked_items = list(
                    PedidoProduccionItem.objects.select_for_update()
                    .filter(pedido=pedido)
                    .select_related("producto", "orden_herreria")
                    .order_by("id")
                )
                item_map = {int(it.id): it for it in locked_items}

                errors = False
                herr_order_ids = set()
                for f in formset:
                    cd = getattr(f, "cleaned_data", None) or {}
                    item_id = cd.get("item_id")
                    prod = cd.get("producto")
                    qty = int(cd.get("cantidad_total") or 0)
                    to_delete = bool(cd.get("DELETE"))

                    if not item_id and (not prod) and qty <= 0:
                        continue

                    if item_id:
                        it = item_map.get(int(item_id))
                        if not it:
                            f.add_error(None, "El item ya no existe.")
                            errors = True
                            continue
                        if it.orden_herreria_id:
                            herr_order_ids.add(int(it.orden_herreria_id))
                        min_allowed = int(it.apartado or 0) + int(it.enviado or 0)
                        if to_delete:
                            if min_allowed > 0 or it.orden_herreria_id:
                                f.add_error(None, "No se puede eliminar: ya tiene movimientos o ya fue aceptado en Herrería.")
                                errors = True
                                continue
                            it.delete()
                            continue
                        if qty <= 0:
                            f.add_error("cantidad_total", "Cantidad requerida.")
                            errors = True
                            continue
                        if qty < min_allowed:
                            f.add_error("cantidad_total", f"No puede ser menor a {min_allowed} (apartado+enviado).")
                            errors = True
                            continue
                        if it.orden_herreria_id and qty < int(it.cantidad_total or 0):
                            f.add_error("cantidad_total", "No se puede reducir: ya fue aceptado en Herrería.")
                            errors = True
                            continue
                        if qty != int(it.cantidad_total or 0):
                            it.cantidad_total = max(1, int(qty))
                            it.save(update_fields=["cantidad_total", "actualizado_en"])
                    else:
                        if not prod or qty <= 0:
                            continue
                        PedidoProduccionItem.objects.create(
                            pedido=pedido,
                            producto=prod,
                            cantidad_total=max(1, int(qty)),
                            apartado=0,
                            enviado=0,
                            estado_herreria="Pendiente",
                            aceptado_por="",
                        )

                if errors:
                    transaction.set_rollback(True, using="mes")
                else:
                    old_fecha = pedido.fecha_compromiso
                    pedido.cliente = new_cliente
                    pedido.telefono = new_tel
                    pedido.fecha_compromiso = new_fecha
                    pedido.comentarios = new_comentarios
                    pedido.save(update_fields=["cliente", "telefono", "fecha_compromiso", "comentarios", "actualizado_en"])
                    if new_fecha != old_fecha and herr_order_ids:
                        HerrOrdenProduccion.objects.filter(id__in=list(herr_order_ids)).update(fecha_compromiso=new_fecha, actualizado_en=timezone.now())
                    messages.success(request, "Orden actualizada.")
                    return redirect("catalogos:pedidos_ordenes")

        messages.error(request, "Revisa los campos marcados.")

    return render(
        request,
        "catalogos/pedidos_orden_editar.html",
        {
            "pedido": pedido,
            "form": form,
            "formset": formset,
        },
    )


@login_required
@require_http_methods(["POST"])
def pedidos_orden_cancelar(request, pk: int):
    if not _can_pedidos(request.user):
        return redirect("/")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    with transaction.atomic(using="mes"):
        pedido = get_object_or_404(PedidoProduccion.objects.select_for_update(), pk=int(pk))
        prev_estado = (pedido.estado or "").strip()
        if prev_estado not in {"Activa", "Cancelada"}:
            return redirect("catalogos:pedidos_ordenes")

        now = timezone.now()
        items = list(PedidoProduccionItem.objects.select_for_update().filter(pedido=pedido).select_related("producto"))
        for it in items:
            apartado = int(getattr(it, "apartado", 0) or 0)
            if apartado > 0:
                servicio_almacen.devolver_al_disponible(it.producto, apartado, actor, linea_pedido=it)
                it.apartado = 0
            if not it.orden_herreria_id:
                it.estado_herreria = "Cancelado"
            it.save(update_fields=["apartado", "estado_herreria", "actualizado_en"])

        if prev_estado == "Activa":
            pedido.estado = "Cancelada"
            pedido.save(update_fields=["estado", "actualizado_en"])
    messages.success(request, "Orden cancelada.")
    return redirect("catalogos:pedidos_ordenes")


@login_required
@require_http_methods(["POST"])
def pedidos_orden_eliminar(request, pk: int):
    if not _can_pedidos(request.user):
        return redirect("/")
    try:
        with transaction.atomic(using="mes"):
            pedido = get_object_or_404(PedidoProduccion.objects.select_for_update(), pk=int(pk))
            if (pedido.estado or "") != "Cancelada":
                messages.error(request, "Solo se puede eliminar una orden cancelada.")
                return redirect("catalogos:pedidos_ordenes")
            items = list(PedidoProduccionItem.objects.select_for_update().filter(pedido=pedido))
            item_ids = [int(it.id) for it in items if it and getattr(it, "id", None)]
            if item_ids and LogisticaEnvioItem.objects.filter(pedido_item_id__in=item_ids).exists():
                messages.error(request, "No se puede eliminar: la orden tiene envíos registrados.")
                return redirect("catalogos:pedidos_ordenes")
            can_delete = True
            for it in items:
                if (
                    int(getattr(it, "apartado", 0) or 0) > 0
                    or int(getattr(it, "enviado", 0) or 0) > 0
                    or bool(getattr(it, "orden_herreria_id", None))
                    or (getattr(it, "estado_herreria", "") or "").strip() in {"Espera de material", "En producción"}
                ):
                    can_delete = False
                    break
            if not can_delete:
                messages.error(request, "No se puede eliminar: la orden tiene movimientos o está en proceso.")
                return redirect("catalogos:pedidos_ordenes")
            PedidoProduccionItem.objects.filter(pedido=pedido).delete()
            pedido.delete()
    except ProtectedError:
        messages.error(request, "No se puede eliminar: la orden tiene registros protegidos (envíos/historial).")
        return redirect("catalogos:pedidos_ordenes")
    messages.success(request, "Orden eliminada.")
    return redirect("catalogos:pedidos_ordenes")


@login_required
@require_http_methods(["GET", "POST"])
def pedidos_logistica(request):
    if not _can_pedidos(request.user):
        return redirect("/")

    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "apartar":
            item_id = int(request.POST.get("item_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            if item_id and qty > 0:
                with transaction.atomic(using="mes"):
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update(), pk=item_id)
                    if it.saldo <= 0:
                        messages.error(request, "El pedido ya no tiene saldo.")
                    else:
                        qty = min(qty, int(it.saldo))
                        try:
                            servicio_almacen.apartar(it.producto, qty, actor, linea_pedido=it)
                        except ErrorDeDominio as e:
                            messages.error(request, e.mensaje)
                        else:
                            it.apartado = int(it.apartado or 0) + qty
                            it.save(update_fields=["apartado", "actualizado_en"])
                            messages.success(request, "Apartado aplicado.")
                return redirect("catalogos:pedidos_logistica")

        if action == "enviar":
            item_id = int(request.POST.get("item_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            if item_id and qty > 0:
                with transaction.atomic(using="mes"):
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update().select_related("pedido", "producto"), pk=item_id)
                    if int(it.apartado or 0) <= 0:
                        messages.error(request, "No hay apartado para enviar.")
                        return redirect("catalogos:pedidos_logistica")
                    pdf = request.FILES.get("comprobante_pdf")
                    if not pdf:
                        messages.error(request, "PDF requerido para registrar el envío.")
                        return redirect("catalogos:pedidos_logistica")
                    else:
                        qty = min(qty, int(it.apartado or 0))
                        now = timezone.now()
                        envio = LogisticaEnvio.objects.create(
                            pedido=it.pedido,
                            fecha=timezone.localdate(),
                            actor=actor,
                            comprobante_pdf=pdf if pdf else None,
                        )
                        envio_item = LogisticaEnvioItem.objects.create(envio=envio, pedido_item=it, producto=it.producto, cantidad=qty, revertido=0)
                        it.apartado = int(it.apartado or 0) - qty
                        it.enviado = int(it.enviado or 0) + qty
                        it.save(update_fields=["apartado", "enviado", "actualizado_en"])
                        LogisticaMovimiento.objects.create(tipo="enviar", producto=it.producto, pedido_item=it, cantidad=-qty, actor=actor)
                    messages.success(request, f"Envío registrado (#{int(envio_item.id)}).")
                    items_all = list(
                        PedidoProduccionItem.objects.select_for_update()
                        .select_related("producto")
                        .filter(pedido_id=int(it.pedido_id))
                        .order_by("id")
                    )
                    if _pedido_is_completo(items_all):
                        _ensure_logistica_expediente(it.pedido, actor=actor, now=now)
                return redirect("catalogos:pedidos_logistica")

        if action == "revertir_envio":
            envio_item_id = int(request.POST.get("envio_item_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            destino = (request.POST.get("destino") or "stock").strip()
            if envio_item_id and qty > 0:
                with transaction.atomic(using="mes"):
                    ei = get_object_or_404(LogisticaEnvioItem.objects.select_for_update(), pk=envio_item_id)
                    it = get_object_or_404(PedidoProduccionItem.objects.select_for_update(), pk=int(ei.pedido_item_id))
                    disponible = int(ei.cantidad or 0) - int(ei.revertido or 0)
                    if disponible <= 0:
                        messages.error(request, "No hay cantidad disponible para revertir.")
                    else:
                        qty = min(qty, disponible)
                        ei.revertido = int(ei.revertido or 0) + qty
                        ei.save(update_fields=["revertido"])
                        it.enviado = max(0, int(it.enviado or 0) - qty)
                        if destino == "apartado":
                            it.apartado = int(it.apartado or 0) + qty
                            it.save(update_fields=["enviado", "apartado", "actualizado_en"])
                            LogisticaMovimiento.objects.create(tipo="revertir_a_apartado", producto=it.producto, pedido_item=it, cantidad=qty, actor=actor)
                            messages.success(request, "Reversión aplicada a apartado.")
                        else:
                            it.save(update_fields=["enviado", "actualizado_en"])
                            servicio_almacen.devolver_al_disponible(it.producto, qty, actor, linea_pedido=it)
                            messages.success(request, "Reversión aplicada a stock.")
                return redirect("catalogos:pedidos_logistica")

        if action == "decote_delete":
            pedido_id = int(request.POST.get("pedido_id") or 0)
            if pedido_id:
                today = timezone.localdate()
                with transaction.atomic(using="mes"):
                    pedido = get_object_or_404(PedidoProduccion.objects.using("mes").select_for_update(), pk=pedido_id)
                    items = list(PedidoProduccionItem.objects.using("mes").select_for_update().filter(pedido_id=int(pedido.id)).order_by("id"))
                    if not _pedido_is_completo(items):
                        messages.error(request, "No se puede eliminar: el pedido no está completamente enviado.")
                        return redirect("catalogos:pedidos_logistica")
                    last_fecha = (
                        LogisticaEnvio.objects.using("mes")
                        .filter(pedido_id=int(pedido.id))
                        .aggregate(mx=Max("fecha"))
                        .get("mx")
                    )
                    age_days = (today - last_fecha).days if last_fecha else -1
                    if not last_fecha or age_days < 5:
                        messages.error(request, "No se puede eliminar: aún no cumple 5 días en Decote.")
                        return redirect("catalogos:pedidos_logistica")
                    exp = LogisticaExpediente.objects.using("mes").select_for_update().filter(pedido_id=int(pedido.id)).first()
                    if not exp or not getattr(exp, "generado_en", None):
                        messages.error(request, "No se puede eliminar: el expediente no ha sido generado.")
                        return redirect("catalogos:pedidos_logistica")
                    if int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                        messages.error(request, "No se puede eliminar: el expediente no ha sido descargado.")
                        return redirect("catalogos:pedidos_logistica")

                    # LogisticaAcuseEntrega apunta a PedidoProduccionItem con
                    # PROTECT y se crea sola cuando herrería termina las piezas,
                    # así que casi cualquier pedido que haya pasado por producción
                    # tiene acuses. Borrar los ítems más abajo lanzaba entonces un
                    # ProtectedError sin capturar: un 500 en el flujo normal.
                    #
                    # Esta comprobación va aquí, antes del primer borrado, y no
                    # junto al delete que falla: para cuando se borran los ítems ya
                    # se han eliminado envíos, movimientos, expediente y, sobre
                    # todo, los PDF de comprobante del disco, que no los recupera
                    # ningún rollback.
                    #
                    # Un acuse firmado es un documento, así que se rechaza la purga
                    # en lugar de arrastrarlo. Queda por decidir con el taller si el
                    # Decote debe poder archivarlos.
                    n_acuses = (
                        LogisticaAcuseEntrega.objects.using("mes")
                        .filter(pedido_item__pedido_id=int(pedido.id))
                        .count()
                    )
                    if n_acuses:
                        messages.error(
                            request,
                            f"No se puede eliminar: el pedido tiene {n_acuses} acuse(s) de entrega "
                            "firmados. El expediente ya descargado conserva la documentación; "
                            "el pedido se mantiene por trazabilidad.",
                        )
                        return redirect("catalogos:pedidos_logistica")

                    envios_to_delete = list(LogisticaEnvio.objects.using("mes").filter(pedido_id=int(pedido.id)))
                    for e in envios_to_delete:
                        try:
                            if getattr(e, "comprobante_pdf", None):
                                e.comprobante_pdf.delete(save=False)
                        except Exception:
                            logger.exception("Error ignorado en pedidos_logistica()")
                    LogisticaEnvioItem.objects.using("mes").filter(envio__pedido_id=int(pedido.id)).delete()
                    LogisticaEnvio.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                    LogisticaMovimiento.objects.using("mes").filter(pedido_item__pedido_id=int(pedido.id)).delete()
                    LogisticaExpedienteDescarga.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                    LogisticaExpediente.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                    try:
                        PedidoProduccionItem.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                        pedido.delete(using="mes")
                    except ProtectedError as e:
                        protegidos = list(getattr(e, "protected_objects", []) or [])
                        detalle = ", ".join(str(o) for o in protegidos[:3])
                        extra = f" (+{len(protegidos) - 3} más)" if len(protegidos) > 3 else ""
                        messages.error(
                            request,
                            f"No se puede eliminar: hay registros que dependen del pedido"
                            f"{f' ({detalle}{extra})' if detalle else ''}.",
                        )
                        return redirect("catalogos:pedidos_logistica")
                    messages.success(request, "Pedido eliminado de Decote.")
                return redirect("catalogos:pedidos_logistica")

        if action == "revertir_apartado":
            item_id = int(request.POST.get("item_id") or 0)
            cantidad = int(request.POST.get("cantidad") or 0)
            if item_id and cantidad > 0:
                with transaction.atomic(using="mes"):
                    item = get_object_or_404(PedidoProduccionItem.objects.using("mes").select_for_update(), pk=item_id)
                    if int(getattr(item, "apartado", 0) or 0) < cantidad:
                        messages.error(request, "Cantidad a liberar excede lo apartada.")
                        return redirect("catalogos:pedidos_logistica")
                    item.apartado = max(0, int(getattr(item, "apartado", 0) or 0) - cantidad)
                    item.save(update_fields=["apartado", "actualizado_en"])
                    servicio_almacen.devolver_al_disponible(item.producto, cantidad, actor, linea_pedido=item)
                    messages.success(request, f"Cantidad liberada: {cantidad}.")
                return redirect("catalogos:pedidos_logistica")

        if action == "purge_enviado":
            pedido_id = int(request.POST.get("pedido_id") or 0)
            if pedido_id:
                today = timezone.localdate()
                with transaction.atomic(using="mes"):
                    pedido = get_object_or_404(PedidoProduccion.objects.using("mes").select_for_update(), pk=pedido_id)
                    items = list(
                        PedidoProduccionItem.objects.using("mes")
                        .select_for_update()
                        .filter(pedido_id=int(pedido.id))
                        .exclude(estado_herreria="Cancelado")
                    )
                    all_done = True
                    for it in items:
                        saldo = max(0, int(it.cantidad_total or 0) - int(it.apartado or 0) - int(it.enviado or 0))
                        if saldo != 0:
                            all_done = False
                            break
                    last_fecha = (
                        LogisticaEnvio.objects.using("mes")
                        .filter(pedido_id=int(pedido.id))
                        .aggregate(mx=Max("fecha"))
                        .get("mx")
                    )
                    age_days = (today - last_fecha).days if last_fecha else -1
                    if not items or not all_done:
                        messages.error(request, "No se puede eliminar: el pedido no está completamente enviado.")
                    elif not last_fecha or age_days <= 5:
                        messages.error(request, "No se puede eliminar: aún no cumple más de 5 días en Enviados.")
                    else:
                        exp = LogisticaExpediente.objects.using("mes").select_for_update().filter(pedido_id=int(pedido.id)).first()
                        if not exp or not getattr(exp, "generado_en", None):
                            messages.error(request, "No se puede eliminar: el expediente no ha sido generado.")
                        elif int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                            messages.error(request, "No se puede eliminar: el expediente no ha sido descargado.")
                        else:
                            LogisticaEnvioItem.objects.using("mes").filter(envio__pedido_id=int(pedido.id)).delete()
                            LogisticaEnvio.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                            LogisticaMovimiento.objects.using("mes").filter(pedido_item__pedido_id=int(pedido.id)).delete()
                            LogisticaExpedienteDescarga.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                            LogisticaExpediente.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                            PedidoProduccionItem.objects.using("mes").filter(pedido_id=int(pedido.id)).delete()
                            pedido.delete(using="mes")
                            messages.success(request, "Pedido eliminado de Enviados.")
                return redirect("catalogos:pedidos_logistica")

    stocks = list(
        LogisticaStock.objects.select_related("producto")
        .filter(producto__activo=True, stock__gt=0)
        .order_by("producto__nombre")[:2000]
    )

    items_all = list(
        PedidoProduccionItem.objects.select_related("pedido", "producto")
        .filter(pedido__estado="Activa")
        .exclude(estado_herreria="Cancelado")
        .order_by("-creado_en")[:2000]
    )
    pedido_by_id: dict[int, PedidoProduccion] = {}
    items_by_pedido_id: dict[int, list[PedidoProduccionItem]] = {}
    for it in items_all:
        it.saldo_calc = max(0, int(it.cantidad_total or 0) - int(it.apartado or 0) - int(it.enviado or 0))
        pid = int(getattr(getattr(it, "pedido", None), "id", 0) or 0)
        if pid:
            pedido_by_id[pid] = it.pedido
            items_by_pedido_id.setdefault(pid, []).append(it)
    activos = [
        it
        for it in items_all
        if int(getattr(it, "saldo_calc", 0) or 0) > 0 and int(getattr(it, "apartado", 0) or 0) == 0
    ]

    complete_ids: list[int] = []
    for pid, items in items_by_pedido_id.items():
        if items and all(int(getattr(x, "saldo_calc", 0) or 0) == 0 for x in items):
            if all(int(getattr(x, "apartado", 0) or 0) == 0 for x in items):
                if any(int(getattr(x, "enviado", 0) or 0) > 0 for x in items):
                    complete_ids.append(int(pid))

    apartados_items: list[PedidoProduccionItem] = []
    apartados_items = [it for it in items_all if int(getattr(it, "apartado", 0) or 0) > 0]
    enviados_pedidos: list[PedidoProduccion] = []
    decote_pedidos: list[PedidoProduccion] = []
    if complete_ids:
        last_map = {
        int(r["pedido_id"]): r["last_fecha"]
        for r in LogisticaEnvio.objects.filter(pedido_id__in=complete_ids)
        .values("pedido_id")
        .annotate(last_fecha=Max("fecha"))
        }
        today = timezone.localdate()
        exp_map = {int(e.pedido_id): e for e in LogisticaExpediente.objects.filter(pedido_id__in=complete_ids)}
        for pid in complete_ids:
            p = pedido_by_id.get(int(pid))
            if not p:
                continue
            last_fecha = last_map.get(int(pid))
            p.last_envio_fecha = last_fecha
            p.enviados_age_days = (today - last_fecha).days if last_fecha else None
            exp = exp_map.get(int(pid))
            p.expediente_generado = bool(exp and getattr(exp, "generado_en", None))
            p.expediente_descargado = bool(exp and int(getattr(exp, "descargas_count", 0) or 0) > 0)
            p.__dict__["expediente_descargas_count"] = int(getattr(exp, "descargas_count", 0) or 0) if exp else 0
            can_decote = bool(p.expediente_generado and p.expediente_descargado)
            is_decote_age = bool(last_fecha and (today - last_fecha).days >= DECOTE_DAYS)
            p.can_delete_decote = bool(last_fecha and (today - last_fecha).days >= 5)
            if is_decote_age and can_decote:
                decote_pedidos.append(p)
            else:
                enviados_pedidos.append(p)
        enviados_pedidos.sort(key=lambda x: (getattr(x, "last_envio_fecha", None) or datetime.min.date()), reverse=True)
        decote_pedidos.sort(key=lambda x: (getattr(x, "last_envio_fecha", None) or datetime.min.date()), reverse=True)

    envios_items = list(LogisticaEnvioItem.objects.select_related("envio", "pedido_item", "producto")[:250])
    envios = list(LogisticaEnvio.objects.select_related("pedido").order_by("-creado_en")[:200])

    return render(
        request,
        "catalogos/pedidos_logistica.html",
        {
        "stocks": stocks,
        "activos": activos,
        "apartados_items": apartados_items,
        "enviados_pedidos": enviados_pedidos,
        "decote_pedidos": decote_pedidos,
        "decote_days": DECOTE_DAYS,
        "envios_items": envios_items,
        "envios": envios,
        "logistica_tab": "op",
        },
    )


@login_required
@require_http_methods(["GET"])
def pedidos_expediente_zip(request, pk: int):
    if not _can_pedidos(request.user):
        return redirect("/")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    zip_bytes = _build_pedido_expediente_zip_bytes(int(pk))
    now = timezone.now()
    with transaction.atomic(using="mes"):
        pedido = get_object_or_404(PedidoProduccion.objects.using("mes").select_for_update(), pk=int(pk))
        exp = _ensure_logistica_expediente(pedido, actor=actor, now=now)
        first = not bool(int(getattr(exp, "descargas_count", 0) or 0) > 0)
        exp.descargas_count = int(getattr(exp, "descargas_count", 0) or 0) + 1
        if first:
            exp.primera_descarga_en = now
            exp.primera_descarga_por = actor
        exp.ultima_descarga_en = now
        exp.ultima_descarga_por = actor
        exp.save(
            update_fields=[
                "descargas_count",
                "primera_descarga_en",
                "primera_descarga_por",
                "ultima_descarga_en",
                "ultima_descarga_por",
                "actualizado_en",
            ]
        )
        LogisticaExpedienteDescarga.objects.create(expediente=exp, pedido=pedido, descargado_en=now, usuario=actor)

    folio = ""
    try:
        folio = (pedido.folio or "").strip()
    except Exception:
        folio = ""
    filename = f"Expediente_{folio or int(pk)}.zip"
    resp = HttpResponse(zip_bytes, content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_http_methods(["GET"])
def corta_expediente_zip(request, pk: int):
    if not (_can_pedidos(request.user) or _can_corte_laser(request.user) or _is_admin_user(request.user)):
        return redirect("/")
    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    now = timezone.now()
    with transaction.atomic(using="mes"):
        orden = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_for_update(), pk=int(pk))
        if not _corta_is_completo(orden):
            return redirect("catalogos:logistica_corta")
        exp = _ensure_logistica_expediente_corta(orden, actor=actor, now=now)
        first = not bool(int(getattr(exp, "descargas_count", 0) or 0) > 0)
        exp.descargas_count = int(getattr(exp, "descargas_count", 0) or 0) + 1
        if first:
            exp.primera_descarga_en = now
            exp.primera_descarga_por = actor
        exp.ultima_descarga_en = now
        exp.ultima_descarga_por = actor
        exp.save(
            update_fields=[
                "descargas_count",
                "primera_descarga_en",
                "primera_descarga_por",
                "ultima_descarga_en",
                "ultima_descarga_por",
                "actualizado_en",
            ]
        )
        LogisticaExpedienteDescarga.objects.create(expediente=exp, orden_corta=orden, descargado_en=now, usuario=actor)

    zip_bytes = _build_corta_expediente_zip_bytes(int(pk))
    folio = ""
    try:
        folio = ((getattr(orden, "folio_externo", "") or getattr(orden, "codigo", "") or getattr(orden, "folio", "")) or "").strip()
    except Exception:
        folio = ""
    filename = f"Expediente_{folio or int(pk)}.zip"
    resp = HttpResponse(zip_bytes, content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    return resp


@login_required
@require_http_methods(["GET", "POST"])
def logistica_corta(request):
    if not (_can_pedidos(request.user) or _can_corte_laser(request.user) or _is_admin_user(request.user)):
        return redirect("/")

    actor = ""
    try:
        actor = request.user.get_username() or ""
    except Exception:
        actor = ""
    now = timezone.now()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "apartar":
            orden_id = int(request.POST.get("orden_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            if orden_id and qty > 0:
                with transaction.atomic(using="mes"):
                    o = get_object_or_404(LaserOrdenProduccion.objects.select_for_update(), pk=orden_id)
                    if (o.estado or "").strip() != "Abierta":
                        messages.error(request, "La orden está cerrada/cancelada.")
                    else:
                        total = int(getattr(o, "total_piezas", 0) or 0)
                        saldo = max(0, total - int(getattr(o, "apartado", 0) or 0) - int(getattr(o, "enviado", 0) or 0))
                        if saldo <= 0:
                            messages.error(request, "El pedido ya no tiene saldo.")
                        else:
                            producto = _corta_producto_from_orden(o)
                            if not producto:
                                messages.error(request, "Producto inválido en la orden.")
                            else:
                                qty = min(qty, saldo)
                                try:
                                    servicio_almacen.apartar_corta(producto, qty, actor, orden=o)
                                except ErrorDeDominio as e:
                                    messages.error(request, e.mensaje)
                                else:
                                    o.apartado = int(getattr(o, "apartado", 0) or 0) + qty
                                    o.save(update_fields=["apartado", "actualizado_en"])
                                    messages.success(request, "Apartado aplicado.")
            return redirect("catalogos:logistica_corta")

        if action == "revertir_apartado_corta":
            orden_id = int(request.POST.get("orden_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            if orden_id and qty > 0:
                with transaction.atomic(using="mes"):
                    o = get_object_or_404(LaserOrdenProduccion.objects.select_for_update(), pk=orden_id)
                    if int(getattr(o, "apartado", 0) or 0) < qty:
                        messages.error(request, "Cantidad a liberar excede lo apartada.")
                    else:
                        o.apartado = max(0, int(getattr(o, "apartado", 0) or 0) - qty)
                        o.save(update_fields=["apartado", "actualizado_en"])
                        producto = _corta_producto_from_orden(o)
                        if producto:
                            servicio_almacen.devolver_al_disponible_corta(producto, qty, actor, orden=o)
                        messages.success(request, f"Apartado liberado: {qty}.")
            return redirect("catalogos:logistica_corta")

        if action == "decote_delete":
            orden_id = int(request.POST.get("orden_id") or 0)
            if orden_id:
                today = timezone.localdate()
                with transaction.atomic(using="mes"):
                    o = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_for_update(), pk=orden_id)
                    if not _corta_is_completo(o):
                        messages.error(request, "No se puede eliminar: el pedido no está completamente enviado.")
                        return redirect("catalogos:logistica_corta")
                    last_fecha = (
                        LogisticaEnvioCorta.objects.using("mes")
                        .filter(orden_id=int(o.id))
                        .aggregate(mx=Max("fecha"))
                        .get("mx")
                    )
                    age_days = (today - last_fecha).days if last_fecha else -1
                    if not last_fecha or age_days < 5:
                        messages.error(request, "No se puede eliminar: aún no cumple 5 días en Decote.")
                        return redirect("catalogos:logistica_corta")
                    exp = LogisticaExpediente.objects.using("mes").select_for_update().filter(orden_corta_id=int(o.id)).first()
                    if not exp or not getattr(exp, "generado_en", None):
                        messages.error(request, "No se puede eliminar: el expediente no ha sido generado.")
                        return redirect("catalogos:logistica_corta")
                    if int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                        messages.error(request, "No se puede eliminar: el expediente no ha sido descargado.")
                        return redirect("catalogos:logistica_corta")
                    envios_to_delete = list(LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(o.id)))
                    for e in envios_to_delete:
                        try:
                            if getattr(e, "comprobante_pdf", None):
                                e.comprobante_pdf.delete(save=False)
                        except Exception:
                            logger.exception("Error ignorado en logistica_corta()")
                    LogisticaMovimientoCorta.objects.using("mes").filter(orden_id=int(o.id)).delete()
                    LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(o.id)).delete()
                    LogisticaExpedienteDescarga.objects.using("mes").filter(orden_corta_id=int(o.id)).delete()
                    LogisticaExpediente.objects.using("mes").filter(orden_corta_id=int(o.id)).delete()
                    LaserProduccion.objects.using("mes").filter(orden_item__orden_id=int(o.id)).delete()
                    LaserAsignacion.objects.using("mes").filter(orden_id=int(o.id)).delete()
                    o.delete(using="mes")
                    messages.success(request, "Pedido eliminado de Decote.")
            return redirect("catalogos:logistica_corta")

        if action == "purge_enviado":
            orden_id = int(request.POST.get("orden_id") or 0)
            if orden_id:
                today = timezone.localdate()
                with transaction.atomic(using="mes"):
                    o = get_object_or_404(LaserOrdenProduccion.objects.using("mes").select_for_update(), pk=orden_id)
                    last_fecha = (
                        LogisticaEnvioCorta.objects.using("mes")
                        .filter(orden_id=int(o.id))
                        .aggregate(mx=Max("fecha"))
                        .get("mx")
                    )
                    age_days = (today - last_fecha).days if last_fecha else -1
                    if not _corta_is_completo(o):
                        messages.error(request, "No se puede eliminar: el pedido no está completamente enviado.")
                    elif not last_fecha or age_days <= 5:
                        messages.error(request, "No se puede eliminar: aún no cumple más de 5 días en Enviados.")
                    else:
                        exp = LogisticaExpediente.objects.using("mes").select_for_update().filter(orden_corta_id=int(o.id)).first()
                        if not exp or not getattr(exp, "generado_en", None):
                            messages.error(request, "No se puede eliminar: el expediente no ha sido generado.")
                        elif int(getattr(exp, "descargas_count", 0) or 0) <= 0:
                            messages.error(request, "No se puede eliminar: el expediente no ha sido descargado.")
                        else:
                            envios_to_delete = list(LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(o.id)))
                            for e in envios_to_delete:
                                try:
                                    if getattr(e, "comprobante_pdf", None):
                                        e.comprobante_pdf.delete(save=False)
                                except Exception:
                                    logger.exception("Error ignorado en logistica_corta()")
                            LogisticaMovimientoCorta.objects.using("mes").filter(orden_id=int(o.id)).delete()
                            LogisticaEnvioCorta.objects.using("mes").filter(orden_id=int(o.id)).delete()
                            LogisticaExpedienteDescarga.objects.using("mes").filter(orden_corta_id=int(o.id)).delete()
                            LogisticaExpediente.objects.using("mes").filter(orden_corta_id=int(o.id)).delete()
                            LaserProduccion.objects.using("mes").filter(orden_item__orden_id=int(o.id)).delete()
                            LaserAsignacion.objects.using("mes").filter(orden_id=int(o.id)).delete()
                            o.delete(using="mes")
                            messages.success(request, "Pedido eliminado de Enviados.")
            return redirect("catalogos:logistica_corta")

        if action == "enviar":
            orden_id = int(request.POST.get("orden_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            if orden_id and qty > 0:
                with transaction.atomic(using="mes"):
                    o = get_object_or_404(LaserOrdenProduccion.objects.select_for_update(), pk=orden_id)
                    if (o.estado or "").strip() != "Abierta":
                        messages.error(request, "La orden está cerrada/cancelada.")
                    else:
                        disponible = int(getattr(o, "apartado", 0) or 0)
                        if disponible <= 0:
                            messages.error(request, "No hay apartado para enviar.")
                        else:
                            producto = _corta_producto_from_orden(o)
                            if not producto:
                                messages.error(request, "Producto inválido en la orden.")
                            else:
                                qty = min(qty, disponible)
                                pdf = request.FILES.get("comprobante_pdf")
                                if not pdf:
                                    messages.error(request, "PDF requerido para registrar el envío.")
                                    return redirect("catalogos:logistica_corta")
                                envio = LogisticaEnvioCorta.objects.create(
                                    orden=o,
                                    fecha=timezone.localdate(),
                                    actor=actor,
                                    comprobante_pdf=pdf if pdf else None,
                                    producto=producto,
                                    cantidad=qty,
                                    revertido=0,
                                )
                                o.apartado = max(0, int(getattr(o, "apartado", 0) or 0) - qty)
                                o.enviado = int(getattr(o, "enviado", 0) or 0) + qty
                                o.save(update_fields=["apartado", "enviado", "actualizado_en"])
                                LogisticaMovimientoCorta.objects.create(
                                    tipo="enviar",
                                    producto=producto,
                                    orden=o,
                                    cantidad=-qty,
                                    actor=actor,
                                )
                                if _corta_is_completo(o):
                                    _ensure_logistica_expediente_corta(o, actor=actor, now=now)
                                messages.success(request, f"Envío registrado (#{int(envio.id)}).")
            return redirect("catalogos:logistica_corta")

        if action == "revertir_envio":
            envio_id = int(request.POST.get("envio_id") or 0)
            qty = int(request.POST.get("cantidad") or 0)
            destino = (request.POST.get("destino") or "stock").strip()
            if envio_id and qty > 0:
                with transaction.atomic(using="mes"):
                    e = get_object_or_404(LogisticaEnvioCorta.objects.select_for_update(), pk=envio_id)
                    o = get_object_or_404(LaserOrdenProduccion.objects.select_for_update(), pk=int(e.orden_id))
                    disponible = int(e.cantidad or 0) - int(e.revertido or 0)
                    if disponible <= 0:
                        messages.error(request, "No hay cantidad disponible para revertir.")
                    else:
                        qty = min(qty, disponible)
                        e.revertido = int(e.revertido or 0) + qty
                        e.save(update_fields=["revertido"])
                        o.enviado = max(0, int(getattr(o, "enviado", 0) or 0) - qty)
                        if destino == "apartado":
                            o.apartado = int(getattr(o, "apartado", 0) or 0) + qty
                            o.save(update_fields=["enviado", "apartado", "actualizado_en"])
                            LogisticaMovimientoCorta.objects.create(tipo="revertir_a_apartado", producto=e.producto, orden=o, cantidad=qty, actor=actor)
                            messages.success(request, "Reversión aplicada a apartado.")
                        else:
                            o.save(update_fields=["enviado", "actualizado_en"])
                            stock = _get_or_create_stock_corta_locked(e.producto)
                            stock.stock = int(stock.stock or 0) + qty
                            stock.save(update_fields=["stock", "actualizado_en"])
                            LogisticaMovimientoCorta.objects.create(tipo="revertir_a_stock", producto=e.producto, orden=o, cantidad=qty, actor=actor)
                            messages.success(request, "Reversión aplicada a stock.")
            return redirect("catalogos:logistica_corta")

    stocks = list(LogisticaStockCorta.objects.filter(stock__gt=0).order_by("producto")[:2000])

    ordenes = list(
        LaserOrdenProduccion.objects.select_related("corta_cliente_proyecto")
        .filter(estado="Abierta")
        .order_by("-id")[:2000]
    )
    today = timezone.localdate()
    last_map: dict[int, object] = {}
    ids = [int(o.id) for o in ordenes]
    if ids:
        last_map = {
            int(r["orden_id"]): r["last_fecha"]
            for r in LogisticaEnvioCorta.objects.filter(orden_id__in=ids).values("orden_id").annotate(last_fecha=Max("fecha"))
        }
    exp_map_c: dict[int, object] = {}
    if ids:
        exp_map_c = {int(e.orden_corta_id): e for e in LogisticaExpediente.objects.filter(orden_corta_id__in=ids)}
    apartados_ordenes: list[LaserOrdenProduccion] = []
    activos_ordenes: list[LaserOrdenProduccion] = []
    enviados_ordenes: list[LaserOrdenProduccion] = []
    decote_ordenes: list[LaserOrdenProduccion] = []
    for o in ordenes:
        total = int(getattr(o, "total_piezas", 0) or 0)
        apartada = int(getattr(o, "apartado", 0) or 0)
        enviado = int(getattr(o, "enviado", 0) or 0)
        saldo = max(0, total - apartada - enviado)
        if total > 0 and apartada > 0:
            o.folio_display = (getattr(o, "folio_externo", "") or "").strip() or (getattr(o, "codigo", "") or "").strip() or o.folio
            o.cliente_display = getattr(getattr(o, "corta_cliente_proyecto", None), "nombre", "") or ""
            o.producto_display = _corta_producto_from_orden(o)
            o.total_display = total
            o.apartado_display = apartada
            o.enviado_display = enviado
            o.saldo_display = saldo
            apartados_ordenes.append(o)
        elif saldo > 0:
            o.folio_display = (getattr(o, "folio_externo", "") or "").strip() or (getattr(o, "codigo", "") or "").strip() or o.folio
            o.cliente_display = getattr(getattr(o, "corta_cliente_proyecto", None), "nombre", "") or ""
            o.producto_display = _corta_producto_from_orden(o)
            o.total_display = total
            o.apartado_display = apartada
            o.enviado_display = enviado
            o.saldo_display = saldo
            activos_ordenes.append(o)
        elif enviado > 0:
            o.folio_display = (getattr(o, "folio_externo", "") or "").strip() or (getattr(o, "codigo", "") or "").strip() or o.folio
            o.cliente_display = getattr(getattr(o, "corta_cliente_proyecto", None), "nombre", "") or ""
            o.producto_display = _corta_producto_from_orden(o)
            o.total_display = total
            o.apartado_display = apartada
            o.enviado_display = enviado
            o.saldo_display = saldo
            last_fecha = last_map.get(int(o.id))
            o.last_envio_fecha = last_fecha
            o.enviados_age_days = (today - last_fecha).days if last_fecha else None
            exp = exp_map_c.get(int(o.id))
            o.expediente_generado = bool(exp and getattr(exp, "generado_en", None))
            o.expediente_descargado = bool(exp and int(getattr(exp, "descargas_count", 0) or 0) > 0)
            o.__dict__["expediente_descargas_count"] = int(getattr(exp, "descargas_count", 0) or 0) if exp else 0
            can_decote = bool(o.expediente_generado and o.expediente_descargado)
            is_decote_age = bool(last_fecha and (today - last_fecha).days >= DECOTE_DAYS)
            o.can_delete_decote = bool(last_fecha and (today - last_fecha).days >= 5)
            if is_decote_age and can_decote:
                decote_ordenes.append(o)
            else:
                enviados_ordenes.append(o)

    envios = list(LogisticaEnvioCorta.objects.select_related("orden").order_by("-creado_en")[:250])
    for e in envios:
        e.orden_folio = (getattr(getattr(e, "orden", None), "folio_externo", "") or "").strip() or (
            getattr(getattr(e, "orden", None), "codigo", "") or ""
        ).strip() or getattr(getattr(e, "orden", None), "folio", "")

    return render(
        request,
        "catalogos/pedidos_logistica_corta.html",
        {
            "stocks": stocks,
            "activos": activos_ordenes,
            "apartados": apartados_ordenes,
            "enviados_ordenes": enviados_ordenes,
            "decote_ordenes": decote_ordenes,
            "decote_days": DECOTE_DAYS,
            "envios": envios,
            "logistica_tab": "corta",
        },
    )
