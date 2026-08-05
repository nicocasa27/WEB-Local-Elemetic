import json
import csv
import logging
import io
import base64
import re
import statistics
from datetime import datetime, time, timedelta, timezone as dt_timezone
from pathlib import Path

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.files.base import ContentFile
from django.contrib.staticfiles import finders
from django.db import connections, transaction
from django.db.models import Avg, Case, Count, DateTimeField, DurationField, ExpressionWrapper, F, FloatField, IntegerField, Q, Sum, Value, When, Window
from django.db.models.functions import Upper
from django.db.models.functions import Coalesce, Greatest, Least
from django.db.models.functions.window import Lead
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
try:
    from pypdf import PdfReader
except ModuleNotFoundError:
    PdfReader = None
from django.views.decorators.http import require_POST

from core import estados as core_estados
from core import metricas, paginacion, roles
from core.estados import clase as clase_de_estado
from core.servicios import inventario as servicio_inventario
from core.servicios import ruta as servicio_ruta
from core.servicios import trabajo as servicio_trabajo

from catalogos.models import Proyecto
from catalogos.models import EquipoTrabajo
from catalogos.models import PdfExtractionTemplate
from catalogos.models import (
    Colaborador,
    Maquina,
    MaquinaFalla,
    MaquinaParo,
    PlantaEvento,
    VigaAsignacion,
)
from catalogos.models import SeguimientoDespacho
from catalogos.models import RobotProduccion
from catalogos.models import HerrOrdenAsignacion
from catalogos.models import HerrAsignacion
from catalogos.models import HerrOrdenItem
from catalogos.models import HerrOrdenProduccion
from catalogos.models import LaserOrdenProduccion
from catalogos.models import LaserEstadoCambio
from catalogos.models import LogisticaEnvioCorta
from catalogos.models import LogisticaEnvioItem
from catalogos.models import RobotOrdenAsignacion
from catalogos.models import RobotOrdenItem
from catalogos.models import RobotOrdenProduccion
from catalogos.models import VigaPlano
from catalogos.models import WeeklyReportSnapshot

from .forms import StatusChangeForm, VigaBatchCreateForm, VigaForm, VigaImportUploadForm
from .models import ESTADOS, ProductionLog, Viga

logger = logging.getLogger("mes.produccion")

STATUS_COLORS = {
    "Espera de corte": "#ff8f00",
    "Corte": "#f39c12",
    "Espera de armado": "#d35400",
    "Armado": "#3498db",
    "Espera de soldadura": "#5d6d7e",
    "Soldadura": "#9b59b6",
    "Espera de pintura": "#7f8c8d",
    "Pintura": "#16a085",
    "Terminado": "#2dce89",
    "Enviado": "#11cdef",
    "Espera Armado": "#d35400",
    "Espera Soldadura": "#5d6d7e",
    "Espera Pintura": "#7f8c8d",
}

ESTADO_ALIASES = {
    "ESPERA ARMADO": "Espera de armado",
    "ESPERA SOLDADURA": "Espera de soldadura",
    "ESPERA PINTURA": "Espera de pintura",
}


def _norm_estado(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    up = s.upper()
    if up in ESTADO_ALIASES:
        return ESTADO_ALIASES[up]
    return s


def _estado_variants(value: str) -> list[str]:
    s = _norm_estado(value)
    if not s:
        return []
    out = {s}
    for k, v in ESTADO_ALIASES.items():
        if v == s:
            out.add(k.title())
    return list(out)

DECOTE_DAYS = 5


def _format_utc_naive_dt_as_local(dt) -> str:
    if not dt:
        return ""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, dt_timezone.utc)
    return timezone.localtime(dt, timezone.get_default_timezone()).strftime("%Y-%m-%d %H:%M")


def _utc_naive_dt_to_local(dt):
    if not dt:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, dt_timezone.utc)
    return timezone.localtime(dt, timezone.get_default_timezone())


def _labor_seconds_between(start_utc_naive, end_utc_naive) -> float:
    start_local = _utc_naive_dt_to_local(start_utc_naive)
    end_local = _utc_naive_dt_to_local(end_utc_naive)
    if not start_local or not end_local or end_local <= start_local:
        return 0.0

    day_start = start_local.date()
    day_end = end_local.date()
    total = 0.0
    cur = day_start
    while cur <= day_end:
        if cur.weekday() < 5:
            w1_start = timezone.make_aware(datetime.combine(cur, time(7, 30)), timezone.get_default_timezone())
            w1_end = timezone.make_aware(datetime.combine(cur, time(13, 0)), timezone.get_default_timezone())
            w2_start = timezone.make_aware(datetime.combine(cur, time(13, 30)), timezone.get_default_timezone())
            w2_end = timezone.make_aware(datetime.combine(cur, time(17, 0)), timezone.get_default_timezone())

            s = max(start_local, w1_start)
            e = min(end_local, w1_end)
            if e > s:
                total += (e - s).total_seconds()

            s = max(start_local, w2_start)
            e = min(end_local, w2_end)
            if e > s:
                total += (e - s).total_seconds()
        cur = cur + timedelta(days=1)
    return total

def _safe_next(value: str) -> str:
    value = (value or "").strip()
    if value.startswith("/"):
        return value
    return ""


#: Tipos por extensión, para incrustar un archivo dentro del HTML exportado.
_TIPOS_MIME = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "webp": "image/webp",
    "gif": "image/gif",
    "svg": "image/svg+xml",
    "woff2": "font/woff2",
    "woff": "font/woff",
    "ttf": "font/ttf",
}


def _ruta_de_estatico(url: str):
    """Del `/static/...` de una página al archivo en disco, o `None`.

    Se prueban las dos formas porque conviven: con `collectstatic` el archivo
    está en `STATIC_ROOT` y su nombre lleva un hash; sin él, hay que buscarlo
    en las carpetas `static/` de las aplicaciones.
    """
    ruta = (url or "").split("?", 1)[0].split("#", 1)[0]
    prefijo = (settings.STATIC_URL or "/static/").rstrip("/")
    if prefijo and ruta.startswith(prefijo):
        ruta = ruta[len(prefijo) :]
    ruta = ruta.lstrip("/")
    if not ruta:
        return None

    recogido = Path(settings.STATIC_ROOT or "") / ruta
    if settings.STATIC_ROOT and recogido.is_file():
        return recogido

    encontrado = finders.find(ruta)
    return Path(encontrado) if encontrado else None


def _leer_estatico(url: str) -> str:
    """Contenido de texto de un archivo estático propio.

    Antes esto descargaba el CSS y el JavaScript de un CDN **durante la
    petición**, con quince segundos de espera. Ahora todo está en el disco del
    servidor, que es lo que permite que el taller trabaje sin internet.
    """
    ruta = _ruta_de_estatico(url)
    if ruta is None:
        return ""
    try:
        return ruta.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        logger.exception("No se pudo leer el estático %s", url)
        return ""


def _static_data_uri(static_path: str) -> str:
    ruta = _ruta_de_estatico(static_path)
    if ruta is None:
        return ""
    try:
        crudo = ruta.read_bytes()
    except OSError:
        logger.exception("No se pudo leer el estático %s", static_path)
        return ""
    mime = _TIPOS_MIME.get(ruta.suffix.lstrip(".").lower(), "application/octet-stream")
    return f"data:{mime};base64,{base64.b64encode(crudo).decode('ascii')}"


def _incrustar_urls_de_css(css: str, ruta_css: Path) -> str:
    """Mete las fuentes dentro del propio CSS.

    El HTML que se exporta se abre fuera de la aplicación, muchas veces desde
    una carpeta o un correo. Ahí una `url(fonts/...)` relativa no resuelve, y
    sin esto el archivo exportado sale sin ningún icono.
    """

    def sustituir(m):
        destino = m.group(1).strip("\"'")
        if destino.startswith(("data:", "http://", "https://", "#")):
            return m.group(0)
        archivo = (ruta_css.parent / destino.split("?", 1)[0].split("#", 1)[0]).resolve()
        if not archivo.is_file():
            return m.group(0)
        mime = _TIPOS_MIME.get(archivo.suffix.lstrip(".").lower(), "application/octet-stream")
        datos = base64.b64encode(archivo.read_bytes()).decode("ascii")
        return f'url("data:{mime};base64,{datos}")'

    return re.sub(r"url\(([^)]+)\)", sustituir, css)


def _inline_html_assets(html: str) -> str:
    """Convierte una página en un HTML suelto que se abre sin servidor.

    Se usa para los reportes que el taller descarga y manda por correo. Lo que
    la página pide con un enlace, aquí se mete dentro del archivo.
    """
    if not html:
        return html

    def repl_link(m):
        tag = m.group(0)
        href = m.group(1)
        if not href or 'rel="stylesheet"' not in tag:
            return tag
        ruta = _ruta_de_estatico(href)
        if ruta is None:
            return tag
        css = _incrustar_urls_de_css(_leer_estatico(href), ruta)
        return f"<style>\n{css}\n</style>"

    def repl_script(m):
        src = m.group(1)
        if not src:
            return m.group(0)
        js = _leer_estatico(src)
        if not js:
            return m.group(0)
        # Un `</script>` dentro del código cerraría la etiqueta que lo envuelve
        # y partiría el archivo por la mitad.
        return "<script>\n" + js.replace("</script>", "<\\/script>") + "\n</script>"

    def repl_img(m):
        before = m.group(1)
        src = m.group(2)
        after = m.group(3)
        if not src:
            return m.group(0)
        data_uri = _static_data_uri(src)
        if data_uri:
            return f'<img{before}src="{data_uri}"{after}>'
        return m.group(0)

    html = re.sub(r'<meta name="viewport" content="[^"]*"\s*/?>', '<meta name="viewport" content="width=1200, initial-scale=1" />', html)
    html = re.sub(r"<link[^>]+href=\"([^\"]+)\"[^>]*>", repl_link, html)
    html = re.sub(r"<script[^>]+src=\"([^\"]+)\"[^>]*>\s*</script>", repl_script, html)
    html = re.sub(r"<img([^>]*?)src=\"([^\"]+)\"([^>]*)>", repl_img, html)
    return html


def _dashboard_quien_detalle_payload(colab_id: int, etapa: str) -> dict:
    etapa = (etapa or "").strip()
    if etapa not in {"Corte", "Armado", "Soldadura", "Pintura"}:
        return {"ok": False, "error": "Etapa inválida."}

    viga_ids = list(
        VigaAsignacion.objects.filter(
            vigente=True,
            etapa=etapa,
            colaborador_id=int(colab_id),
        )
        .values_list("viga_internal_id", flat=True)
        .distinct()
    )

    qs = (
        Viga.objects.filter(internal_id__in=viga_ids)
        .exclude(estado__in=["Enviado", "Terminado"])
        .order_by("estado", "codigo_viga", "pieza_no")
    )
    items = []
    for v in qs[:250]:
        kg = float(getattr(v, "peso_kg", 0.0) or 0.0)
        items.append(
            {
                "area": "Producción",
                "id": int(v.internal_id),
                "codigo": getattr(v, "codigo_viga", ""),
                "pieza": f"{getattr(v, 'pieza_no', '')}/{getattr(v, 'total_piezas', '')}",
                "proyecto": getattr(v, "proyecto", ""),
                "estado": getattr(v, "estado", ""),
                "ton": round(kg / 1000.0, 3),
                "fecha_compromiso": v.fecha_compromiso.isoformat() if getattr(v, "fecha_compromiso", None) else "",
            }
        )

    r_asigs = list(
        RobotOrdenAsignacion.objects.filter(etapa=etapa, colaborador_id=int(colab_id))
        .values_list("orden_id", flat=True)
        .distinct()
    )
    if r_asigs:
        r_kg_rows = (
            RobotOrdenItem.objects.filter(orden_id__in=r_asigs)
            .values("orden_id")
            .annotate(
                kg=Sum(
                    ExpressionWrapper(
                        F("cantidad_requerida")
                        * Coalesce(F("pieza__peso_kg"), F("pieza_custom_peso_kg"), 0.0),
                        output_field=FloatField(),
                    )
                )
            )
        )
        r_kg_map = {int(r["orden_id"]): float(r["kg"] or 0.0) for r in r_kg_rows}
        r_orders = (
            RobotOrdenProduccion.objects.filter(id__in=r_asigs, estado="Abierta")
            .select_related("proyecto")
            .order_by("-id")[:250]
        )
        for o in r_orders:
            kg = float(r_kg_map.get(int(o.id), 0.0) or 0.0)
            items.append(
                {
                    "area": "Robótica",
                    "id": int(o.id),
                    "codigo": o.folio,
                    "pieza": (o.nombre or o.producto or "").strip(),
                    "proyecto": (o.proyecto.nombre if getattr(o, "proyecto_id", None) else "") or "",
                    "estado": f"{etapa}",
                    "ton": round(kg / 1000.0, 3),
                    "fecha_compromiso": "",
                }
            )

    h_asigs_old = list(
        HerrOrdenAsignacion.objects.filter(etapa=etapa, colaborador_id=int(colab_id))
        .values_list("orden_id", flat=True)
        .distinct()
    )
    h_asigs_new = list(
        HerrAsignacion.objects.filter(
            vigente=True,
            etapa=etapa,
            colaborador_id=int(colab_id),
        )
        .values_list("orden_id", flat=True)
        .distinct()
    )
    h_asigs = list(dict.fromkeys([*h_asigs_old, *h_asigs_new]))
    if h_asigs:
        h_kg_rows = (
            HerrOrdenItem.objects.filter(orden_id__in=h_asigs)
            .values("orden_id")
            .annotate(
                kg=Sum(
                    ExpressionWrapper(
                        F("cantidad_requerida")
                        * Coalesce(F("pieza__peso_kg"), F("pieza_custom_peso_kg"), 0.0),
                        output_field=FloatField(),
                    )
                )
            )
        )
        h_kg_map = {int(r["orden_id"]): float(r["kg"] or 0.0) for r in h_kg_rows}
        h_orders = (
            HerrOrdenProduccion.objects.filter(id__in=h_asigs, estado="Abierta")
            .exclude(estado_etapa="Terminado")
            .select_related("proyecto")
            .order_by("-id")[:250]
        )
        for o in h_orders:
            kg = float(h_kg_map.get(int(o.id), 0.0) or 0.0)
            items.append(
                {
                    "area": "Herrería",
                    "id": int(o.id),
                    "codigo": o.folio,
                    "pieza": (o.nombre or "").strip(),
                    "proyecto": (o.proyecto.nombre if getattr(o, "proyecto_id", None) else "") or "",
                    "estado": getattr(o, "estado_etapa", "") or "",
                    "ton": round(kg / 1000.0, 3),
                    "fecha_compromiso": "",
                }
            )

    colab = Colaborador.objects.filter(id=int(colab_id)).select_related("equipo").first()
    return {
        "ok": True,
        "colaborador": {
            "id": int(colab.id) if colab else int(colab_id),
            "nombre": colab.nombre if colab else "",
            "equipo": getattr(colab.equipo, "nombre", "") if colab else "",
        },
        "etapa": etapa,
        "count": len(items),
        "items": items,
    }


def _user_in_group(user, name: str) -> bool:
    try:
        return bool(user and user.is_authenticated and user.groups.filter(name=name).exists())
    except Exception:
        return False


def _is_admin_user(user) -> bool:
    return bool(
        user
        and user.is_authenticated
        and (
            getattr(user, "is_superuser", False)
            or getattr(user, "is_staff", False)
            or _user_in_group(user, "admin_general")
            or _user_in_group(user, "ingenieria_civil")
        )
    )


def _user_role(user) -> str:
    if _is_admin_user(user):
        return "admin"
    if _user_in_group(user, "corte"):
        return "corte"
    if _user_in_group(user, "soldadura"):
        return "soldadura"
    if _user_in_group(user, "robotica"):
        return "robotica"
    if _user_in_group(user, "herreria"):
        return "herreria"
    return ""



#: Qué etapas puede mover cada grupo del piso.
#:
#: Estaba escrita dos veces, palabra por palabra, en las dos vistas que
#: cambian el estado de una pieza. Dos copias de una regla de permisos es una
#: que se va a quedar vieja sin que nadie lo note.
#:
#: Corte y soldadura se solapan en «Espera de armado», y soldadura y pintura
#: en «Espera de pintura», a propósito: la etapa de espera es el punto de
#: entrega entre dos áreas y las dos tienen que poder tocarla. Sin el solape,
#: la pieza se queda en tierra de nadie.
ETAPAS_POR_GRUPO = {
    "corte": {"Espera de corte", "Corte", "Espera de armado"},
    "soldadura": {
        "Espera de armado",
        "Armado",
        "Espera de soldadura",
        "Soldadura",
        "Espera de pintura",
    },
    "pintura": {"Espera de pintura", "Pintura", "Terminado"},
}

#: Lo que «soldadura» cubría antes de que existiera el grupo de pintura.
ETAPAS_DE_PINTURA = {"Espera de pintura", "Pintura", "Terminado"}


def _etapas_permitidas(user):
    """Las etapas que esta cuenta puede mover. Vacío es «ninguna».

    La red de seguridad: mientras no exista **ninguna** cuenta de pintura,
    «soldadura» sigue cubriendo pintura y terminado, como hacía antes. Sin
    ella, el día del despliegue las piezas en pintura dejarían de poder
    avanzar hasta que alguien se acordara de mover a los pintores de grupo, y
    nadie relacionaría una cosa con la otra: el síntoma sería «el sistema ya
    no me deja» tres días después del cambio.

    En cuanto haya un solo pintor con cuenta, el reparto se separa solo.
    """
    permitidas = set()
    for grupo, etapas in ETAPAS_POR_GRUPO.items():
        if _user_in_group(user, grupo):
            permitidas |= etapas
    if _user_in_group(user, "soldadura") and not roles.hay_cuentas_de_pintura():
        permitidas |= ETAPAS_DE_PINTURA
    return permitidas


def _equipo_for_etapa(etapa: str):
    """El equipo que cubre una etapa.

    Se busca en dos pasadas. Primero por `estados`, que es la lista explícita
    de etapas del equipo y manda cuando está capturada. Si nadie la capturó
    —que es el caso de los cuatro equipos que hay hoy: los cuatro tienen
    `estados` vacío— se cae al `area`, que dice exactamente lo mismo y sí
    está puesta.

    Sin esa segunda pasada la función devolvía `None` siempre, así que el
    diálogo «Asignaciones por etapa» abría con las tres listas en blanco y
    no se podía asignar a nadie: el aviso decía «marca las casillas» y no
    había ninguna casilla.
    """
    etapa = (etapa or "").strip()
    if not etapa:
        return None
    equipos = list(
        EquipoTrabajo.objects.filter(activo=True).order_by("area", "sub_area", "nombre")
    )
    for equipo in equipos:
        if etapa in equipo.estados:
            return equipo
    clave = etapa.casefold()
    for equipo in equipos:
        if clave in {(equipo.area or "").casefold(), (equipo.sub_area or "").casefold()}:
            return equipo
    return None


def _build_asignacion_payload():
    payload = {}
    armado_equipo = _equipo_for_etapa("Armado")
    if armado_equipo:
        qs = Colaborador.objects.filter(activo=True, equipo=armado_equipo).order_by("rol", "nombre")
        payload["Armado"] = {
            "equipo": armado_equipo.nombre,
            "soldadores": [{"id": c.id, "nombre": c.nombre} for c in qs.filter(rol="Soldador")],
            "auxiliares": [{"id": c.id, "nombre": c.nombre} for c in qs.filter(rol="Auxiliar")],
        }
    pintura_equipo = _equipo_for_etapa("Pintura")
    if pintura_equipo:
        qs = Colaborador.objects.filter(activo=True, equipo=pintura_equipo).order_by("rol", "nombre")
        payload["Pintura"] = {
            "equipo": pintura_equipo.nombre,
            "pintores": [{"id": c.id, "nombre": c.nombre} for c in qs.filter(rol="Pintor")],
        }
    return payload


def _build_participantes_payload():
    def build_for(etapa_key: str, etapa_lookup: str):
        equipo = _equipo_for_etapa(etapa_lookup)
        if not equipo and etapa_lookup == "Soldadura":
            equipo = _equipo_for_etapa("Armado")
        if not equipo:
            return None
        qs = Colaborador.objects.filter(activo=True, equipo=equipo).order_by("rol", "nombre")
        maquinas = []
        operadores = []
        if etapa_key == "Corte":
            maquinas = list(
                Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre").values("id", "nombre")
            )
            operadores = [{"id": c.id, "nombre": c.nombre} for c in qs]
        return {
            "equipo": equipo.nombre,
            "items": [{"id": c.id, "nombre": c.nombre, "rol": c.rol} for c in qs],
            "operadores": operadores,
            "maquinas": [{"id": int(m["id"]), "nombre": m["nombre"]} for m in maquinas],
        }

    payload = {}
    for key in ("Corte", "Soldadura", "Pintura"):
        block = build_for(key, key)
        if block:
            payload[key] = block
    return payload


def _parse_ids_csv(value: str):
    raw = (value or "").strip()
    if not raw:
        return []
    parts = [p.strip() for p in raw.split(",")]
    ids = []
    for p in parts:
        if p.isdigit():
            ids.append(int(p))
    return list(dict.fromkeys(ids))


def _save_asignaciones_for_etapa(viga_internal_id: int, etapa: str, post_data, actor: str = ""):
    """Guarda a quién se asigna una etapa. Devuelve `(salió bien, error)`.

    La comprobación del equipo se hace **después** de saber si hay alguien a
    quien asignar, y ése era un fallo que paraba el taller:

    El equipo se busca por el área, y no hay ninguno con área «Armado» —los
    cuatro del taller son Corte, Soldadura, Soldadura y Pintura—. Como la
    comprobación estaba arriba del todo, cualquier intento de pasar una pieza
    a Armado moría con un 400 «No hay equipo configurado para la etapa
    Armado», **aunque no se estuviera asignando a nadie**. Las once piezas que
    esperan armado no se podían mover, ni desde la lista ni desde el celular,
    y el mensaje no daba ninguna pista de que se arreglaba dando de alta un
    equipo en Configuración.

    Sin nadie a quien asignar no hace falta ningún equipo: no hay nada que
    validar contra él.
    """
    etapa = (etapa or "").strip()
    actor = (actor or "").strip()

    def con_equipo():
        equipo = _equipo_for_etapa(etapa)
        if not equipo:
            return None, (
                f"No hay ningún equipo del área de {etapa}. Se da de alta en "
                "Configuración de planta → Equipos."
            )
        return equipo, ""

    if etapa == "Armado":
        soldador_id = int(post_data.get("soldador_id") or 0)
        auxiliares_ids = []
        try:
            if hasattr(post_data, "getlist"):
                auxiliares_ids = [int(x) for x in post_data.getlist("auxiliar_ids") if str(x).isdigit()]
        except Exception:
            auxiliares_ids = []
        if not auxiliares_ids:
            auxiliares_ids = _parse_ids_csv(post_data.get("auxiliar_ids") or "")
        if not soldador_id and not auxiliares_ids:
            return True, ""
        equipo, error = con_equipo()
        if error:
            return False, error
        if not soldador_id:
            return False, "Debes asignar 1 soldador para Armado."
        if len(auxiliares_ids) < 1 or len(auxiliares_ids) > 2:
            return False, "Debes asignar 1 o 2 auxiliares para Armado."
        soldador = Colaborador.objects.using("mes").filter(id=soldador_id, activo=True, equipo=equipo, rol="Soldador").first()
        if not soldador:
            return False, "Soldador inválido para Armado."
        auxiliares = list(
            Colaborador.objects.using("mes").filter(id__in=auxiliares_ids, activo=True, equipo=equipo, rol="Auxiliar")
        )
        if len(auxiliares) != len(auxiliares_ids):
            return False, "Auxiliares inválidos para Armado."

        VigaAsignacion.objects.using("mes").filter(viga_internal_id=viga_internal_id, etapa=etapa, vigente=True).update(vigente=False)
        VigaAsignacion.objects.using("mes").create(
            viga_internal_id=viga_internal_id,
            etapa=etapa,
            rol="Soldador",
            colaborador=soldador,
            vigente=True,
            asignado_por=actor,
        )
        for a in auxiliares:
            VigaAsignacion.objects.using("mes").create(
                viga_internal_id=viga_internal_id,
                etapa=etapa,
                rol="Auxiliar",
                colaborador=a,
                vigente=True,
                asignado_por=actor,
            )
        return True, ""

    if etapa == "Pintura":
        pintor_id = int(post_data.get("pintor_id") or 0)
        if not pintor_id:
            return True, ""
        equipo, error = con_equipo()
        if error:
            return False, error
        pintor = Colaborador.objects.using("mes").filter(id=pintor_id, activo=True, equipo=equipo, rol="Pintor").first()
        if not pintor:
            return False, "Pintor inválido para Pintura."
        VigaAsignacion.objects.using("mes").filter(viga_internal_id=viga_internal_id, etapa=etapa, vigente=True).update(vigente=False)
        VigaAsignacion.objects.using("mes").create(
            viga_internal_id=viga_internal_id,
            etapa=etapa,
            rol="Pintor",
            colaborador=pintor,
            vigente=True,
            asignado_por=actor,
        )
        return True, ""

    return True, ""


def _delete_asignaciones_for_vigas(internal_ids):
    ids = [int(x) for x in (internal_ids or []) if str(x).isdigit()]
    if not ids:
        return
    VigaAsignacion.objects.using("mes").filter(viga_internal_id__in=ids).delete()


def _equipos_de_corte():
    """Los seis equipos de corte, para el selector del avance.

    Se ordenan por nombre y no por antigüedad: el operador los busca por
    nombre, que es como los llama en el piso.
    """
    return list(
        Maquina.objects.using("mes")
        .filter(activo=True, tipo="Corte", es_robot=False)
        .order_by("nombre")
        .values("id", "nombre", "funcion")
    )


def _maquina_asignada(viga_internal_id, etapa):
    """El equipo ya asignado a la pieza para esa etapa, si lo hay.

    Existe para que quien ya asignó máquina desde la pantalla de asignaciones
    no tenga que volver a elegirla al avanzar. Pedir dos veces el mismo dato es
    la forma más rápida de que la gente deje de darlo.
    """
    identificador = (
        VigaAsignacion.objects.using("mes")
        .filter(
            viga_internal_id=int(viga_internal_id),
            etapa=etapa,
            vigente=True,
            maquina_id__isnull=False,
        )
        .values_list("maquina_id", flat=True)
        .first()
    )
    if not identificador:
        return None
    return Maquina.objects.using("mes").filter(id=int(identificador)).first()


def _validate_corte_asignacion(corte_operadores_ids, corte_maquina_ids):
    corte_operadores_ids = [int(x) for x in (corte_operadores_ids or []) if str(x).isdigit()]
    corte_maquina_ids = [int(x) for x in (corte_maquina_ids or []) if str(x).isdigit()]
    corte_operadores_ids = list(dict.fromkeys(corte_operadores_ids))
    corte_maquina_ids = list(dict.fromkeys(corte_maquina_ids))

    if not corte_operadores_ids and not corte_maquina_ids:
        return True, "", [], []
    if not corte_operadores_ids:
        return False, "Debes seleccionar al menos 1 operador en Corte.", [], []

    corte_equipo = _equipo_for_etapa("Corte")
    if not corte_equipo:
        return False, "No hay equipo configurado para la etapa Corte.", [], []

    allowed_ops = set(
        Colaborador.objects.filter(activo=True, equipo=corte_equipo).values_list("id", flat=True)
    )
    if any(i not in allowed_ops for i in corte_operadores_ids):
        return False, "Operadores inválidos en Corte. Revisa colaboradores activos del equipo.", [], []

    if corte_maquina_ids:
        allowed_m = set(
            Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False, id__in=corte_maquina_ids).values_list("id", flat=True)
        )
        if any(i not in allowed_m for i in corte_maquina_ids):
            return False, "Máquinas de corte inválidas.", [], []

    return True, "", corte_operadores_ids, corte_maquina_ids


def _save_corte_asignaciones_for_vigas(internal_ids, corte_operadores_ids, corte_maquina_ids, actor: str = ""):
    actor = (actor or "").strip()
    ids = [int(x) for x in (internal_ids or []) if str(x).isdigit()]
    if not ids:
        return
    for vid in ids:
        VigaAsignacion.objects.using("mes").filter(viga_internal_id=vid, etapa="Corte", vigente=True).update(vigente=False)
        for cid in corte_operadores_ids:
            VigaAsignacion.objects.using("mes").create(
                viga_internal_id=vid,
                etapa="Corte",
                rol="Operador",
                colaborador_id=cid,
                maquina=None,
                vigente=True,
                asignado_por=actor,
            )
        for mid in corte_maquina_ids:
            VigaAsignacion.objects.using("mes").create(
                viga_internal_id=vid,
                etapa="Corte",
                rol="Maquina",
                colaborador=None,
                maquina_id=mid,
                vigente=True,
                asignado_por=actor,
            )


def _save_plano_for_vigas(internal_ids, uploaded_file):
    if not uploaded_file:
        return
    name = (getattr(uploaded_file, "name", "") or "").lower()
    content_type = (getattr(uploaded_file, "content_type", "") or "").lower()
    if not (name.endswith(".pdf") or "pdf" in content_type):
        return
    data = uploaded_file.read()
    if not data:
        return
    for internal_id in internal_ids:
        if not internal_id:
            continue
        filename = f"viga_{internal_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        defaults = {
            "nombre_original": getattr(uploaded_file, "name", "") or "",
        }
        plano, created = VigaPlano.objects.using("mes").get_or_create(viga_internal_id=internal_id, defaults=defaults)
        if not created and plano.archivo_pdf:
            try:
                plano.archivo_pdf.delete(save=False)
            except Exception:
                logger.exception("Error ignorado en _save_plano_for_vigas()")
        plano.nombre_original = getattr(uploaded_file, "name", "") or ""
        plano.archivo_pdf.save(filename, ContentFile(data), save=True)


def _delete_plano_for_vigas(internal_ids):
    ids = [int(x) for x in (internal_ids or []) if str(x).isdigit()]
    if not ids:
        return
    for plano in VigaPlano.objects.filter(viga_internal_id__in=ids).only("id", "archivo_pdf"):
        try:
            if plano.archivo_pdf:
                plano.archivo_pdf.delete(save=False)
        except Exception:
            logger.exception("Error ignorado en _delete_plano_for_vigas()")
        try:
            plano.delete()
        except Exception:
            logger.exception("Error ignorado en _delete_plano_for_vigas()")


@login_required
def home(request):
    """La portada: qué hay que atender hoy.

    Antes era un muro de diecisiete mosaicos sobre una foto —«selecciona tu
    grupo»— que sólo servía para entrar. Con la barra lateral eso sobra: el
    menú ya está siempre a la vista, así que la portada puede decir algo en vez
    de preguntar a dónde vas.

    Los números salen de lo que ya se deduce en cada módulo. Ninguno se guarda
    aparte: un contador guardado se desactualiza y entonces la portada miente,
    que es peor que no tenerla.
    """
    u = request.user
    can_admin = bool(_is_admin_user(u) or getattr(u, "is_staff", False))

    return render(
        request,
        "produccion/home.html",
        {
            "can_admin": can_admin,
            "pendientes": _pendientes_del_dia(u),
            "hoy": timezone.localdate(),
        },
    )


def _pendientes_del_dia(usuario):
    """Los avisos de la portada, cada uno con a dónde ir a resolverlo.

    Un aviso sin destino obliga a buscar la pantalla; con destino, ver el
    problema y entrar a arreglarlo es el mismo gesto.

    Si algo falla al contar, ese aviso no sale y los demás sí. Una portada en
    blanco por un módulo caído deja al taller sin saber qué hacer.
    """
    from catalogos.despacho import cuantos_esperan_despacho
    from catalogos.models import Cuadrilla
    from core import roles
    from inventario.compras import cuantos_hay_que_comprar

    avisos = []

    def contar(titulo, detalle, icono, ruta, funcion, tono="neutro", cuando_cero=None):
        try:
            cuantos = funcion()
        except Exception:
            logger.exception("no se pudo contar «%s» para la portada", titulo)
            return
        if not cuantos and cuando_cero is None:
            return
        avisos.append({
            "titulo": titulo,
            "cuantos": cuantos,
            "detalle": cuando_cero if not cuantos else detalle,
            "icono": icono,
            "ruta": ruta,
            "tono": tono if cuantos else "bien",
        })

    contar(
        "Listo para salir", "esperan que Logística los despache",
        "bi-box-seam", "catalogos:despacho", cuantos_esperan_despacho,
        cuando_cero="Nada terminado esperando camión",
    )
    contar(
        "Por comprar", "materiales en o bajo su mínimo",
        "bi-cart-plus", "inventario:compras", cuantos_hay_que_comprar,
        tono="aviso", cuando_cero="Ningún material bajo mínimo",
    )
    contar(
        "Material apartado", "obras con material esperando surtirse",
        "bi-buildings", "inventario:por_proyecto",
        lambda: len(servicio_inventario.proyectos_por_surtir()),
    )

    # La cuadrilla es distinta: lo que importa es que falte, no que haya.
    if roles.puede_administrar_usuarios(usuario) or can_ver_cuadrillas(usuario):
        try:
            armadas = Cuadrilla.objects.using("mes").filter(
                fecha=timezone.localdate()
            ).count()
        except Exception:
            logger.exception("no se pudo contar las cuadrillas de hoy")
        else:
            avisos.append({
                "titulo": "Cuadrillas de hoy",
                "cuantos": armadas,
                "detalle": "armadas" if armadas else "Sin armar: el trabajo de hoy no se atribuirá a nadie",
                "icono": "bi-people-fill",
                "ruta": "catalogos:cuadrillas",
                "tono": "bien" if armadas else "aviso",
            })

    return avisos


def can_ver_cuadrillas(usuario):
    return bool(getattr(usuario, "is_staff", False) or _is_admin_user(usuario))

def _sync_projects() -> None:
    names = (
        Viga.objects.values_list("proyecto", flat=True)
        .distinct()
    )
    for raw in names:
        name = (raw or "").strip()
        if not name:
            continue
        norm = name.upper()
        Proyecto.objects.get_or_create(
            nombre_normalizado=norm,
            defaults={"nombre": norm, "activo": True},
        )


def _viga_queryset(request):
    qs = Viga.objects.all()

    active_projects = list(
        Proyecto.objects.filter(activo=True).values_list("nombre_normalizado", flat=True)
    )
    if active_projects:
        qs = qs.annotate(proyecto_norm=Upper("proyecto")).filter(proyecto_norm__in=active_projects)

    estado = request.GET.get("estado", "").strip()
    if estado == "Enviado":
        estado = ""
    qs = qs.exclude(estado="Enviado")
    if estado and estado != "Todos":
        variants = _estado_variants(estado)
        if variants:
            qs = qs.filter(estado__in=variants)
        else:
            qs = qs.filter(estado=estado)

    proyecto = request.GET.get("proyecto", "").strip()
    if proyecto and proyecto != "Todos":
        qs = qs.filter(proyecto__iexact=proyecto)

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(codigo_viga__icontains=q)
            | Q(proyecto__icontains=q)
            | Q(descripcion__icontains=q)
        )

    order = request.GET.get("order", "").strip() or "ultimo_mov_desc"
    order_map = {
        "ultimo_mov_desc": ["-ultimo_cambio", "prioridad", "fecha_compromiso", "proyecto", "codigo_viga", "pieza_no"],
        "ultimo_mov_asc": ["ultimo_cambio", "prioridad", "fecha_compromiso", "proyecto", "codigo_viga", "pieza_no"],
        "prioridad_asc": ["prioridad", "fecha_compromiso", "proyecto", "codigo_viga", "pieza_no"],
        "prioridad_desc": ["-prioridad", "fecha_compromiso", "proyecto", "codigo_viga", "pieza_no"],
        "fecha_asc": ["fecha_compromiso", "prioridad", "codigo_viga", "pieza_no"],
        "fecha_desc": ["-fecha_compromiso", "prioridad", "codigo_viga", "pieza_no"],
        "proyecto": ["proyecto", "prioridad", "codigo_viga", "pieza_no"],
    }
    if order == "estado":
        whens = []
        for i, s in enumerate(ESTADOS):
            for v in _estado_variants(s):
                whens.append(When(estado=v, then=Value(i)))
        qs = qs.annotate(
            estado_rank=Case(
                *whens,
                default=Value(len(ESTADOS)),
                output_field=IntegerField(),
            )
        ).order_by("estado_rank", "prioridad", "fecha_compromiso", "proyecto", "codigo_viga", "pieza_no")
    else:
        qs = qs.order_by(*order_map.get(order, order_map["prioridad_asc"]))

    return qs, {"estado": estado, "proyecto": proyecto, "q": q, "order": order}


@login_required
def viga_list(request):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    _sync_projects()
    qs, filters = _viga_queryset(request)
    projects = list(
        Proyecto.objects.filter(activo=True)
        .values_list("nombre", flat=True)
        .order_by("nombre")
    )
    # Antes: `qs[:2000]`. El recorte era silencioso y el navegador tenía que
    # dibujar hasta dos mil filas antes de enseñar nada.
    pagina = paginacion.paginar(request, qs)
    vigas = list(pagina.object_list)
    plano_map = {}
    ids = [v.internal_id for v in vigas if getattr(v, "internal_id", None)]
    estados_index = {s: i for i, s in enumerate(ESTADOS)}
    for v in vigas:
        v.estado = _norm_estado(getattr(v, "estado", "") or "")
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        idx = estados_index.get(v.estado)
        if idx is None or idx >= len(ESTADOS) - 1:
            v.next_estado = ""
        else:
            v.next_estado = ESTADOS[idx + 1]
        if v.estado == "Terminado" and v.next_estado == "Enviado":
            v.next_estado = ""

    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    enviado_fecha_sq = Subquery(
        ProductionLog.objects.filter(viga_internal_id=OuterRef("pk"), estado_nuevo="Enviado")
        .order_by("-fecha_operacion", "-timestamp")
        .values("fecha_operacion")[:1]
    )
    enviados_qs = Viga.objects.filter(estado="Enviado")

    active_projects = list(
        Proyecto.objects.filter(activo=True).values_list("nombre_normalizado", flat=True)
    )
    if active_projects:
        enviados_qs = enviados_qs.annotate(proyecto_norm=Upper("proyecto")).filter(
            proyecto_norm__in=active_projects
        )

    if filters.get("proyecto") and filters["proyecto"] != "Todos":
        enviados_qs = enviados_qs.filter(proyecto__iexact=filters["proyecto"])

    if filters.get("q"):
        q = filters["q"]
        enviados_qs = enviados_qs.filter(
            Q(codigo_viga__icontains=q)
            | Q(proyecto__icontains=q)
            | Q(descripcion__icontains=q)
        )

    enviados_qs = enviados_qs.annotate(enviado_fecha=enviado_fecha_sq).filter(enviado_fecha__isnull=False)

    enviados_recent_qs = (
        enviados_qs.filter(enviado_fecha__gt=cutoff)
        .order_by("proyecto", "-enviado_fecha", "codigo_viga", "pieza_no")
    )
    decote_qs = (
        enviados_qs.filter(enviado_fecha__lte=cutoff)
        .order_by("enviado_fecha", "proyecto", "codigo_viga", "pieza_no")
    )

    enviados_vigas = list(enviados_recent_qs[:2000])
    decote_total = decote_qs.count()
    decote_vigas = list(decote_qs[:500])

    extra_ids = [v.internal_id for v in (enviados_vigas + decote_vigas) if getattr(v, "internal_id", None)]
    all_ids = list({int(x) for x in (ids + extra_ids) if x})
    if all_ids:
        for viga_id, archivo in VigaPlano.objects.filter(viga_internal_id__in=all_ids).values_list(
            "viga_internal_id", "archivo_pdf"
        ):
            if archivo:
                plano_map[int(viga_id)] = f"{settings.MEDIA_URL}{archivo}"

    enviados_por_proyecto = {}
    for v in enviados_vigas:
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        v.plano_url = plano_map.get(int(v.internal_id), "")
        v.enviado_dias = int((today - v.enviado_fecha).days) if getattr(v, "enviado_fecha", None) else 0
        key = (v.proyecto or "").strip().upper() or "SIN PROYECTO"
        enviados_por_proyecto.setdefault(key, []).append(v)

    for v in decote_vigas:
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        v.plano_url = plano_map.get(int(v.internal_id), "")
        v.enviado_dias = int((today - v.enviado_fecha).days) if getattr(v, "enviado_fecha", None) else 0

    for v in vigas:
        v.plano_url = plano_map.get(int(v.internal_id), "")

    asigns = (
        VigaAsignacion.objects.filter(viga_internal_id__in=ids, vigente=True, etapa__in=["Corte", "Soldadura", "Pintura"])
        .values_list("viga_internal_id", "etapa", "rol", "colaborador_id", "maquina_id")
    )
    asign_map = {}
    for vid, etapa, rol, cid, mid in asigns:
        entry = asign_map.setdefault(int(vid), {}).setdefault(str(etapa), {"ids": [], "operadores": [], "maquinas": []})
        if mid:
            entry["maquinas"].append(int(mid))
        if cid:
            if (rol or "") == "Operador":
                entry["operadores"].append(int(cid))
            else:
                entry["ids"].append(int(cid))
    for v in vigas:
        m = asign_map.get(int(v.internal_id), {})
        v.asig_corte = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("ids", []))
        v.asig_soldadura = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("ids", []))
        v.asig_pintura = ",".join(str(x) for x in (m.get("Pintura", {}) or {}).get("ids", []))
        v.asig_corte_maquinas = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("maquinas", []))
        v.asig_soldadura_maquinas = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("maquinas", []))
        v.asig_corte_operadores = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("operadores", []))
        v.asig_soldadura_operadores = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("operadores", []))

    maquina_ids = set()
    for v in vigas:
        for part in (getattr(v, "asig_corte_maquinas", "") or "").split(","):
            part = (part or "").strip()
            if part.isdigit():
                maquina_ids.add(int(part))
    maquinas_info = {}
    maquinas_estado = {}
    if maquina_ids:
        for m in Maquina.objects.filter(id__in=list(maquina_ids)).values("id", "nombre", "tipo", "es_robot"):
            maquinas_info[int(m["id"])] = {"nombre": m["nombre"], "tipo": m["tipo"], "es_robot": bool(m["es_robot"])}
        for p in MaquinaParo.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("motivo"):
            maquinas_estado[int(p.maquina_id)] = {
                "paro": True,
                "falla": False,
                "motivo": getattr(getattr(p, "motivo", None), "nombre", "") or "",
                "paro_inicio": timezone.localtime(p.inicio).strftime("%Y-%m-%d %H:%M") if getattr(p, "inicio", None) else "",
            }
        for f in MaquinaFalla.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("tipo"):
            entry = maquinas_estado.setdefault(int(f.maquina_id), {"paro": False, "falla": False, "motivo": "", "paro_inicio": ""})
            entry["falla"] = True
            entry["tipo_falla"] = getattr(getattr(f, "tipo", None), "nombre", "") or ""
            entry["falla_inicio"] = timezone.localtime(f.inicio).strftime("%Y-%m-%d %H:%M") if getattr(f, "inicio", None) else ""

    return render(
        request,
        "produccion/viga_list.html",
        {
            "vigas": vigas,
            "pagina": pagina,
            "mode": "admin",
            "reset_url": reverse("produccion:viga_list"),
            "estados_filtro": ["Todos", *[s for s in ESTADOS if s != "Enviado"]],
            "estados_operables": [s for s in ESTADOS if s != "Enviado"],
            "estados_order": ESTADOS,
            "proyectos": ["Todos", *projects],
            "filters": filters,
            "status_colors": STATUS_COLORS,
            "participantes_payload": _build_participantes_payload(),
            "equipos_de_corte": _equipos_de_corte(),
            "maquinas_info_payload": maquinas_info,
            "maquinas_estado_payload": maquinas_estado,
            "enviados_por_proyecto": sorted(enviados_por_proyecto.items(), key=lambda x: x[0]),
            "enviados_total": sum(len(items) for _p, items in enviados_por_proyecto.items()),
            "decote_vigas": decote_vigas,
            "decote_days": DECOTE_DAYS,
            "decote_total": decote_total,
        },
    )


@login_required
def area_corte(request):
    if not (_is_admin_user(request.user) or _user_in_group(request.user, "corte")):
        return redirect("produccion:home")

    u = request.user
    if not (_is_admin_user(u) or getattr(u, "is_staff", False)):
        src = (request.GET.get("src") or "").strip().lower()
        if src not in {"menu", "nav"}:
            can_corte = bool(_user_in_group(u, "corte"))
            can_soldadura = bool(_user_in_group(u, "soldadura"))
            can_robotica = bool(_user_in_group(u, "robotica"))
            can_herreria = bool(_user_in_group(u, "herreria") or _user_in_group(u, "herreria_supervision"))
            if sum([1 if can_corte else 0, 1 if can_soldadura else 0, 1 if can_robotica else 0, 1 if can_herreria else 0]) > 1:
                return redirect("produccion:home")

    _sync_projects()
    qs, filters = _viga_queryset(request)
    allowed_norm = {"Espera de corte", "Corte", "Espera de armado"}
    allowed_variants = set()
    for s in allowed_norm:
        allowed_variants.update(_estado_variants(s))
    if filters.get("estado") and filters["estado"] not in {"", "Todos"} and _norm_estado(filters["estado"]) not in allowed_norm:
        filters["estado"] = "Todos"
    qs = qs.filter(estado__in=list(allowed_variants))

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    # Antes: `qs[:2000]`, un recorte silencioso.
    pagina = paginacion.paginar(request, qs)
    vigas = list(pagina.object_list)
    ids = [v.internal_id for v in vigas if getattr(v, "internal_id", None)]
    plano_map = {}
    if ids:
        for viga_id, archivo in VigaPlano.objects.filter(viga_internal_id__in=ids).values_list("viga_internal_id", "archivo_pdf"):
            if archivo:
                plano_map[int(viga_id)] = f"{settings.MEDIA_URL}{archivo}"

    for v in vigas:
        v.estado = _norm_estado(getattr(v, "estado", "") or "")
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        if v.estado == "Espera de corte":
            v.next_estado = "Corte"
        elif v.estado == "Corte":
            v.next_estado = "Espera de armado"
            v.show_badge = ""
        elif v.estado == "Espera de armado":
            v.next_estado = ""
            v.show_badge = "En área de soldadura"
        else:
            v.next_estado = ""
            v.show_badge = ""
        v.plano_url = plano_map.get(int(v.internal_id), "")

    asigns = (
        VigaAsignacion.objects.using("mes")
        .filter(viga_internal_id__in=ids, vigente=True, etapa__in=["Corte"])
        .values_list("viga_internal_id", "etapa", "rol", "colaborador_id", "maquina_id")
    )
    asign_map = {}
    for vid, etapa, rol, cid, mid in asigns:
        entry = asign_map.setdefault(int(vid), {}).setdefault(str(etapa), {"ids": [], "operadores": [], "maquinas": []})
        if mid:
            entry["maquinas"].append(int(mid))
        if cid:
            if (rol or "") == "Operador":
                entry["operadores"].append(int(cid))
            else:
                entry["ids"].append(int(cid))
    for v in vigas:
        m = asign_map.get(int(v.internal_id), {})
        v.asig_corte = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("ids", []))
        v.asig_corte_maquinas = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("maquinas", []))
        v.asig_corte_operadores = ",".join(str(x) for x in (m.get("Corte", {}) or {}).get("operadores", []))
        v.asig_soldadura = ""
        v.asig_pintura = ""
        v.asig_soldadura_maquinas = ""
        v.asig_soldadura_operadores = ""

    maquina_ids = set()
    for v in vigas:
        for part in (getattr(v, "asig_corte_maquinas", "") or "").split(","):
            part = (part or "").strip()
            if part.isdigit():
                maquina_ids.add(int(part))
    maquinas_info = {}
    maquinas_estado = {}
    if maquina_ids:
        for m in Maquina.objects.filter(id__in=list(maquina_ids)).values("id", "nombre", "tipo", "es_robot"):
            maquinas_info[int(m["id"])] = {"nombre": m["nombre"], "tipo": m["tipo"], "es_robot": bool(m["es_robot"])}
        for p in MaquinaParo.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("motivo"):
            maquinas_estado[int(p.maquina_id)] = {
                "paro": True,
                "falla": False,
                "motivo": getattr(getattr(p, "motivo", None), "nombre", "") or "",
                "paro_inicio": timezone.localtime(p.inicio).strftime("%Y-%m-%d %H:%M") if getattr(p, "inicio", None) else "",
            }
        for f in MaquinaFalla.objects.filter(maquina_id__in=list(maquina_ids), fin__isnull=True).select_related("tipo"):
            entry = maquinas_estado.setdefault(int(f.maquina_id), {"paro": False, "falla": False, "motivo": "", "paro_inicio": ""})
            entry["falla"] = True
            entry["tipo_falla"] = getattr(getattr(f, "tipo", None), "nombre", "") or ""
            entry["falla_inicio"] = timezone.localtime(f.inicio).strftime("%Y-%m-%d %H:%M") if getattr(f, "inicio", None) else ""

    return render(
        request,
        "produccion/viga_list.html",
        {
            "mode": "corte",
            "reset_url": reverse("produccion:area_corte"),
            "vigas": vigas,
            "pagina": pagina,
            "estados_filtro": ["Todos", "Espera de corte", "Corte", "Espera de armado"],
            "estados_operables": ["Espera de corte", "Corte", "Espera de armado"],
            "estados_order": ESTADOS,
            "proyectos": ["Todos", *projects],
            "filters": filters,
            "status_colors": STATUS_COLORS,
            "participantes_payload": _build_participantes_payload(),
            "equipos_de_corte": _equipos_de_corte(),
            "maquinas_info_payload": maquinas_info,
            "maquinas_estado_payload": maquinas_estado,
            "enviados_por_proyecto": [],
            "enviados_total": 0,
            "decote_vigas": [],
            "decote_days": DECOTE_DAYS,
            "decote_total": 0,
        },
    )


@login_required
def area_soldadura(request):
    if not (_is_admin_user(request.user) or _user_in_group(request.user, "soldadura")):
        return redirect("produccion:home")

    u = request.user
    if not (_is_admin_user(u) or getattr(u, "is_staff", False)):
        src = (request.GET.get("src") or "").strip().lower()
        if src not in {"menu", "nav"}:
            can_corte = bool(_user_in_group(u, "corte"))
            can_soldadura = bool(_user_in_group(u, "soldadura"))
            can_robotica = bool(_user_in_group(u, "robotica"))
            can_herreria = bool(_user_in_group(u, "herreria") or _user_in_group(u, "herreria_supervision"))
            if sum([1 if can_corte else 0, 1 if can_soldadura else 0, 1 if can_robotica else 0, 1 if can_herreria else 0]) > 1:
                return redirect("produccion:home")

    _sync_projects()
    qs, filters = _viga_queryset(request)
    allowed_norm = {"Espera de armado", "Armado", "Espera de soldadura", "Soldadura", "Espera de pintura", "Pintura", "Terminado"}
    allowed_variants = set()
    for s in allowed_norm:
        allowed_variants.update(_estado_variants(s))
    if filters.get("estado") and filters["estado"] not in {"", "Todos"} and _norm_estado(filters["estado"]) not in allowed_norm:
        filters["estado"] = "Todos"
    qs = qs.filter(estado__in=list(allowed_variants))

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    # Antes: `qs[:2000]`, un recorte silencioso.
    pagina = paginacion.paginar(request, qs)
    vigas = list(pagina.object_list)
    ids = [v.internal_id for v in vigas if getattr(v, "internal_id", None)]
    plano_map = {}
    if ids:
        for viga_id, archivo in VigaPlano.objects.filter(viga_internal_id__in=ids).values_list("viga_internal_id", "archivo_pdf"):
            if archivo:
                plano_map[int(viga_id)] = f"{settings.MEDIA_URL}{archivo}"

    estados_index = {s: i for i, s in enumerate(ESTADOS)}
    for v in vigas:
        v.estado = _norm_estado(getattr(v, "estado", "") or "")
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        idx = estados_index.get(v.estado)
        if idx is None or idx >= len(ESTADOS) - 1:
            v.next_estado = ""
        else:
            v.next_estado = ESTADOS[idx + 1]
        if v.estado == "Terminado" and v.next_estado == "Enviado":
            v.next_estado = ""
        v.plano_url = plano_map.get(int(v.internal_id), "")

    asigns = (
        VigaAsignacion.objects.using("mes")
        .filter(viga_internal_id__in=ids, vigente=True, etapa__in=["Soldadura", "Pintura"])
        .values_list("viga_internal_id", "etapa", "rol", "colaborador_id", "maquina_id")
    )
    asign_map = {}
    for vid, etapa, rol, cid, mid in asigns:
        entry = asign_map.setdefault(int(vid), {}).setdefault(str(etapa), {"ids": [], "operadores": [], "maquinas": []})
        if mid:
            entry["maquinas"].append(int(mid))
        if cid:
            entry["ids"].append(int(cid))
    for v in vigas:
        m = asign_map.get(int(v.internal_id), {})
        v.asig_corte = ""
        v.asig_corte_maquinas = ""
        v.asig_corte_operadores = ""
        v.asig_soldadura = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("ids", []))
        v.asig_pintura = ",".join(str(x) for x in (m.get("Pintura", {}) or {}).get("ids", []))
        v.asig_soldadura_maquinas = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("maquinas", []))
        v.asig_soldadura_operadores = ",".join(str(x) for x in (m.get("Soldadura", {}) or {}).get("operadores", []))

    return render(
        request,
        "produccion/viga_list.html",
        {
            "mode": "soldadura",
            "reset_url": reverse("produccion:area_soldadura"),
            "vigas": vigas,
            "pagina": pagina,
            "estados_filtro": ["Todos", "Espera de armado", "Armado", "Espera de soldadura", "Soldadura", "Espera de pintura", "Pintura", "Terminado"],
            "estados_operables": ["Espera de armado", "Armado", "Espera de soldadura", "Soldadura", "Espera de pintura", "Pintura", "Terminado"],
            "estados_order": ESTADOS,
            "proyectos": ["Todos", *projects],
            "filters": filters,
            "status_colors": STATUS_COLORS,
            "participantes_payload": _build_participantes_payload(),
            "equipos_de_corte": _equipos_de_corte(),
            "maquinas_info_payload": {},
            "maquinas_estado_payload": {},
            "enviados_por_proyecto": [],
            "enviados_total": 0,
            "decote_vigas": [],
            "decote_days": DECOTE_DAYS,
            "decote_total": 0,
        },
    )


@login_required
def viga_global(request):
    return redirect("produccion:solo_lectura_produccion")


@login_required
def solo_lectura_produccion(request):
    _sync_projects()
    qs, filters = _viga_queryset(request)
    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    # Antes: `qs[:2000]`, un recorte silencioso.
    pagina = paginacion.paginar(request, qs)
    vigas = list(pagina.object_list)
    ids = [v.internal_id for v in vigas if getattr(v, "internal_id", None)]
    plano_map = {}
    if ids:
        for viga_id, archivo in VigaPlano.objects.filter(viga_internal_id__in=ids).values_list("viga_internal_id", "archivo_pdf"):
            if archivo:
                plano_map[int(viga_id)] = f"{settings.MEDIA_URL}{archivo}"
    for v in vigas:
        v.estado_color = STATUS_COLORS.get(v.estado, "#8898aa")
        v.ultimo_mov_str = _format_utc_naive_dt_as_local(getattr(v, "ultimo_cambio", None))
        v.plano_url = plano_map.get(int(v.internal_id), "")
    return render(
        request,
        "produccion/solo_lectura_produccion.html",
        {
            "vigas": vigas,
            "pagina": pagina,
            "filters": filters,
            "proyectos": ["Todos", *projects],
            "status_colors": STATUS_COLORS,
            "reset_url": reverse("produccion:solo_lectura_produccion"),
            "solo_lectura_active": "produccion",
        },
    )


@login_required
def solo_lectura_herreria(request):
    estado_doc = (request.GET.get("estado") or "Abierta").strip() or "Abierta"
    proyecto = (request.GET.get("proyecto") or "Todos").strip() or "Todos"
    q = (request.GET.get("q") or "").strip()
    filters = {"estado": estado_doc, "proyecto": proyecto, "q": q}

    qs = HerrOrdenProduccion.objects.select_related("proyecto")
    if estado_doc and estado_doc not in {"Todos", ""}:
        qs = qs.filter(estado=estado_doc)
    if proyecto and proyecto not in {"Todos", ""}:
        qs = qs.filter(proyecto__nombre=proyecto)
    if q:
        qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(descripcion__icontains=q) | Q(proyecto__nombre__icontains=q))
    qs = qs.order_by("-ultimo_cambio", "-id")

    orders = list(qs[:2000])
    for o in orders:
        o.codigo_show = (o.codigo or o.folio or "").strip()
        o.proyecto_nombre = getattr(getattr(o, "proyecto", None), "nombre", "") or ""
        o.estado_etapa_show = (o.estado_etapa or "").strip()
        o.ultimo_mov_str = timezone.localtime(o.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if getattr(o, "ultimo_cambio", None) else ""
        o.ton = round(float(getattr(o, "peso_kg", 0.0) or 0.0) / 1000.0, 3)

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    estados_doc = ["Todos", "Abierta", "Cerrada", "Cancelada"]
    return render(
        request,
        "produccion/solo_lectura_herreria.html",
        {
            "orders": orders,
            "filters": filters,
            "proyectos": ["Todos", *projects],
            "estados_doc": estados_doc,
            "reset_url": reverse("produccion:solo_lectura_herreria"),
            "solo_lectura_active": "herreria",
        },
    )


@login_required
def solo_lectura_corte_laser(request):
    estado_doc = (request.GET.get("estado") or "Abierta").strip() or "Abierta"
    proyecto = (request.GET.get("proyecto") or "Todos").strip() or "Todos"
    q = (request.GET.get("q") or "").strip()
    filters = {"estado": estado_doc, "proyecto": proyecto, "q": q}

    qs = LaserOrdenProduccion.objects.select_related("proyecto", "material")
    if estado_doc and estado_doc not in {"Todos", ""}:
        qs = qs.filter(estado=estado_doc)
    if proyecto and proyecto not in {"Todos", ""}:
        qs = qs.filter(proyecto__nombre=proyecto)
    if q:
        qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(descripcion__icontains=q) | Q(proyecto__nombre__icontains=q))
    qs = qs.order_by("-ultimo_cambio", "-id")

    orders = list(qs[:2000])
    for o in orders:
        o.codigo_show = (o.codigo or o.folio or "").strip()
        o.proyecto_nombre = getattr(getattr(o, "proyecto", None), "nombre", "") or ""
        o.estado_etapa_show = (o.estado_etapa or "").strip()
        o.ultimo_mov_str = timezone.localtime(o.ultimo_cambio).strftime("%Y-%m-%d %H:%M") if getattr(o, "ultimo_cambio", None) else ""
        o.ton = round(float(getattr(o, "peso_kg", 0.0) or 0.0) / 1000.0, 3)

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    estados_doc = ["Todos", "Abierta", "Cerrada", "Cancelada"]
    return render(
        request,
        "produccion/solo_lectura_corte_laser.html",
        {
            "orders": orders,
            "filters": filters,
            "proyectos": ["Todos", *projects],
            "estados_doc": estados_doc,
            "reset_url": reverse("produccion:solo_lectura_corte_laser"),
            "solo_lectura_active": "corte_laser",
        },
    )


@login_required
def solo_lectura_robotica(request):
    estado_doc = (request.GET.get("estado") or "Abierta").strip() or "Abierta"
    proyecto = (request.GET.get("proyecto") or "Todos").strip() or "Todos"
    q = (request.GET.get("q") or "").strip()
    filters = {"estado": estado_doc, "proyecto": proyecto, "q": q}

    qs = RobotOrdenProduccion.objects.select_related("proyecto")
    if estado_doc and estado_doc not in {"Todos", ""}:
        qs = qs.filter(estado=estado_doc)
    if proyecto and proyecto not in {"Todos", ""}:
        qs = qs.filter(proyecto__nombre=proyecto)
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(producto__icontains=q) | Q(proyecto__nombre__icontains=q))
    qs = qs.order_by("-creado_en", "-id")

    orders = list(qs[:2000])
    ids = [int(o.id) for o in orders if getattr(o, "id", None)]
    kg_map = {}
    if ids:
        kg_rows = (
            RobotOrdenItem.objects.filter(orden_id__in=ids)
            .values("orden_id")
            .annotate(
                kg=Sum(
                    ExpressionWrapper(
                        F("cantidad_requerida")
                        * Coalesce(
                            F("pieza__peso_kg"),
                            F("pieza_custom_peso_kg"),
                            0.0,
                        ),
                        output_field=FloatField(),
                    )
                )
            )
        )
        kg_map = {int(r["orden_id"]): float(r["kg"] or 0.0) for r in kg_rows}

    for o in orders:
        o.folio_show = o.folio
        o.proyecto_nombre = getattr(getattr(o, "proyecto", None), "nombre", "") or ""
        o.nombre_show = (o.nombre or o.producto or "").strip()
        kg = float(kg_map.get(int(o.id), 0.0) or 0.0)
        o.ton = round(kg / 1000.0, 3)

    prod_qs = RobotProduccion.objects.select_related(
        "robot",
        "operador",
        "pieza",
        "orden_item",
        "orden_item__pieza",
        "orden_item__orden",
        "orden_item__orden__proyecto",
    )
    if proyecto and proyecto not in {"Todos", ""}:
        prod_qs = prod_qs.filter(orden_item__orden__proyecto__nombre=proyecto)
    if q:
        prod_qs = prod_qs.filter(
            Q(pieza__nombre__icontains=q)
            | Q(pieza_custom_nombre__icontains=q)
            | Q(orden_item__pieza_custom_nombre__icontains=q)
            | Q(orden_item__pieza__nombre__icontains=q)
            | Q(robot__nombre__icontains=q)
            | Q(operador__nombre__icontains=q)
            | Q(orden_item__orden__nombre__icontains=q)
            | Q(orden_item__orden__producto__icontains=q)
            | Q(orden_item__orden__proyecto__nombre__icontains=q)
        )
    producciones = list(prod_qs.order_by("-fecha", "-id")[:500])
    for p in producciones:
        try:
            p.op_folio = p.orden_item.orden.folio if p.orden_item_id and getattr(p.orden_item, "orden_id", None) else ""
        except Exception:
            p.op_folio = ""
        try:
            p.proyecto_nombre = (
                p.orden_item.orden.proyecto.nombre
                if p.orden_item_id and getattr(p.orden_item, "orden_id", None) and getattr(p.orden_item.orden, "proyecto_id", None)
                else ""
            )
        except Exception:
            p.proyecto_nombre = ""
        try:
            if p.orden_item_id:
                p.pieza_nombre_show = (p.orden_item.pieza_nombre or "").strip()
            elif p.pieza_id:
                p.pieza_nombre_show = (p.pieza.nombre or "").strip()
            else:
                p.pieza_nombre_show = (p.pieza_custom_nombre or "").strip()
        except Exception:
            p.pieza_nombre_show = ""
        p.kg = round(float(getattr(p, "total_kg", 0.0) or 0.0), 3)
        p.ton = round(p.kg / 1000.0, 3)

    projects = list(Proyecto.objects.filter(activo=True).values_list("nombre", flat=True).order_by("nombre"))
    estados_doc = ["Todos", "Abierta", "Cerrada", "Cancelada"]
    return render(
        request,
        "produccion/solo_lectura_robotica.html",
        {
            "orders": orders,
            "producciones": producciones,
            "filters": filters,
            "proyectos": ["Todos", *projects],
            "estados_doc": estados_doc,
            "reset_url": reverse("produccion:solo_lectura_robotica"),
            "solo_lectura_active": "robotica",
        },
    )


@login_required
def viga_create(request):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    _sync_projects()
    next_url = _safe_next(request.GET.get("next", ""))
    corte_equipo = _equipo_for_etapa("Corte")
    corte_operadores = []
    if corte_equipo:
        corte_operadores = list(
            Colaborador.objects.filter(activo=True, equipo=corte_equipo).order_by("rol", "nombre").values("id", "nombre", "rol")
        )
    corte_maquinas = list(
        Maquina.objects.filter(activo=True, tipo="Corte", es_robot=False).order_by("nombre").values("id", "nombre")
    )
    selected_corte_ops = []
    selected_corte_maqs = []
    if request.method == "POST":
        form = VigaBatchCreateForm(request.POST)
        selected_corte_ops = [int(x) for x in request.POST.getlist("corte_operador_ids") if str(x).isdigit()]
        selected_corte_maqs = [int(x) for x in request.POST.getlist("corte_maquina_ids") if str(x).isdigit()]
        ok, err, selected_corte_ops, selected_corte_maqs = _validate_corte_asignacion(selected_corte_ops, selected_corte_maqs)
        if not ok:
            form.add_error(None, err)
        if form.is_valid():
            now = timezone.now()
            cantidad = int(form.cleaned_data["cantidad_piezas"] or 1)
            base = form.save(commit=False)
            plano_file = request.FILES.get("plano_pdf")
            created_ids = []
            actor = ""
            try:
                if getattr(request, "user", None) and request.user.is_authenticated:
                    actor = request.user.get_username() or ""
            except Exception:
                actor = ""
            try:
                with transaction.atomic(using="mes"):
                    base_codigo = (base.codigo_viga or "").strip()
                    for i in range(1, cantidad + 1):
                        v = Viga.objects.create(
                            codigo_viga=base_codigo,
                            pieza_no=i,
                            total_piezas=cantidad,
                            proyecto=base.proyecto,
                            descripcion=base.descripcion,
                            fecha_compromiso=base.fecha_compromiso,
                            estado=base.estado,
                            observaciones=base.observaciones,
                            prioridad=base.prioridad,
                            peso_kg=base.peso_kg,
                            fecha_creacion=now,
                            ultimo_cambio=now,
                        )
                        created_ids.append(v.internal_id)
            except Exception as e:
                form.add_error(None, f"No se pudieron crear las piezas. ({type(e).__name__})")
                return render(
                    request,
                    "produccion/viga_form.html",
                    {
                        "form": form,
                        "title": "Nueva pieza",
                        "next_url": next_url,
                        "plano_url": "",
                        "corte_operadores": corte_operadores,
                        "corte_maquinas": corte_maquinas,
                        "selected_corte_operadores": selected_corte_ops,
                        "selected_corte_maquinas": selected_corte_maqs,
                    },
                )

            try:
                if selected_corte_ops or selected_corte_maqs:
                    with transaction.atomic(using="mes"):
                        _save_corte_asignaciones_for_vigas(created_ids, selected_corte_ops, selected_corte_maqs, actor=actor)
            except Exception as e:
                if created_ids:
                    with transaction.atomic(using="mes"):
                        _delete_asignaciones_for_vigas(created_ids)
                        Viga.objects.filter(internal_id__in=created_ids).delete()
                form.add_error(None, f"No se pudieron guardar las asignaciones iniciales de Corte. ({type(e).__name__})")
                return render(
                    request,
                    "produccion/viga_form.html",
                    {
                        "form": form,
                        "title": "Nueva pieza",
                        "next_url": next_url,
                        "plano_url": "",
                        "corte_operadores": corte_operadores,
                        "corte_maquinas": corte_maquinas,
                        "selected_corte_operadores": selected_corte_ops,
                        "selected_corte_maquinas": selected_corte_maqs,
                    },
                )

            try:
                if plano_file and created_ids:
                    with transaction.atomic(using="mes"):
                        _save_plano_for_vigas(created_ids, plano_file)
            except Exception as e:
                if created_ids:
                    with transaction.atomic(using="mes"):
                        _delete_asignaciones_for_vigas(created_ids)
                        Viga.objects.filter(internal_id__in=created_ids).delete()
                form.add_error(None, f"No se pudo guardar el plano PDF. ({type(e).__name__})")
                return render(
                    request,
                    "produccion/viga_form.html",
                    {
                        "form": form,
                        "title": "Nueva pieza",
                        "next_url": next_url,
                        "plano_url": "",
                        "corte_operadores": corte_operadores,
                        "corte_maquinas": corte_maquinas,
                        "selected_corte_operadores": selected_corte_ops,
                        "selected_corte_maquinas": selected_corte_maqs,
                    },
                )
            next_post = _safe_next(request.POST.get("next", ""))
            if next_post:
                return redirect(next_post)
            return redirect(f"{reverse('produccion:viga_list')}?q={base.codigo_viga}")
    else:
        form = VigaBatchCreateForm(initial={"estado": ESTADOS[0], "cantidad_piezas": 1})

    return render(
        request,
        "produccion/viga_form.html",
        {
            "form": form,
            "title": "Nueva pieza",
            "next_url": next_url,
            "plano_url": "",
            "corte_operadores": corte_operadores,
            "corte_maquinas": corte_maquinas,
            "selected_corte_operadores": selected_corte_ops,
            "selected_corte_maquinas": selected_corte_maqs,
        },
    )


@login_required
def viga_update(request, pk: int):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    _sync_projects()
    viga = get_object_or_404(Viga, pk=pk)
    next_url = _safe_next(request.GET.get("next", ""))
    if request.method == "POST":
        form = VigaForm(request.POST, instance=viga)
        if form.is_valid():
            viga = form.save(commit=False)
            viga.ultimo_cambio = timezone.now()
            viga.save()
            plano_file = request.FILES.get("plano_pdf")
            if plano_file:
                _save_plano_for_vigas([viga.internal_id], plano_file)
            next_post = _safe_next(request.POST.get("next", ""))
            if next_post:
                return redirect(next_post)
            return redirect("produccion:viga_list")
    else:
        form = VigaForm(instance=viga)
    plano = VigaPlano.objects.filter(viga_internal_id=viga.internal_id).first()
    plano_url = plano.archivo_pdf.url if (plano and plano.archivo_pdf) else ""
    return render(
        request,
        "produccion/viga_form.html",
        {"form": form, "title": "Editar pieza", "next_url": next_url, "plano_url": plano_url},
    )


@login_required
def viga_delete(request, pk: int):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    viga = get_object_or_404(Viga, pk=pk)
    next_url = _safe_next(request.GET.get("next", ""))
    if request.method == "POST":
        internal_id = viga.internal_id
        with transaction.atomic(using="mes"):
            ProductionLog.objects.filter(viga_internal_id=viga.internal_id).delete()
            viga.delete()
        _delete_plano_for_vigas([internal_id])
        _delete_asignaciones_for_vigas([internal_id])
        next_post = _safe_next(request.POST.get("next", ""))
        if next_post:
            return redirect(next_post)
        return redirect("produccion:viga_list")
    return render(
        request,
        "produccion/viga_confirm_delete.html",
        {"viga": viga, "next_url": next_url},
    )


@login_required
def viga_enviar(request, pk: int):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    if request.method != "POST":
        return redirect("produccion:viga_list")

    viga = get_object_or_404(Viga, pk=pk)
    if viga.estado != "Terminado":
        return redirect("produccion:viga_list")

    now = timezone.now()
    with transaction.atomic(using="mes"):
        estado_anterior = viga.estado
        viga.estado = "Enviado"
        viga.ultimo_cambio = now
        viga.save(update_fields=["estado", "ultimo_cambio"])
        ProductionLog.objects.create(
            viga_internal=viga,
            fecha_operacion=timezone.localdate(),
            estado_anterior=estado_anterior,
            estado_nuevo="Enviado",
            comentario="",
            timestamp=now,
        )

    next_post = _safe_next(request.POST.get("next", ""))
    if next_post:
        return redirect(next_post)
    return redirect("produccion:viga_list")


@login_required
def viga_regresar_produccion(request, pk: int):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    if request.method != "POST":
        return redirect("produccion:viga_list")

    viga = get_object_or_404(Viga, pk=pk)
    if viga.estado != "Enviado":
        return redirect("produccion:viga_list")

    now = timezone.now()
    with transaction.atomic(using="mes"):
        estado_anterior = viga.estado
        viga.estado = "Terminado"
        viga.ultimo_cambio = now
        viga.save(update_fields=["estado", "ultimo_cambio"])
        ProductionLog.objects.create(
            viga_internal=viga,
            fecha_operacion=timezone.localdate(),
            estado_anterior=estado_anterior,
            estado_nuevo="Terminado",
            comentario="",
            timestamp=now,
        )

    next_post = _safe_next(request.POST.get("next", ""))
    if next_post:
        return redirect(next_post)
    return redirect("produccion:viga_list")


@login_required
def viga_delete_decote(request, pk: int):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    if request.method != "POST":
        return redirect("produccion:viga_list")

    viga = get_object_or_404(Viga, pk=pk)
    if viga.estado != "Enviado":
        return redirect("produccion:viga_list")

    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    last_term = (
        ProductionLog.objects.filter(viga_internal_id=viga.internal_id, estado_nuevo="Enviado")
        .order_by("-fecha_operacion", "-timestamp")
        .values_list("fecha_operacion", flat=True)
        .first()
    )
    if not last_term or last_term > cutoff:
        return redirect("produccion:viga_list")

    internal_id = viga.internal_id
    with transaction.atomic(using="mes"):
        ProductionLog.objects.filter(viga_internal_id=viga.internal_id).delete()
        viga.delete()
    _delete_plano_for_vigas([internal_id])
    _delete_asignaciones_for_vigas([internal_id])

    next_post = _safe_next(request.POST.get("next", ""))
    if next_post:
        return redirect(next_post)
    return redirect("produccion:viga_list")


@login_required
def viga_delete_decote_all(request):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    if request.method != "POST":
        return redirect("produccion:viga_list")

    today = timezone.localdate()
    cutoff = today - timedelta(days=DECOTE_DAYS)
    filters = {
        "proyecto": request.POST.get("proyecto", "").strip(),
        "q": request.POST.get("q", "").strip(),
    }

    terminado_fecha_sq = Subquery(
        ProductionLog.objects.filter(viga_internal_id=OuterRef("pk"), estado_nuevo="Enviado")
        .order_by("-fecha_operacion", "-timestamp")
        .values("fecha_operacion")[:1]
    )

    decote_qs = Viga.objects.filter(estado="Enviado").annotate(terminado_fecha=terminado_fecha_sq)

    active_projects = list(
        Proyecto.objects.filter(activo=True).values_list("nombre_normalizado", flat=True)
    )
    if active_projects:
        decote_qs = decote_qs.annotate(proyecto_norm=Upper("proyecto")).filter(
            proyecto_norm__in=active_projects
        )

    if filters["proyecto"] and filters["proyecto"] != "Todos":
        decote_qs = decote_qs.filter(proyecto__iexact=filters["proyecto"])

    if filters["q"]:
        q = filters["q"]
        decote_qs = decote_qs.filter(
            Q(codigo_viga__icontains=q)
            | Q(proyecto__icontains=q)
            | Q(descripcion__icontains=q)
        )

    decote_qs = decote_qs.filter(terminado_fecha__isnull=False, terminado_fecha__lte=cutoff)
    ids = list(decote_qs.values_list("internal_id", flat=True)[:5000])
    if not ids:
        next_post = _safe_next(request.POST.get("next", ""))
        if next_post:
            return redirect(next_post)
        return redirect("produccion:viga_list")

    with transaction.atomic(using="mes"):
        ProductionLog.objects.filter(viga_internal_id__in=ids).delete()
        Viga.objects.filter(internal_id__in=ids).delete()
    _delete_plano_for_vigas(ids)
    _delete_asignaciones_for_vigas(ids)

    next_post = _safe_next(request.POST.get("next", ""))
    if next_post:
        return redirect(next_post)
    return redirect("produccion:viga_list")


def _guess_mapping(headers):
    keys = [h.strip() for h in headers]
    lo = [h.lower() for h in keys]
    def find(candidates):
        for name in candidates:
            for i, h in enumerate(lo):
                if h == name or name in h:
                    return keys[i]
        return keys[0] if keys else ""
    return {
        "codigo": find(["marca", "codigo_viga", "código", "codigo", "code", "viga"]),
        "proyecto": find(["proyecto", "proy"]),
        "descripcion": find(["descripcion", "descripción", "desc", "perfil", "titulo"]),
        "peso": find(["kg/pza", "kg_pza", "kg por pieza", "peso_kg", "peso", "kg"]),
        "cantidad": find(["cantidad", "numero de piezas", "número de piezas", "piezas", "total"]),
        "estado": "",
        "prioridad": "",
        "fecha": "",
    }


def _normalize_row(row, map_, today):
    def get(name):
        col = map_.get(name) or ""
        val = row.get(col, "")
        if val is None:
            return ""
        return str(val).strip()
    codigo = get("codigo")
    proyecto = get("proyecto").upper()
    descripcion = get("descripcion")
    try:
        peso = float(get("peso").replace(",", "."))
    except Exception:
        peso = 0.0
    try:
        cantidad = int(float(get("cantidad") or "1"))
    except Exception:
        cantidad = 1
    estado = get("estado")
    if estado and estado not in ESTADOS:
        estado = ""
    prioridad = get("prioridad")
    fecha_txt = get("fecha")
    fecha = today
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            fecha = datetime.strptime(fecha_txt, fmt).date()
            break
        except Exception:
            logger.exception("Error ignorado en _normalize_row()")
    if not proyecto:
        proyecto = "SIN PROYECTO"
    return {
        "codigo": codigo,
        "proyecto": proyecto,
        "descripcion": descripcion,
        "peso": peso,
        "cantidad": max(cantidad, 1),
        "estado": estado or ESTADOS[0],
        "prioridad": prioridad or "",
        "fecha": fecha,
    }

def _find_top_marca(lines):
    joined = " ".join([ln.strip() for ln in lines[:40] if (ln or "").strip()]).upper()
    joined = re.sub(r"\s+", " ", joined)
    pat = re.compile(r"\b\d{1,2}[A-Z]{1,6}\d{1,4}-[0-9]{1,3}[A-Z]?\b")
    m = pat.search(joined)
    if m:
        return m.group(0).strip()
    pat2 = re.compile(r"\b[A-Z]{1,6}\d{1,4}-[0-9]{1,3}[A-Z]?\b")
    m2 = pat2.search(joined)
    if m2:
        return m2.group(0).strip()
    return ""

def _to_canonical_rows(rows, headers):
    today = timezone.localdate()
    map_ = _guess_mapping(headers)
    out = []
    for r in rows:
        d = _normalize_row(r, map_, today)
        out.append(
            {
                "codigo": d["codigo"],
                "proyecto": d["proyecto"],
                "descripcion": d["descripcion"],
                "peso": d["peso"],
                "cantidad": d["cantidad"],
                "estado": d["estado"],
                "prioridad": d["prioridad"] or "1",
                "observaciones": "",
                "fecha": d["fecha"].isoformat(),
            }
        )
    return out


def _pdf_default_config():
    return {
        "big_mark_regex": r"\b\d{1,2}[A-Z]{1,6}\d{1,4}-\d{1,3}[A-Z]?\b",
        "mark_regex": r"\b[A-Z]{1,6}\d{1,4}-\d{1,3}[A-Z]?\b",
        "profile_regex": r"\b(IPR|IPE|IPN|HEB|HEA|W)\b\s*([0-9A-ZxX.,/]+(?:KG/ML)?)",
        "total_regex": r"\bTOTAL\b.*?(\d+(?:[.,]\d+)?)\b",
        "num_piezas_regex": r"\b(\d{1,4})\b",
        "prefer_total": True,
    }


def _extract_pdf_lines(content: bytes):
    if PdfReader is None:
        return []
    reader = PdfReader(io.BytesIO(content))
    lines = []
    for page in reader.pages[:10]:
        text = page.extract_text() or ""
        text = text.replace("\u00a0", " ")
        for ln in text.splitlines():
            ln = ln.strip()
            if ln:
                lines.append(ln)
    return lines


def _guess_proyecto_from_lines(lines):
    up_all = re.sub(r"\s+", " ", "\n".join(lines)).upper()
    if "ALMAERA" in up_all:
        return "ALMAERA"
    m = re.search(r"\bPROYECTO\b\s*[:\-]\s*([A-Z0-9 \-_/]{3,})", up_all)
    if m:
        return m.group(1).strip().upper()
    return ""


def _extract_rows_from_lines(lines, config):
    cfg = _pdf_default_config()
    cfg.update({k: v for k, v in (config or {}).items() if v is not None})

    up_all = re.sub(r"\s+", " ", "\n".join(lines)).upper()
    proyecto = _guess_proyecto_from_lines(lines) or "SIN PROYECTO"

    big_mark_re = re.compile(cfg["big_mark_regex"])
    mark_re = re.compile(cfg["mark_regex"])
    profile_re = re.compile(cfg["profile_regex"])
    total_re = re.compile(cfg["total_regex"], flags=re.IGNORECASE)
    num_re = re.compile(cfg["num_piezas_regex"])

    current_big = ""
    for ln in lines[:60]:
        m = big_mark_re.search(ln.upper())
        if m:
            current_big = m.group(0).strip()
            break

    total_doc = None
    for ln in lines:
        if "TOTAL" in ln.upper():
            m = total_re.search(ln)
            if m:
                try:
                    total_doc = float(m.group(1).replace(",", "."))
                    break
                except Exception:
                    logger.exception("Error ignorado en _extract_rows_from_lines()")

    num_piezas = None
    for ln in lines:
        up = ln.upper()
        if ("NUMERO DE PIEZAS" in up) or ("NÚMERO DE PIEZAS" in up):
            m = num_re.search(up)
            if m:
                try:
                    num_piezas = int(m.group(1))
                    break
                except Exception:
                    logger.exception("Error ignorado en _extract_rows_from_lines()")

    keywords = ("IPR", "IPE", "IPN", "HEB", "HEA", "W ")
    seen = set()
    out = []
    for ln in lines:
        up = ln.upper()
        if not any(k in up for k in keywords):
            continue
        prof_m = profile_re.search(up)
        if not prof_m:
            continue
        mm = mark_re.search(up)
        if not mm:
            continue

        marca = mm.group(0).strip()
        if marca in seen:
            continue
        seen.add(marca)

        perfil = f"{prof_m.group(1)} {prof_m.group(2)}".strip()
        nums = re.findall(r"\d+(?:[.,]\d+)?", up)
        kg_pza = None
        kg_total = None
        if len(nums) >= 2:
            try:
                kg_pza = float(nums[-2].replace(",", "."))
                kg_total = float(nums[-1].replace(",", "."))
            except Exception:
                kg_pza = None
                kg_total = None

        peso_out = None
        if cfg.get("prefer_total") and total_doc:
            peso_out = total_doc
        elif kg_total:
            peso_out = kg_total
        elif kg_pza:
            peso_out = kg_pza
        else:
            peso_out = 0.0

        out.append(
            {
                "codigo": current_big or marca,
                "proyecto": proyecto,
                "descripcion": perfil,
                "peso": round(float(peso_out or 0.0), 2),
                "cantidad": int(num_piezas or 1),
                "estado": ESTADOS[0],
                "prioridad": "1",
                "observaciones": "",
                "fecha": timezone.localdate().isoformat(),
            }
        )

    return out


def _extract_rows_from_pdf(content: bytes):
    lines = _extract_pdf_lines(content)
    text_all = "\n".join(lines)
    up_all = re.sub(r"\s+", " ", "\n".join(lines)).upper()

    proyecto = ""
    m = re.search(r"\bPROYECTO\b\s*[:\-]\s*([A-Z0-9 \-_/]{3,})", up_all)
    if m:
        proyecto = m.group(1).strip()
    if "ALMAERA" in up_all:
        proyecto = "ALMAERA"
    if not proyecto:
        proyecto = "SIN PROYECTO"

    descripcion = ""
    m = re.search(r"\bTITULO\s+DEL\s+PLANO\b\s*[:\-]\s*(.+)", text_all, flags=re.IGNORECASE)
    if m:
        descripcion = m.group(1).strip()
    if not descripcion:
        m = re.search(r"\bVIGA\s+PRINCIPAL\b", up_all)
        if m:
            descripcion = "VIGA PRINCIPAL"

    keywords = ("IPR", "IPE", "IPN", "HEB", "HEA", "W ")
    big_mark_re = re.compile(r"\b\d{1,2}[A-Z]{1,6}\d{1,4}-\d{1,3}[A-Z]?\b")
    mark_re = re.compile(r"\b[A-Z]{1,6}\d{1,4}-\d{1,3}[A-Z]?\b")
    profile_re = re.compile(r"\b(IPR|IPE|IPN|HEB|HEA|W)\b\s*([0-9A-ZxX.,/]+(?:KG/ML)?)")

    seen = set()
    out_rows = []
    current_big = ""
    current_num_piezas = None
    current_total = None

    for ln in lines:
        up = ln.upper()

        bm = big_mark_re.search(up)
        if bm:
            current_big = bm.group(0).strip()

        if ("NUMERO DE PIEZAS" in up) or ("NÚMERO DE PIEZAS" in up):
            ints = re.findall(r"\b\d{1,4}\b", up)
            if ints:
                try:
                    current_num_piezas = int(ints[-1])
                except Exception:
                    logger.exception("Error ignorado en _extract_rows_from_pdf()")

        sm = mark_re.search(up)
        if sm and current_num_piezas is None and "EMBARQUE" in up_all:
            ints = re.findall(r"\b\d{1,4}\b", up)
            if ints:
                try:
                    current_num_piezas = int(ints[0])
                except Exception:
                    logger.exception("Error ignorado en _extract_rows_from_pdf()")

        if "TOTAL" in up:
            nums = re.findall(r"\d+(?:[.,]\d+)?", up)
            if nums:
                try:
                    current_total = float(nums[-1].replace(",", "."))
                except Exception:
                    logger.exception("Error ignorado en _extract_rows_from_pdf()")

        if "LISTA DE MATERIALES" not in up_all and not any(k in up for k in keywords):
            continue

        prof_m = profile_re.search(up)
        if not prof_m:
            continue

        mm = mark_re.search(up)
        if not mm:
            continue
        marca = mm.group(0).strip()

        nums = re.findall(r"\d+(?:[.,]\d+)?", up)
        kg_pza = None
        kg_total = None
        if len(nums) >= 2:
            try:
                kg_pza = float(nums[-2].replace(",", "."))
                kg_total = float(nums[-1].replace(",", "."))
            except Exception:
                kg_pza = None
                kg_total = None

        if marca in seen:
            continue
        seen.add(marca)

        cantidad = int(current_num_piezas or 1)
        if not current_total and kg_total:
            current_total = kg_total

        perfil = f"{prof_m.group(1)} {prof_m.group(2)}".strip()
        codigo_out = current_big or marca
        peso_out = float(current_total or 0.0)
        if peso_out <= 0 and kg_total:
            peso_out = float(kg_total)
        if peso_out <= 0 and kg_pza:
            peso_out = float(kg_pza)

        out_rows.append(
            {
                "codigo": codigo_out,
                "proyecto": proyecto,
                "descripcion": perfil,
                "peso": round(peso_out, 2),
                "cantidad": cantidad,
                "estado": ESTADOS[0],
                "prioridad": "1",
                "observaciones": "",
                "fecha": timezone.localdate().isoformat(),
            }
        )

    if out_rows:
        return out_rows, []

    codigo = ""
    m = big_mark_re.search(up_all) or mark_re.search(up_all)
    if m:
        codigo = m.group(0).strip()
    peso = ""
    m = re.search(r"\bPESO\b\s*[:\-]?\s*([0-9]+(?:[.,][0-9]+)?)\s*KG", up_all)
    if m:
        peso = m.group(1)
    cantidad = ""
    m = re.search(r"(?:PZAS|PIEZAS|CANTIDAD)\s*[:\-]?\s*([0-9]{1,4})", up_all)
    if m:
        cantidad = m.group(1)
    rows = []
    if codigo or proyecto or descripcion or peso or cantidad:
        rows = [
            {
                "codigo": codigo,
                "proyecto": proyecto,
                "descripcion": descripcion,
                "peso": float(str(peso).replace(",", ".") or 0),
                "cantidad": int(cantidad or "1"),
                "estado": ESTADOS[0],
                "prioridad": "1",
                "observaciones": "",
                "fecha": timezone.localdate().isoformat(),
            }
        ]
    return rows, []


@login_required
def viga_import(request):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    pdf_notice = None
    if request.method == "POST":
        stage = request.POST.get("stage", "upload")
        if stage == "upload":
            form = VigaImportUploadForm(request.POST, request.FILES)
            if form.is_valid():
                f = form.cleaned_data["archivo"]
                name = f.name.lower()
                content = f.read()
                rows = []
                headers = []
                pdf_lines = []
                template_cfg = {}
                template_id = ""
                guessed_proyecto = ""
                if name.endswith(".csv"):
                    text = content.decode("utf-8", errors="ignore")
                    reader = csv.DictReader(io.StringIO(text))
                    headers = reader.fieldnames or []
                    for r in reader:
                        rows.append({k: r.get(k, "") for k in headers})
                elif name.endswith(".xlsx"):
                    bio = io.BytesIO(content)
                    wb = load_workbook(bio, read_only=True, data_only=True)
                    ws = wb.active
                    it = ws.iter_rows(values_only=True)
                    headers = [str(x or "").strip() for x in next(it, [])]
                    for row in it:
                        d = {}
                        for i, h in enumerate(headers):
                            d[h] = row[i] if i < len(row) else ""
                        rows.append(d)
                elif name.endswith(".pdf"):
                    try:
                        pdf_lines = _extract_pdf_lines(content)
                        guessed_proyecto = _guess_proyecto_from_lines(pdf_lines)
                        tmpl = (
                            PdfExtractionTemplate.objects.filter(activo=True, proyecto_normalizado=guessed_proyecto)
                            .order_by("-actualizado_en")
                            .first()
                            if guessed_proyecto
                            else PdfExtractionTemplate.objects.filter(activo=True, proyecto_normalizado="")
                            .order_by("-actualizado_en")
                            .first()
                        )
                        if tmpl:
                            template_id = str(tmpl.pk)
                            try:
                                template_cfg = json.loads(tmpl.config_json or "{}") if tmpl.config_json else {}
                            except Exception:
                                template_cfg = {}
                        rows = _extract_rows_from_lines(pdf_lines, template_cfg)
                        if not rows:
                            pdf_notice = "No se detectó una tabla o datos suficientes en el PDF. Usa “Ver texto extraído” y ajusta la plantilla."
                    except Exception:
                        pdf_notice = "No se pudo leer el PDF. Intenta exportar la tabla a Excel/CSV y súbela aquí."
                else:
                    pdf_notice = "Formato no reconocido. Usa PDF, CSV o XLSX."
                if name.endswith(".pdf") and pdf_lines:
                    return render(
                        request,
                        "produccion/viga_import.html",
                        {
                            "form": VigaImportUploadForm(),
                            "rows": rows,
                            "estados": ESTADOS,
                            "pdf_lines_json": json.dumps(pdf_lines) if pdf_lines else "",
                            "pdf_text": "\n".join(pdf_lines) if pdf_lines else "",
                            "template_id": template_id,
                            "template_cfg": {**_pdf_default_config(), **(template_cfg or {})},
                            "template_proyecto": guessed_proyecto,
                            "pdf_notice": pdf_notice,
                        },
                    )
                if rows:
                    headers = [str(h) for h in headers]
                    canon_rows = _to_canonical_rows(rows, headers)
                    return render(
                        request,
                        "produccion/viga_import.html",
                        {
                            "form": VigaImportUploadForm(),
                            "rows": canon_rows,
                            "estados": ESTADOS,
                            "pdf_notice": pdf_notice,
                        },
                    )
            return render(
                request,
                "produccion/viga_import.html",
                {"form": form, "pdf_notice": pdf_notice},
            )
        elif stage in {"reparse", "save_template"}:
            pdf_lines = json.loads(request.POST.get("pdf_lines_json") or "[]")
            template_id = (request.POST.get("template_id") or "").strip()
            template_proyecto = (request.POST.get("template_proyecto") or "").strip().upper()
            template_cfg = {
                "big_mark_regex": request.POST.get("big_mark_regex") or _pdf_default_config()["big_mark_regex"],
                "mark_regex": request.POST.get("mark_regex") or _pdf_default_config()["mark_regex"],
                "profile_regex": request.POST.get("profile_regex") or _pdf_default_config()["profile_regex"],
                "total_regex": request.POST.get("total_regex") or _pdf_default_config()["total_regex"],
                "num_piezas_regex": request.POST.get("num_piezas_regex") or _pdf_default_config()["num_piezas_regex"],
                "prefer_total": request.POST.get("prefer_total") == "1",
            }
            if stage == "save_template":
                nombre = (request.POST.get("template_nombre") or "").strip() or f"Plantilla {template_proyecto or 'PDF'}"
                if template_id:
                    PdfExtractionTemplate.objects.filter(pk=template_id).update(
                        nombre=nombre,
                        proyecto_normalizado=template_proyecto,
                        config_json=json.dumps(template_cfg),
                        activo=True,
                    )
                else:
                    t = PdfExtractionTemplate.objects.create(
                        nombre=nombre,
                        proyecto_normalizado=template_proyecto,
                        config_json=json.dumps(template_cfg),
                        activo=True,
                    )
                    template_id = str(t.pk)
                pdf_notice = "Plantilla guardada."
            rows = _extract_rows_from_lines(pdf_lines, template_cfg)
            return render(
                request,
                "produccion/viga_import.html",
                {
                    "form": VigaImportUploadForm(),
                    "rows": rows,
                    "estados": ESTADOS,
                    "pdf_lines_json": json.dumps(pdf_lines) if pdf_lines else "",
                    "pdf_text": "\n".join(pdf_lines) if pdf_lines else "",
                    "template_id": template_id,
                    "template_cfg": {**_pdf_default_config(), **(template_cfg or {})},
                    "template_proyecto": template_proyecto,
                    "pdf_notice": pdf_notice,
                },
            )
        elif stage == "commit":
            total_rows = int(request.POST.get("total_rows") or "0")
            created = 0
            first_code = ""
            observaciones_global = (request.POST.get("observaciones_global") or "").strip()
            with transaction.atomic(using="mes"):
                now = timezone.now()
                for i in range(total_rows):
                    codigo = (request.POST.get(f"codigo_{i}") or "").strip()
                    if not codigo:
                        continue
                    proyecto = (request.POST.get(f"proyecto_{i}") or "SIN PROYECTO").strip().upper()
                    descripcion = (request.POST.get(f"descripcion_{i}") or "").strip()
                    try:
                        peso = float((request.POST.get(f"peso_{i}") or "0").replace(",", "."))
                    except Exception:
                        peso = 0.0
                    try:
                        cantidad = int(request.POST.get(f"cantidad_{i}") or "1")
                    except Exception:
                        cantidad = 1
                    cantidad = max(cantidad, 1)
                    estado = (request.POST.get(f"estado_{i}") or ESTADOS[0]).strip()
                    if estado not in ESTADOS:
                        estado = ESTADOS[0]
                    try:
                        prioridad = int(request.POST.get(f"prioridad_{i}") or "1")
                    except Exception:
                        prioridad = 1
                    prioridad = max(prioridad, 1)
                    observaciones = (request.POST.get(f"observaciones_{i}") or "").strip() or observaciones_global
                    fecha_txt = (request.POST.get(f"fecha_{i}") or "").strip()
                    fecha = timezone.localdate()
                    if fecha_txt:
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                            try:
                                fecha = datetime.strptime(fecha_txt, fmt).date()
                                break
                            except Exception:
                                logger.exception("Error ignorado en viga_import()")

                    if not first_code:
                        first_code = codigo

                    Proyecto.objects.get_or_create(
                        nombre_normalizado=proyecto.upper(),
                        defaults={"nombre": proyecto.upper(), "activo": True},
                    )

                    for n_pieza in range(1, cantidad + 1):
                        Viga.objects.create(
                            codigo_viga=codigo,
                            pieza_no=n_pieza,
                            total_piezas=cantidad,
                            proyecto=proyecto,
                            descripcion=descripcion,
                            fecha_compromiso=fecha,
                            estado=estado,
                            observaciones=observaciones,
                            prioridad=prioridad,
                            peso_kg=peso,
                            fecha_creacion=now,
                            ultimo_cambio=now,
                        )
                        created += 1
            if first_code:
                return redirect(f"{reverse('produccion:viga_list')}?q={first_code}")
            return redirect("produccion:viga_list")
    form = VigaImportUploadForm()
    return render(
        request,
        "produccion/viga_import.html",
        {"form": form, "pdf_notice": pdf_notice},
    )

@login_required
def viga_change_status(request, pk: int):
    viga = get_object_or_404(Viga, pk=pk)
    role = _user_role(request.user)
    if role != "admin":
        allowed = _etapas_permitidas(request.user)
        if not allowed or _norm_estado(viga.estado) not in allowed:
            return redirect("produccion:home")
    viga.estado = _norm_estado(getattr(viga, "estado", "") or "")
    viga.estado_color = STATUS_COLORS.get(viga.estado, "#8898aa")
    estados_index = {s: i for i, s in enumerate(ESTADOS)}
    idx = estados_index.get(viga.estado)
    if idx is None or idx >= len(ESTADOS) - 1:
        next_estado = ""
    else:
        next_estado = ESTADOS[idx + 1]
    if request.method == "POST":
        form = StatusChangeForm(request.POST)
        if form.is_valid():
            estado_nuevo = _norm_estado(form.cleaned_data["estado_nuevo"])
            if role != "admin" and estado_nuevo not in allowed:
                form.add_error(None, "Sin permiso para ese estado.")
                next_url = request.POST.get("next", "")
                if not next_url.startswith("/"):
                    next_url = ""
                return render(
                    request,
                    "produccion/status_form.html",
                    {
                        "form": form,
                        "viga": viga,
                        "next_url": next_url,
                        "today": timezone.localdate().isoformat(),
                        "estados": ESTADOS,
                        "status_colors": STATUS_COLORS,
                        "next_estado": next_estado,
                    },
                )
            fecha_operacion = form.cleaned_data["fecha_operacion"]
            comentario = form.cleaned_data["comentario"] or ""
            motivo_retroceso = (request.POST.get("motivo_retroceso") or "").strip()
            if estado_nuevo == viga.estado:
                next_url = request.POST.get("next", "")
                if next_url.startswith("/"):
                    return redirect(next_url)
                return redirect("produccion:viga_list" if role == "admin" else ("produccion:area_corte" if role == "corte" else "produccion:area_soldadura"))

            idx_actual = estados_index.get(viga.estado)
            idx_nuevo = estados_index.get(estado_nuevo)
            retroceso = idx_actual is not None and idx_nuevo is not None and idx_nuevo < idx_actual
            if retroceso:
                if motivo_retroceso not in {"error_dedo", "retrabajo"}:
                    form.add_error(None, "Debes seleccionar el motivo del retroceso.")
                    next_url = request.POST.get("next", "")
                    if not next_url.startswith("/"):
                        next_url = ""
                    return render(
                        request,
                        "produccion/status_form.html",
                        {
                            "form": form,
                            "viga": viga,
                            "next_url": next_url,
                            "today": timezone.localdate().isoformat(),
                            "estados": ESTADOS,
                            "status_colors": STATUS_COLORS,
                            "next_estado": next_estado,
                        },
                    )
                tag = "[MOTIVO=RETRABAJO]" if motivo_retroceso == "retrabajo" else "[MOTIVO=ERROR_DE_DEDO]"
                comentario = (f"{tag} {comentario}").strip()

            now = timezone.now()
            actor = ""
            try:
                if getattr(request, "user", None) and request.user.is_authenticated:
                    actor = request.user.get_username() or ""
            except Exception:
                actor = ""

            try:
                with transaction.atomic(using="mes"):
                    if estado_nuevo in ("Armado", "Pintura"):
                        ok, err = _save_asignaciones_for_etapa(viga.internal_id, estado_nuevo, request.POST, actor=actor)
                        if not ok:
                            raise ValueError(err)
                    estado_anterior = viga.estado
                    viga.estado = estado_nuevo
                    viga.ultimo_cambio = now
                    viga.save(update_fields=["estado", "ultimo_cambio"])
                    ProductionLog.objects.create(
                        viga_internal=viga,
                        fecha_operacion=fecha_operacion,
                        estado_anterior=estado_anterior,
                        estado_nuevo=estado_nuevo,
                        comentario=comentario,
                        timestamp=now,
                    )
            except ValueError as e:
                form.add_error(None, str(e))
                next_url = request.POST.get("next", "")
                if not next_url.startswith("/"):
                    next_url = ""
                return render(
                    request,
                    "produccion/status_form.html",
                    {
                        "form": form,
                        "viga": viga,
                        "next_url": next_url,
                        "today": timezone.localdate().isoformat(),
                        "estados": ESTADOS,
                        "status_colors": STATUS_COLORS,
                        "next_estado": next_estado,
                    },
                )
            next_url = request.POST.get("next", "")
            if next_url.startswith("/"):
                return redirect(next_url)
            return redirect("produccion:viga_list" if role == "admin" else ("produccion:area_corte" if role == "corte" else "produccion:area_soldadura"))
    else:
        form = StatusChangeForm(initial={"estado_nuevo": viga.estado})
    next_url = request.GET.get("next", "")
    if not next_url.startswith("/"):
        next_url = ""
    return render(
        request,
        "produccion/status_form.html",
        {
            "form": form,
            "viga": viga,
            "next_url": next_url,
            "today": timezone.localdate().isoformat(),
            "estados": ESTADOS,
            "status_colors": STATUS_COLORS,
            "next_estado": next_estado,
        },
    )


@login_required
@require_POST
def viga_avanzar_grupo(request):
    """«De estas cincuenta, hice treinta y dos.»

    En Estructuras una orden de cincuenta piezas son **cincuenta renglones**,
    uno por pieza, porque el avance se lleva por etapas y no por contadores.
    Para un soldador que hizo cuarenta, eso eran cuarenta toques con guantes
    en un teléfono: en la práctica se apuntaba en papel y alguien lo capturaba
    por la tarde, que es exactamente por lo que el sistema iba siempre por
    detrás del taller.

    Aquí se avanzan N piezas del mismo código y la misma etapa de una vez.

    **No se reimplementa nada.** Se llama al mismo endpoint de una pieza, una
    vez por pieza, así que se aplican sus mismas comprobaciones: permisos de
    etapa, ruta de la orden, equipo de corte obligatorio, máquina en paro,
    motivo de retroceso, apunte de trabajo y cuadrilla. Copiar ese bloque para
    la versión en lote habría sido la sexta copia de la misma lógica en este
    archivo, y la primera en separarse.
    """
    codigo = (request.POST.get("codigo") or "").strip()
    etapa_actual = _norm_estado(request.POST.get("estado_actual") or "")
    try:
        cuantas = int(request.POST.get("cantidad") or 0)
    except (TypeError, ValueError):
        cuantas = 0

    if not codigo or not etapa_actual or cuantas <= 0:
        return JsonResponse({"ok": False, "error": "Datos incompletos."}, status=400)

    escrituras = core_estados.variantes(etapa_actual)
    piezas = list(
        Viga.objects.filter(codigo_viga=codigo, estado__in=escrituras)
        .order_by("pieza_no", "internal_id")[:cuantas]
    )
    if not piezas:
        return JsonResponse(
            {"ok": False, "error": f"Ya no hay piezas de {codigo} en {etapa_actual}."},
            status=409,
        )

    hechas, error = 0, ""
    for pieza in piezas:
        respuesta = viga_change_status_json(request, pieza.internal_id)
        if respuesta.status_code == 200:
            hechas += 1
            continue
        # Se para en el primer fallo en vez de seguir: si falta el equipo de
        # corte o la máquina está en paro, va a fallar en las cincuenta, y
        # cincuenta errores iguales no dicen más que uno.
        try:
            error = json.loads(respuesta.content).get("error") or ""
        except Exception:
            error = "No se pudo registrar el avance."
        break

    if not hechas:
        return JsonResponse({"ok": False, "error": error}, status=400)

    quedan = Viga.objects.filter(codigo_viga=codigo, estado__in=escrituras).count()
    return JsonResponse(
        {
            "ok": True,
            "hechas": hechas,
            "quedan": quedan,
            "codigo": codigo,
            "aviso": error,
        }
    )


@login_required
def viga_change_status_json(request, pk: int):
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método inválido."}, status=405)

    viga = get_object_or_404(Viga, pk=pk)
    role = _user_role(request.user)
    if role != "admin":
        allowed = _etapas_permitidas(request.user)
    else:
        allowed = set(ESTADOS)
    form = StatusChangeForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"ok": False, "error": "Datos inválidos."}, status=400)

    estado_nuevo = form.cleaned_data["estado_nuevo"]
    cur_estado = _norm_estado(getattr(viga, "estado", "") or "")
    new_estado = _norm_estado(estado_nuevo)
    # Quien es dueño de una etapa puede **completarla**, lleve su ruta a donde
    # la lleve.
    #
    # Sin esto, una pieza que no pasa por pintura deja al soldador atascado: su
    # ruta manda de «Soldadura» a «Terminado», y «Terminado» es un permiso del
    # área de pintura. El botón saldría y el servidor lo rechazaría con un «Sin
    # permiso» que en el piso no explica nada, y nadie podría cerrar esa pieza
    # sin llamar a un administrador.
    #
    # Es un permiso concreto y no un agujero: vale sólo cuando la ruta de esa
    # pieza **se salta** algo, y sólo para el destino exacto al que la manda.
    #
    # La condición de que el destino sea distinto del normal es lo que lo
    # cierra. Sin ella se colaba lo contrario de lo que se quiere: una pieza en
    # «Espera de pintura» tiene como siguiente «Pintura» tanto en su ruta como
    # en la secuencia general, así que un soldador —que puede tocar la espera,
    # porque es el punto de entrega— habría podido meterse a pintar.
    siguiente_de_su_ruta = servicio_ruta.siguiente("Viga", viga.internal_id, cur_estado)
    posicion_normal = core_estados.posicion(cur_estado)
    siguiente_normal = (
        core_estados.SECUENCIA[posicion_normal + 1]
        if posicion_normal is not None
        and posicion_normal + 1 < len(core_estados.SECUENCIA)
        else ""
    )
    cierra_su_etapa = (
        cur_estado in allowed
        and new_estado == siguiente_de_su_ruta
        and new_estado != siguiente_normal
    )
    if role != "admin" and not (
        allowed and cur_estado in allowed and (new_estado in allowed or cierra_su_etapa)
    ):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    estado_nuevo = new_estado

    # El equipo con el que se hace la etapa. Puede venir en la petición —el
    # selector de la pantalla— o estar ya asignado a la pieza.
    maquina_apunte = None
    if servicio_trabajo.exige_maquina(new_estado):
        pedida = request.POST.get("maquina_id")
        if pedida:
            maquina_apunte, error_maquina = servicio_trabajo.maquina_valida(
                pedida, etapa=new_estado
            )
            if error_maquina:
                return JsonResponse({"ok": False, "error": error_maquina}, status=400)
        else:
            maquina_apunte = _maquina_asignada(viga.internal_id, new_estado)
        # Si el taller todavía no ha dado de alta ningún equipo de corte, no se
        # puede exigir elegir uno: se quedaría sin poder mover una sola pieza,
        # y el requisito de medición habría parado la producción.
        if maquina_apunte is None and not _equipos_de_corte():
            logger.warning(
                "no se exige equipo en %s: no hay ninguno dado de alta", new_estado
            )
        elif maquina_apunte is None:
            # Sin esto el avance se registra «en corte» a secas y la
            # producción de los seis equipos queda en un solo montón: no se
            # puede saber cuál va saturado ni cuál está parado.
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Elige en qué equipo de corte se va a trabajar.",
                    "falta": "maquina",
                },
                status=400,
            )

    if cur_estado == "Espera de corte" and new_estado == "Corte":
        mids = list(
            VigaAsignacion.objects.using("mes")
            .filter(viga_internal_id=int(viga.internal_id), vigente=True, etapa="Corte", maquina_id__isnull=False)
            .values_list("maquina_id", flat=True)
        )
        mids = [int(x) for x in mids if x]
        if maquina_apunte is not None and maquina_apunte.id not in mids:
            mids.append(int(maquina_apunte.id))
        if mids:
            if MaquinaFalla.objects.filter(maquina_id__in=mids, fin__isnull=True).exists():
                return JsonResponse(
                    {"ok": False, "error": "No se puede avanzar: la máquina asignada tiene falla activa. Revisa Paros.", "block": "falla"},
                    status=409,
                )
            if MaquinaParo.objects.filter(maquina_id__in=mids, fin__isnull=True).exists():
                return JsonResponse(
                    {"ok": False, "error": "No se puede avanzar: la máquina asignada está en paro. Reanuda la máquina.", "block": "paro"},
                    status=409,
                )
    fecha_operacion = form.cleaned_data["fecha_operacion"]
    comentario = form.cleaned_data["comentario"] or ""
    motivo_retroceso = (request.POST.get("motivo_retroceso") or "").strip()

    estados_index = {s: i for i, s in enumerate(ESTADOS)}
    if estado_nuevo == cur_estado:
        idx = estados_index.get(cur_estado)
        if role == "corte":
            if cur_estado == "Espera de corte":
                next_estado = "Corte"
            elif cur_estado == "Corte":
                next_estado = "Espera de armado"
            else:
                next_estado = ""
        else:
            next_estado = "" if idx is None or idx >= len(ESTADOS) - 1 else ESTADOS[idx + 1]
        next_label = "En área de soldadura" if (role == "corte" and cur_estado == "Espera de armado") else ""
        ultimo_mov = _format_utc_naive_dt_as_local(getattr(viga, "ultimo_cambio", None))
        return JsonResponse(
            {
                "ok": True,
                "id": int(viga.internal_id),
                "estado": cur_estado,
                "estado_color": STATUS_COLORS.get(cur_estado, "#8898aa"),
                "estado_clase": clase_de_estado(cur_estado),
                "next_estado": next_estado,
                "next_label": next_label,
                "ultimo_mov": ultimo_mov,
            }
        )

    idx_actual = estados_index.get(cur_estado)
    idx_nuevo = estados_index.get(estado_nuevo)
    retroceso = idx_actual is not None and idx_nuevo is not None and idx_nuevo < idx_actual
    if retroceso:
        if motivo_retroceso not in {"error_dedo", "retrabajo"}:
            return JsonResponse({"ok": False, "error": "Debes seleccionar el motivo del retroceso."}, status=400)
        tag = "[MOTIVO=RETRABAJO]" if motivo_retroceso == "retrabajo" else "[MOTIVO=ERROR_DE_DEDO]"
        comentario = (f"{tag} {comentario}").strip()

    now = timezone.now()
    actor = ""
    try:
        if getattr(request, "user", None) and request.user.is_authenticated:
            actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    try:
        with transaction.atomic(using="mes"):
            if estado_nuevo in ("Armado", "Pintura"):
                ok, err = _save_asignaciones_for_etapa(viga.internal_id, estado_nuevo, request.POST, actor=actor)
                if not ok:
                    return JsonResponse({"ok": False, "error": err}, status=400)
            estado_anterior = viga.estado
            viga.estado = estado_nuevo
            viga.ultimo_cambio = now
            viga.save(update_fields=["estado", "ultimo_cambio"])
            ProductionLog.objects.create(
                viga_internal=viga,
                fecha_operacion=fecha_operacion,
                estado_anterior=estado_anterior,
                estado_nuevo=estado_nuevo,
                comentario=comentario,
                timestamp=now,
            )
            # En la misma transacción: un apunte de trabajo sin su cambio de
            # estado mediría trabajo que no ocurrió.
            servicio_trabajo.anotar(
                linea=SeguimientoDespacho.Linea.VIGAS,
                referencia=int(viga.internal_id),
                codigo=getattr(viga, "codigo_viga", "") or "",
                etapa=estado_nuevo,
                etapa_anterior=_norm_estado(estado_anterior or ""),
                maquina=maquina_apunte or _maquina_asignada(viga.internal_id, estado_nuevo),
                actor=actor,
                ocurrido_en=now,
            )
    except Exception:
        logger.exception("no se pudo cambiar el estado de la viga %s", viga.pk)
        return JsonResponse({"ok": False, "error": "No se pudo cambiar el estado."}, status=500)

    idx = estados_index.get(viga.estado)
    if role == "corte":
        if viga.estado == "Espera de corte":
            next_estado = "Corte"
        elif viga.estado == "Corte":
            next_estado = "Espera de armado"
        else:
            next_estado = ""
    else:
        next_estado = "" if idx is None or idx >= len(ESTADOS) - 1 else ESTADOS[idx + 1]
    next_label = "En área de soldadura" if (role == "corte" and viga.estado == "Espera de armado") else ""
    ultimo_mov = _format_utc_naive_dt_as_local(getattr(viga, "ultimo_cambio", None))
    return JsonResponse(
        {
            "ok": True,
            "id": int(viga.internal_id),
            "estado": viga.estado,
            "estado_color": STATUS_COLORS.get(viga.estado, "#8898aa"),
            "estado_clase": clase_de_estado(viga.estado),
            "next_estado": next_estado,
            "next_label": next_label,
            "ultimo_mov": ultimo_mov,
        }
    )


@login_required
def viga_asignaciones(request, pk: int):
    viga = get_object_or_404(Viga, pk=pk)
    if request.method != "POST":
        next_url = _safe_next(request.GET.get("next", ""))
        if next_url:
            return redirect(next_url)
        return redirect("produccion:viga_list")

    wants_json = "application/json" in (request.headers.get("Accept") or "") or (
        (request.headers.get("X-Requested-With") or "").lower() == "xmlhttprequest"
    )

    role = _user_role(request.user)
    if role not in {"admin", "corte", "soldadura"}:
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403) if wants_json else redirect("produccion:home")

    actor = ""
    try:
        if getattr(request, "user", None) and request.user.is_authenticated:
            actor = request.user.get_username() or ""
    except Exception:
        actor = ""

    next_url = _safe_next(request.POST.get("next", "")) or reverse("produccion:viga_list")
    if role == "soldadura":
        etapas = {"soldadura_ids": "Soldadura", "pintura_ids": "Pintura"}
    elif role == "corte":
        etapas = {}
    else:
        etapas = {"soldadura_ids": "Soldadura", "pintura_ids": "Pintura"}

    def get_ids(key):
        ids = []
        try:
            if hasattr(request.POST, "getlist"):
                ids = [int(x) for x in request.POST.getlist(key) if str(x).isdigit()]
        except Exception:
            ids = []
        if not ids:
            ids = _parse_ids_csv(request.POST.get(key) or "")
        return list(dict.fromkeys(ids))

    def fail(msg: str, status: int = 400):
        if wants_json:
            return JsonResponse({"ok": False, "error": msg}, status=status)
        messages.error(request, msg)
        return redirect(next_url)

    try:
        with transaction.atomic(using="mes"):
            do_corte = ("corte_operador_ids" in request.POST) or ("corte_maquina_ids" in request.POST)
            if role in {"admin", "corte"} and do_corte:
                etapa = "Corte"
                corte_equipo = _equipo_for_etapa(etapa)
                if not corte_equipo:
                    return fail("No hay equipo configurado para la etapa Corte.")
                corte_operadores_ids = get_ids("corte_operador_ids")
                if not corte_operadores_ids:
                    return fail("Debes seleccionar al menos 1 operador en Corte.")
                if len(corte_operadores_ids) > 20:
                    return fail("Demasiados operadores en Corte.")
                corte_allowed = set(
                    Colaborador.objects.using("mes").filter(activo=True, equipo=corte_equipo).values_list("id", flat=True)
                )
                if any(i not in corte_allowed for i in corte_operadores_ids):
                    return fail("Operadores inválidos en Corte. Revisa colaboradores activos del equipo.")
                corte_maquina_ids = get_ids("corte_maquina_ids")
                if len(corte_maquina_ids) > 20:
                    return fail("Demasiadas máquinas en Corte.")
                if corte_maquina_ids:
                    ok_m = set(
                        Maquina.objects.using("mes").filter(activo=True, tipo="Corte", es_robot=False, id__in=corte_maquina_ids).values_list("id", flat=True)
                    )
                    if any(i not in ok_m for i in corte_maquina_ids):
                        return fail("Máquinas de corte inválidas.")

                VigaAsignacion.objects.using("mes").filter(viga_internal_id=viga.internal_id, etapa=etapa, vigente=True).update(vigente=False)
                colabs = {
                    c.id: c
                    for c in Colaborador.objects.using("mes").filter(id__in=corte_operadores_ids, activo=True, equipo=corte_equipo)
                }
                for cid in corte_operadores_ids:
                    c = colabs.get(cid)
                    if not c:
                        continue
                    VigaAsignacion.objects.using("mes").create(
                        viga_internal_id=viga.internal_id,
                        etapa=etapa,
                        rol="Operador",
                        colaborador=c,
                        maquina=None,
                        vigente=True,
                        asignado_por=actor,
                    )
                for mid in corte_maquina_ids:
                    VigaAsignacion.objects.using("mes").create(
                        viga_internal_id=viga.internal_id,
                        etapa=etapa,
                        rol="Maquina",
                        colaborador=None,
                        maquina_id=mid,
                        vigente=True,
                        asignado_por=actor,
                    )

            for form_key, etapa in etapas.items():
                ids = get_ids(form_key)
                if len(ids) > 20:
                    return fail(f"Demasiados asignados en {etapa}.")

                equipo = _equipo_for_etapa(etapa)
                if not equipo and etapa == "Soldadura":
                    equipo = _equipo_for_etapa("Armado")
                if not equipo:
                    return fail(f"No hay equipo configurado para la etapa {etapa}.")

                VigaAsignacion.objects.using("mes").filter(viga_internal_id=viga.internal_id, etapa=etapa, vigente=True).update(vigente=False)
                allowed = set(
                    Colaborador.objects.using("mes").filter(activo=True, equipo=equipo).values_list("id", flat=True)
                )

                if any(i not in allowed for i in ids):
                    return fail(f"Selección inválida en {etapa}. Revisa colaboradores activos del equipo.")

                if ids:
                    colabs = {
                        c.id: c
                        for c in Colaborador.objects.using("mes").filter(id__in=ids, activo=True, equipo=equipo)
                    }
                    for cid in ids:
                        c = colabs.get(cid)
                        if not c:
                            continue
                        VigaAsignacion.objects.using("mes").create(
                            viga_internal_id=viga.internal_id,
                            etapa=etapa,
                            rol=c.rol,
                            colaborador=c,
                            maquina=None,
                            vigente=True,
                            asignado_por=actor,
                        )
    except Exception:
        return fail("No se pudieron guardar las asignaciones.", status=500)

    if wants_json:
        asigns = list(
            VigaAsignacion.objects.using("mes").filter(viga_internal_id=viga.internal_id, vigente=True)
            .values("etapa", "rol", "colaborador_id", "maquina_id")
        )
        corte_oper = sorted({int(a["colaborador_id"]) for a in asigns if a["etapa"] == "Corte" and a["rol"] == "Operador" and a["colaborador_id"]})
        corte_maq = sorted({int(a["maquina_id"]) for a in asigns if a["etapa"] == "Corte" and a["rol"] == "Maquina" and a["maquina_id"]})
        sold = sorted({int(a["colaborador_id"]) for a in asigns if a["etapa"] == "Soldadura" and a["colaborador_id"]})
        pint = sorted({int(a["colaborador_id"]) for a in asigns if a["etapa"] == "Pintura" and a["colaborador_id"]})
        return JsonResponse(
            {
                "ok": True,
                "id": int(viga.internal_id),
                "corte_operador_ids": corte_oper,
                "corte_maquina_ids": corte_maq,
                "soldadura_ids": sold,
                "pintura_ids": pint,
            }
        )

    messages.success(request, "Asignaciones guardadas.")
    return redirect(next_url)


@login_required
def viga_update_meta_json(request, pk: int):
    if not _is_admin_user(request.user):
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método inválido."}, status=405)

    viga = get_object_or_404(Viga, pk=pk)
    raw_fecha = (request.POST.get("fecha_compromiso") or "").strip()
    raw_prio = (request.POST.get("prioridad") or "").strip()
    try:
        fecha = datetime.strptime(raw_fecha, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "error": "Fecha compromiso inválida."}, status=400)

    try:
        prioridad = int(raw_prio)
    except Exception:
        return JsonResponse({"ok": False, "error": "Prioridad inválida."}, status=400)
    if prioridad < 1 or prioridad > 5:
        return JsonResponse({"ok": False, "error": "Prioridad debe ser 1 a 5."}, status=400)

    try:
        with transaction.atomic(using="mes"):
            viga.fecha_compromiso = fecha
            viga.prioridad = prioridad
            viga.save(update_fields=["fecha_compromiso", "prioridad"])
    except Exception:
        return JsonResponse({"ok": False, "error": "No se pudo guardar."}, status=500)

    return JsonResponse(
        {
            "ok": True,
            "id": int(viga.internal_id),
            "fecha_compromiso": viga.fecha_compromiso.isoformat() if viga.fecha_compromiso else "",
            "prioridad": int(viga.prioridad or 0),
        }
    )


@login_required
def dashboard(request):
    base_qs = Viga.objects.exclude(estado="Enviado")
    estados_reporte = [s for s in ESTADOS if s != "Enviado"]
    agg = base_qs.values("estado").annotate(piezas=Count("internal_id"), kg=Sum("peso_kg"))
    by_estado = {}
    for row in agg:
        st = _norm_estado(str(row.get("estado") or ""))
        cur = by_estado.get(st) or {"piezas": 0, "kg": 0.0}
        cur["piezas"] = int(cur.get("piezas") or 0) + int(row.get("piezas") or 0)
        cur["kg"] = float(cur.get("kg") or 0.0) + float(row.get("kg") or 0.0)
        by_estado[st] = cur
    resumen = []
    total_piezas = 0
    total_kg = 0.0
    for estado in estados_reporte:
        row = by_estado.get(estado) or {"piezas": 0, "kg": 0}
        piezas = int(row["piezas"] or 0)
        kg = float(row["kg"] or 0)
        total_piezas += piezas
        total_kg += kg
        resumen.append(
            {
                "estado": estado,
                "piezas": piezas,
                "kg": kg,
                "ton": kg / 1000.0,
                "pct": 0.0,
                "pct_ton": 0.0,
                "pct_css": "0",
                "pct_ton_css": "0",
                "color": STATUS_COLORS.get(estado, "#8898aa"),
            }
        )

    total_ton = total_kg / 1000.0
    for r in resumen:
        r["pct"] = (r["piezas"] / total_piezas * 100.0) if total_piezas else 0.0
        r["pct_ton"] = (r["ton"] / total_ton * 100.0) if total_ton else 0.0
        r["pct_css"] = f"{r['pct']:.1f}"
        r["pct_ton_css"] = f"{r['pct_ton']:.1f}"

    terminado = next((r for r in resumen if r["estado"] == "Terminado"), None)
    terminado_piezas = int(terminado["piezas"]) if terminado else 0
    terminado_ton = float(terminado["ton"]) if terminado else 0.0
    pct_terminado = (terminado_piezas / total_piezas * 100.0) if total_piezas else 0.0
    labels = [r["estado"] for r in resumen]
    piezas_data = [r["piezas"] for r in resumen]
    ton_data = [round(r["ton"], 3) for r in resumen]
    colors = [r["color"] for r in resumen]

    # Antes: `comentario__icontains="retrabajo"`, que daba por retrabajo
    # cualquier comentario donde apareciera la palabra, incluido «no hubo
    # retrabajo». Es la misma definición que usa el bloque semanal, que
    # hasta ahora era distinta en la misma pantalla.
    retrabajo_vigas_qs = base_qs.filter(
        metricas.filtro_de_retrabajo("logs__comentario")
    ).distinct()
    retrabajo_piezas = retrabajo_vigas_qs.count()
    retrabajo_kg = float(retrabajo_vigas_qs.aggregate(total=Sum("peso_kg"))["total"] or 0.0)
    retrabajo_ton = retrabajo_kg / 1000.0
    retrabajo_pct = (retrabajo_piezas / total_piezas * 100.0) if total_piezas else 0.0

    equipos_activos = list(EquipoTrabajo.objects.filter(activo=True))
    total_equipos = len(equipos_activos)
    colab_counts = dict(
        Colaborador.objects.filter(activo=True, equipo__activo=True)
        .values_list("equipo_id")
        .annotate(cnt=Count("id"))
        .values_list("equipo_id", "cnt")
    )
    total_integrantes = sum(int(colab_counts.get(e.id, 0)) for e in equipos_activos)
    # Producción por persona: **flujo**, no inventario.
    #
    # Antes era `toneladas en estado Terminado / integrantes`, comparado contra
    # una meta de media tonelada por persona **a la semana**. Arriba había un
    # inventario y abajo un flujo: el número subía solo mientras no se enviara
    # nada y se desplomaba el día que salía un camión, o sea justo al revés de
    # lo que significa producir.
    #
    # Ahora se mide lo terminado en los últimos siete días, que es la misma
    # dimensión que la meta.
    ton_por_persona_meta = float(getattr(settings, "TON_POR_PERSONA_META", 0.5) or 0.5)
    # `today` se define más abajo en esta misma vista; aquí se calcula aparte
    # para no depender del orden de mil cuatrocientas líneas.
    hoy_kpi = timezone.localdate()
    semana_desde = hoy_kpi - timedelta(days=6)
    semana_hasta = hoy_kpi + timedelta(days=1)
    ton_terminadas_semana = metricas.toneladas_terminadas(semana_desde, semana_hasta)
    prod_ton_por_equipo_global = (
        (ton_terminadas_semana / total_equipos) if total_equipos else 0.0
    )
    prod_ton_por_integrante_global = metricas.toneladas_por_persona(
        semana_desde, semana_hasta, total_integrantes
    )

    equipos_detalle = [
        {"nombre": e.nombre, "integrantes": int(colab_counts.get(e.id, 0)), "estados": e.estados_texto}
        for e in equipos_activos
    ]
    equipos_labels = [e["nombre"] for e in equipos_detalle]
    equipos_integrantes = [e["integrantes"] for e in equipos_detalle]
    equipos_colors = [STATUS_COLORS.get(_norm_estado((e["estados"].split(",")[0] or "").strip()), "#5e72e4") for e in equipos_detalle]

    kpi_maquinas_corte = []
    corte_asigs = (
        VigaAsignacion.objects.filter(etapa="Corte", rol="Maquina", vigente=True, maquina__tipo="Corte", maquina__es_robot=False)
        .select_related("maquina")
        .values_list("maquina_id", "maquina__nombre", "viga_internal_id")
    )
    corte_map = {}
    for mid, mname, vid in corte_asigs:
        if not mid or not vid:
            continue
        corte_map.setdefault(int(mid), {"maquina": mname or "", "vigas": set()})["vigas"].add(int(vid))
    all_viga_ids = []
    for row in corte_map.values():
        all_viga_ids.extend(list(row["vigas"]))
    all_viga_ids = list(dict.fromkeys(all_viga_ids))
    peso_map = dict(Viga.objects.filter(internal_id__in=all_viga_ids).values_list("internal_id", "peso_kg"))
    for mid, row in corte_map.items():
        vigas_ids = list(row["vigas"])
        piezas = len(vigas_ids)
        kg_total = sum(float(peso_map.get(vid, 0.0) or 0.0) for vid in vigas_ids)
        ton_total = kg_total / 1000.0
        kpi_maquinas_corte.append(
            {
                "maquina": row["maquina"] or f"#{mid}",
                "piezas": piezas,
                "ton": round(ton_total, 3),
                "ton_promedio": round((ton_total / piezas) if piezas else 0.0, 3),
            }
        )
    kpi_maquinas_corte.sort(key=lambda x: x["ton"], reverse=True)

    today = timezone.localdate()
    wip_qs = base_qs.exclude(estado="Terminado")
    wip_piezas = int(wip_qs.count())
    wip_kg = float(wip_qs.aggregate(total=Sum("peso_kg"))["total"] or 0.0)
    wip_ton = wip_kg / 1000.0

    vencidas_qs = wip_qs.filter(fecha_compromiso__lt=today)
    vencidas_piezas = int(vencidas_qs.count())
    vencidas_kg = float(vencidas_qs.aggregate(total=Sum("peso_kg"))["total"] or 0.0)
    vencidas_ton = vencidas_kg / 1000.0

    vencidas_por_proyecto_estado = list(
        vencidas_qs.values("proyecto", "estado")
        .annotate(piezas=Count("internal_id"), kg=Sum("peso_kg"))
        .order_by("-kg")[:15]
    )
    for r in vencidas_por_proyecto_estado:
        r["ton"] = float(r.get("kg") or 0.0) / 1000.0

    age_expr = ExpressionWrapper(timezone.now() - F("ultimo_cambio"), output_field=DurationField())
    age_by_estado_qs = (
        base_qs.values("estado")
        .annotate(avg_age=Avg(age_expr), piezas=Count("internal_id"))
        .order_by()
    )
    avg_age_map = {}
    for row in age_by_estado_qs:
        st = _norm_estado(str(row.get("estado") or ""))
        avg_age_map[st] = row
    aging_labels = []
    aging_days = []
    aging_colors = []
    for estado in estados_reporte:
        row = avg_age_map.get(estado) or {}
        avg_age = row.get("avg_age")
        days = float(avg_age.total_seconds() / 86400.0) if avg_age else 0.0
        aging_labels.append(estado)
        aging_days.append(round(days, 2))
        aging_colors.append(STATUS_COLORS.get(estado, "#8898aa"))

    oldest_wip = []
    for v in wip_qs.order_by("ultimo_cambio", "prioridad", "codigo_viga", "pieza_no")[:12]:
        last = v.ultimo_cambio.date() if getattr(v, "ultimo_cambio", None) else today
        oldest_wip.append(
            {
                "id": v.internal_id,
                "codigo": v.codigo_viga,
                "pieza": f"{v.pieza_no}/{v.total_piezas}",
                "proyecto": v.proyecto,
                "estado": v.estado,
                "dias": int((today - last).days),
            }
        )

    top_backlog = list(
        vencidas_qs.values("proyecto")
        .annotate(piezas=Count("internal_id"), kg=Sum("peso_kg"))
        .order_by("-kg")[:10]
    )
    for r in top_backlog:
        r["ton"] = float(r["kg"] or 0.0) / 1000.0

    top_wip = list(
        wip_qs.values("proyecto")
        .annotate(piezas=Count("internal_id"), kg=Sum("peso_kg"))
        .order_by("-kg")[:10]
    )
    for r in top_wip:
        r["ton"] = float(r["kg"] or 0.0) / 1000.0

    last_days = 30
    start = today - timedelta(days=last_days - 1)
    term_events = (
        ProductionLog.objects.filter(estado_nuevo="Terminado", fecha_operacion__gte=start, fecha_operacion__lte=today)
        .exclude(viga_internal__estado="Enviado")
        .values("fecha_operacion", "viga_internal_id")
        .distinct()
        .values("fecha_operacion")
        .annotate(
            piezas=Count("viga_internal_id"),
            kg=Sum("viga_internal__peso_kg"),
        )
        .order_by("fecha_operacion")
    )
    term_map = {row["fecha_operacion"]: row for row in term_events}
    throughput_labels = []
    throughput_ton = []
    throughput_piezas = []
    for i in range(last_days):
        d = start + timedelta(days=i)
        row = term_map.get(d) or {}
        throughput_labels.append(d.isoformat())
        throughput_piezas.append(int(row.get("piezas") or 0))
        throughput_ton.append(round(float(row.get("kg") or 0.0) / 1000.0, 3))

    term_logs = (
        ProductionLog.objects.filter(estado_nuevo="Terminado", fecha_operacion__gte=start, fecha_operacion__lte=today)
        .exclude(viga_internal__estado="Enviado")
        .select_related("viga_internal")
        .order_by("viga_internal_id", "-fecha_operacion", "-timestamp")
    )
    seen_ids = set()
    on_time = 0
    late = 0
    no_due = 0
    late_days = []
    for lg in term_logs:
        vid = int(lg.viga_internal_id)
        if vid in seen_ids:
            continue
        seen_ids.add(vid)
        due = getattr(lg.viga_internal, "fecha_compromiso", None)
        if not due:
            no_due += 1
            continue
        if lg.fecha_operacion <= due:
            on_time += 1
        else:
            late += 1
            late_days.append(int((lg.fecha_operacion - due).days))
    avg_late_days = round(sum(late_days) / len(late_days), 2) if late_days else 0.0

    enviados_total_qs = Viga.objects.filter(estado="Enviado")
    enviados_total = int(enviados_total_qs.count())
    enviados_total_kg = float(enviados_total_qs.aggregate(total=Sum("peso_kg"))["total"] or 0.0)
    enviados_total_ton = enviados_total_kg / 1000.0

    def _to_thursday_start(d):
        delta = (d.weekday() - 3) % 7
        return d - timedelta(days=delta)

    retention_weeks = int(getattr(settings, "WEEKLY_SNAPSHOT_RETENTION_WEEKS", 156) or 156)
    retention_cutoff = today - timedelta(days=retention_weeks * 7)
    WeeklyReportSnapshot.objects.filter(week_start__lt=retention_cutoff).delete()

    selected_week_start = request.GET.get("week_start", "").strip()
    refresh_weekly = request.GET.get("weekly_refresh", "").strip() == "1"
    week_start = _to_thursday_start(today)
    if selected_week_start:
        try:
            parsed = datetime.strptime(selected_week_start, "%Y-%m-%d").date()
            week_start = _to_thursday_start(parsed)
        except Exception:
            logger.exception("Error ignorado en dashboard()")
    week_end = week_start + timedelta(days=7)
    week_start_local = timezone.make_aware(datetime.combine(week_start, time.min), timezone.get_default_timezone())
    week_end_local = timezone.make_aware(datetime.combine(week_end, time.min), timezone.get_default_timezone())
    week_start_dt = week_start_local.astimezone(dt_timezone.utc).replace(tzinfo=None)
    week_end_dt = week_end_local.astimezone(dt_timezone.utc).replace(tzinfo=None)

    logistica_op_kg_expr = ExpressionWrapper(
        F("cantidad") * Coalesce(F("producto__peso_kg"), 0.0),
        output_field=FloatField(),
    )
    logistica_op_week = LogisticaEnvioItem.objects.filter(
        envio__fecha__gte=week_start,
        envio__fecha__lt=week_end,
    )
    logistica_op_week_agg = logistica_op_week.aggregate(pzs=Sum("cantidad"), kg=Sum(logistica_op_kg_expr))
    logistica_op_piezas_week = int(logistica_op_week_agg.get("pzs") or 0)
    logistica_op_kg_week = float(logistica_op_week_agg.get("kg") or 0.0)
    logistica_op_ton_week = logistica_op_kg_week / 1000.0

    logistica_corta_week = list(
        LogisticaEnvioCorta.objects.select_related("orden").filter(
            fecha__gte=week_start,
            fecha__lt=week_end,
        )
    )
    logistica_corta_piezas_week = 0
    logistica_corta_kg_week = 0.0
    for e in logistica_corta_week:
        qty = int(getattr(e, "cantidad", 0) or 0)
        if qty <= 0:
            continue
        logistica_corta_piezas_week += qty
        o = getattr(e, "orden", None)
        total = int(getattr(o, "total_piezas", 0) or 0) if o else 0
        peso_total = float(getattr(o, "peso_kg", 0.0) or 0.0) if o else 0.0
        per_piece = (peso_total / float(total)) if (total > 0 and peso_total > 0) else 0.0
        logistica_corta_kg_week += float(per_piece) * float(qty)
    logistica_corta_kg_week = float(round(logistica_corta_kg_week, 3))
    logistica_corta_ton_week = logistica_corta_kg_week / 1000.0

    robot_peso_expr = Coalesce(
        F("orden_item__pieza__peso_kg"),
        F("orden_item__pieza_custom_peso_kg"),
        F("pieza__peso_kg"),
        F("pieza_custom_peso_kg"),
        0.0,
    )
    robot_expr_kg_week = ExpressionWrapper(F("cantidad") * robot_peso_expr, output_field=FloatField())
    robot_week_qs = RobotProduccion.objects.filter(fecha__gte=week_start, fecha__lt=week_end)
    robot_total_kg_week = float(robot_week_qs.aggregate(total=Sum(robot_expr_kg_week))["total"] or 0.0)
    robot_total_ton_week = robot_total_kg_week / 1000.0
    robot_total_piezas_week = int(robot_week_qs.aggregate(total=Sum("cantidad"))["total"] or 0)
    robot_day_rows = (
        robot_week_qs.values("fecha")
        .annotate(piezas=Sum("cantidad"), kg=Sum(robot_expr_kg_week))
        .order_by("fecha")
    )
    robot_day_map = {row["fecha"]: row for row in robot_day_rows}

    # Antes esto salía de `HerrProduccion`, multiplicando la cantidad por el
    # peso de la pieza del renglón de la orden. En la base sólo hay una fila
    # de esa tabla y tiene el renglón vacío, así que el peso resolvía a cero:
    # **las toneladas de Herrería llevaban saliendo cero desde siempre.** El
    # avance real se lleva por contadores y deja rastro en HerrAvanceCambio.
    herr_day_map = metricas.produccion_de_herreria(week_start, week_end)
    herr_total_kg_week = sum(f["kg"] for f in herr_day_map.values())
    herr_total_ton_week = herr_total_kg_week / 1000.0
    herr_total_piezas_week = sum(f["piezas"] for f in herr_day_map.values())

    corte_term_qs = LaserEstadoCambio.objects.filter(
        estado_nuevo="Terminado",
        fecha_operacion__gte=week_start,
        fecha_operacion__lt=week_end,
    )
    corte_term_rows = list(
        corte_term_qs.order_by("orden_id", "-fecha_operacion", "-creado_en")
        .distinct("orden_id")
        .values("orden_id", "fecha_operacion")[:10000]
    )
    corte_term_ids = [int(r["orden_id"]) for r in corte_term_rows if r.get("orden_id")]
    corte_order_rows = list(
        LaserOrdenProduccion.objects.filter(id__in=corte_term_ids).values("id", "peso_kg", "total_piezas")
    ) if corte_term_ids else []
    corte_order_map = {
        int(r["id"]): {"kg": float(r.get("peso_kg") or 0.0), "piezas": int(r.get("total_piezas") or 0)}
        for r in corte_order_rows
        if r.get("id")
    }
    corte_total_kg_week = sum((corte_order_map.get(int(i), {}) or {}).get("kg", 0.0) for i in corte_term_ids)
    corte_total_ton_week = float(corte_total_kg_week or 0.0) / 1000.0
    corte_total_piezas_week = sum((corte_order_map.get(int(i), {}) or {}).get("piezas", 0) for i in corte_term_ids)
    corte_day_map = {}
    for r in corte_term_rows:
        d = r.get("fecha_operacion")
        oid = int(r.get("orden_id") or 0)
        if not d or oid <= 0:
            continue
        meta = corte_order_map.get(oid) or {}
        entry = corte_day_map.setdefault(d, {"fecha": d, "piezas": 0, "kg": 0.0})
        entry["piezas"] = int(entry.get("piezas") or 0) + int(meta.get("piezas") or 0)
        entry["kg"] = float(entry.get("kg") or 0.0) + float(meta.get("kg") or 0.0)

    weekly_term_ids = list(
        ProductionLog.objects.filter(
            estado_nuevo="Terminado",
            fecha_operacion__gte=week_start,
            fecha_operacion__lt=week_end,
        )
        .values_list("viga_internal_id", flat=True)
        .distinct()
    )
    # El peso de esas piezas se pedía aquí y no se usaba: una consulta por
    # cada carga del tablero, para nada. El KPI de toneladas de la semana se
    # calcula más abajo, a partir de la bitácora.
    weekly_retrabajo_piezas = int(
        ProductionLog.objects.filter(
            viga_internal_id__in=weekly_term_ids,
            fecha_operacion__gte=week_start,
            fecha_operacion__lt=week_end,
        )
        .filter(metricas.filtro_de_retrabajo())
        .values("viga_internal_id")
        .distinct()
        .count()
    )
    weekly_retrabajo_any_piezas = int(
        ProductionLog.objects.filter(
            fecha_operacion__gte=week_start,
            fecha_operacion__lt=week_end,
        )
        .filter(metricas.filtro_de_retrabajo())
        .values("viga_internal_id")
        .distinct()
        .count()
    )
    retrabajo_any_list = []
    try:
        retrab_logs = list(
            ProductionLog.objects.filter(
                fecha_operacion__gte=week_start,
                fecha_operacion__lt=week_end,
            )
            .filter(metricas.filtro_de_retrabajo())
            .order_by("viga_internal_id", "-timestamp", "-id")
            .distinct("viga_internal_id")
            .values("viga_internal_id", "fecha_operacion", "estado_anterior", "estado_nuevo", "timestamp")
        )
        retrab_ids = [int(r["viga_internal_id"]) for r in retrab_logs if r.get("viga_internal_id")]
        vrows = list(
            Viga.objects.filter(internal_id__in=retrab_ids)
            .values(
                "internal_id",
                "codigo_viga",
                "pieza_no",
                "total_piezas",
                "proyecto",
                "estado",
                "fecha_compromiso",
                "prioridad",
            )
        )
        vmap = {int(v["internal_id"]): v for v in vrows if v.get("internal_id")}
        for r in retrab_logs:
            vid = int(r["viga_internal_id"])
            v = vmap.get(vid)
            if not v:
                continue
            retrabajo_any_list.append(
                {
                    "id": vid,
                    "codigo": v.get("codigo_viga") or "",
                    "pieza_no": int(v.get("pieza_no") or 0),
                    "total_piezas": int(v.get("total_piezas") or 0),
                    "proyecto": v.get("proyecto") or "",
                    "estado": v.get("estado") or "",
                    "fecha_compromiso": v.get("fecha_compromiso"),
                    "prioridad": int(v.get("prioridad") or 0),
                    "retrabajo_fecha": r.get("fecha_operacion"),
                    "retrabajo_de": r.get("estado_anterior") or "",
                    "retrabajo_a": r.get("estado_nuevo") or "",
                    "retrabajo_ts": r.get("timestamp"),
                }
            )
        retrabajo_any_list.sort(key=lambda x: (x.get("retrabajo_ts") or datetime.min), reverse=True)
        retrabajo_any_list = retrabajo_any_list[:200]
    except Exception:
        retrabajo_any_list = []
    weekly_retrabajo_pct = (weekly_retrabajo_piezas / len(weekly_term_ids) * 100.0) if weekly_term_ids else 0.0

    weekly_snapshot = WeeklyReportSnapshot.objects.filter(week_start=week_start).first()
    weekly_payload = {}
    if (not weekly_snapshot) or refresh_weekly:
        term_week_qs = ProductionLog.objects.filter(
            estado_nuevo="Terminado",
            fecha_operacion__gte=week_start,
            fecha_operacion__lt=week_end,
        )
        term_ids = list(term_week_qs.values_list("viga_internal_id", flat=True).distinct())
        term_vigas = list(Viga.objects.filter(internal_id__in=term_ids).values("internal_id", "peso_kg", "proyecto"))
        peso_map = {int(r["internal_id"]): float(r["peso_kg"] or 0.0) for r in term_vigas}
        proyecto_map = {int(r["internal_id"]): (r["proyecto"] or "SIN PROYECTO") for r in term_vigas}

        term_piezas = len(term_ids)
        term_kg = sum(peso_map.get(int(i), 0.0) for i in term_ids)
        term_ton = term_kg / 1000.0

        enviados_set = set(
            ProductionLog.objects.filter(estado_nuevo="Enviado", viga_internal_id__in=term_ids)
            .values_list("viga_internal_id", flat=True)
            .distinct()
        )
        term_enviadas_ids = [i for i in term_ids if i in enviados_set]
        term_enviadas_piezas = len(term_enviadas_ids)
        term_enviadas_kg = sum(peso_map.get(int(i), 0.0) for i in term_enviadas_ids)
        term_enviadas_ton = term_enviadas_kg / 1000.0

        day_rows = (
            term_week_qs.values("fecha_operacion", "viga_internal_id")
            .distinct()
            .values("fecha_operacion")
            .annotate(piezas=Count("viga_internal_id"), kg=Sum("viga_internal__peso_kg"))
            .order_by("fecha_operacion")
        )
        daily = []
        day_map = {row["fecha_operacion"]: row for row in day_rows}
        for i in range(7):
            d = week_start + timedelta(days=i)
            row = day_map.get(d) or {}
            robot_row = robot_day_map.get(d) or {}
            herr_row = herr_day_map.get(d) or {}
            manual_piezas = int(row.get("piezas") or 0)
            manual_ton = float(row.get("kg") or 0.0) / 1000.0
            robot_piezas = int(robot_row.get("piezas") or 0)
            robot_ton = float(robot_row.get("kg") or 0.0) / 1000.0
            herr_piezas = int(herr_row.get("piezas") or 0)
            herr_ton = float(herr_row.get("kg") or 0.0) / 1000.0
            corte_row = corte_day_map.get(d) or {}
            corte_piezas = int(corte_row.get("piezas") or 0)
            corte_ton = float(corte_row.get("kg") or 0.0) / 1000.0
            daily.append(
                {
                    "date": d.isoformat(),
                    "manual_piezas": manual_piezas,
                    "manual_ton": round(manual_ton, 3),
                    "robot_piezas": robot_piezas,
                    "robot_ton": round(robot_ton, 3),
                    "herr_piezas": herr_piezas,
                    "herr_ton": round(herr_ton, 3),
                    "corte_piezas": corte_piezas,
                    "corte_ton": round(corte_ton, 3),
                    "piezas": manual_piezas + robot_piezas + herr_piezas + corte_piezas,
                    "ton": round(manual_ton + robot_ton + herr_ton + corte_ton, 3),
                }
            )

        by_proyecto = {}
        for vid in term_ids:
            proy = (proyecto_map.get(int(vid)) or "SIN PROYECTO").strip().upper() or "SIN PROYECTO"
            by_proyecto.setdefault(proy, 0.0)
            by_proyecto[proy] += peso_map.get(int(vid), 0.0)
        top_proyectos = [
            {"proyecto": k, "ton": round(v / 1000.0, 3)}
            for k, v in sorted(by_proyecto.items(), key=lambda x: x[1], reverse=True)[:12]
        ]

        integrantes_snapshot = int(total_integrantes or 0)
        ton_por_persona_sem = (term_ton / integrantes_snapshot) if integrantes_snapshot else 0.0

        weekly_payload = {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "terminadas_piezas": term_piezas,
            "terminadas_ton": round(term_ton, 3),
            "robot_piezas": int(robot_total_piezas_week),
            "robot_ton": round(robot_total_ton_week, 3),
            "herr_piezas": int(herr_total_piezas_week),
            "herr_ton": round(herr_total_ton_week, 3),
            "corte_piezas": int(corte_total_piezas_week),
            "corte_kg": round(float(corte_total_kg_week or 0.0), 3),
            "corte_ton": round(corte_total_ton_week, 3),
            "logistica_op_piezas": int(logistica_op_piezas_week),
            "logistica_op_kg": round(float(logistica_op_kg_week or 0.0), 3),
            "logistica_op_ton": round(float(logistica_op_ton_week or 0.0), 3),
            "logistica_corta_piezas": int(logistica_corta_piezas_week),
            "logistica_corta_kg": round(float(logistica_corta_kg_week or 0.0), 3),
            "logistica_corta_ton": round(float(logistica_corta_ton_week or 0.0), 3),
            "total_ton": round(term_ton + robot_total_ton_week + herr_total_ton_week + corte_total_ton_week, 3),
            "terminadas_enviadas_piezas": term_enviadas_piezas,
            "terminadas_enviadas_ton": round(term_enviadas_ton, 3),
            "integrantes_total": integrantes_snapshot,
            "ton_por_persona": round(ton_por_persona_sem, 3),
            "retrabajo_piezas": int(weekly_retrabajo_piezas),
            "retrabajo_pct": round(float(weekly_retrabajo_pct), 2),
            "retrabajo_any_piezas": int(weekly_retrabajo_any_piezas),
            "daily": daily,
            "top_proyectos": top_proyectos,
        }
        weekly_snapshot, _created = WeeklyReportSnapshot.objects.update_or_create(
            week_start=week_start,
            defaults={
                "week_end": week_end,
                "integrantes_total": integrantes_snapshot,
                "payload_json": json.dumps(weekly_payload),
            },
        )
    else:
        try:
            weekly_payload = json.loads(weekly_snapshot.payload_json or "{}") if weekly_snapshot.payload_json else {}
        except Exception:
            weekly_payload = {}
    if isinstance(weekly_payload, dict):
        weekly_payload.setdefault("robot_piezas", int(robot_total_piezas_week))
        weekly_payload.setdefault("robot_ton", round(robot_total_ton_week, 3))
        weekly_payload.setdefault("herr_piezas", int(herr_total_piezas_week))
        weekly_payload.setdefault("herr_ton", round(herr_total_ton_week, 3))
        weekly_payload.setdefault("corte_piezas", int(corte_total_piezas_week))
        weekly_payload.setdefault("corte_kg", round(float(corte_total_kg_week or 0.0), 3))
        weekly_payload.setdefault("corte_ton", round(corte_total_ton_week, 3))
        weekly_payload.setdefault("logistica_op_piezas", int(logistica_op_piezas_week))
        weekly_payload.setdefault("logistica_op_kg", round(float(logistica_op_kg_week or 0.0), 3))
        weekly_payload.setdefault("logistica_op_ton", round(float(logistica_op_ton_week or 0.0), 3))
        weekly_payload.setdefault("logistica_corta_piezas", int(logistica_corta_piezas_week))
        weekly_payload.setdefault("logistica_corta_kg", round(float(logistica_corta_kg_week or 0.0), 3))
        weekly_payload.setdefault("logistica_corta_ton", round(float(logistica_corta_ton_week or 0.0), 3))
        weekly_payload["total_ton"] = round(
            float(weekly_payload.get("terminadas_ton") or 0.0)
            + float(weekly_payload.get("robot_ton") or 0.0)
            + float(weekly_payload.get("herr_ton") or 0.0)
            + float(weekly_payload.get("corte_ton") or 0.0),
            3,
        )
        weekly_payload.setdefault("retrabajo_any_piezas", int(weekly_retrabajo_any_piezas))
        daily_src = weekly_payload.get("daily") or []
        daily_out = []
        daily_map = {d.get("date"): d for d in daily_src if isinstance(d, dict)}
        for i in range(7):
            d = week_start + timedelta(days=i)
            k = d.isoformat()
            row = daily_map.get(k) or {}
            manual_piezas = int(row.get("manual_piezas") or row.get("piezas") or 0)
            manual_ton = float(row.get("manual_ton") or row.get("ton") or 0.0)
            robot_row = robot_day_map.get(d) or {}
            robot_piezas = int(robot_row.get("piezas") or 0)
            robot_ton = float(robot_row.get("kg") or 0.0) / 1000.0
            herr_row = herr_day_map.get(d) or {}
            herr_piezas = int(herr_row.get("piezas") or 0)
            herr_ton = float(herr_row.get("kg") or 0.0) / 1000.0
            corte_row = corte_day_map.get(d) or {}
            corte_piezas = int(corte_row.get("piezas") or 0)
            corte_ton = float(corte_row.get("kg") or 0.0) / 1000.0
            daily_out.append(
                {
                    "date": k,
                    "manual_piezas": manual_piezas,
                    "manual_ton": round(manual_ton, 3),
                    "robot_piezas": robot_piezas,
                    "robot_ton": round(robot_ton, 3),
                    "herr_piezas": herr_piezas,
                    "herr_ton": round(herr_ton, 3),
                    "corte_piezas": corte_piezas,
                    "corte_ton": round(corte_ton, 3),
                    "piezas": manual_piezas + robot_piezas + herr_piezas + corte_piezas,
                    "ton": round(manual_ton + robot_ton + herr_ton + corte_ton, 3),
                }
            )
        weekly_payload["daily"] = daily_out

    week_start_utc = week_start_local.astimezone(dt_timezone.utc)
    week_end_utc = week_end_local.astimezone(dt_timezone.utc)
    now_utc = timezone.now()

    def _overlap_seconds(a_start, a_end, b_start, b_end):
        if not a_start:
            return 0.0
        s = a_start
        e = a_end or now_utc
        if s >= b_end or e <= b_start:
            return 0.0
        ss = s if s > b_start else b_start
        ee = e if e < b_end else b_end
        try:
            return max(0.0, (ee - ss).total_seconds())
        except Exception:
            return 0.0

    paros_qs = (
        MaquinaParo.objects.filter(inicio__lt=week_end_utc)
        .filter(Q(fin__isnull=True) | Q(fin__gt=week_start_utc))
        .select_related("motivo")
    )
    fallas_qs = (
        MaquinaFalla.objects.filter(inicio__lt=week_end_utc)
        .filter(Q(fin__isnull=True) | Q(fin__gt=week_start_utc))
        .select_related("tipo")
    )
    energia_qs = (
        PlantaEvento.objects.filter(tipo="Energia", inicio__lt=week_end_utc)
        .filter(Q(fin__isnull=True) | Q(fin__gt=week_start_utc))
    )

    paro_seconds = 0.0
    falla_seconds = 0.0
    energia_seconds = 0.0
    paro_by = {}
    falla_by = {}

    for p in paros_qs:
        secs = _overlap_seconds(p.inicio, p.fin, week_start_utc, week_end_utc)
        paro_seconds += secs
        k = (p.motivo.nombre if getattr(p, "motivo", None) else "Sin motivo") or "Sin motivo"
        paro_by[k] = paro_by.get(k, 0.0) + secs
    for f in fallas_qs:
        secs = _overlap_seconds(f.inicio, f.fin, week_start_utc, week_end_utc)
        falla_seconds += secs
        k = (f.tipo.nombre if getattr(f, "tipo", None) else "Sin tipo") or "Sin tipo"
        falla_by[k] = falla_by.get(k, 0.0) + secs
    for e in energia_qs:
        energia_seconds += _overlap_seconds(e.inicio, e.fin, week_start_utc, week_end_utc)

    weekly_payload["paros_horas"] = round(paro_seconds / 3600.0, 2)
    weekly_payload["fallas_horas"] = round(falla_seconds / 3600.0, 2)
    weekly_payload["energia_horas"] = round(energia_seconds / 3600.0, 2)
    weekly_payload["top_paros"] = [
        {"motivo": k, "horas": round(v / 3600.0, 2)}
        for k, v in sorted(paro_by.items(), key=lambda x: x[1], reverse=True)[:7]
    ]
    weekly_payload["top_fallas"] = [
        {"tipo": k, "horas": round(v / 3600.0, 2)}
        for k, v in sorted(falla_by.items(), key=lambda x: x[1], reverse=True)[:7]
    ]

    work_windows = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        if d.weekday() > 4:
            continue
        tz = timezone.get_default_timezone()
        a1 = timezone.make_aware(datetime.combine(d, time(7, 30)), tz).astimezone(dt_timezone.utc)
        b1 = timezone.make_aware(datetime.combine(d, time(13, 0)), tz).astimezone(dt_timezone.utc)
        a2 = timezone.make_aware(datetime.combine(d, time(13, 30)), tz).astimezone(dt_timezone.utc)
        b2 = timezone.make_aware(datetime.combine(d, time(17, 0)), tz).astimezone(dt_timezone.utc)
        work_windows.append((a1, b1))
        work_windows.append((a2, b2))

    planned_per_machine_seconds = 0.0
    for a, b in work_windows:
        try:
            planned_per_machine_seconds += max(0.0, (b - a).total_seconds())
        except Exception:
            logger.exception("Error ignorado en dashboard()")

    maquinas_activas = list(Maquina.objects.filter(activo=True).values("id", "nombre", "tipo", "es_robot"))
    ids_total = {int(m["id"]) for m in maquinas_activas}
    ids_corte = {int(m["id"]) for m in maquinas_activas if (m["tipo"] == "Corte" and (not bool(m["es_robot"])))}
    ids_soldadura = {int(m["id"]) for m in maquinas_activas if (m["tipo"] == "Soldadura" and (not bool(m["es_robot"])))}
    ids_robot = {int(m["id"]) for m in maquinas_activas if bool(m["es_robot"])}
    maquinas_by_id = {int(m["id"]): m for m in maquinas_activas}

    def _overlap_work_seconds(a_start, a_end):
        total = 0.0
        for w_start, w_end in work_windows:
            total += _overlap_seconds(a_start, a_end, w_start, w_end)
        return total

    down_total = 0.0
    down_corte = 0.0
    down_soldadura = 0.0
    down_robot = 0.0
    down_by_machine = {}
    for p in paros_qs:
        mid = int(getattr(p, "maquina_id", 0) or 0)
        if not mid:
            continue
        ov = _overlap_work_seconds(p.inicio, p.fin)
        if ov <= 0:
            continue
        down_by_machine[mid] = down_by_machine.get(mid, 0.0) + ov
        if mid in ids_total:
            down_total += ov
        if mid in ids_corte:
            down_corte += ov
        if mid in ids_soldadura:
            down_soldadura += ov
        if mid in ids_robot:
            down_robot += ov

    def _hours(v):
        return round(float(v or 0.0) / 3600.0, 2)

    def _planned_hours(n):
        return _hours(planned_per_machine_seconds * float(n or 0))

    def _uptime_hours(planned_h, down_h):
        return round(max(0.0, float(planned_h or 0.0) - float(down_h or 0.0)), 2)

    planned_per_machine_h = _hours(planned_per_machine_seconds)

    def _machine_rows(ids_set):
        out = []
        for mid in sorted(list(ids_set)):
            meta = maquinas_by_id.get(int(mid)) or {}
            down_h = _hours(down_by_machine.get(int(mid), 0.0))
            if down_h > planned_per_machine_h:
                down_h = planned_per_machine_h
            out.append(
                {
                    "id": int(mid),
                    "nombre": (meta.get("nombre") or f"Máquina #{mid}"),
                    "planned_h": planned_per_machine_h,
                    "down_h": down_h,
                    "up_h": _uptime_hours(planned_per_machine_h, down_h),
                }
            )
        out.sort(key=lambda x: x["nombre"])
        return out

    weekly_payload["paros_planned_horas_total"] = _planned_hours(len(ids_total))
    weekly_payload["paros_downtime_horas_total"] = _hours(down_total)
    weekly_payload["paros_uptime_horas_total"] = _uptime_hours(
        weekly_payload["paros_planned_horas_total"], weekly_payload["paros_downtime_horas_total"]
    )
    weekly_payload["paros_planned_horas_corte"] = _planned_hours(len(ids_corte))
    weekly_payload["paros_downtime_horas_corte"] = _hours(down_corte)
    weekly_payload["paros_uptime_horas_corte"] = _uptime_hours(
        weekly_payload["paros_planned_horas_corte"], weekly_payload["paros_downtime_horas_corte"]
    )
    weekly_payload["paros_planned_horas_soldadura"] = _planned_hours(len(ids_soldadura))
    weekly_payload["paros_downtime_horas_soldadura"] = _hours(down_soldadura)
    weekly_payload["paros_uptime_horas_soldadura"] = _uptime_hours(
        weekly_payload["paros_planned_horas_soldadura"], weekly_payload["paros_downtime_horas_soldadura"]
    )
    weekly_payload["paros_planned_horas_robot"] = _planned_hours(len(ids_robot))
    weekly_payload["paros_downtime_horas_robot"] = _hours(down_robot)
    weekly_payload["paros_uptime_horas_robot"] = _uptime_hours(
        weekly_payload["paros_planned_horas_robot"], weekly_payload["paros_downtime_horas_robot"]
    )
    weekly_payload["paros_maquinas_corte"] = _machine_rows(ids_corte)
    weekly_payload["paros_maquinas_soldadura"] = _machine_rows(ids_soldadura)
    weekly_payload["paros_maquinas_robot"] = _machine_rows(ids_robot)

    avg_time_rows = []
    try:
        if not weekly_term_ids:
            raise Exception("No hay piezas terminadas en la semana seleccionada.")

        sql = """
            WITH ordered AS (
                SELECT
                    viga_internal_id,
                    estado_nuevo,
                    "timestamp" AS ts,
                    LEAD("timestamp") OVER (PARTITION BY viga_internal_id ORDER BY "timestamp") AS next_ts
                FROM production_log
                WHERE viga_internal_id = ANY(%(ids)s::bigint[])
                  AND "timestamp" < %(we)s
            ),
            intervals AS (
                SELECT
                    viga_internal_id,
                    estado_nuevo,
                    GREATEST(ts, %(ws)s) AS start_ts,
                    LEAST(COALESCE(next_ts, %(we)s), %(we)s) AS end_ts
                FROM ordered
                WHERE estado_nuevo <> 'Enviado'
                  AND LEAST(COALESCE(next_ts, %(we)s), %(we)s) > GREATEST(ts, %(ws)s)
            )
            SELECT viga_internal_id, estado_nuevo, start_ts, end_ts
            FROM intervals
        """
        params = {"ws": week_start_dt, "we": week_end_dt, "ids": weekly_term_ids}
        with connections["mes"].cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
        by_piece_state_seconds = {}
        for viga_internal_id, estado_nuevo, start_ts, end_ts in rows:
            sec = _labor_seconds_between(start_ts, end_ts)
            if sec <= 0:
                continue
            st = _norm_estado(str(estado_nuevo or ""))
            vid = int(viga_internal_id or 0)
            if vid:
                per_state = by_piece_state_seconds.get(vid)
                if not per_state:
                    per_state = {}
                    by_piece_state_seconds[vid] = per_state
                per_state[st] = per_state.get(st, 0) + sec

        total_seconds = 0
        for per_state in by_piece_state_seconds.values():
            total_seconds += sum(per_state.values())
        for st in ESTADOS:
            if st == "Enviado":
                continue
            secs = []
            for per_state in by_piece_state_seconds.values():
                v = per_state.get(st)
                if v:
                    secs.append(v)
            if not secs:
                continue
            avg_hours = (sum(secs) / len(secs)) / 3600.0
            med_hours = statistics.median(secs) / 3600.0 if secs else 0.0
            pct_total = (sum(secs) / total_seconds * 100.0) if total_seconds else 0.0
            avg_time_rows.append(
                {
                    "estado": st,
                    "avg_hours": round(avg_hours, 2),
                    "median_hours": round(med_hours, 2),
                    "pct_total": round(pct_total, 1),
                    "n": int(len(secs)),
                }
            )

        piece_secs = [sum(per_state.values()) for per_state in by_piece_state_seconds.values() if per_state]
        if piece_secs:
            avg_hours = (sum(piece_secs) / len(piece_secs)) / 3600.0
            med_hours = statistics.median(piece_secs) / 3600.0
            avg_time_rows.append(
                {
                    "estado": "Proceso completo",
                    "avg_hours": round(avg_hours, 2),
                    "median_hours": round(med_hours, 2),
                    "pct_total": 100.0,
                    "n": int(len(piece_secs)),
                }
            )
    except Exception:
        avg_time_rows = []

    avg_stage = {"Corte": 0.0, "Armado": 0.0, "Soldadura": 0.0, "Pintura": 0.0}
    for row in avg_time_rows:
        st = (row.get("estado") or "").strip()
        if st in avg_stage:
            avg_stage[st] = float(row.get("avg_hours") or 0.0)
    weekly_payload["avg_stage_hours"] = {
        "Corte": round(avg_stage["Corte"], 2),
        "Armado": round(avg_stage["Armado"], 2),
        "Soldadura": round(avg_stage["Soldadura"], 2),
        "Pintura": round(avg_stage["Pintura"], 2),
    }

    retrabajo_por_proyecto = []
    try:
        if weekly_term_ids:
            term_by_proj = {
                (r["proyecto"] or "").strip(): int(r["piezas"] or 0)
                for r in Viga.objects.filter(internal_id__in=weekly_term_ids)
                .values("proyecto")
                .annotate(piezas=Count("internal_id"))
            }
            rqs = (
                ProductionLog.objects.filter(
                    viga_internal_id__in=weekly_term_ids,
                    fecha_operacion__gte=week_start,
                    fecha_operacion__lt=week_end,
                )
                .filter(metricas.filtro_de_retrabajo())
                .values("viga_internal__proyecto")
                .annotate(piezas=Count("viga_internal_id", distinct=True))
            )
            for row in rqs:
                proj = (row.get("viga_internal__proyecto") or "").strip()
                retrab_piezas = int(row.get("piezas") or 0)
                term_piezas = int(term_by_proj.get(proj) or 0)
                pct = (retrab_piezas / term_piezas * 100.0) if term_piezas else 0.0
                retrabajo_por_proyecto.append(
                    {
                        "proyecto": proj or "SIN PROYECTO",
                        "retrabajo_piezas": retrab_piezas,
                        "terminadas_piezas": term_piezas,
                        "pct": round(pct, 1),
                    }
                )
            retrabajo_por_proyecto.sort(key=lambda x: (x["pct"], x["retrabajo_piezas"]), reverse=True)
            retrabajo_por_proyecto = retrabajo_por_proyecto[:10]
    except Exception:
        retrabajo_por_proyecto = []

    weekly_equipo_participacion = []

    # Armado entraba a trabajar y no salía en ninguna tabla.
    #
    # Las tres consultas de «Detalle por persona» miraban Corte, Soldadura y
    # Pintura, y Armado es una etapa como las otras: gente asignada, horas
    # hechas y kilos movidos. Con los datos de hoy son 0.3 toneladas de
    # trabajo que la pantalla atribuía a nadie, y nadie que mirara la tabla
    # tenía forma de notar que faltaba una columna.
    etapas_del_detalle = ["Corte", "Armado", "Soldadura", "Pintura"]

    agg = {}

    asigns_vig = list(
        VigaAsignacion.objects.filter(vigente=True, etapa__in=etapas_del_detalle, colaborador__isnull=False)
        .values_list("viga_internal_id", "etapa", "colaborador_id")
    )
    asig_viga_ids = list(dict.fromkeys([int(a[0]) for a in asigns_vig if a and a[0]]))
    wip_peso_map = dict(
        Viga.objects.filter(internal_id__in=asig_viga_ids)
        .exclude(estado__in=["Enviado", "Terminado"])
        .values_list("internal_id", "peso_kg")
    )
    # El peso se reparte entre quienes hicieron la etapa, no se le da entero a
    # cada uno.
    #
    # Antes, tres personas que armaron una viga de cien kilos salían con cien
    # kilos cada una: el tablero decía que el taller había producido
    # trescientos. Robótica y Herrería, dos bloques más abajo, siempre lo
    # repartieron; Estructuras no, así que las tres tablas de «Detalle por
    # persona» de la misma pantalla no eran comparables entre sí.
    #
    # A partes iguales, que es lo único defendible sin medir el tiempo de cada
    # quien. Cuando la línea corra sobre el motor unificado, la fracción será
    # un campo de la asignación y se podrá poner a mano cuando no sea mitad y
    # mitad.
    por_pieza_etapa = {}
    for vid, etapa, cid in asigns_vig:
        por_pieza_etapa.setdefault((int(vid), str(etapa)), set()).add(int(cid))

    for (vid, etapa), cids in por_pieza_etapa.items():
        kg = wip_peso_map.get(vid)
        if kg is None or not cids:
            continue
        parte = float(kg or 0.0) / float(len(cids))
        for cid in cids:
            key = (etapa, cid)
            row = agg.get(key)
            if not row:
                row = {"items": set(), "kg": 0.0}
                agg[key] = row
            token = ("P", vid)
            if token in row["items"]:
                continue
            row["items"].add(token)
            row["kg"] += parte

    r_asigs = list(
        RobotOrdenAsignacion.objects.filter(etapa__in=etapas_del_detalle).values_list("orden_id", "etapa", "colaborador_id")
    )
    r_ids = list(dict.fromkeys([int(a[0]) for a in r_asigs if a and a[0]]))
    if r_ids:
        r_kg_rows = (
            RobotOrdenItem.objects.filter(orden_id__in=r_ids)
            .values("orden_id")
            .annotate(
                kg=Sum(
                    ExpressionWrapper(
                        F("cantidad_requerida") * Coalesce(F("pieza__peso_kg"), F("pieza_custom_peso_kg"), 0.0),
                        output_field=FloatField(),
                    )
                )
            )
        )
        r_kg_map = {int(r["orden_id"]): float(r["kg"] or 0.0) for r in r_kg_rows}
        active_r_ids = set(RobotOrdenProduccion.objects.filter(id__in=r_ids, estado="Abierta").values_list("id", flat=True))
        by_order_stage = {}
        for oid, etapa, cid in r_asigs:
            oid = int(oid)
            cid = int(cid)
            if oid not in active_r_ids:
                continue
            k = (oid, str(etapa))
            by_order_stage.setdefault(k, set()).add(cid)
        for (oid, etapa), cids in by_order_stage.items():
            kg = r_kg_map.get(int(oid))
            if kg is None:
                continue
            cids = list(cids)
            if not cids:
                continue
            share = float(kg or 0.0) / float(len(cids))
            for cid in cids:
                key = (str(etapa), int(cid))
                row = agg.get(key)
                if not row:
                    row = {"items": set(), "kg": 0.0}
                    agg[key] = row
                token = ("R", int(oid))
                if token in row["items"]:
                    continue
                row["items"].add(token)
                row["kg"] += share

    h_asigs_old = list(
        HerrOrdenAsignacion.objects.filter(etapa__in=etapas_del_detalle).values_list("orden_id", "etapa", "colaborador_id")
    )
    h_asigs_new = list(
        HerrAsignacion.objects.filter(
            vigente=True,
            etapa__in=etapas_del_detalle,
            colaborador_id__isnull=False,
        ).values_list("orden_id", "etapa", "colaborador_id")
    )
    h_asigs = [*h_asigs_old, *h_asigs_new]
    h_ids = list(dict.fromkeys([int(a[0]) for a in h_asigs if a and a[0]]))
    if h_ids:
        h_kg_rows = (
            HerrOrdenItem.objects.filter(orden_id__in=h_ids)
            .values("orden_id")
            .annotate(
                kg=Sum(
                    ExpressionWrapper(
                        F("cantidad_requerida") * Coalesce(F("pieza__peso_kg"), F("pieza_custom_peso_kg"), 0.0),
                        output_field=FloatField(),
                    )
                )
            )
        )
        h_kg_map = {int(r["orden_id"]): float(r["kg"] or 0.0) for r in h_kg_rows}
        active_h_ids = set(
            HerrOrdenProduccion.objects.filter(id__in=h_ids, estado="Abierta").exclude(estado_etapa="Terminado").values_list("id", flat=True)
        )
        by_order_stage = {}
        for oid, etapa, cid in h_asigs:
            oid = int(oid)
            cid = int(cid)
            if oid not in active_h_ids:
                continue
            k = (oid, str(etapa))
            by_order_stage.setdefault(k, set()).add(cid)
        for (oid, etapa), cids in by_order_stage.items():
            kg = h_kg_map.get(int(oid))
            if kg is None:
                continue
            cids = list(cids)
            if not cids:
                continue
            share = float(kg or 0.0) / float(len(cids))
            for cid in cids:
                key = (str(etapa), int(cid))
                row = agg.get(key)
                if not row:
                    row = {"items": set(), "kg": 0.0}
                    agg[key] = row
                token = ("H", int(oid))
                if token in row["items"]:
                    continue
                row["items"].add(token)
                row["kg"] += share

    colab_ids = list(dict.fromkeys([cid for (_e, cid) in agg.keys()]))
    colabs = Colaborador.objects.filter(id__in=colab_ids).select_related("equipo")
    colab_map = {c.id: c for c in colabs}

    quien_hace_que = {etapa: [] for etapa in etapas_del_detalle}
    for (etapa, cid), row in agg.items():
        c = colab_map.get(cid)
        if not c:
            continue
        piezas = len(row["items"])
        ton = float(row["kg"] or 0.0) / 1000.0
        quien_hace_que.setdefault(etapa, []).append(
            {
                "colaborador_id": int(cid),
                "nombre": c.nombre,
                "equipo": getattr(c.equipo, "nombre", ""),
                "piezas": piezas,
                "ton": round(ton, 3),
            }
        )
    for etapa in quien_hace_que.keys():
        quien_hace_que[etapa].sort(key=lambda r: (-float(r["ton"]), -int(r["piezas"]), r["nombre"]))
        quien_hace_que[etapa] = quien_hace_que[etapa][:50]

    robot_week_qs = RobotProduccion.objects.filter(fecha__gte=week_start, fecha__lt=week_end).select_related("pieza", "robot", "operador")
    robot_expr_kg = ExpressionWrapper(F("cantidad") * F("pieza__peso_kg"), output_field=FloatField())
    robot_total_kg = float(robot_week_qs.aggregate(total=Sum(robot_expr_kg))["total"] or 0.0)
    robot_total_ton = robot_total_kg / 1000.0
    robot_total_piezas = int(robot_week_qs.aggregate(total=Sum("cantidad"))["total"] or 0)
    robot_top = list(
        robot_week_qs.values("robot__nombre")
        .annotate(kg=Sum(robot_expr_kg))
        .order_by("-kg")[:5]
    )
    robot_weekly = {
        "ton": round(robot_total_ton, 3),
        "piezas": robot_total_piezas,
        "top_robots": [{"robot": r["robot__nombre"], "ton": round(float(r["kg"] or 0.0) / 1000.0, 3)} for r in robot_top],
    }

    term_hist = []
    term_logs = (
        ProductionLog.objects.filter(estado_nuevo="Terminado")
        .select_related("viga_internal")
        .order_by("-fecha_operacion", "-timestamp")
    )
    seen_ids = set()
    for lg in term_logs[:500]:
        vid = int(lg.viga_internal_id)
        if vid in seen_ids:
            continue
        seen_ids.add(vid)
        v = lg.viga_internal
        term_hist.append(
            {
                "id": vid,
                "codigo": getattr(v, "codigo_viga", ""),
                "pieza": f"{getattr(v, 'pieza_no', '')}/{getattr(v, 'total_piezas', '')}",
                "proyecto": getattr(v, "proyecto", ""),
                "peso_kg": float(getattr(v, "peso_kg", 0.0) or 0.0),
                "peso_ton": float(getattr(v, "peso_kg", 0.0) or 0.0) / 1000.0,
                "fecha_terminado": lg.fecha_operacion,
                "corte": "",
                "soldadura": "",
                "pintura": "",
            }
        )
        if len(term_hist) >= 50:
            break
    term_ids = [x["id"] for x in term_hist]
    if term_ids:
        asigns = (
            VigaAsignacion.objects.filter(viga_internal_id__in=term_ids, vigente=True, etapa__in=["Corte", "Soldadura", "Pintura"])
            .select_related("colaborador", "maquina")
            .order_by("viga_internal_id", "etapa", "rol", "-asignado_en")
        )
        a_map = {}
        for a in asigns:
            if a.colaborador_id:
                name = a.colaborador.nombre
            elif a.maquina_id:
                name = f"[{a.maquina.nombre}]"
            else:
                name = ""
            if not name:
                continue
            a_map.setdefault(int(a.viga_internal_id), {}).setdefault(a.etapa, []).append(name)
        for row in term_hist:
            vid = int(row["id"])
            row["corte"] = ", ".join(a_map.get(vid, {}).get("Corte") or [])
            row["soldadura"] = ", ".join(a_map.get(vid, {}).get("Soldadura") or [])
            row["pintura"] = ", ".join(a_map.get(vid, {}).get("Pintura") or [])

    prev_week_start = week_start - timedelta(days=7)
    prev_week_end = week_start
    prev_week_ids = list(
        ProductionLog.objects.filter(
            estado_nuevo="Terminado",
            fecha_operacion__gte=prev_week_start,
            fecha_operacion__lt=prev_week_end,
        )
        .values_list("viga_internal_id", flat=True)
        .distinct()
    )
    prev_week_kg = float(Viga.objects.filter(internal_id__in=prev_week_ids).aggregate(total=Sum("peso_kg"))["total"] or 0.0)
    prev_week_ton = prev_week_kg / 1000.0
    prev_robot_kg = float(
        RobotProduccion.objects.filter(fecha__gte=prev_week_start, fecha__lt=prev_week_end).aggregate(total=Sum(robot_expr_kg_week))["total"]
        or 0.0
    )
    prev_total_ton = (prev_week_ton + (prev_robot_kg / 1000.0))
    curr_week_ton = float(weekly_payload.get("total_ton") or 0.0)
    rendimiento_delta_pct = ((curr_week_ton - prev_total_ton) / prev_total_ton * 100.0) if prev_total_ton else 0.0

    max_oldest_dias = 0
    try:
        max_oldest_dias = max(int(v.get("dias") or 0) for v in (oldest_wip or []))
    except Exception:
        max_oldest_dias = 0

    alerts = []
    if float(vencidas_ton or 0.0) > 0.0:
        alerts.append(
            {
                "title": "Retrasos (WIP vencido)",
                "value": f"{float(vencidas_ton or 0.0):.2f} ton",
                "detail": f"{int(vencidas_piezas or 0)} piezas",
                "level": "warning" if float(vencidas_ton or 0.0) < 5.0 else "danger",
            }
        )
    else:
        alerts.append({"title": "Retrasos (WIP vencido)", "value": "0.00 ton", "detail": "Sin atraso", "level": "success"})

    if max_oldest_dias >= 10:
        alerts.append(
            {
                "title": "Desviación de tiempo",
                "value": f"{max_oldest_dias} días",
                "detail": "Pieza más antigua en proceso",
                "level": "danger" if max_oldest_dias >= 20 else "warning",
            }
        )
    else:
        alerts.append({"title": "Desviación de tiempo", "value": f"{max_oldest_dias} días", "detail": "Rango sano", "level": "success"})

    if prev_week_ton > 0:
        level = "success"
        if rendimiento_delta_pct <= -20:
            level = "danger"
        elif rendimiento_delta_pct <= -10:
            level = "warning"
        alerts.append(
            {
                "title": "Rendimiento vs semana anterior",
                "value": f"{rendimiento_delta_pct:+.1f}%",
                "detail": f"{curr_week_ton:.3f} ton vs {prev_week_ton:.3f} ton",
                "level": level,
            }
        )
    else:
        alerts.append(
            {
                "title": "Rendimiento vs semana anterior",
                "value": "—",
                "detail": "Sin base previa",
                "level": "secondary",
            }
        )

    weekly_retrabajo_pct = float(weekly_payload.get("retrabajo_pct") or 0.0)
    weekly_retrabajo_piezas = int(weekly_payload.get("retrabajo_piezas") or 0)
    if weekly_retrabajo_pct >= 5.0:
        alerts.append(
            {
                "title": "Retrabajo (semana)",
                "value": f"{weekly_retrabajo_pct:.1f}%",
                "detail": f"{weekly_retrabajo_piezas} piezas",
                "level": "danger" if weekly_retrabajo_pct >= 10.0 else "warning",
            }
        )
    else:
        alerts.append(
            {
                "title": "Retrabajo (semana)",
                "value": f"{weekly_retrabajo_pct:.1f}%",
                "detail": f"{weekly_retrabajo_piezas} piezas",
                "level": "success",
            }
        )

    export_mode = bool(getattr(request, "_dashboard_export", False))
    quien_detalle_cache = {}
    if export_mode:
        for etapa, rows in (quien_hace_que or {}).items():
            for row in (rows or []):
                cid = int(row.get("colaborador_id") or 0) if isinstance(row, dict) else 0
                if not cid:
                    continue
                key = f"{cid}|{str(etapa)}"
                if key in quien_detalle_cache:
                    continue
                quien_detalle_cache[key] = _dashboard_quien_detalle_payload(cid, str(etapa))

    ctx = {
            "resumen": resumen,
            "total_piezas": total_piezas,
            "total_kg": total_kg,
            "total_ton": total_ton,
            "terminado_piezas": terminado_piezas,
            "terminado_ton": terminado_ton,
            "pct_terminado": pct_terminado,
            "terminado_progress_payload": {
                "terminado_piezas": terminado_piezas,
                "total_piezas": total_piezas,
                "pct": round(float(pct_terminado or 0.0), 2),
            },
            "chart_payload": {
                "labels": labels,
                "piezas": piezas_data,
                "ton": ton_data,
                "pct_piezas": [round(r["pct"], 2) for r in resumen],
                "pct_ton": [round(r["pct_ton"], 2) for r in resumen],
                "colors": colors,
            },
            "retrabajo_piezas": retrabajo_piezas,
            "retrabajo_ton": retrabajo_ton,
            "retrabajo_pct": retrabajo_pct,
            "total_equipos": total_equipos,
            "total_integrantes": total_integrantes,
            "prod_ton_por_equipo_global": prod_ton_por_equipo_global,
            "prod_ton_por_integrante_global": prod_ton_por_integrante_global,
            "ton_por_persona_meta": ton_por_persona_meta,
            "equipos_detalle": equipos_detalle,
            "wip_piezas": wip_piezas,
            "wip_ton": wip_ton,
            "vencidas_piezas": vencidas_piezas,
            "vencidas_ton": vencidas_ton,
            "oldest_wip": oldest_wip,
            "top_backlog": top_backlog,
            "top_wip": top_wip,
            "enviados_total": enviados_total,
            "enviados_total_ton": enviados_total_ton,
            "sla": {"on_time": on_time, "late": late, "no_due": no_due, "avg_late_days": avg_late_days},
            "weekly": weekly_payload,
            "weekly_selected_start": week_start.isoformat(),
            "weekly_selected_end": (week_start + timedelta(days=7)).isoformat(),
            "weekly_payload": {
                "labels": [d["date"] for d in (weekly_payload.get("daily") or [])],
                "ton": [d["ton"] for d in (weekly_payload.get("daily") or [])],
                "piezas": [d["piezas"] for d in (weekly_payload.get("daily") or [])],
                "paros_planned_horas_total": weekly_payload.get("paros_planned_horas_total", 0),
                "paros_downtime_horas_total": weekly_payload.get("paros_downtime_horas_total", 0),
                "paros_uptime_horas_total": weekly_payload.get("paros_uptime_horas_total", 0),
                "paros_planned_horas_corte": weekly_payload.get("paros_planned_horas_corte", 0),
                "paros_downtime_horas_corte": weekly_payload.get("paros_downtime_horas_corte", 0),
                "paros_uptime_horas_corte": weekly_payload.get("paros_uptime_horas_corte", 0),
                "paros_planned_horas_soldadura": weekly_payload.get("paros_planned_horas_soldadura", 0),
                "paros_downtime_horas_soldadura": weekly_payload.get("paros_downtime_horas_soldadura", 0),
                "paros_uptime_horas_soldadura": weekly_payload.get("paros_uptime_horas_soldadura", 0),
                "paros_planned_horas_robot": weekly_payload.get("paros_planned_horas_robot", 0),
                "paros_downtime_horas_robot": weekly_payload.get("paros_downtime_horas_robot", 0),
                "paros_uptime_horas_robot": weekly_payload.get("paros_uptime_horas_robot", 0),
                "paros_maquinas_corte": weekly_payload.get("paros_maquinas_corte", []),
                "paros_maquinas_soldadura": weekly_payload.get("paros_maquinas_soldadura", []),
                "paros_maquinas_robot": weekly_payload.get("paros_maquinas_robot", []),
                "avg_stage_hours": weekly_payload.get("avg_stage_hours", {}),
            },
            "alerts": alerts,
            "retrabajo_any_list": retrabajo_any_list,
            "avg_time_rows": avg_time_rows,
            "vencidas_por_proyecto_estado": vencidas_por_proyecto_estado,
            "retrabajo_por_proyecto": retrabajo_por_proyecto,
            "weekly_equipo_participacion": weekly_equipo_participacion,
            "quien_hace_que": quien_hace_que,
            "robot_weekly": robot_weekly,
            "kpi_maquinas_corte": kpi_maquinas_corte,
            "terminado_historial": term_hist,
            "productividad_payload": {
                "labels": equipos_labels,
                "integrantes": equipos_integrantes,
                "ton_terminadas": round(float(terminado_ton or 0.0), 3),
                "ton_por_integrante_global": round(float(prod_ton_por_integrante_global or 0.0), 3),
                "ton_por_persona_meta": round(float(ton_por_persona_meta), 3),
                "colors": equipos_colors,
            },
            "flow_payload": {
                "labels": throughput_labels,
                "ton_terminadas": throughput_ton,
                "piezas_terminadas": throughput_piezas,
            },
            "sla_payload": {
                "on_time": on_time,
                "late": late,
                "no_due": no_due,
                "avg_late_days": avg_late_days,
            },
            "aging_payload": {
                "labels": aging_labels,
                "days": aging_days,
                "colors": aging_colors,
            },
            "retrabajo_payload": {
                "donut": {
                    "labels": ["Retrabajos", "No retrabajos"],
                    "data": [retrabajo_piezas, max(total_piezas - retrabajo_piezas, 0)],
                    "colors": ["#f5365c", "#2dce89"],
                },
            },
            "export_mode": export_mode,
            "quien_detalle_cache": quien_detalle_cache,
    }

    if bool(getattr(request, "_dashboard_context_only", False)):
        return ctx

    return render(request, "produccion/dashboard.html", ctx)


@login_required
def dashboard_quien_detalle(request, colab_id: int, etapa: str):
    payload = _dashboard_quien_detalle_payload(int(colab_id), str(etapa))
    status = 200 if payload.get("ok") else 400
    return JsonResponse(payload, status=status)


@login_required
def dashboard_export_html(request):
    request._dashboard_export = True
    resp = dashboard(request)
    html = resp.content.decode("utf-8", errors="ignore")
    html = _inline_html_assets(html)
    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    out = HttpResponse(html, content_type="text/html; charset=utf-8")
    out["Content-Disposition"] = f'attachment; filename="dashboard_{ts}.html"'
    return out


@login_required
def dashboard_export_xlsx(request):
    request._dashboard_context_only = True
    ctx = dashboard(request)
    if not isinstance(ctx, dict):
        return redirect("produccion:dashboard")

    weekly = ctx.get("weekly") or {}
    start = str(ctx.get("weekly_selected_start") or weekly.get("week_start") or "")
    end = str(ctx.get("weekly_selected_end") or weekly.get("week_end") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    center = Alignment(vertical="center")
    header = Font(bold=True)

    ws.append(["Dashboard", ""])
    ws["A1"].font = bold
    ws.append(["Periodo", f"{start} → {end}".strip()])
    ws.append([])

    ws.append(["KPI", "Valor", "Unidad"])
    for cell in ws[ws.max_row]:
        cell.font = header
        cell.alignment = center

    def k(label, val, unit):
        ws.append([label, val, unit])

    k("Ton total (semana)", float(weekly.get("total_ton") or weekly.get("terminadas_ton") or 0.0), "ton")
    k("Piezas terminadas (semana)", int(weekly.get("terminadas_piezas") or 0), "pzs")
    k("Ton/persona (semana)", float(weekly.get("ton_por_persona") or 0.0), "ton")
    k("Robot (semana) ton", float(weekly.get("robot_ton") or 0.0), "ton")
    k("Robot (semana) pzs", int(weekly.get("robot_piezas") or 0), "pzs")
    k("Herrería (semana) ton", float(weekly.get("herr_ton") or 0.0), "ton")
    k("Herrería (semana) pzs", int(weekly.get("herr_piezas") or 0), "pzs")
    k("Corta.mx (semana) kg", float(weekly.get("corte_kg") or 0.0), "kg")
    k("Corta.mx (semana) ton", float(weekly.get("corte_ton") or 0.0), "ton")
    k("Corta.mx (semana) pzs", int(weekly.get("corte_piezas") or 0), "pzs")
    k("Logística OP (semana) kg", float(weekly.get("logistica_op_kg") or 0.0), "kg")
    k("Logística OP (semana) ton", float(weekly.get("logistica_op_ton") or 0.0), "ton")
    k("Logística OP (semana) pzs", int(weekly.get("logistica_op_piezas") or 0), "pzs")
    k("Logística Corta (semana) kg", float(weekly.get("logistica_corta_kg") or 0.0), "kg")
    k("Logística Corta (semana) ton", float(weekly.get("logistica_corta_ton") or 0.0), "ton")
    k("Logística Corta (semana) pzs", int(weekly.get("logistica_corta_piezas") or 0), "pzs")
    k("Retrabajo (semana)", float(weekly.get("retrabajo_pct") or 0.0), "%")
    k("Retrabajos (semana)", int(weekly.get("retrabajo_any_piezas") or 0), "pzs")
    k("Terminadas ya enviadas (ton)", float(weekly.get("terminadas_enviadas_ton") or 0.0), "ton")

    avg_stage = (weekly.get("avg_stage_hours") or {}) if isinstance(weekly, dict) else {}
    k("Promedio en corte", float(avg_stage.get("Corte") or 0.0), "hr")
    k("Promedio en armado", float(avg_stage.get("Armado") or 0.0), "hr")
    k("Promedio en soldadura", float(avg_stage.get("Soldadura") or 0.0), "hr")
    k("Promedio en pintura", float(avg_stage.get("Pintura") or 0.0), "hr")

    ws.append([])
    ws.append(["Tiempo promedio por estado (semana)", "", ""])
    ws["A" + str(ws.max_row)].font = bold
    ws.append(["Estado", "Promedio (h)", "Mediana (h)"])
    for cell in ws[ws.max_row]:
        cell.font = header
        cell.alignment = center
    for r in (ctx.get("avg_time_rows") or []):
        ws.append(
            [
                str(r.get("estado") or ""),
                float(r.get("avg_hours") or 0.0),
                float(r.get("median_hours") or 0.0),
            ]
        )

    ws.append([])
    ws.append(["Detalle por estado", "", ""])
    ws["A" + str(ws.max_row)].font = bold
    ws.append(["Estado", "Piezas", "%", "Ton", "% Ton"])
    for cell in ws[ws.max_row]:
        cell.font = header
        cell.alignment = center
    for r in (ctx.get("resumen") or []):
        ws.append(
            [
                str(r.get("estado") or ""),
                int(r.get("piezas") or 0),
                float(r.get("pct") or 0.0),
                float(r.get("ton") or 0.0),
                float(r.get("pct_ton") or 0.0),
            ]
        )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                continue
            if cell.column == 2 and cell.row > 3:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
            cell.alignment = Alignment(vertical="center")

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("Diario")
    ws2.append(
        [
            "Fecha",
            "Manual pzs",
            "Manual ton",
            "Robot pzs",
            "Robot ton",
            "Herr pzs",
            "Herr ton",
            "Corta pzs",
            "Corta ton",
            "Total pzs",
            "Total ton",
        ]
    )
    for cell in ws2[1]:
        cell.font = header
        cell.alignment = center

    for d in (weekly.get("daily") or []):
        ws2.append(
            [
                str(d.get("date") or ""),
                int(d.get("manual_piezas") or 0),
                float(d.get("manual_ton") or 0.0),
                int(d.get("robot_piezas") or 0),
                float(d.get("robot_ton") or 0.0),
                int(d.get("herr_piezas") or 0),
                float(d.get("herr_ton") or 0.0),
                int(d.get("corte_piezas") or 0),
                float(d.get("corte_ton") or 0.0),
                int(d.get("piezas") or 0),
                float(d.get("ton") or 0.0),
            ]
        )

    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for c in r:
            if c.column in {3, 5, 7, 9, 11}:
                c.number_format = "0.000"
            c.alignment = Alignment(vertical="center")

    for col in range(1, ws2.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws2[letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws2.column_dimensions[letter].width = min(max_len + 2, 35)
    ws2.freeze_panes = "A2"

    out = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
    filename = f"dashboard_{ts}.xlsx"
    out["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(out)
    return out


@login_required
def export_vigas_excel(request):
    if not _is_admin_user(request.user):
        return redirect("produccion:home")
    qs, _filters = _viga_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vigas"

    headers = [
        "ID",
        "Código",
        "Pieza",
        "Total",
        "Proyecto",
        "Descripción",
        "Fecha compromiso",
        "Estado",
        "Prioridad",
        "Peso (kg)",
        "Observaciones",
        "Creación",
        "Último cambio",
    ]
    ws.append(headers)

    for v in qs.iterator(chunk_size=500):
        ws.append(
            [
                v.internal_id,
                v.codigo_viga,
                v.pieza_no,
                v.total_piezas,
                v.proyecto,
                v.descripcion,
                v.fecha_compromiso.strftime("%Y-%m-%d") if v.fecha_compromiso else "",
                v.estado,
                v.prioridad,
                float(v.peso_kg) if v.peso_kg is not None else 0,
                v.observaciones,
                v.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S") if v.fecha_creacion else "",
                v.ultimo_cambio.strftime("%Y-%m-%d %H:%M:%S") if v.ultimo_cambio else "",
            ]
        )

    out = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"vigas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(out)
    return out
